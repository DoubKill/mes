import hmac
import json
import os
import time
import logging
from base64 import standard_b64encode
from datetime import datetime, timedelta
from io import BytesIO

import cx_Oracle
import requests
import xlwt
from operator import itemgetter
from django.db.models import Q, F
from django.http import HttpResponse
from openpyxl import load_workbook, cell
from rest_framework.exceptions import ValidationError
import pandas as pd

from equipment.models import EquipApplyOrder, EquipMaintenanceAreaSetting, EquipInspectionOrder, EquipSpareErp
from mes import settings
from system.models import User, Section

logger = logging.getLogger('error_log')


def gen_template_response(export_fields_dict, data, file_name, sheet_name=None, handle_str=False):
    export_fields = list(export_fields_dict.values())
    sheet_heads = list(export_fields_dict.keys())
    wb = load_workbook('xlsx_template/example.xlsx')
    ws = wb.worksheets[0]
    sheet = wb.copy_worksheet(ws)
    if sheet_name:
        sheet.title = sheet_name
    for idx, sheet_head in enumerate(sheet_heads):
        sheet.cell(1, idx + 1).value = sheet_head

    data_row = 2
    for i in data:
        for col_num, data_key in enumerate(export_fields):
            set_value = i[data_key]
            if handle_str and isinstance(set_value, str):
                set_value = cell.cell.ILLEGAL_CHARACTERS_RE.sub(r'', set_value)
            sheet.cell(data_row, col_num + 1).value = set_value
        data_row += 1

    wb.remove_sheet(ws)
    output = BytesIO()
    wb.save(output)
    # 重新定位到开始
    output.seek(0)
    response = HttpResponse(content_type='application/vnd.ms-excel')
    filename = file_name
    response['Content-Disposition'] = u'attachment;filename= ' + filename.encode('gbk').decode(
        'ISO-8859-1') + '.xls'
    response.write(output.getvalue())
    return response


def gen_excels_response(export_fields_dict, data_list, file_name, sheet_keyword, handle_str=False):
    """同样内容的sheet按日期导出一个execl"""
    export_fields = list(export_fields_dict.values())
    sheet_heads = list(export_fields_dict.keys())
    wb = xlwt.Workbook(encoding='utf8')
    output = BytesIO()
    sheet_name_used = []
    for i, s_data in enumerate(data_list):
        words = itemgetter(*sheet_keyword)(s_data)
        if isinstance(words, tuple):
            sheet_name = '-'.join(list(words))[5:]  # 第一部分是时间取年月
        else:
            if words in sheet_name_used:
                sheet_name = f"{words}({i + 1})"
            else:
                sheet_name = f"{words}"
                sheet_name_used.append(words)
        sheet = wb.add_sheet(sheet_name, cell_overwrite_ok=True)
        for idx, sheet_head in enumerate(sheet_heads):
            sheet.write(0, idx, sheet_head)

        data_row = 1
        for data in s_data.get('table_details'):
            for col_num, data_key in enumerate(export_fields):
                set_value = "否" if data[data_key] is False else ("是" if data[data_key] is True else data[data_key])
                if handle_str and isinstance(set_value, str):
                    set_value = cell.cell.ILLEGAL_CHARACTERS_RE.sub(r'', set_value)
                sheet.write(data_row, col_num, set_value)
            data_row += 1
    wb.save(output)
    # 重新定位到开始
    output.seek(0)
    response = HttpResponse(content_type='application/vnd.ms-excel')
    filename = file_name
    response['Content-Disposition'] = u'attachment;filename= ' + filename.encode('gbk').decode(
        'ISO-8859-1') + '.xls'
    response.write(output.getvalue())
    return response


class DinDinAPI(object):
    if settings.DEBUG:
        APP_KEY = 'dingpygx79tszcicv6jb'
        APP_SECRET = 'LBTg3wzgkDZ5b2l5pqRBrUy8XY8npiaPKYqBk5fCfCKMg_PEB7FJFpYMj8OfiiUu'
        AGENT_ID = '1336171749'
    else:
        APP_KEY = 'dingju0317gj7k9spsw1'
        APP_SECRET = 'rvARM-cMGq93AKSZWrhEyJez0AvU9TaN-kTiaV78aJIwmQyY3TGmmoPOE3eRybEB'
        AGENT_ID = '1144203103'

    def __init__(self):
        self.access_token = self.get_token()

    def get_token(self):
        """
            获取钉钉授权认证token
        @return:
        """
        url = 'https://oapi.dingtalk.com/gettoken'
        ret = requests.get(url, params={'appkey': self.APP_KEY, 'appsecret': self.APP_SECRET})
        data = json.loads(ret.text)
        if not data.get('errcode') == 0:
            raise ValidationError(data.get('errmsg'))
        return data.get('access_token')

    def get_user_id(self, phone_num):
        """
            根据手机号获取钉钉用户id
        @param phone_num:
        """
        url = 'https://oapi.dingtalk.com/topapi/v2/user/getbymobile'
        ret = requests.post(url, params={'access_token': self.access_token}, data={"mobile": phone_num}, timeout=5)
        data = json.loads(ret.text)
        if not data.get('errcode') == 0:
            return ''
        else:
            return data.get('result').get('userid')

    def get_user_attendance(self, user_ids, begin_time, end_time):
        """
            获取用户考勤信息
        @param user_ids: 用户id列表
        @param begin_time: 开始工作日
        @param end_time: 结束工作日
        @return: 考勤打卡列表
        """
        url = 'https://oapi.dingtalk.com/attendance/list'
        data = {
            "workDateFrom": str(begin_time) + ' 00:00:00',
            "offset": 0,
            "userIdList": user_ids,
            "limit": 10,
            "isI18n": False,
            "workDateTo": str(end_time) + ' 00:00:00'
        }
        ret = requests.post(url, params={'access_token': self.access_token}, json=data, timeout=5)
        data = json.loads(ret.text)
        if not data.get('errcode') == 0:
            return []
        else:
            return data.get('recordresult')

    def send_message(self, user_ids, content, order_id=0, inspection=False, attendance=False):
        """
            发送钉钉工作消息给用户
        @param user_ids:
        """
        url = "https://oapi.dingtalk.com/topapi/message/corpconversation/asyncsend_v2"
        message_url = "eapp://pages/workOrderList/workOrderList"
        if order_id:
            message_url = f"eapp://pages/repairOrder/repairOrder?id={order_id}" + ("&isInspection=true" if inspection else "")
        if attendance:
            message_url = "eapp://pages/index/index"
        data = {
            "msg": {
                "msgtype": "oa",
                "oa": {
                    "message_url": message_url,
                    "head": {
                        "bgcolor": "FFBBBBBB",
                        "text": "头部标题"
                    },
                    "body": content
                }
            },
            "agent_id": self.AGENT_ID,
            "userid_list": ",".join(user_ids),
            "to_all_user": False
        }
        ret = requests.post(url, params={'access_token': self.access_token}, json=data, timeout=5)
        data = json.loads(ret.text)
        if not data.get('errcode') == 0:
            logger.error(data.get('errmsg'))

    def auth(self, code):
        """微信认证，获取用户信息"""
        user_info_url = "https://oapi.dingtalk.com/topapi/v2/user/getuserinfo"
        data = {"code": code}
        user_ret = requests.post(user_info_url, params={'access_token': self.access_token}, json=data, timeout=5)
        user_data = json.loads(user_ret.text)
        if not user_data.get('errcode') == 0:
            raise ValidationError(user_data.get('errmsg'))
        return user_data.get('result')

    @staticmethod
    def get_ding_talk_signature(app_secret, utc_timestamp):
        """
        :param app_secret: 钉钉开发者文档创建的app密钥
        :param utc_timestamp: 官方文档中要签名的数据，单位是毫秒时间戳
        :return: 为所需要的签名值，此值为可逆的
        """
        digest = hmac.HMAC(key=app_secret.encode('utf8'),
                           msg=utc_timestamp.encode('utf8'),
                           digestmod=hmac._hashlib.sha256).digest()
        signature = standard_b64encode(digest).decode('utf8')
        return signature

    def get_union_id(self, code):
        """
            根据扫码信息获取的临时登录码请求接口，换取钉钉用户名
        """
        payload = {'tmp_auth_code': code}
        headers = {'Content-Type': 'application/json'}
        t = time.time()
        timestamp = str((int(round(t * 1000))))
        signature = self.get_ding_talk_signature(self.APP_SECRET, timestamp)
        query_params = {'accessKey': self.APP_KEY, 'timestamp': timestamp, 'signature': signature}
        res = requests.post('https://oapi.dingtalk.com/sns/getuserinfo_bycode',
                            params=query_params,
                            json=payload,
                            headers=headers)
        res_dict = json.loads(res.text)
        if not res_dict.get('errcode') == 0:
            raise ValidationError(res_dict.get('errmsg'))
        return res_dict.get('user_info').get('unionid')

    def get_user_id_through_union_id(self, union_id):
        """
            通过unionid获取钉钉用户id
        """
        user_url = 'https://oapi.dingtalk.com/topapi/user/getbyunionid'
        user_query_params = {'access_token': self.access_token}
        user_data = {
            "unionid": union_id
        }
        res = requests.post(user_url, params=user_query_params, json=user_data)
        res_dict = json.loads(res.text)
        if not res_dict.get('errcode') == 0:
            raise ValidationError(res_dict.get('errmsg'))
        dd_user_id = res_dict.get('result').get('userid')
        return dd_user_id


def get_children_section(init_section, include_self=True):
    """获取所有可指派的部门(默认设备科及其下)"""
    r = []
    if include_self:
        r.append(init_section.name)
    for c_section in init_section.children_sections.all():
        _r = get_children_section(c_section, include_self=True)
        if len(_r) > 0:
            r.extend(_r)
    return r


def get_staff_status(ding_api, section_name, group=''):
    """section_name: 设备部"""
    result = []
    filter_kwargs = {'section_users__repair_group': group} if group else {'section_users__repair_group__isnull': False}
    # 获取所有下级部门
    init_section = Section.objects.filter(name=section_name).last()
    if not init_section:
        return result
    section_list = get_children_section(init_section)
    # 获取部门所有员工信息
    staffs = Section.objects.filter(name__in=section_list, **filter_kwargs) \
        .annotate(username=F('section_users__username'), phone_number=F('section_users__phone_number'),
                  group=F('section_users__repair_group'), is_active=F('section_users__is_active'),
                  uid=F('section_users__id'), leader=F('in_charge_user__username'),
                  leader_phone_number=F('in_charge_user__phone_number')) \
        .values('username', 'phone_number', 'uid', 'leader', 'leader_phone_number', 'group', 'name', 'is_active')
    for staff in staffs:
        # 去除已经删除的员工
        if staff.get('is_active') == 0:
            continue
        staff_dict = {'id': staff.get('uid'), 'phone_number': staff.get('phone_number'), 'optional': False,
                      'username': staff.get('username'), 'group': staff.get('group'), 'leader': staff.get('leader'),
                      'leader_phone_number': staff.get('leader_phone_number'), 'section_name': staff.get('name')}
        # 根据手机号获取用户钉钉uid
        ding_uid = ding_api.get_user_id(staff.get('phone_number'))
        if ding_uid:
            # 查询考勤记录
            staff_dict['ding_uid'] = ding_uid
            if settings.DEBUG:
                staff_dict['optional'] = True
            else:
                records = ding_api.get_user_attendance([ding_uid], begin_time=datetime.now().date(), end_time=datetime.now().date())
                if records and len([i for i in records if i['checkType'] != 'OnDuty' and i['timeResult'] != 'NotSigned']) == 0:
                    staff_dict['optional'] = True
                staff_dict['records'] = records
        result.append(staff_dict)
    return result


def get_maintenance_status(ding_api, equip_no, maintenance_type):
    result = []
    """获取包干人员信息"""
    if maintenance_type == '通用':
        query_set = EquipMaintenanceAreaSetting.objects.filter(equip__equip_no=equip_no)
    else:
        query_set = EquipMaintenanceAreaSetting.objects.filter(equip__equip_no=equip_no, maintenance_user__workshop__icontains=maintenance_type)
    maintenances = query_set.annotate(username=F('maintenance_user__username'),
                                      phone_number=F('maintenance_user__phone_number'),
                                      group=F('maintenance_user__repair_group'),
                                      uid=F('maintenance_user__id'),
                                      leader=F('maintenance_user__section__in_charge_user__username'),
                                      leader_phone_number=F('maintenance_user__section__in_charge_user__phone_number'),
                                      is_active=F('maintenance_user__is_active'))\
        .values('username', 'phone_number', 'uid', 'leader', 'leader_phone_number', 'group', 'is_active').distinct()

    for staff in maintenances:
        # 去除已经删除的员工
        if staff.get('is_active') == 0:
            continue
        staff_dict = {'id': staff.get('uid'), 'phone_number': staff.get('phone_number'), 'optional': False,
                      'username': staff.get('username'), 'group': staff.get('group'), 'leader': staff.get('leader'),
                      'leader_phone_number': staff.get('leader_phone_number')}
        # 根据手机号获取用户钉钉uid
        ding_uid = ding_api.get_user_id(staff.get('phone_number'))
        if ding_uid:
            # 查询考勤记录
            staff_dict['ding_uid'] = ding_uid
            if settings.DEBUG:
                staff_dict['optional'] = True
            else:
                records = ding_api.get_user_attendance([ding_uid], begin_time=datetime.now().date(), end_time=datetime.now().date())
                if records and len([i for i in records if i['checkType'] != 'OnDuty' and i['timeResult'] != 'NotSigned']) == 0:
                    staff_dict['optional'] = True
                staff_dict['records'] = records
        result.append(staff_dict)
    return result


def get_ding_uids(ding_api, pks=None, names=None, check_type=None):
    """pks：维修单id列表"""
    user_ids, assign_user_list = [], []
    if pks:
        if not check_type:
            assign_user_list = EquipApplyOrder.objects.filter(id__in=pks).values_list('assign_user').distinct()
        else:
            assign_user_list = EquipInspectionOrder.objects.filter(id__in=pks).values_list('assign_user').distinct()
    if names:
        assign_user_list = names
    phone_numbers = User.objects.filter(username__in=assign_user_list).values_list('phone_number')
    for phone in phone_numbers:
        res = ding_api.get_user_id(phone)
        if res:
            user_ids.append(res)
    return user_ids


def pd_export_xls(sql,
                  host=settings.DATABASES['default']['HOST'],
                  user=settings.DATABASES['default']['USER'],
                  name=settings.DATABASES['default']['NAME'],
                  password=settings.DATABASES['default']['PASSWORD']
                  ):
    os.environ['NLS_LANG'] = 'SIMPLIFIED CHINESE_CHINA.UTF8'
    conn = cx_Oracle.connect(user, password, '{}:1521/{}'.format(host, name))
    df = pd.read_sql(sql, con=conn)
    bio = BytesIO()
    writer = pd.ExcelWriter(bio, engine='xlsxwriter')  # 注意安装这个包 pip install xlsxwriter
    df.to_excel(writer, sheet_name='Sheet1', index=False)
    writer.save()
    bio.seek(0)
    from django.http import FileResponse
    response = FileResponse(bio)
    response['Content-Type'] = 'application/octet-stream'
    response['Content-Disposition'] = 'attachment;filename="mm.xlsx"'
    return response


def handle_spare(last_time=None, wlxxid=None):
    params = {}
    if wlxxid:
        params['wlxxid'] = wlxxid
    else:
        if not last_time:
            last = EquipSpareErp.objects.filter(sync_date__isnull=False).order_by('sync_date').last()  # 第一次先在数据库插入一条假数据
            if not last:
                return False, '未找到最新一次同步时间数据'
            last_time = (last.sync_date + timedelta(seconds=1)).strftime('%Y-%m-%d %H:%M:%S')
        params['syncDate'] = last_time
    url = 'http://10.1.10.136/zcxjws_web/zcxjws/pc/jc/getbjwlxx.io'
    try:
        res = requests.post(url=url, json=params, timeout=3)
    except Exception:
        return False, '网络异常'
    if res.status_code != 200:
        return False, '请求失败'
    data = json.loads(res.content)
    if not data.get('flag'):
        return False, data.get('message')
    return True, data.get('obj')
