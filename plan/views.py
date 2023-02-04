import copy
import datetime
import json
import re
from io import BytesIO
from operator import itemgetter

import pandas as pd
import requests
import xlrd
from django.db import connection, IntegrityError
from django.db.models import Sum, Max, Q, Min
from django.db.transaction import atomic
from django.http import HttpResponse
from django.utils.decorators import method_decorator
from django_filters.rest_framework import DjangoFilterBackend
from openpyxl import load_workbook
from rest_framework import status, mixins
from rest_framework.decorators import action
from rest_framework.exceptions import ValidationError
from rest_framework.filters import OrderingFilter
from rest_framework.generics import ListAPIView, CreateAPIView, UpdateAPIView, GenericAPIView, get_object_or_404
from rest_framework.permissions import IsAuthenticated, IsAuthenticatedOrReadOnly
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.viewsets import ModelViewSet, GenericViewSet

from basics.models import WorkSchedulePlan, GlobalCode, Equip
from basics.views import CommonDeleteMixin
from equipment.utils import gen_template_response
from inventory.models import ProductStockDailySummary
from mes.common_code import get_weekdays
from mes.conf import JZ_EQUIP_NO
from mes.derorators import api_recorder
from mes.paginations import SinglePageNumberPagination
from mes.permissions import PermissionClass
from mes.sync import ProductClassesPlanSyncInterface
from plan.filters import ProductDayPlanFilter, MaterialDemandedFilter, ProductClassesPlanFilter, \
    BatchingClassesPlanFilter, SchedulingRecipeMachineSettingFilter, SchedulingProductDemandedDeclareSummaryFilter, \
    SchedulingProductSafetyParamsFilter, SchedulingProductDemandedDeclareFilter, ProductStockDailySummaryFilter
from plan.models import ProductDayPlan, ProductClassesPlan, MaterialDemanded, BatchingClassesPlan, \
    BatchingClassesEquipPlan, SchedulingParamsSetting, SchedulingRecipeMachineSetting, SchedulingEquipCapacity, \
    SchedulingWashRule, SchedulingWashPlaceKeyword, SchedulingWashPlaceOperaKeyword, SchedulingProductDemandedDeclare, \
    SchedulingProductDemandedDeclareSummary, SchedulingProductSafetyParams, SchedulingResult, \
    SchedulingEquipShutDownPlan, WeightPackageDailyTimeConsume
from plan.serializers import ProductDayPlanSerializer, ProductClassesPlanManyCreateSerializer, \
    ProductBatchingSerializer, ProductBatchingDetailSerializer, ProductDayPlansySerializer, \
    ProductClassesPlansySerializer, MaterialsySerializer, BatchingClassesPlanSerializer, \
    IssueBatchingClassesPlanSerializer, BatchingClassesEquipPlanSerializer, PlantImportSerializer, \
    SchedulingParamsSettingSerializer, SchedulingRecipeMachineSettingSerializer, SchedulingEquipCapacitySerializer, \
    SchedulingWashRuleSerializer, SchedulingWashPlaceKeywordSerializer, SchedulingWashPlaceOperaKeywordSerializer, \
    RecipeMachineWeightSerializer, SchedulingProductDemandedDeclareSerializer, \
    SchedulingProductDemandedDeclareSummarySerializer, SchedulingProductSafetyParamsSerializer, \
    SchedulingResultSerializer, SchedulingEquipShutDownPlanSerializer, ProductStockDailySummarySerializer, \
    WeightPackageDailyTimeConsumeSerializer
from plan.utils import calculate_product_plan_trains, extend_last_aps_result, APSLink, \
    calculate_equip_recipe_avg_mixin_time, plan_sort, calculate_product_stock, convert_fm_weight
from production.models import PlanStatus, TrainsFeedbacks, MaterialTankStatus
from quality.utils import get_cur_sheet, get_sheet_data
from recipe.models import ProductBatching, ProductBatchingDetail, Material, MaterialAttribute, WeighBatchingDetail
from system.serializers import PlanReceiveSerializer
from terminal.models import JZRecipePre, RecipePre, JZRecipeMaterial, RecipeMaterial, WeightPackageManual
from terminal.utils import get_current_factory_date


@method_decorator([api_recorder], name="dispatch")
class ProductDayPlanViewSet(CommonDeleteMixin, ModelViewSet):
    """
    list:
        胶料日计划列表
    create:
        新建胶料日计划（单增），暂且不用，
    update:
        修改原胶料日计划
    destroy:
        删除胶料日计划
    """
    queryset = ProductDayPlan.objects.filter(delete_flag=False).select_related(
        'equip__category', 'plan_schedule', 'product_batching').prefetch_related(
        'pdp_product_classes_plan__work_schedule_plan', 'pdp_product_batching_day_plan')
    serializer_class = ProductDayPlanSerializer
    permission_classes = (IsAuthenticated,)
    filter_backends = (DjangoFilterBackend, OrderingFilter)
    filter_class = ProductDayPlanFilter
    ordering_fields = ['id', 'equip__category__equip_type__global_name']

    def destroy(self, request, *args, **kwargs):
        """"胶料计划删除 先删除胶料计划，随后删除胶料计划对应的班次日计划和原材料需求量表"""
        instance = self.get_object()
        MaterialDemanded.objects.filter(
            product_classes_plan__product_day_plan=instance).delete()
        ProductClassesPlan.objects.filter(product_day_plan=instance).update(delete_flag=True, delete_user=request.user)
        instance.delete_flag = True
        instance.delete_user = request.user
        instance.save()
        return Response(status=status.HTTP_204_NO_CONTENT)


@method_decorator([api_recorder], name="dispatch")
class ProductDayPlanManyCreate(APIView):
    """胶料计划群增接口"""
    permission_classes = (IsAuthenticated,)

    def post(self, request, *args, **kwargs):
        if isinstance(request.data, dict):
            many = False
        elif isinstance(request.data, list):
            many = True
        else:
            return Response(data={'detail': '数据有误'}, status=400)
        s = ProductDayPlanSerializer(data=request.data, many=many, context={'request': request})
        s.is_valid(raise_exception=True)
        s.save()
        return Response('新建成功')


@method_decorator([api_recorder], name="dispatch")
class MaterialDemandedAPIView(ListAPIView):
    """原材料需求量展示，plan_date参数必填"""

    queryset = MaterialDemanded.objects.filter(delete_flag=False).all()
    filter_backends = (DjangoFilterBackend,)
    filter_class = MaterialDemandedFilter
    permission_classes = (IsAuthenticated,)

    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        data = queryset.filter(**kwargs).select_related('material', 'classes__classes'). \
            values('material__material_name', 'material__material_no',
                   'material__material_type__global_name', 'work_schedule_plan__classes__global_name'
                   ).annotate(num=Sum('material_demanded'))
        materials = []
        ret = {}
        for item in data:
            if item['material__material_name'] not in materials:
                ret[item['material__material_name']] = {
                    'material_no': item['material__material_no'],
                    'material_name': item['material__material_name'],
                    "material_type": item['material__material_type__global_name'],
                    "class_details": {item['work_schedule_plan__classes__global_name']: item['num']}}
                materials.append(item['material__material_name'])
            else:
                ret[item['material__material_name']
                ]['class_details'][item['work_schedule_plan__classes__global_name']] = item['num']
        page = self.paginate_queryset(list(ret.values()))
        return self.get_paginated_response(page)


@method_decorator([api_recorder], name="dispatch")
class ProductDayPlanAPiView(APIView):
    """计划数据下发至上辅机"""
    permission_classes = ()
    authentication_classes = ()

    @atomic()
    def post(self, request):
        equip = request.data.get('equip', None)
        work_schedule_plan = request.data.get('work_schedule_plan', None)
        if not equip or not work_schedule_plan:
            raise ValidationError('缺失参数')
        try:
            product_classes_plan_set = ProductClassesPlan.objects.filter(equip_id=equip,
                                                                         work_schedule_plan_id=work_schedule_plan,
                                                                         delete_flag=False)
        except Exception:
            raise ValidationError('该计划不存在')
        for product_classes_plan in product_classes_plan_set:

            # 当前班次之前的计划不准现在下发
            work_schedule_plan = product_classes_plan.work_schedule_plan
            end_time = work_schedule_plan.end_time
            now_time = datetime.datetime.now()
            if now_time > end_time:
                raise ValidationError(
                    f'{end_time.strftime("%Y-%m-%d")}的{work_schedule_plan.classes.global_name}的计划不允许现在下发给上辅机')

            if product_classes_plan.status not in ['已保存', '等待']:
                continue
            else:
                interface = ProductClassesPlanSyncInterface(instance=product_classes_plan)
                try:
                    interface.request()
                except Exception as e:
                    raise ValidationError(e)
                product_classes_plan.status = '等待'
                product_classes_plan.save()
                PlanStatus.objects.filter(plan_classes_uid=product_classes_plan.plan_classes_uid).update(status='等待')
        return Response('下达成功', status=status.HTTP_200_OK)


@method_decorator([api_recorder], name="dispatch")
class MaterialDemandedView(APIView):
    """计划原材料需求列表"""
    permission_classes = (IsAuthenticated,)

    def get(self, request):
        # 条件筛选
        params = request.query_params
        filter_dict = {'delete_flag': False}
        plan_date = params.get('plan_date')
        classes = params.get('classes')
        product_no = params.get('product_no')
        if plan_date:
            filter_dict['work_schedule_plan__plan_schedule__day_time'] = plan_date
        if classes:
            filter_dict['work_schedule_plan__classes__global_name'] = classes
        if product_no:
            filter_dict[
                'product_classes_plan__product_batching__stage_product_batch_no__icontains'] = product_no

        # 库存请求
        material_inventory_dict = {}
        try:
            ret = requests.get("http://49.235.45.128:8169/storageSpace/GetInventoryCount")
            ret_json = json.loads(ret.text)
            for i in ret_json.get("datas"):
                material_inventory_dict[i['materialCode']] = i
        except Exception as e:
            pass
        try:
            page = int(params.get("page", 1))
            page_size = int(params.get("page_size", 10))
        except Exception as e:
            return Response("page和page_size必须是int", status=400)
        md_list = MaterialDemanded.objects.filter(**filter_dict).values(
            'product_classes_plan__product_batching__stage_product_batch_no',
            'work_schedule_plan',
            'material__material_no',
            'material__material_name',
            'material__material_type__global_name',
        ).annotate(demanded=Sum('material_demanded'))
        counts = md_list.count()
        md_list = md_list[(page - 1) * page_size:page_size * page]
        res = []
        for md_detail_list in md_list:
            md = {}
            md['product_no'] = md_detail_list[
                'product_classes_plan__product_batching__stage_product_batch_no']
            md['classes'] = WorkSchedulePlan.objects.get(id=md_detail_list['work_schedule_plan']).classes.global_name
            md['material_no'] = md_detail_list['material__material_no']
            md['material_name'] = md_detail_list['material__material_name']
            md['material_type'] = md_detail_list['material__material_type__global_name']
            md['material_demanded'] = md_detail_list['demanded']
            # 库存
            inventory_detail = material_inventory_dict.get(md.get('material_no'))
            if inventory_detail:
                quantity = inventory_detail.get('quantity')
                weightOfActual = inventory_detail.get('weightOfActual')
                unit_weight = weightOfActual / quantity  # TODO 单位重量到底是总重量除以总数量还是计件数量 这个计件数量到底是什么意思
                md['qty'] = quantity
                md['total_weight'] = weightOfActual
                md['unit_weight'] = unit_weight
                md['need_unit_weight'] = unit_weight
                md['need_qty'] = md['material_demanded'] / unit_weight
            else:
                md['qty'] = None
                md['total_weight'] = None
                md['unit_weight'] = None
                md['need_unit_weight'] = None
                md['need_qty'] = None
            res.append(md)
        return Response({'results': res, 'count': counts})


@method_decorator([api_recorder], name="dispatch")
class ProductClassesPlanManyCreate(APIView):
    """胶料日班次计划群增接口"""

    permission_classes = (IsAuthenticated,)

    @atomic()
    def post(self, request, *args, **kwargs):
        if isinstance(request.data, dict):
            day_time = WorkSchedulePlan.objects.filter(
                id=request.data['work_schedule_plan']).first().plan_schedule.day_time
            schedule_no = WorkSchedulePlan.objects.filter(
                id=request.data['work_schedule_plan']).first().plan_schedule.work_schedule.schedule_no
            pcp_set = ProductClassesPlan.objects.filter(work_schedule_plan__plan_schedule__day_time=day_time,
                                                        equip_id=request.data['equip'], delete_flag=False,
                                                        work_schedule_plan__plan_schedule__work_schedule__schedule_no=schedule_no).all()
            for pcp_obj in pcp_set:
                pcp_obj.delete_flag = True
                pcp_obj.save()
                PlanStatus.objects.filter(plan_classes_uid=pcp_obj.plan_classes_uid).update(delete_flag=True)
                MaterialDemanded.objects.filter(product_classes_plan=pcp_obj).update(delete_flag=True)
            return Response('操作成功')
        elif isinstance(request.data, list):
            many = True
            work_list = []
            plan_list = []
            equip_list = []
            for class_dict in request.data:
                work_list.append(class_dict['work_schedule_plan'])
                schedule_no = WorkSchedulePlan.objects.filter(
                    id=class_dict['work_schedule_plan']).first().plan_schedule.work_schedule.schedule_no
                plan_list.append(class_dict['plan_classes_uid'])
                equip_list.append(class_dict['equip'])
                class_dict['status'] = '已保存'
                # 判断胶料日计划是否存在
                wsp_obj = WorkSchedulePlan.objects.filter(id=class_dict['work_schedule_plan']).first()
                pdp_obj = ProductDayPlan.objects.filter(equip_id=class_dict['equip'],
                                                        product_batching_id=class_dict['product_batching'],
                                                        plan_schedule=wsp_obj.plan_schedule, delete_flag=False).first()
                if pdp_obj:
                    class_dict['product_day_plan'] = pdp_obj.id

                else:
                    class_dict['product_day_plan'] = ProductDayPlan.objects.create(equip_id=class_dict['equip'],
                                                                                   product_batching_id=class_dict[
                                                                                       'product_batching'],
                                                                                   plan_schedule=wsp_obj.plan_schedule,
                                                                                   last_updated_date=datetime.datetime.now(),
                                                                                   created_date=datetime.datetime.now()).id
            # 举例说明：本来有四条 前端只传了三条 就会删掉多余的一条
            day_time = WorkSchedulePlan.objects.filter(
                id=class_dict['work_schedule_plan']).first().plan_schedule.day_time
            pcp_set = ProductClassesPlan.objects.filter(work_schedule_plan__plan_schedule__day_time=day_time,
                                                        equip_id__in=equip_list, delete_flag=False,
                                                        work_schedule_plan__plan_schedule__work_schedule__schedule_no=schedule_no).exclude(
                plan_classes_uid__in=plan_list)
            for pcp_obj in pcp_set:
                # 删除前要先判断该数据的状态是不是非等待，只要等待中的计划才可以删除
                plan_status = PlanStatus.objects.filter(plan_classes_uid=pcp_obj.plan_classes_uid,
                                                        delete_flag=False).order_by(
                    'created_date').last()
                if plan_status:
                    if plan_status.status not in ['等待', '已保存']:
                        raise ValidationError("只要等待中或者已保存的计划才可以删除")
                else:
                    pass

                # 删除多余的数据已经以及向关联的计划状态变更表和原材料需求量表需求量表
                pcp_obj.delete_flag = True
                pcp_obj.save()
                PlanStatus.objects.filter(plan_classes_uid=pcp_obj.plan_classes_uid).update(delete_flag=True)
                MaterialDemanded.objects.filter(product_classes_plan=pcp_obj).update(delete_flag=True)
            s = ProductClassesPlanManyCreateSerializer(data=request.data, many=many,
                                                       context={'request': request})
            s.is_valid(raise_exception=True)
            s.save()
            return Response('保存成功')
        else:
            return Response(data={'detail': '数据有误'}, status=400)


@method_decorator([api_recorder], name="dispatch")
class ProductClassesPlanList(mixins.ListModelMixin, GenericViewSet):
    """计划新增展示数据"""
    queryset = ProductClassesPlan.objects.filter(delete_flag=False).order_by('sn')
    serializer_class = ProductClassesPlanManyCreateSerializer
    permission_classes = (IsAuthenticatedOrReadOnly,)
    pagination_class = SinglePageNumberPagination
    filter_backends = (DjangoFilterBackend,)
    filter_class = ProductClassesPlanFilter


@method_decorator([api_recorder], name="dispatch")
class PlanReceive(CreateAPIView):
    """
        接受上辅机计划数据接口
        """
    # permission_classes = ()
    # authentication_classes = ()
    permission_classes = (IsAuthenticated,)
    serializer_class = PlanReceiveSerializer
    queryset = ProductDayPlan.objects.all()

    @atomic()
    def post(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        headers = self.get_success_headers(serializer.data)
        return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)


@method_decorator([api_recorder], name="dispatch")
class ProductBatchingReceive(CreateAPIView):
    """胶料配料标准同步"""
    serializer_class = ProductBatchingSerializer
    queryset = ProductBatching.objects.all()


@method_decorator([api_recorder], name="dispatch")
class ProductBatchingDetailReceive(CreateAPIView):
    """胶料配料标准详情同步"""
    serializer_class = ProductBatchingDetailSerializer
    queryset = ProductBatchingDetail.objects.all()


@method_decorator([api_recorder], name="dispatch")
class ProductDayPlanReceive(CreateAPIView):
    """胶料日计划表同步"""
    serializer_class = ProductDayPlansySerializer
    queryset = ProductDayPlan.objects.all()


@method_decorator([api_recorder], name="dispatch")
class ProductClassesPlanReceive(CreateAPIView):
    """胶料日班次计划表同步"""
    serializer_class = ProductClassesPlansySerializer
    queryset = ProductClassesPlan.objects.all()


@method_decorator([api_recorder], name="dispatch")
class MaterialReceive(CreateAPIView):
    """原材料表同步"""
    serializer_class = MaterialsySerializer
    queryset = Material.objects.all()


@method_decorator([api_recorder], name="dispatch")
class BatchingClassesPlanView(ModelViewSet):
    """配料日班次计划"""
    queryset = BatchingClassesPlan.objects.filter(delete_flag=False).order_by('-created_date')
    serializer_class = BatchingClassesPlanSerializer
    filter_backends = (DjangoFilterBackend,)
    filter_class = BatchingClassesPlanFilter

    def perform_update(self, serializer):
        serializer.save(package_changed=True)


@method_decorator([api_recorder], name="dispatch")
class BatchingClassesEquipPlanViewSet(ModelViewSet):
    """配料日班次计划"""
    queryset = BatchingClassesEquipPlan.objects.order_by('-id')
    serializer_class = BatchingClassesEquipPlanSerializer
    filter_backends = (DjangoFilterBackend,)
    filter_fields = ('batching_class_plan', )
    pagination_class = None

    def create(self, request, *args, **kwargs):
        if not isinstance(request.data, list):
            raise ValidationError('参数错误')
        for item in request.data:
            s = BatchingClassesEquipPlanSerializer(data=item, context={'request': request})
            if not s.is_valid():
                raise ValidationError(s.errors)
            s.save()
        return Response('新建成功')

    def perform_update(self, serializer):
        serializer.save(package_changed=True)


@method_decorator([api_recorder], name="dispatch")
class IssueBatchingClassesPlanView(UpdateAPIView):
    queryset = BatchingClassesPlan.objects.filter(delete_flag=False)
    serializer_class = IssueBatchingClassesPlanSerializer
    permission_classes = (IsAuthenticated,)

    def perform_update(self, serializer):
        serializer.save(send_user=self.request.user)


@method_decorator([api_recorder], name="dispatch")
class PlantImportView(CreateAPIView):
    serializer_class = PlantImportSerializer
    queryset = ProductClassesPlan.objects.all()
    permission_classes = (IsAuthenticated, PermissionClass({'add': 'add_productdayplan'}))


@method_decorator([api_recorder], name="dispatch")
class LabelPlanInfo(APIView):
    """根据计划编号，获取工厂日期和班组"""

    def get(self, request):
        plan_classes_uid = self.request.query_params.get('planid')  # 计划uid
        produce_time = self.request.query_params.get('producetime')  # 生产时间
        try:
            produce_time = datetime.datetime.strptime(produce_time, "%Y-%m-%d %H:%M:%S")
        except Exception:
            raise ValidationError('produce_time日期格式错误')
        ret = {'factory_date': '', 'group': '', 'expire_time': ''}
        plan = ProductClassesPlan.objects.filter(plan_classes_uid=plan_classes_uid).first()
        if plan:
            ret['factory_date'] = plan.work_schedule_plan.plan_schedule.day_time
            ret['group'] = plan.work_schedule_plan.group.global_name
            material_detail = MaterialAttribute.objects.filter(
                material__material_no=plan.product_batching.stage_product_batch_no).first()
            if material_detail:
                if material_detail.period_of_validity:
                    unit = material_detail.validity_unit
                    if unit in ["天", "days", "day"]:
                        param = {"days": material_detail.period_of_validity}
                    elif unit in ["小时", "hours", "hour"]:
                        param = {"hours": material_detail.period_of_validity}
                    else:
                        param = {"days": material_detail.period_of_validity}
                    expire_time = (produce_time + datetime.timedelta(**param)).strftime('%Y-%m-%d %H:%M:%S')
                    ret['expire_time'] = expire_time
        elif produce_time:
            current_work_schedule_plan = WorkSchedulePlan.objects.filter(
                start_time__lte=produce_time,
                end_time__gte=produce_time,
                plan_schedule__work_schedule__work_procedure__global_name='密炼'
            ).first()
            if current_work_schedule_plan:
                ret['factory_date'] = current_work_schedule_plan.plan_schedule.day_time
                ret['group'] = current_work_schedule_plan.group.global_name
        return Response(ret)


@method_decorator([api_recorder], name="dispatch")
class SchedulingParamsSettingView(ModelViewSet):
    queryset = SchedulingParamsSetting.objects.all()
    serializer_class = SchedulingParamsSettingSerializer
    permission_classes = (IsAuthenticated,)
    pagination_class = None


@method_decorator([api_recorder], name="dispatch")
class SchedulingRecipeMachineSettingView(ModelViewSet):
    queryset = SchedulingRecipeMachineSetting.objects.order_by('rubber_type', 'product_no', 'version')
    serializer_class = SchedulingRecipeMachineSettingSerializer
    pagination_class = None
    # permission_classes = (IsAuthenticated,)
    filter_backends = (DjangoFilterBackend,)
    filter_class = SchedulingRecipeMachineSettingFilter

    @action(methods=['post'], detail=False, url_path='confirm')
    def confirm(self, request):
        obj_id = request.data.get('id')
        SchedulingRecipeMachineSetting.objects.filter(id=obj_id).update(confirmed=True)
        return Response('OK')

    @action(methods=['post'], detail=False)
    def import_xlsx(self, request):
        excel_file = request.FILES.get('file', None)
        if not excel_file:
            raise ValidationError('文件不可为空！')
        cur_sheet = get_cur_sheet(excel_file)
        # if cur_sheet.ncols != 18:
        #     raise ValidationError('导入文件数据错误！')
        data = get_sheet_data(cur_sheet, start_row=2)
        for item in data:
            if not all([item[0], item[1], item[2], item[3]]):
                raise ValidationError('必填数据缺失')
            try:
                SchedulingRecipeMachineSetting.objects.update_or_create(
                    defaults={
                              "rubber_type": item[0],
                              "stages": item[3],
                              "main_machine_HMB": item[4],
                              "vice_machine_HMB": item[5],
                              "main_machine_CMB": item[6],
                              "vice_machine_CMB": item[7],
                              "main_machine_1MB": item[8],
                              "vice_machine_1MB": item[9],
                              "main_machine_2MB": item[10],
                              "vice_machine_2MB": item[11],
                              "main_machine_3MB": item[12],
                              "vice_machine_3MB": item[13],
                              "main_machine_4MB": item[14],
                              "vice_machine_4MB": item[15],
                              "main_machine_FM": item[16],
                              "vice_machine_FM": item[17],
                              },
                    **{"product_no": item[1], "version": item[2]}
                )
            except Exception:
                raise
        return Response('ok')


@method_decorator([api_recorder], name="dispatch")
class RecipeMachineWeight(ListAPIView):
    queryset = ProductBatching.objects.all()
    serializer_class = RecipeMachineWeightSerializer

    def get_queryset(self):
        equip_no = self.request.query_params.get('equip_no')
        dev_type = self.request.query_params.get('dev_type')
        product_no = self.request.query_params.get('product_no')
        query_kwargs = {}
        if equip_no:
            query_kwargs['equip__equip_no'] = equip_no
        if dev_type:
            query_kwargs['dev_type__category_name'] = dev_type
        if product_no:
            query_kwargs['stage_product_batch_no__icontains'] = product_no
        return ProductBatching.objects.using('SFJ').exclude(
            used_type=6).filter(**query_kwargs).filter(
            batching_type=1).values('id', 'equip__equip_no', 'stage_product_batch_no', 'batching_weight', 'equip__category__category_no').order_by('equip__equip_no')


@method_decorator([api_recorder], name="dispatch")
class MaterialTankStatusView(APIView):

    def get(self, request):
        tank_type = self.request.query_params.get('tank_type', '1')
        data = MaterialTankStatus.objects.filter(
            tank_no__in=[str(i) for i in range(11)],
            tank_type=tank_type,
            equip_no__in=['Z01', 'Z02', 'Z03', 'Z04', 'Z05', 'Z06', 'Z07', 'Z08', 'Z09', 'Z10']
        ).using('SFJ').values('equip_no', 'tank_no', 'material_name').order_by('equip_no', 'tank_no')
        ret = {}
        equip_nos = set()
        for item in data:
            equip_nos.add(item['equip_no'])
            if item['tank_no'] not in ret:
                ret[item['tank_no']] = {'tank_no': item['tank_no'], item['equip_no']: item['material_name']}
            else:
                ret[item['tank_no']][item['equip_no']] = item['material_name']
        return Response({'equip_nos': sorted(list(equip_nos)), 'data': list(ret.values())})


@method_decorator([api_recorder], name="dispatch")
class SchedulingEquipCapacityViewSet(ModelViewSet):
    queryset = SchedulingEquipCapacity.objects.order_by('equip_no', 'product_no')
    serializer_class = SchedulingEquipCapacitySerializer
    permission_classes = (IsAuthenticated,)
    filter_backends = (DjangoFilterBackend, OrderingFilter)
    filter_fields = ('equip_no', 'product_no')
    ordering_fields = ['avg_mixing_time', 'avg_interval_time', 'avg_rubbery_quantity']
    FILE_NAME = '机台设备生产能力'
    EXPORT_FIELDS_DICT = {'机台': 'equip_no',
                          '胶料编码': 'product_no',
                          '平均工作时间(秒）': 'avg_mixing_time',
                          '平均间隔时间(秒）': 'avg_interval_time',
                          '平均加胶量(kg)': 'avg_rubbery_quantity',
                          '录入者': 'created_username',
                          '录入时间': 'created_date',
                          }

    def list(self, request, *args, **kwargs):
        export = self.request.query_params.get('export')
        queryset = self.filter_queryset(self.get_queryset())
        if export:
            data = self.get_serializer(queryset, many=True).data
            return gen_template_response(self.EXPORT_FIELDS_DICT, data, self.FILE_NAME)
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)

        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)


@method_decorator([api_recorder], name="dispatch")
class SchedulingWashRuleViewSet(CommonDeleteMixin, ModelViewSet):
    queryset = SchedulingWashRule.objects.order_by('-id')
    serializer_class = SchedulingWashRuleSerializer
    permission_classes = (IsAuthenticated,)
    filter_backends = (DjangoFilterBackend,)
    pagination_class = None


@method_decorator([api_recorder], name="dispatch")
class SchedulingWashPlaceKeywordViewSet(ModelViewSet):
    queryset = SchedulingWashPlaceKeyword.objects.order_by('id')
    serializer_class = SchedulingWashPlaceKeywordSerializer
    permission_classes = (IsAuthenticated,)
    filter_backends = (DjangoFilterBackend,)
    pagination_class = None


@method_decorator([api_recorder], name="dispatch")
class SchedulingWashPlaceOperaKeywordViewSet(ModelViewSet):
    queryset = SchedulingWashPlaceOperaKeyword.objects.order_by('id')
    serializer_class = SchedulingWashPlaceOperaKeywordSerializer
    permission_classes = (IsAuthenticated,)
    filter_backends = (DjangoFilterBackend,)
    pagination_class = None


@method_decorator([api_recorder], name="dispatch")
class SchedulingProductDemandedDeclareViewSet(ModelViewSet):
    queryset = SchedulingProductDemandedDeclare.objects.order_by('id')
    serializer_class = SchedulingProductDemandedDeclareSerializer
    permission_classes = (IsAuthenticated,)
    filter_backends = (DjangoFilterBackend,)
    filter_class = SchedulingProductDemandedDeclareFilter

    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        page = self.paginate_queryset(queryset)
        serializer = self.get_serializer(page, many=True)
        data = self.get_paginated_response(serializer.data).data
        sum_data = queryset.aggregate(total_today_demanded=Sum('today_demanded'),
                                      total_tomorrow_demanded=Sum('tomorrow_demanded'),
                                      total_current_stock=Sum('current_stock'),
                                      total_underway_stock=Sum('underway_stock'))
        data['total_today_demanded'] = sum_data['total_today_demanded']
        data['total_tomorrow_demanded'] = sum_data['total_tomorrow_demanded']
        data['total_current_stock'] = sum_data['total_current_stock']
        data['total_underway_stock'] = sum_data['total_underway_stock']
        return Response(data)

    def create(self, request, *args, **kwargs):
        data = self.request.data
        order_no = 'FC{}'.format(datetime.datetime.now().strftime('%Y%m%d%H%M%S'))
        if not isinstance(data, list):
            raise ValidationError('data error')
        s = self.serializer_class(data=data, many=True, context={'request': request, 'order_no': order_no})
        s.is_valid(raise_exception=True)
        s.save()
        return Response('ok')


@method_decorator([api_recorder], name="dispatch")
class SchedulingProductSafetyParamsViewSet(CommonDeleteMixin, ModelViewSet):
    queryset = SchedulingProductSafetyParams.objects.order_by('id')
    serializer_class = SchedulingProductSafetyParamsSerializer
    permission_classes = (IsAuthenticated,)
    filter_backends = (DjangoFilterBackend,)
    filter_class = SchedulingProductSafetyParamsFilter

    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        page = self.paginate_queryset(queryset)
        serializer = self.get_serializer(page, many=True)
        data = self.get_paginated_response(serializer.data).data
        sum_data = queryset.aggregate(total_safety_stock=Sum('safety_stock'),
                                      total_daily_usage=Sum('daily_usage'))
        data['total_safety_stock'] = sum_data['total_safety_stock']
        data['total_daily_usage'] = sum_data['total_daily_usage']
        return Response(data)


@method_decorator([api_recorder], name="dispatch")
class ProductDeclareSummaryViewSet(ModelViewSet):
    queryset = SchedulingProductDemandedDeclareSummary.objects.order_by('available_time')
    serializer_class = SchedulingProductDemandedDeclareSummarySerializer
    permission_classes = (IsAuthenticated,)
    filter_backends = (DjangoFilterBackend,)
    filter_class = SchedulingProductDemandedDeclareSummaryFilter
    pagination_class = None

    @action(methods=['post'], detail=False, url_path='up-sequence')
    def up_sequence(self, request):
        try:
            instance = SchedulingProductDemandedDeclareSummary.objects.get(id=self.request.data.get('id'))
        except Exception:
            raise ValidationError('object does not exits!')
        sn = instance.sn
        previous_instance = SchedulingProductDemandedDeclareSummary.objects.filter(
            factory_date=instance.factory_date, sn__lt=instance.sn).order_by('sn').last()
        if previous_instance:
            instance.sn = previous_instance.sn
            previous_instance.sn = sn
            instance.save()
            previous_instance.save()
        return Response('成功')

    @action(methods=['post'], detail=False, url_path='down-sequence')
    def down_sequence(self, request):
        try:
            instance = SchedulingProductDemandedDeclareSummary.objects.get(id=self.request.data.get('id'))
        except Exception:
            raise ValidationError('object does not exits!')
        sn = instance.sn
        next_instance = SchedulingProductDemandedDeclareSummary.objects.filter(
            factory_date=instance.factory_date, sn__gt=instance.sn).order_by('sn').first()
        if next_instance:
            instance.sn = next_instance.sn
            next_instance.sn = sn
            instance.save()
            next_instance.save()
        return Response('成功')

    @action(methods=['post'], detail=False, permission_classes=[], url_path='import_xlsx',
            url_name='import_xlsx')
    def import_xlx(self, request):
        factory_date = self.request.data.get('factory_date', datetime.datetime.now().strftime('%Y-%m-%d'))
        date_splits = factory_date.split('-')
        m = date_splits[1] if not date_splits[1].startswith('0') else date_splits[1].lstrip('0')
        d = date_splits[2]
        excel_file = request.FILES.get('file', None)
        if not excel_file:
            raise ValidationError('文件不可为空！')
        if not excel_file.name.split('.')[-1] in ['xls', 'xlsx', 'xlsm']:
            raise ValidationError('文件格式错误,仅支持 xls、xlsx、xlsm文件')
        try:
            data = xlrd.open_workbook(filename=None, file_contents=excel_file.read())
            cur_sheet = data.sheet_by_name(sheet_name='{}.{}'.format(m, d))
        except Exception:
            raise ValidationError('未找到{}.{}库存excel文档！'.format(m, d))
        data = get_sheet_data(cur_sheet, start_row=5)

        area_list = []
        c = SchedulingProductDemandedDeclareSummary.objects.filter(
            factory_date=factory_date).count()
        sn = c + 1
        for idx, item in enumerate(data):
            product_no = re.sub(r'[\u4e00-\u9fa5]+', '', item[4])
            if not product_no or not item[5]:
                continue
            pbs = list(ProductBatching.objects.using('SFJ').filter(
                used_type=4,
                stage_product_batch_no__icontains='-FM-{}-'.format(product_no)
            ).order_by('used_time').values_list('stage_product_batch_no', flat=True))
            if not pbs:
                raise ValidationError('未找到该规格FM启用配方：{}'.format(product_no))
            try:
                if len(set(pbs)) == 1:  # 启用规格只有一种
                    version = pbs[0].split('-')[-1]
                    area_list.append({'factory_date': factory_date,
                                      'sn': sn,
                                      'product_no': product_no,
                                      'version': version,
                                      'plan_weight': item[5],
                                      'workshop_weight': round(item[16], 1) if item[16] else 0,
                                      'current_stock': round(item[17], 1) if item[17] else 0,
                                      'desc': '',
                                      # 'target_stock': float(item[1]) * 1.5,
                                      # 'demanded_weight': float(item[1]) * 1.5 - float(item[2]) - float(item[3])
                                      })
                else:
                    old_version_recipe = pbs[0]
                    new_version_recipe = pbs[-1]
                    old_version_recipe_split_data = old_version_recipe.split('-')
                    new_version_recipe_split_data = new_version_recipe.split('-')
                    old_pb_no, old_version = old_version_recipe_split_data[2], old_version_recipe_split_data[3]
                    new_pb_no, new_version = new_version_recipe_split_data[2], new_version_recipe_split_data[3]
                    old_recipe_weight = convert_fm_weight(old_pb_no, old_version, factory_date)
                    if not old_recipe_weight:
                        area_list.append({'factory_date': factory_date,
                                          'sn': sn,
                                          'product_no': product_no,
                                          'version': new_version,
                                          'plan_weight': item[5],
                                          'workshop_weight': 0 if not item[16] else round(item[16], 1),
                                          'current_stock': 0 if not item[17] else round(item[17], 1),
                                          'desc': '',
                                          })
                    else:
                        if item[5] <= old_recipe_weight:
                            area_list.append({'factory_date': factory_date,
                                              'sn': sn,
                                              'product_no': product_no,
                                              'version': old_version,
                                              'plan_weight': old_recipe_weight,
                                              'workshop_weight': 0 if not item[16] else round(item[16], 1),
                                              'current_stock': 0 if not item[17] else round(item[17], 1),
                                              'desc': '',
                                              'demanded_weight': old_recipe_weight
                                              })
                        else:
                            area_list.append({'factory_date': factory_date,
                                              'sn': sn,
                                              'product_no': product_no,
                                              'version': old_version,
                                              'plan_weight': old_recipe_weight,
                                              'workshop_weight': 0,
                                              'current_stock': 0,
                                              'desc': '',
                                              'demanded_weight': old_recipe_weight
                                              })
                            area_list.append({'factory_date': factory_date,
                                              'sn': sn,
                                              'product_no': product_no,
                                              'version': new_version,
                                              'plan_weight': round(item[5]-old_recipe_weight, 2),
                                              'workshop_weight': 0,
                                              'current_stock': 0,
                                              'desc': '',
                                              })
            except Exception:
                raise ValidationError('第{}行数据有错，请检查后重试!'.format(6+idx))
            sn += 1
        s = self.serializer_class(data=area_list, many=True)
        if not s.is_valid():
            raise ValidationError('导入数据有误，请检查后重试!')
        ps = SchedulingParamsSetting.objects.first()
        small_ton_stock_days = ps.small_ton_stock_days
        middle_ton_stock_days = ps.middle_ton_stock_days
        big_ton_stock_days = ps.big_ton_stock_days
        for item in s.validated_data:
            if 'demanded_weight' not in item:
                if item['plan_weight'] < 5:
                    min_stock_day = float(small_ton_stock_days)
                elif 5 <= item['plan_weight'] <= 10:
                    min_stock_day = float(middle_ton_stock_days)
                else:
                    min_stock_day = float(big_ton_stock_days)
                item['target_stock'] = round(item['plan_weight'] * min_stock_day, 1)
                demanded_weight = round(item['plan_weight'] * min_stock_day - item['workshop_weight'] - item['current_stock'], 1)
                item['demanded_weight'] = demanded_weight if demanded_weight >= 0 else 0
        s.save()
        return Response('ok')


class SchedulingResultViewSet(ModelViewSet):
    queryset = SchedulingResult.objects.order_by('id')
    serializer_class = SchedulingResultSerializer
    permission_classes = (IsAuthenticated,)
    filter_backends = (DjangoFilterBackend,)
    pagination_class = None

    @atomic()
    def create(self, request, *args, **kwargs):
        schedule_no = self.request.data.get('schedule_no')
        plan_data = self.request.data.get('plan_data')
        if not all([schedule_no, plan_data]):
            raise ValidationError('参数缺失！')
        sr = SchedulingResult.objects.filter(schedule_no=schedule_no).first()
        factory_date = sr.factory_date
        SchedulingResult.objects.filter(schedule_no=schedule_no).delete()
        equip_aps_st = {}
        aps_st = datetime.datetime(year=factory_date.year, month=factory_date.month, day=factory_date.day, hour=8)
        for key, value in plan_data.items():
            for idx, item in enumerate(value):
                if not item.get('recipe_name') or not item.get('plan_trains'):
                    continue
                equip_st = equip_aps_st.get(key)
                if not equip_st:
                    equip_st = aps_st
                equip_et = equip_st + datetime.timedelta(hours=float(item['time_consume']))
                equip_aps_st[key] = equip_et
                SchedulingResult.objects.create(
                    factory_date=factory_date,
                    schedule_no=schedule_no,
                    equip_no=key,
                    sn=idx+1,
                    recipe_name=item['recipe_name'],
                    time_consume=item['time_consume'] if item['time_consume'] else 0,
                    plan_trains=item['plan_trains'],
                    desc=item['desc'],
                    start_time=equip_st,
                    end_time=equip_et
                )
        return Response('成功')

    @action(methods=['get'], detail=False)
    def schedule_nos(self, request):
        factory_date = self.request.query_params.get('factory_date')
        query_set = SchedulingResult.objects.all()
        if factory_date:
            query_set = query_set.filter(factory_date=factory_date)
        return Response(query_set.values('schedule_no').annotate(a=Max('id')).order_by('-a').values_list('schedule_no', flat=True))
        # return Response(sorted(list(set(query_set.values_list('schedule_no', flat=True))), reverse=True))

    def list(self, request, *args, **kwargs):
        schedule_no = self.request.query_params.get('schedule_no')
        if not schedule_no:
            raise ValidationError('请输入排程单号！')
        ret = {}
        for equip in Equip.objects.filter(
                category__equip_type__global_name='密炼设备'
        ).order_by('equip_no'):
            ret[equip.equip_no] = {'data': [], 'dev_type': equip.category.category_name}
        queryset = self.get_queryset().filter(schedule_no=schedule_no)
        for instance in queryset:
            ret[instance.equip_no]['data'].append(self.get_serializer(instance).data)
        return Response(ret)

    @action(methods=['post'], detail=False)
    def import_xlx(self, request):
        factory_date = self.request.data.get('factory_date')
        if not factory_date:
            raise ValidationError('请选择日期！')
        date_splits = factory_date.split('-')
        m = date_splits[1] if not date_splits[1].startswith('0') else date_splits[1].lstrip('0')
        d = date_splits[2]
        schedule_no = 'APS1{}'.format(datetime.datetime.now().strftime('%Y%m%d%H%M%S'))
        excel_file = request.FILES.get('file', None)
        if not excel_file:
            raise ValidationError('文件不可为空！')
        if not excel_file.name.split('.')[-1] in ['xls', 'xlsx', 'xlsm']:
            raise ValidationError('文件格式错误,仅支持 xls、xlsx、xlsm文件')
        try:
            data = xlrd.open_workbook(filename=None, file_contents=excel_file.read())
        except Exception as e:
            raise ValidationError('打开文件失败，请用文档另存为xlsx文件后导入！'.format(e))
        try:
            cur_sheet = data.sheet_by_name(sheet_name='{}.{}'.format(m, d))
        except Exception:
            raise ValidationError('未找到{}.{}日排程结果excel文档！'.format(m, d))
        data = get_sheet_data(cur_sheet, start_row=3)
        i = 0
        ret = []
        aps_st = datetime.datetime(year=int(date_splits[0]), month=int(date_splits[1]), day=int(date_splits[1]), hour=8)
        equip_aps_st = {}
        for idx, item in enumerate(data):
            if idx < 20:
                equip_nos = ['Z01', 'Z02', 'Z03', 'Z04', 'Z05', 'Z06', 'Z07', 'Z08']
            elif 22 < idx < 43:
                equip_nos = ['Z09', 'Z10', 'Z11', 'Z12', 'Z13', 'Z14', 'Z15']
            else:
                continue
            for j, equip_no in enumerate(equip_nos):
                try:
                    product_no = item[j * 4 + 1]
                    if not product_no:
                        continue
                    try:
                        plan_trains = int(item[j * 4 + 2])
                    except Exception:
                        continue
                    pb = ProductBatching.objects.using('SFJ').exclude(used_type=6).filter(
                        stage_product_batch_no__icontains='-{}'.format(product_no)).first()
                    if pb:
                        product_no = pb.stage_product_batch_no
                    train_time_consume = calculate_equip_recipe_avg_mixin_time(equip_no, product_no)
                    equip_st = equip_aps_st.get(equip_no)
                    if not equip_st:
                        equip_st = aps_st
                    equip_et = equip_st + datetime.timedelta(hours=round(train_time_consume*plan_trains/3600, 1))
                    equip_aps_st[equip_no] = equip_et
                    ret.append(SchedulingResult(**{'factory_date': factory_date,
                                                    'schedule_no': schedule_no,
                                                    'equip_no': equip_no,
                                                    'sn': i + 1,
                                                    'recipe_name': product_no,
                                                    'plan_trains': plan_trains,
                                                    'time_consume': round(train_time_consume*plan_trains/3600, 1),
                                                    'desc': item[j * 4 + 4],
                                                    'start_time': equip_st,
                                                    'end_time': equip_et}))
                except Exception:
                    raise ValidationError('导入数据有误，请检查后重试!')
            i += 1
        SchedulingResult.objects.bulk_create(ret)
        return Response('导入排程结果成功!')


@method_decorator([api_recorder], name="dispatch")
class SchedulingEquipShutDownPlanViewSet(ModelViewSet):
    queryset = SchedulingEquipShutDownPlan.objects.all()
    serializer_class = SchedulingEquipShutDownPlanSerializer
    permission_classes = (IsAuthenticated,)
    filter_backends = (DjangoFilterBackend,)

    def get_queryset(self):
        queryset = self.queryset
        factory_date = self.request.query_params.get('factory_date')
        if factory_date:
            queryset = queryset.filter(begin_time__date=factory_date)
        return queryset


@method_decorator([api_recorder], name="dispatch")
class SchedulingProceduresView(APIView):
    permission_classes = (IsAuthenticated,)

    @atomic()
    def post(self, request):
        factory_date = self.request.data.get('factory_date')
        if not factory_date:
            raise ValidationError('参数缺失！')
        try:
            factory_date = datetime.datetime.strptime(factory_date, "%Y-%m-%d")
        except Exception:
            raise ValidationError('参数错误！')
        while 1:
            schedule_no = 'APS1{}'.format(datetime.datetime.now().strftime('%Y%m%d%H%M%S'))
            if not SchedulingResult.objects.filter(schedule_no=schedule_no).exists():
                break

        # 继承前一天排程未完成的计划
        equip_tree_data = extend_last_aps_result(factory_date, schedule_no)

        # equip_nos = Equip.objects.filter(category__equip_type__global_name='密炼设备').values_list('equip_no', flat=True).order_by('equip_no')
        # links = [APSLink(equip_no, schedule_no) for equip_no in range(len(equip_nos))]
        # link_dict = {equip_no: links[idx] for idx, equip_no in enumerate(equip_nos)}

        product_demanded_rains = {}
        for dec in SchedulingProductDemandedDeclareSummary.objects.filter(
                factory_date=factory_date, demanded_weight__gt=0).order_by('available_time'):
            try:
                data = calculate_product_plan_trains(factory_date,
                                                     dec.product_no,
                                                     dec.demanded_weight)
            except Exception as e:
                raise ValidationError(e)
            product_demanded_rains[dec.product_no] = data
        equip_tree = plan_sort(product_demanded_rains, equip_tree_data)
        for equip, items in equip_tree.items():
            i = 1
            for item in items:
                if not item['plan_trains'] > 0:
                    continue
                SchedulingResult.objects.create(factory_date=factory_date,
                                                schedule_no=schedule_no,
                                                equip_no=equip,
                                                sn=i,
                                                recipe_name=item['recipe_name'],
                                                time_consume=item['time_consume'],
                                                plan_trains=item['plan_trains'],
                                                desc=item.get('desc'))
                i += 1
        #     for item in data:
        #         instance = link_dict[item['equip_no']]
        #         instance.append(item)
        # for i in links:
        #     ret = i.travel()
        #     for item in ret:
        #         # SchedulingRecipeMachineRelationHistory.objects.create(
        #         #     schedule_no=schedule_no,
        #         #     equip_no=item['equip_no'],
        #         #     recipe_name=item['product_no'],
        #         #     batching_weight=item['batching_weight'],
        #         #     devoted_weight=item['devoted_weight'],
        #         #     dev_type=item['dev_type'],
        #         # )
        #         SchedulingResult.objects.create(
        #             sn=1,
        #             factory_date=factory_date,
        #             schedule_no=schedule_no,
        #             equip_no=item['equip_no'],
        #             recipe_name=item['product_no'],
        #             time_consume=item['consume_time'],
        #             plan_trains=item['plan_trains']
        #         )
        return Response('ok')


@method_decorator([api_recorder], name="dispatch")
class SchedulingStockSummary(ModelViewSet):
    queryset = ProductStockDailySummary.objects.filter(stage__in=('HMB', 'CMB', '1MB', '2MB', '3MB')).order_by('product_no')
    serializer_class = ProductStockDailySummarySerializer
    permission_classes = (IsAuthenticated,)
    filter_backends = (DjangoFilterBackend,)
    filter_class = ProductStockDailySummaryFilter

    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        serializer = self.get_serializer(queryset, many=True)
        data = serializer.data
        ret = {}
        for item in data:
            k = item['product_no'] + '-' + item['version']
            if k not in ret:
                ret[k] = {'product_no': item['product_no'],
                          'version': item['version'],
                          'stock_weight_{}'.format(item['stage']): item['stock_weight'],
                          'area_weight_{}'.format(item['stage']): item['area_weight']}
            else:
                ret[k]['stock_weight_{}'.format(item['stage'])] = item['stock_weight']
                ret[k]['area_weight_{}'.format(item['stage'])] = item['area_weight']
        return Response(ret.values())

    def create(self, request, *args, **kwargs):
        data = self.request.data
        factory_date = data.get('factory_date')
        stock_data = data.get('stock_data')
        if not isinstance(stock_data, list):
            raise ValidationError('data error!')
        for stock in stock_data:
            product_no = stock.pop('product_no')
            version = stock.pop('version')
            if not all([product_no, version]):
                raise ValidationError('数据不全！')
            if ProductStockDailySummary.objects.filter(
                    factory_date=factory_date, product_no=product_no, version=version).exists():
                raise ValidationError('该规格库存数据已存在，请勿重复添加！')
            s_data = {'factory_date': factory_date, 'product_no': product_no, 'version': version}
            for key, value in stock.items():
                dt = dict()
                split_data = key.split('_')
                s_data['stage'] = split_data[2]
                dt['_'.join([split_data[0], split_data[1]])] = value
                ProductStockDailySummary.objects.update_or_create(defaults=dt, **s_data)
        return Response('ok')

    @action(methods=['post'], detail=False)
    def confirm(self, request):
        data = self.request.data
        factory_date = data.get('factory_date')
        stock_data = data.get('stock_data')
        if not isinstance(stock_data, list):
            raise ValidationError('data error!')
        ProductStockDailySummary.objects.filter(factory_date=factory_date).delete()
        for stock in stock_data:
            product_no = stock.pop('product_no')
            version = stock.pop('version')
            if not all([product_no, version]):
                raise ValidationError('数据不全！')
            s_data = {'factory_date': factory_date, 'product_no': product_no, 'version': version}
            for key, value in stock.items():
                dt = dict()
                split_data = key.split('_')
                s_data['stage'] = split_data[2]
                dt['_'.join([split_data[0], split_data[1]])] = value
                ProductStockDailySummary.objects.update_or_create(defaults=dt, **s_data)
        return Response('ok')

    @atomic()
    @action(methods=['post'], detail=False, permission_classes=[], url_path='import-xlsx',
            url_name='import-xlsx')
    def import_xlx(self, request):
        factory_date = self.request.data.get('factory_date', datetime.datetime.now().strftime('%Y-%m-%d'))
        date_splits = factory_date.split('-')
        m = date_splits[1] if not date_splits[1].startswith('0') else date_splits[1].lstrip('0')
        d = date_splits[2]
        excel_file = request.FILES.get('file', None)
        if not excel_file:
            raise ValidationError('文件不可为空！')
        if not excel_file.name.split('.')[-1] in ['xls', 'xlsx', 'xlsm']:
            raise ValidationError('文件格式错误,仅支持 xls、xlsx、xlsm文件')
        try:
            data = xlrd.open_workbook(filename=None, file_contents=excel_file.read())
            cur_sheet = data.sheet_by_name(sheet_name='{}.({})'.format(m, d))
        except Exception:
            raise ValidationError('未找到{}.({})库存excel文档！'.format(m, d))
        data = get_sheet_data(cur_sheet, start_row=7)
        for item in data:
            version = item[1]
            product_no = item[2]
            stock_weight_hmb = 0 if not item[4] else item[4]
            stock_weight_cmb = 0 if not item[6] else item[6]
            stock_weight_1mb = 0 if not item[8] else item[8]
            stock_weight_2mb = 0 if not item[10] else item[10]
            stock_weight_3mb = 0 if not item[12] else item[12]
            if not all([version, product_no]):
                continue
            # if stock_weight_hmb:
            dt = {'area_weight': stock_weight_hmb}
            s_data = {'factory_date': factory_date, 'product_no': product_no, 'stage': 'HMB', 'version': version}
            ProductStockDailySummary.objects.update_or_create(defaults=dt, **s_data)
            # if stock_weight_cmb:
            dt = {'area_weight': stock_weight_cmb}
            s_data = {'factory_date': factory_date, 'product_no': product_no, 'stage': 'CMB', 'version': version}
            ProductStockDailySummary.objects.update_or_create(defaults=dt, **s_data)
            # if stock_weight_1mb:
            dt = {'area_weight': stock_weight_1mb}
            s_data = {'factory_date': factory_date, 'product_no': product_no, 'stage': '1MB', 'version': version}
            ProductStockDailySummary.objects.update_or_create(defaults=dt, **s_data)
            # if stock_weight_2mb:
            dt = {'area_weight': stock_weight_2mb}
            s_data = {'factory_date': factory_date, 'product_no': product_no, 'stage': '2MB', 'version': version}
            ProductStockDailySummary.objects.update_or_create(defaults=dt, **s_data)
            # if stock_weight_3mb:
            dt = {'area_weight': stock_weight_3mb}
            s_data = {'factory_date': factory_date, 'product_no': product_no, 'stage': '3MB', 'version': version}
            ProductStockDailySummary.objects.update_or_create(defaults=dt, **s_data)
        return Response('ok')

    @action(methods=['get'], detail=False)
    def export(self, request):
        data = self.request.query_params
        st = data.get('st')
        et = data.get('et')
        queryset = self.get_queryset().filter(factory_date__gte=st, factory_date__lte=et).order_by('factory_date')
        serializer = self.get_serializer(queryset, many=True)
        data = serializer.data
        ret = {}
        for item in data:
            k = item['product_no'] + '-' + item['version']
            if item['factory_date'] not in ret:
                ret[item['factory_date']] = {k: {'product_no': item['product_no'],
                                                 'version': item['version'],
                                                 'stock_weight_{}'.format(item['stage']): item['stock_weight'],
                                                 'area_weight_{}'.format(item['stage']): item['area_weight']}}
            else:
                if k not in ret[item['factory_date']]:
                    ret[item['factory_date']][k] = {'product_no': item['product_no'],
                                                    'version': item['version'],
                                                    'stock_weight_{}'.format(item['stage']): item['stock_weight'],
                                                    'area_weight_{}'.format(item['stage']): item['area_weight']}
                else:
                    ret[item['factory_date']][k]['stock_weight_{}'.format(item['stage'])] = item['stock_weight']
                    ret[item['factory_date']][k]['area_weight_{}'.format(item['stage'])] = item['area_weight']
        if not ret:
            raise ValidationError('时间范围内无数据可以导出')
        bio = BytesIO()
        writer = pd.ExcelWriter(bio, engine='xlsxwriter')  # 注意安装这个包 pip install xlsxwriter
        for k, v in ret.items():
            df = pd.DataFrame(v.values(), columns=['version', 'product_no', 'stock_weight_HMB', 'area_weight_HMB', 'stock_weight_CMB', 'area_weight_CMB', 'stock_weight_1MB', 'area_weight_1MB', 'stock_weight_2MB', 'area_weight_2MB', 'stock_weight_3MB', 'area_weight_3MB'])
            try:
                df = df.rename(columns={'version': '版本号', 'product_no': '规格',
                                        'stock_weight_HMB': 'HMB(库内)', 'area_weight_HMB': 'HMB(现场)',
                                        'stock_weight_CMB': 'CMB(库内)', 'area_weight_CMB': 'CMB(现场)',
                                        'stock_weight_1MB': '1MB(库内)', 'area_weight_1MB': '1MB(现场)',
                                        'stock_weight_2MB': '2MB(库内)', 'area_weight_2MB': '2MB(现场)',
                                        'stock_weight_3MB': '3MB(库内)', 'area_weight_3MB': '3MB(现场)'})
            except:
                pass
            df.to_excel(writer, sheet_name=k, index=False, encoding='SIMPLIFIED CHINESE_CHINA.UTF8')
            worksheet = writer.sheets[k]
            worksheet.set_column(1, 10, 15)
        writer.save()
        bio.seek(0)
        from django.http import FileResponse
        response = FileResponse(bio)
        response['Content-Type'] = 'application/octet-stream'
        response['Content-Disposition'] = 'attachment;filename="mm.xlsx"'
        return response


@method_decorator([api_recorder], name="dispatch")
class SchedulingMaterialDemanded(APIView):

    def get(self, request):
        interval_type = self.request.query_params.get('interval_type', '1')
        filter_material_type = self.request.query_params.get('material_type')
        filter_material_name = self.request.query_params.get('material_name')
        equip_no = self.request.query_params.get('equip_no')

        if datetime.datetime.now().hour < 8:
            now_time_str = datetime.datetime.now().strftime('%Y-%m-%d 08:00:00')
            now_time = datetime.datetime.strptime(now_time_str, "%Y-%m-%d %H:%M:%S")
        else:
            now_time = datetime.datetime.now()

        if interval_type == '1':
            st = now_time
            et = now_time + datetime.timedelta(hours=4)
        elif interval_type == '2':
            st = now_time + datetime.timedelta(hours=4)
            et = now_time + datetime.timedelta(hours=8)
        else:
            st = now_time + datetime.timedelta(hours=8)
            et = now_time + datetime.timedelta(hours=12)

        factory_date = get_current_factory_date().get('factory_date', datetime.datetime.now().date())
        if not factory_date:
            raise ValidationError('参数缺失！')
        last_aps_result = SchedulingResult.objects.filter(factory_date=factory_date).order_by('id').last()
        if not last_aps_result:
            return Response([])

        else:
            if not equip_no:
                equip_nos = Equip.objects.filter(
                    category__equip_type__global_name='密炼设备'
                ).values_list('equip_no', flat=True).order_by('equip_no')
            else:
                equip_nos = [equip_no]
            plan_data = []
            for equip_no in equip_nos:
                scheduling_plans = SchedulingResult.objects.filter(schedule_no=last_aps_result.schedule_no,
                                                                   equip_no=equip_no).order_by('sn')
                plan_start_time = last_aps_result.created_time
                for plan in scheduling_plans:
                    previous_plan_time_consume = SchedulingResult.objects.filter(
                        schedule_no=last_aps_result.schedule_no,
                        equip_no=equip_no,
                        sn__lt=plan.sn).aggregate(s=Sum('time_consume'))['s']
                    previous_plan_time_consume = previous_plan_time_consume if previous_plan_time_consume else 0
                    current_plan_start_time = plan_start_time + datetime.timedelta(hours=previous_plan_time_consume)
                    current_plan_finish_time = plan_start_time + datetime.timedelta(hours=previous_plan_time_consume + plan.time_consume)
                    if current_plan_finish_time <= st:
                        continue
                    if current_plan_start_time <= st < current_plan_finish_time <= et:
                        trains = (plan.time_consume - (st - current_plan_start_time).total_seconds()/3600)/plan.time_consume * plan.plan_trains
                        plan_data.append({'equip_no': equip_no, 'recipe_name': plan.recipe_name, 'plan_trains': int(trains)})
                    elif current_plan_start_time >= st and current_plan_finish_time <= et:
                        plan_data.append({'equip_no': equip_no, 'recipe_name': plan.recipe_name, 'plan_trains': plan.plan_trains})
                    elif st <= current_plan_start_time <= et < current_plan_finish_time:
                        trains = (plan.time_consume - (current_plan_finish_time - et).total_seconds() / 3600) / plan.time_consume * plan.plan_trains
                        plan_data.append(
                            {'equip_no': equip_no, 'recipe_name': plan.recipe_name, 'plan_trains': int(trains)})
                    elif st >= current_plan_start_time and et <= current_plan_finish_time:
                        trains = int((et - st).total_seconds() / 3600 / plan.time_consume * plan.plan_trains)
                        plan_data.append({'equip_no': equip_no, 'recipe_name': plan.recipe_name, 'plan_trains': trains})
            ret = {}
            stages = list(GlobalCode.objects.filter(use_flag=True, global_type__use_flag=True,
                                                    global_type__type_name="胶料段次").values_list("global_name", flat=True))

            for instance in plan_data:
                if instance['equip_no'] == 'Z04':
                    equip_type = Equip.objects.filter(equip_no=instance['equip_no']).first().category
                    pb = ProductBatching.objects.exclude(used_type=6).filter(batching_type=2,
                                                                             stage_product_batch_no=instance['recipe_name'],
                                                                             dev_type=equip_type).first()
                    if not pb:
                        continue
                    details = pb.batching_details.filter(
                        type=1, delete_flag=False).exclude(
                        Q(material__material_name__in=('细料', '硫磺')) | Q(material__material_type__global_name__in=stages)
                    ).values('material__material_name', 'material__material_type__global_name', 'actual_weight')
                else:
                    pb = ProductBatching.objects.using('SFJ').exclude(used_type=6).filter(
                        batching_type=1,
                        stage_product_batch_no=instance['recipe_name'],
                        equip__equip_no=instance['equip_no']).first()
                    if not pb:
                        continue
                    details = pb.batching_details.using('SFJ').filter(
                        type=1, delete_flag=False).exclude(
                        Q(material__material_name__in=('细料', '硫磺')) | Q(material__material_type__global_name__in=stages)
                    ).values('material__material_name', 'material__material_type__global_name', 'actual_weight')
                for item in details:
                    material_name = item['material__material_name'].rstrip('-C').rstrip('-X')
                    weight = item['actual_weight'] * instance['plan_trains']
                    material_type = item['material__material_type__global_name']
                    if material_name not in ret:
                        ret[material_name] = {
                            'material_type': material_type,
                            'material_name': material_name,
                            instance['equip_no']: weight,
                            'total_weight': weight,
                            'detail': [{'equip_no': instance['equip_no'],
                                        'recipe_name': instance['recipe_name'],
                                        'material_type': material_type,
                                        'material_name': material_name,
                                        'weight': weight
                                        }]
                        }
                    else:
                        if instance['equip_no'] in ret[material_name]:
                            ret[material_name][instance['equip_no']] += weight
                        else:
                            ret[material_name][instance['equip_no']] = weight
                        ret[material_name]['total_weight'] += weight
                        ret[material_name]['detail'].append({'equip_no': instance['equip_no'],
                                                             'recipe_name': instance['recipe_name'],
                                                             'material_type': material_type,
                                                             'material_name': material_name,
                                                             'weight': weight
                                                             })
            data = ret.values()
            if filter_material_type:
                data = filter(lambda x: filter_material_type in x['material_type'], data)
            if filter_material_name:
                data = filter(lambda x: filter_material_name in x['material_name'], data)
            data = sorted(data, key=itemgetter('material_type', 'material_name'))  # 按多个字段排序
            return Response({'st': st.strftime('%H: %M'), 'et': et.strftime('%H: %M'), 'data': data})


@method_decorator([api_recorder], name="dispatch")
class RecipeStages(APIView):

    def get(self, request):
        product_no = self.request.query_params.get('product_no')
        version = self.request.query_params.get('version')
        if not all([product_no, version]):
            raise ValidationError('bad request')
        pbs = ProductBatching.objects.using('SFJ').exclude(used_type=6).filter(
            stage_product_batch_no__icontains='-{}-{}'.format(product_no, version),
            stage__global_name__in=('HMB', 'CMB', '1MB', '2MB', '3MB', '4MB', 'FM')
        ).values_list('stage__global_name', flat=True)
        idx_keys = {'HMB': 1, 'CMB': 2, '1MB': 3, '2MB': 4, '3MB': 5, '4MB': 6, 'FM': 7}
        pbs = sorted(list(set(pbs)), key=lambda d: idx_keys[d])
        return Response(pbs)


@method_decorator([api_recorder], name="dispatch")
class MaterialPlanConsumeView(APIView):

    def get(self, request):
        filter_material_type = self.request.query_params.get('material_type')
        material_name = self.request.query_params.get('material_name')
        equip_no = self.request.query_params.get('equip_no')
        product_no = self.request.query_params.get('product_no')
        factory_date = get_current_factory_date().get('factory_date', datetime.datetime.now().date())
        filter_kwargs = {}
        filter_kwargs2 = {}
        if equip_no:
            filter_kwargs['equip__equip_no'] = equip_no.strip()
        if product_no:
            filter_kwargs['product_batching__stage_product_batch_no'] = product_no.strip()
        if material_name:
            filter_kwargs2['material__material_name'] = material_name.strip()
        if filter_material_type:
            filter_kwargs2['material__material_type__global_name'] = filter_material_type.strip()
        plan_date = ProductClassesPlan.objects.filter(**filter_kwargs).filter(
            work_schedule_plan__plan_schedule__day_time=factory_date
        ).values('equip__equip_no',
                 'product_batching__stage_product_batch_no',
                 'equip__category_id').annotate(s=Sum('plan_trains'))

        ret = []
        material_weight_dict = {}
        for item in plan_date:
            equip_no = item['equip__equip_no']
            recipe_no = item['product_batching__stage_product_batch_no']
            dev_type_id = item['equip__category_id']
            trains = item['s']
            mes_pb = ProductBatching.objects.exclude(used_type=6).filter(batching_type=2,
                                                                         stage_product_batch_no=recipe_no,
                                                                         dev_type_id=dev_type_id).first()
            if mes_pb:
                weight_details = WeighBatchingDetail.objects.filter(**filter_kwargs2).filter(
                    weigh_cnt_type__product_batching=mes_pb,
                    delete_flag=False).values('material__material_name',
                                              'material__material_type__global_name',
                                              'standard_weight')
                for detail in weight_details:
                    m_name = detail['material__material_name']
                    weight = detail['standard_weight']*trains
                    ret.append(
                        {'material_type': detail['material__material_type__global_name'],
                         'material_name': m_name,
                         'equip_no': equip_no,
                         'product_no': recipe_no,
                         'total_weight': weight}
                    )
                    material_weight_dict[m_name] = material_weight_dict.get(m_name, 0) + weight
        result = sorted(ret, key=itemgetter('material_type', 'material_name', 'equip_no'))  #  按多个字段排序
        return Response({'result': result, 'material_weight_dict': material_weight_dict})


@method_decorator([api_recorder], name='dispatch')
class WeightPackageDailyTimeConsumeViewSet(ModelViewSet):
    queryset = WeightPackageDailyTimeConsume.objects.order_by('product_no')
    serializer_class = WeightPackageDailyTimeConsumeSerializer
    permission_classes = (IsAuthenticated,)
    pagination_class = None
    filter_fields = ('factory_date', 'product_no', 'product_type')

    @action(methods=['get'], detail=False, permission_classes=[], url_path='find-recipe',
            url_name='find-recipe')
    def find_recipe(self, request):
        product_no = self.request.query_params.get('product_no')  # 胶料代码
        mixin_dev_type = self.request.query_params.get('mixin_dev_type')  # 混炼机型
        final_dev_type = self.request.query_params.get('final_dev_type')  # 终炼机型
        plan_weight = float(self.request.query_params.get('plan_weight'))  # 计划吨数
        # 混炼配方
        mixin_pd = ProductBatching.objects.filter(batching_type=2,
                                                  stage_product_batch_no__contains='-1MB-{}-'.format(product_no),
                                                  used_type=4,
                                                  dev_type__category_no=mixin_dev_type).first()
        # 终炼配方
        final_pd = ProductBatching.objects.filter(batching_type=2,
                                                  stage_product_batch_no__contains='-FM-{}-'.format(product_no),
                                                  used_type=4,
                                                  dev_type__category_no=final_dev_type).first()
        if not mixin_pd:
            raise ValidationError('未找到该胶料启用状态混炼配方数据！')
        if not final_pd:
            raise ValidationError('未找到该胶料启用状态终炼配方数据！')
        # 混炼车次重量
        mixin_weight = float(mixin_pd.batching_weight / 1000)
        stages = list(GlobalCode.objects.filter(global_type__type_name='胶料段次').values_list('global_name', flat=True))
        final_devoted_materials = final_pd.batching_details.filter(
            material__material_type__global_name__in=stages,
            type=1, material__material_no__icontains='-{}-'.format(product_no)).first()
        xl_dp_flag = False  # 细料是否单配
        lh_dp_flag = False  # 终炼是否单配
        mixin_chemical_kind = 0  # 混炼单配化工数量
        final_chemical_kind = 0  # 终炼单配化工数量
        hl_materials = ''  # 合练塑解剂物料及重量
        mixin_materials = ''  # 混炼单配物料及重量
        final_materials = ''  # 混炼单配物料及重量
        xl_split_qty = 1  # 细料分包数量
        lh_split_qty = 1  # 硫磺分包数量
        aw_qty = None  # AW数量
        if final_devoted_materials:
            final_devoted_weight = float(final_devoted_materials.actual_weight / 1000)
        else:
            raise ValidationError('未找到终炼配方投入前段次规格重量！')
        mixin_cl_recipe_name = '{}({})'.format(mixin_pd.stage_product_batch_no, mixin_dev_type)
        for equip_no in Equip.objects.filter(equip_no__startswith='F').values_list('equip_no', flat=True):
            recipe_model = JZRecipePre if equip_no in JZ_EQUIP_NO else RecipePre
            material_detail_model = JZRecipeMaterial if equip_no in JZ_EQUIP_NO else RecipeMaterial
            xl_recipe = recipe_model.objects.using(equip_no).filter(
                name=mixin_cl_recipe_name).first()
            if xl_recipe:
                xl_split_qty = 1 if not xl_recipe.split_count else xl_recipe.split_count
                xl_material_details = dict(material_detail_model.objects.using(equip_no).filter(
                    recipe_name=mixin_cl_recipe_name).values_list('name', 'weight'))
                mes_material_details = dict(WeighBatchingDetail.objects.filter(
                    weigh_cnt_type__product_batching=mixin_pd,
                    delete_flag=False).values_list('material__material_name', 'standard_weight'))
                if len(xl_material_details) != len(mes_material_details):
                    xl_dp_flag = True
                    for k, v in mes_material_details.items():
                        k = k.rstrip('-C').rstrip('-X')
                        if k not in xl_material_details:
                            mixin_chemical_kind += 1
                            mixin_materials += '{}:{}kg;'.format(k, v)
                break

        final_cl_recipe_name = '{}({})'.format(final_pd.stage_product_batch_no, final_dev_type)
        for equip_no in Equip.objects.filter(equip_no__startswith='S').values_list('equip_no', flat=True):
            lh_recipe = RecipePre.objects.using(equip_no).filter(
                name=final_cl_recipe_name).first()
            if lh_recipe:
                lh_split_qty = 1 if not lh_recipe.split_count else lh_recipe.split_count
                lh_material_details = dict(RecipeMaterial.objects.using(equip_no).filter(
                    recipe_name=final_cl_recipe_name).values_list('name', 'weight'))
                mes_material_details = dict(WeighBatchingDetail.objects.filter(
                    weigh_cnt_type__product_batching=final_pd,
                    delete_flag=False).values_list('material__material_name', 'standard_weight'))
                if len(lh_material_details) != len(mes_material_details):
                    lh_dp_flag = True
                    for k, v in mes_material_details.items():
                        k = k.rstrip('-C').rstrip('-X')
                        if k not in lh_material_details:
                            final_chemical_kind += 1
                            final_materials += '{}:{}kg;'.format(k, v)
                break
        xl_plan_qty = int(plan_weight / mixin_weight * xl_split_qty)  # 细料计划包数
        xl_dp_qty = None if not xl_dp_flag else xl_plan_qty  # 细料单配包数
        hl_pd = ProductBatching.objects.filter(batching_type=2,
                                               used_type=4,
                                               dev_type__category_no=mixin_dev_type
                                               ).filter(Q(stage_product_batch_no__contains='-HMB-{}-'.format(product_no)) |
                                                        Q(stage_product_batch_no__contains='-CMB-{}-'.format(product_no))).first()
        if hl_pd:
            hl_materials_data = list(WeighBatchingDetail.objects.filter(
                weigh_cnt_type__product_batching=hl_pd,
                delete_flag=False,
                material__material_name__icontains='塑解剂'
            ).values_list('material__material_name', 'standard_weight'))
            for i in hl_materials_data:
                hl_materials += '{}:{}kg'.format(i[0].rstrip('-C').rstrip('-X'), i[1])
        lh_plan_qty = int(plan_weight / final_devoted_weight * lh_split_qty)  # 硫磺计划包数
        lh_dp_qty = None if not lh_dp_flag else lh_plan_qty  # 硫磺单配包数
        manual_instance = WeightPackageManual.objects.filter(
            product_no=mixin_cl_recipe_name, dev_type=mixin_dev_type).order_by('id').last()
        if manual_instance:
            if manual_instance.package_details.filter(material_name__icontains='AW').exists():
                aw_qty = xl_split_qty

        return Response({'mixin_weight': mixin_weight, 'final_devoted_weight': final_devoted_weight,
                         'xl_split_qty': xl_split_qty, 'lh_split_qty': lh_split_qty,
                         'xl_plan_qty': xl_plan_qty, 'xl_dp_qty': xl_dp_qty,
                         'mixin_chemical_kind': mixin_chemical_kind, 'hl_materials': hl_materials,
                         'mixin_materials': mixin_materials, 'lh_plan_qty': lh_plan_qty,
                         'lh_dp_qty': lh_dp_qty, 'final_materials': final_materials,
                         'final_chemical_kind': final_chemical_kind, "product_no": product_no,
                         'mixin_dev_type': mixin_dev_type, 'final_dev_type': final_dev_type,
                         'plan_weight': plan_weight, 'aw_qty': aw_qty})

    @action(methods=['post'], detail=False, permission_classes=[], url_path='import-xlsx',
            url_name='import-xlsx')
    def import_xlsx(self, request):
        excel_file = request.FILES.get('file', None)
        factory_date = request.data.get('factory_date')
        if not excel_file:
            raise ValidationError('文件不可为空！')
        file_name = excel_file.name
        if not file_name.split('.')[-1] in ['xls', 'xlsx', 'xlsm']:
            raise ValidationError('文件格式错误,仅支持 xls、xlsx、xlsm文件')
        try:
            data = xlrd.open_workbook(filename=None, file_contents=excel_file.read())
            sheet0 = data.sheets()[0]
            sheet1 = data.sheets()[1]
        except Exception:
            raise ValidationError('打开文件错误')
        if sheet0.ncols != 7 and sheet0.ncols != 7:
            raise ValidationError('导入文件数据错误！')
        data0 = get_sheet_data(sheet0)
        data1 = get_sheet_data(sheet1)
        data = {1: data0, 2: data1}
        instance_list = []
        stages = list(
            GlobalCode.objects.filter(global_type__type_name='胶料段次').values_list('global_name', flat=True))
        for product_type, row_data in data.items():
            for i in row_data:
                try:
                    product_no = i[0].strip()
                    plan_weight = float(i[1])
                    mixin_dev_type = i[2].strip()
                    final_dev_type = i[3].strip()
                    hl_time_consume = None if not i[4] else float(i[4])
                    mixin_time_consume = None if not i[5] else float(i[5])
                    final_time_consume = None if not i[6] else float(i[6])
                except Exception:
                    raise ValidationError('导入数据错误，修改后重试！')
                # 混炼配方
                mixin_pd = ProductBatching.objects.filter(batching_type=2,
                                                          stage_product_batch_no__contains='-1MB-{}-'.format(
                                                              product_no),
                                                          used_type=4,
                                                          dev_type__category_no=mixin_dev_type).first()
                # 终炼配方
                final_pd = ProductBatching.objects.filter(batching_type=2,
                                                          stage_product_batch_no__contains='-FM-{}-'.format(product_no),
                                                          used_type=4,
                                                          dev_type__category_no=final_dev_type).first()
                if not mixin_pd:
                    raise ValidationError('未找到该胶料启用状态混炼配方数据：{}！'.format(product_no))
                if not final_pd:
                    raise ValidationError('未找到该胶料启用状态终炼配方数据：{}！'.format(product_no))
                # 混炼车次重量
                mixin_weight = float(mixin_pd.batching_weight / 1000)
                final_devoted_materials = final_pd.batching_details.filter(
                    material__material_type__global_name__in=stages,
                    type=1, material__material_no__icontains='-{}-'.format(product_no)).first()
                xl_dp_flag = False  # 细料是否单配
                lh_dp_flag = False  # 终炼是否单配
                mixin_chemical_kind = 0  # 混炼单配化工数量
                final_chemical_kind = 0  # 终炼单配化工数量
                hl_materials = ''  # 合练塑解剂物料及重量
                mixin_materials = ''  # 混炼单配物料及重量
                final_materials = ''  # 混炼单配物料及重量
                xl_split_qty = 1  # 细料分包数量
                lh_split_qty = 1  # 硫磺分包数量
                aw_qty = None  # AW数量
                if final_devoted_materials:
                    final_devoted_weight = float(final_devoted_materials.actual_weight / 1000)
                else:
                    raise ValidationError('未找到终炼配方投入前段次规格重量！')
                mixin_cl_recipe_name = '{}({})'.format(mixin_pd.stage_product_batch_no, mixin_dev_type)
                for equip_no in Equip.objects.filter(equip_no__startswith='F').values_list('equip_no', flat=True):
                    recipe_model = JZRecipePre if equip_no in JZ_EQUIP_NO else RecipePre
                    material_detail_model = JZRecipeMaterial if equip_no in JZ_EQUIP_NO else RecipeMaterial
                    xl_recipe = recipe_model.objects.using(equip_no).filter(
                        name=mixin_cl_recipe_name).first()
                    if xl_recipe:
                        xl_split_qty = 1 if not xl_recipe.split_count else xl_recipe.split_count
                        xl_material_details = dict(material_detail_model.objects.using(equip_no).filter(
                            recipe_name=mixin_cl_recipe_name).values_list('name', 'weight'))
                        mes_material_details = dict(WeighBatchingDetail.objects.filter(
                            weigh_cnt_type__product_batching=mixin_pd,
                            delete_flag=False).values_list('material__material_name', 'standard_weight'))
                        if len(xl_material_details) != len(mes_material_details):
                            xl_dp_flag = True
                            for k, v in mes_material_details.items():
                                k = k.rstrip('-C').rstrip('-X')
                                if k not in xl_material_details:
                                    mixin_chemical_kind += 1
                                    mixin_materials += '{}:{}kg;'.format(k, v)
                        break

                final_cl_recipe_name = '{}({})'.format(final_pd.stage_product_batch_no, final_dev_type)
                for equip_no in Equip.objects.filter(equip_no__startswith='S').values_list('equip_no', flat=True):
                    lh_recipe = RecipePre.objects.using(equip_no).filter(
                        name=final_cl_recipe_name).first()
                    if lh_recipe:
                        lh_split_qty = 1 if not lh_recipe.split_count else lh_recipe.split_count
                        lh_material_details = dict(RecipeMaterial.objects.using(equip_no).filter(
                            recipe_name=final_cl_recipe_name).values_list('name', 'weight'))
                        mes_material_details = dict(WeighBatchingDetail.objects.filter(
                            weigh_cnt_type__product_batching=final_pd,
                            delete_flag=False).values_list('material__material_name', 'standard_weight'))
                        if len(lh_material_details) != len(mes_material_details):
                            lh_dp_flag = True
                            for k, v in mes_material_details.items():
                                k = k.rstrip('-C').rstrip('-X')
                                if k not in lh_material_details:
                                    final_chemical_kind += 1
                                    final_materials += '{}:{}kg;'.format(k, v)
                        break
                xl_plan_qty = int(plan_weight / mixin_weight * xl_split_qty)  # 细料计划包数
                xl_dp_qty = None if not xl_dp_flag else xl_plan_qty  # 细料单配包数
                hl_pd = ProductBatching.objects.filter(batching_type=2,
                                                       used_type=4,
                                                       dev_type__category_no=mixin_dev_type
                                                       ).filter(
                    Q(stage_product_batch_no__contains='-HMB-{}-'.format(product_no)) |
                    Q(stage_product_batch_no__contains='-CMB-{}-'.format(product_no))).first()
                if hl_pd:
                    hl_materials_data = list(WeighBatchingDetail.objects.filter(
                        weigh_cnt_type__product_batching=hl_pd,
                        delete_flag=False,
                        material__material_name__icontains='塑解剂'
                    ).values_list('material__material_name', 'standard_weight'))
                    for j in hl_materials_data:
                        hl_materials += '{}:{}kg'.format(j[0].rstrip('-C').rstrip('-X'), j[1])
                lh_plan_qty = int(plan_weight / final_devoted_weight * lh_split_qty)  # 硫磺计划包数
                lh_dp_qty = None if not lh_dp_flag else lh_plan_qty  # 硫磺单配包数
                manual_instance = WeightPackageManual.objects.filter(
                    product_no=mixin_cl_recipe_name, dev_type=mixin_dev_type).order_by('id').last()
                if manual_instance:
                    if manual_instance.package_details.filter(material_name__icontains='AW').exists():
                        aw_qty = xl_split_qty
                instance_list.append(WeightPackageDailyTimeConsume(**{
                    'mixin_weight': mixin_weight, 'final_devoted_weight': final_devoted_weight,
                    'xl_split_qty': xl_split_qty, 'lh_split_qty': lh_split_qty,
                    'xl_plan_qty': xl_plan_qty, 'xl_dp_qty': xl_dp_qty,
                    'mixin_chemical_kind': mixin_chemical_kind, 'hl_materials': hl_materials,
                    'mixin_materials': mixin_materials, 'lh_plan_qty': lh_plan_qty,
                    'lh_dp_qty': lh_dp_qty, 'final_materials': final_materials,
                    'final_chemical_kind': final_chemical_kind, "product_no": product_no,
                    'mixin_dev_type': mixin_dev_type, 'final_dev_type': final_dev_type,
                    'plan_weight': plan_weight, 'aw_qty': aw_qty, 'factory_date': factory_date,
                    'product_type': product_type, 'hl_time_consume': hl_time_consume,
                    'mixin_time_consume': mixin_time_consume, 'final_time_consume': final_time_consume}))
        WeightPackageDailyTimeConsume.objects.bulk_create(instance_list)
        return Response('ok')


@method_decorator([api_recorder], name='dispatch')
class APSExportDataView(APIView):

    def extend_last_aps_result(self, factory_date, lock_durations):
        """
        继承前一天未打完的排程计划
        @param lock_durations: 锁定时间
        @param factory_date: 日期
        @return:
        """
        yesterday = factory_date - datetime.timedelta(days=1)
        equip_plan_data = []
        equip_end_time_dict = {}
        now_time = datetime.datetime.now()
        aps_st_time = datetime.datetime(
            year=factory_date.year,
            month=factory_date.month,
            day=factory_date.day,
            hour=8, minute=0, second=0)
        # 当天已经生产的计划，状态为commited
        started_plans = TrainsFeedbacks.objects.exclude(operation_user='Mixer2').filter(
            factory_date=factory_date
        ).values('plan_classes_uid').annotate(st=Min('begin_time'), et=Max('end_time')).order_by('st')
        started_plan_classes_uid = []
        for p in started_plans:
            started_plan_classes_uid.append(p['plan_classes_uid'])
            plan = ProductClassesPlan.objects.using('SFJ').filter(plan_classes_uid=p['plan_classes_uid']).first()
            if not plan:
                continue
            weight = plan.product_batching.batching_weight * plan.plan_trains
            equip_no = plan.equip.equip_no
            recipe_name = plan.product_batching.stage_product_batch_no
            st, et = p['st'], p['et']
            if not st:
                continue
            if et <= aps_st_time:  # 八点之前完成的计划不需要
                continue
            if plan.status in ('完成', '停止'):
                time_consume = round((et - st).total_seconds() / 60, 2)
            else:
                time_consume = round(
                    calculate_equip_recipe_avg_mixin_time(
                        equip_no,
                        recipe_name
                    ) * plan.plan_trains / 60, 2)
                et = st + datetime.timedelta(minutes=int(time_consume))
            if equip_no in equip_end_time_dict:
                if equip_end_time_dict[equip_no] < et:
                    equip_end_time_dict[equip_no] = et
            else:
                equip_end_time_dict[equip_no] = et
            begin_time = round((st - aps_st_time).total_seconds() / 60, 2)
            tc = time_consume + begin_time if begin_time < 0 else time_consume
            equip_plan_data.append({'recipe_name': recipe_name,
                                    'equip_no': equip_no,
                                    'plan_trains': plan.plan_trains,
                                    'time_consume': 0 if tc < 0 else tc,
                                    'status': 'COMMITED',
                                    # 'delivery_time': time_consume + begin_time if begin_time < 0 else time_consume,
                                    'begin_time': 0 if begin_time < 0 else begin_time,
                                    'weight': weight
                                    })

        # 当天已新建，未生产的计划，状态为commited
        for plan in ProductClassesPlan.objects.using('SFJ').exclude(
                plan_classes_uid__in=started_plan_classes_uid).filter(
                work_schedule_plan__plan_schedule__day_time=factory_date,
                delete_flag=False).order_by('id'):
            recipe_name = plan.product_batching.stage_product_batch_no
            weight = plan.product_batching.batching_weight * plan.plan_trains
            equip_no = plan.equip.equip_no
            time_consume = round(
                calculate_equip_recipe_avg_mixin_time(
                    equip_no,
                    plan.product_batching.stage_product_batch_no
                ) * plan.plan_trains / 60, 2)
            if equip_end_time_dict.get(equip_no):
                tt = equip_end_time_dict[equip_no]
            else:
                tt = plan.created_date
            equip_end_time_dict[equip_no] = tt + datetime.timedelta(minutes=int(time_consume))
            begin_time = round((tt - aps_st_time).total_seconds() / 60, 2)
            tc = time_consume + begin_time if begin_time < 0 else time_consume
            if tc <= 0:  # 八点之前完成的计划不需要
                continue
            equip_plan_data.append({'recipe_name': recipe_name,
                                    'equip_no': equip_no,
                                    'plan_trains': plan.plan_trains,
                                    'time_consume': tc,
                                    'status': 'COMMITED',
                                    # 'delivery_time': time_consume + begin_time if begin_time < 0 else time_consume,
                                    'begin_time': 0 if begin_time < 0 else begin_time,
                                    'weight': weight
                                    })

        # 前一天未下达的计划，锁定时间以内的状态为COMMITED，其他为STANDARD
        yesterday_last_res = SchedulingResult.objects.filter(
            factory_date=yesterday).order_by('id').last()
        if yesterday_last_res:
            equip_time_dict = {}
            last_aps_results = SchedulingResult.objects.filter(
                schedule_no=yesterday_last_res.schedule_no,
                status='未下发').order_by('equip_no', 'sn').values()
            for result in last_aps_results:
                equip_no = result['equip_no']
                recipe_name = result['recipe_name']
                plan_trains = result['plan_trains']
                time_consume = result['time_consume']
                equip_time_consume = equip_time_dict.get(equip_no, 0)
                pb = ProductBatching.objects.using('SFJ').filter(
                    stage_product_batch_no=recipe_name, equip__equip_no=equip_no).order_by('id').last()
                if pb:
                    weight = pb.batching_weight * plan_trains
                else:
                    weight = 6.6
                if equip_end_time_dict.get(equip_no):
                    tt = equip_end_time_dict[equip_no]
                else:
                    tt = now_time
                # 锁定时间范围之内的为committed
                locked_end_time = now_time + datetime.timedelta(hours=float(lock_durations))
                if lock_durations != 0 and equip_time_consume <= lock_durations and tt < locked_end_time:
                    equip_end_time_dict[equip_no] = tt + datetime.timedelta(hours=time_consume)
                    begin_time = round((tt - aps_st_time).total_seconds() / 60, 2)
                    equip_plan_data.append({'recipe_name': recipe_name,
                                            'equip_no': equip_no,
                                            'plan_trains': plan_trains,
                                            # 'delivery_time': delivery_time * 60,
                                            'time_consume': time_consume * 60,
                                            'status': 'COMMITED',
                                            'begin_time': 0 if begin_time < 0 else begin_time,
                                            'weight': weight
                                            })
                else:
                    equip_plan_data.append({'recipe_name': recipe_name,
                                            'equip_no': equip_no,
                                            'plan_trains': plan_trains,
                                            'delivery_time': time_consume * 60,
                                            'time_consume': time_consume * 60,
                                            'status': 'STANDARD',
                                            'weight': weight
                                            })
                equip_time_dict[equip_no] = equip_time_consume + time_consume
        return equip_plan_data

    def get(self, request):
        aps_start_time = self.request.query_params.get('start_time')
        factory_date = self.request.query_params.get('factory_date')
        # factory_date = get_current_factory_date().get('factory_date', datetime.datetime.now().date())
        wb = load_workbook('xlsx_template/aps_import_templete.xlsx')
        if not aps_start_time:
            aps_start_time = factory_date + ' 08:00:00'
            # aps_start_time = factory_date.strftime('%Y-%m-%d') + ' 08:00:00'
        sps = SchedulingParamsSetting.objects.first()

        demanded_data = SchedulingProductDemandedDeclareSummary.objects.filter(
            factory_date=factory_date, demanded_weight__gt=0).order_by('available_time')
        equip_stop_plan = SchedulingEquipShutDownPlan.objects.filter(
            begin_time__gte=aps_start_time)
        # summary info sheet
        sheet = wb.worksheets[0]
        sheet.cell(1, 2).value = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')  # 当前时间
        sheet.cell(3, 2).value = 'APS1{}'.format(datetime.datetime.now().strftime('%Y%m%d%H%M%S'))  # 排程编号
        sheet.cell(4, 2).value = Equip.objects.filter(category__equip_type__global_name="密炼设备").count()  # 机台数量
        sheet.cell(5, 2).value = aps_start_time  # 排程开始时间
        sheet.cell(6, 2).value = sps.scheduling_during_time  # 排程持续时间
        sheet.cell(7, 2).value = len(set(equip_stop_plan.values_list('equip_no', flat=True)))  # 停机机台数量
        sheet.cell(9, 2).value = sps.lock_durations  # 排程时间点之后的锁定计划期间
        # project list sheet
        sheet1 = wb.worksheets[1]
        sheet2 = wb.worksheets[2]
        job_list_data = {}  # 待排程结果数据 格式：{'C590-01': {'1MB': {},'2MB': {}}}
        pb_available_time_dict = {}  # 规格可用时间 格式： {'C590-01'：20, 'J290-01'：20}

        # 构建job_list_data数据
        for i in demanded_data:
            pd_ms = SchedulingRecipeMachineSetting.objects.filter(
                product_no=i.product_no, version=i.version).first()
            if not pd_ms:
                raise ValidationError('未找到该规格：{}-{}定机表数据！'.format(i.product_no, i.version))
            if not pd_ms.confirmed:
                raise ValidationError('该规格：{}-{}定机表数据待确认！'.format(i.product_no, i.version))
            pd_stages = pd_ms.stages.split('/')
            need_stages = copy.deepcopy(pd_stages)
            pb_version_name = '{}-{}'.format(pd_ms.product_no, pd_ms.version)
            pb_time_consume = 0
            available_time = int(i.available_time * 60 * 24)  # 最晚完成时间
            # pb_available_time_dict[pb_version_name] = 720 if i.available_time == 0 else int(i.available_time * 60 * 24)  # 最晚完成时间
            equip_nos = set()
            for ps in pd_stages:
                main_machine = getattr(pd_ms, 'main_machine_{}'.format(ps))
                vice_machines = getattr(pd_ms, 'vice_machine_{}'.format(ps))
                equip_nos.add(main_machine)
                if vice_machines:
                    for e in vice_machines.split('/'):
                        equip_nos.add(e)

            # write job list sheet
            pbs = ProductBatching.objects.using('SFJ').filter(
                used_type=4,
                stage_product_batch_no__iendswith='-{}'.format(pb_version_name),
                equip__equip_no__in=list(equip_nos)
            ).values('id', 'batching_weight', 'equip__equip_no', 'stage_product_batch_no',
                     'equip__category__category_name', 'stage__global_name').order_by('batching_weight')
            stage_devoted_weight = {}
            weight_qty = i.demanded_weight * 1000
            stock_trans_weight = convert_fm_weight(i.product_no, i.version, factory_date) * 1000
            if stock_trans_weight <= weight_qty <= stock_trans_weight * 1.15:
                weight_qty = stock_trans_weight
            aps_fm_weight = weight_qty
            # 计算该规格每个段次所投入的重量
            for s in pd_stages[1:][::-1]:
                stage_recipes = list(
                    filter(lambda x: x['stage_product_batch_no'].endswith('-{}-{}'.format(s, pb_version_name)), pbs))
                if not stage_recipes:
                    raise ValidationError('未找到该规格{}：段次：{}配方数据！'.format(pb_version_name, s))
                stage_recipe = stage_recipes[-1]  # 取该段次配方重量最大的一条
                c_pb = ProductBatchingDetail.objects.using('SFJ').filter(
                    product_batching=stage_recipe['id'],
                    delete_flag=False,
                    material__material_no__iendswith='-{}'.format(pb_version_name)).first()
                if not c_pb:
                    raise ValidationError('该规格{}：段次：{}配方错误，未找到投入前段次胶料数据！'.format(pb_version_name, s))
                devoted_weight = float(c_pb.actual_weight)  # 投入前段次重量
                prev_stage = c_pb.material.material_no.split('-')[1]  # 前段次名称
                stock_weight = calculate_product_stock(factory_date, pd_ms.product_no, prev_stage)  # 库存重量
                pd_trains = weight_qty / float(stage_recipe['batching_weight'])
                prev_need_weight = pd_trains * devoted_weight - stock_weight
                # print(pd_trains, devoted_weight, stock_weight)
                if prev_need_weight <= 0:
                    weight_qty = 0
                    stage_devoted_weight[prev_stage] = 0
                    need_stages.remove(prev_stage)
                else:
                    weight_qty = prev_need_weight
                    stage_devoted_weight[prev_stage] = weight_qty
            # print(stage_devoted_weight)
            idx_keys = {'HMB': 1, 'CMB': 2, '1MB': 3, '2MB': 4, '3MB': 5, '4MB': 6, 'FM': 7}
            pbs = sorted(list(pbs), key=lambda x: idx_keys.get(x['stage__global_name'], 999))
            for pb in pbs:
                recipe_name = pb['stage_product_batch_no']  # 配方名称
                stage = recipe_name.split('-')[1]  # 段次
                equip_no = pb['equip__equip_no']  # 机台
                batching_weight = pb['batching_weight']  # 配方重量
                if stage not in need_stages:
                    continue
                if stage == 'HMB':
                    # model_size = 1 if not pd_ms.vice_machine_HMB else len(pd_ms.vice_machine_HMB.split('/')) + 1
                    if equip_no == pd_ms.main_machine_HMB:
                        equip_main_standby = '主'
                    elif equip_no in pd_ms.vice_machine_HMB.split('/'):
                        equip_main_standby = '辅'
                    else:
                        continue
                elif stage == 'CMB':
                    # model_size = 1 if not pd_ms.vice_machine_CMB else len(pd_ms.vice_machine_CMB.split('/')) + 1
                    if equip_no == pd_ms.main_machine_CMB:
                        equip_main_standby = '主'
                    elif equip_no in pd_ms.vice_machine_CMB.split('/'):
                        equip_main_standby = '辅'
                    else:
                        continue
                elif stage == '1MB':
                    # model_size = 1 if not pd_ms.vice_machine_1MB else len(pd_ms.vice_machine_1MB.split('/')) + 1
                    if equip_no == pd_ms.main_machine_1MB:
                        equip_main_standby = '主'
                    elif equip_no in pd_ms.vice_machine_1MB.split('/'):
                        equip_main_standby = '辅'
                    else:
                        continue
                elif stage == '2MB':
                    # model_size = 1 if not pd_ms.vice_machine_2MB else len(pd_ms.vice_machine_2MB.split('/')) + 1
                    if equip_no == pd_ms.main_machine_2MB:
                        equip_main_standby = '主'
                    elif equip_no in pd_ms.vice_machine_2MB.split('/'):
                        equip_main_standby = '辅'
                    else:
                        continue
                elif stage == '3MB':
                    # model_size = 1 if not pd_ms.vice_machine_3MB else len(pd_ms.vice_machine_3MB.split('/')) + 1
                    if equip_no == pd_ms.main_machine_3MB:
                        equip_main_standby = '主'
                    elif equip_no in pd_ms.vice_machine_3MB.split('/'):
                        equip_main_standby = '辅'
                    else:
                        continue
                elif stage == '4MB':
                    # model_size = 1 if not pd_ms.vice_machine_4MB else len(pd_ms.vice_machine_4MB.split('/')) + 1
                    if equip_no == pd_ms.main_machine_4MB:
                        equip_main_standby = '主'
                    elif equip_no in pd_ms.vice_machine_4MB.split('/'):
                        equip_main_standby = '辅'
                    else:
                        continue
                else:
                    # model_size = 1 if not pd_ms.vice_machine_FM else len(pd_ms.vice_machine_FM.split('/')) + 1
                    if equip_no == pd_ms.main_machine_FM:
                        equip_main_standby = '主'
                    elif equip_no in pd_ms.vice_machine_FM.split('/'):
                        equip_main_standby = '辅'
                    else:
                        continue

                if stage == 'FM':
                    weight = round(aps_fm_weight, 2)
                else:
                    try:
                        weight = round(stage_devoted_weight[stage], 2)
                    except Exception:
                        raise ValidationError('定机表该规格段次错误，请检查后重试!:{}'.format(pb_version_name))

                if weight == 0:
                    continue

                plan_trains = weight//float(batching_weight)
                if int(plan_trains) == 0:
                    continue
                train_time_consume = calculate_equip_recipe_avg_mixin_time(equip_no, recipe_name)
                time_consume = plan_trains * train_time_consume/60
                if pb_version_name not in job_list_data:
                    pb_time_consume += time_consume
                    pb_time_consume += int(sps.scheduling_interval_trains * train_time_consume/60)
                    job_list_data[pb_version_name] = {stage: {
                        'project_name': pb_version_name,
                        'stage': stage,
                        'serial_no': need_stages.index(stage) + 1,
                        'job_name': recipe_name,
                        'job_type': 'STANDARD',
                        'weight_qty': weight,
                        # 'model_size': 1,
                        'details': [{'time_consume': time_consume,
                                     'equip_no': int(equip_no[-2:]),
                                     'wait_time': int(train_time_consume/60 * sps.scheduling_interval_trains),
                                     'plan_trains': plan_trains,
                                     'main_equip_flag': 1 if equip_main_standby == '主' else 0
                                     }]}
                    }
                else:
                    if stage not in job_list_data[pb_version_name]:
                        pb_time_consume += time_consume
                        pb_time_consume += int(sps.scheduling_interval_trains * train_time_consume / 60)
                        job_list_data[pb_version_name][stage] = {
                            'project_name': pb_version_name,
                            'stage': stage,
                            'serial_no': need_stages.index(stage) + 1,
                            'job_name': recipe_name,
                            'job_type': 'STANDARD',
                            'weight_qty': weight,
                            # 'model_size': 1,
                            'details': [{'time_consume': time_consume,
                                         'equip_no': int(equip_no[-2:]),
                                         'wait_time': int(train_time_consume/60 * sps.scheduling_interval_trains),
                                         'plan_trains': plan_trains,
                                         'main_equip_flag': 1 if equip_main_standby == '主' else 0
                                         }]}
                    else:
                        # job_list_data[pb_version_name][stage]['model_size'] += 1
                        job_list_data[pb_version_name][stage]['details'].append(
                            {'time_consume': time_consume,
                             'equip_no': int(equip_no[-2:]),
                             'wait_time': int(train_time_consume / 60 * sps.scheduling_interval_trains),
                             'plan_trains': plan_trains,
                             'main_equip_flag': 1 if equip_main_standby == '主' else 0
                             }
                        )

            if available_time < pb_time_consume:
                pb_available_time_dict[pb_version_name] = pb_time_consume
            else:
                pb_available_time_dict[pb_version_name] = available_time

            # sheet1.cell(data_row, 6).value = len(need_stages)
            if pb_version_name not in job_list_data:
                raise ValidationError('该规格启用配方未找到:{}'.format(pb_version_name))

        left_plans = self.extend_last_aps_result(datetime.datetime.strptime(factory_date, "%Y-%m-%d"),
                                                 sps.lock_durations)

        # write excel
        data_row = 2
        data_row1 = 2
        # 待排程和前一天未完成的所有胶料规格数据 ['C590-01', 'J290-01']
        sheet.cell(8, 2).value = len(list(job_list_data.keys())) + len(left_plans)   # 需要排程的规格数量

        release_date = (datetime.datetime.now() -
                        datetime.datetime.strptime(aps_start_time, '%Y-%m-%d %H:%M:%S')
                        ).total_seconds() // 60
        # 结合待排程和未完成计划，写入excel
        for pb_name, item in job_list_data.items():
            sheet1.cell(data_row, 1).value = data_row - 1  # 序号
            sheet1.cell(data_row, 2).value = pb_name  # 规格名称（带版本号）
            sheet1.cell(data_row, 3).value = 0 if release_date <= 0 else release_date  # 胶料代码开始时间
            sheet1.cell(data_row, 4).value = 0  # 关键路径持续时间（暂时无用）
            sheet1.cell(data_row, 5).value = int(pb_available_time_dict.get(pb_name, 720)) + release_date
            sheet1.cell(data_row, 6).value = len(item)  # job_list_size(总共需要打待段次数量)
            for _, data in item.items():
                # 写入job_list sheet
                sheet2.cell(data_row1, 1).value = data_row - 1
                sheet2.cell(data_row1, 2).value = data['project_name']
                sheet2.cell(data_row1, 3).value = data['serial_no']
                sheet2.cell(data_row1, 4).value = data['job_name']
                sheet2.cell(data_row1, 5).value = data['job_type']
                sheet2.cell(data_row1, 6).value = data['weight_qty']
                sheet2.cell(data_row1, 7).value = len(data['details'])
                details = sorted(data['details'], key=lambda x: x['main_equip_flag'], reverse=True)
                data_col1 = 8
                for d in details:
                    sheet2.cell(data_row1, data_col1).value = int(d['time_consume'])
                    sheet2.cell(data_row1, data_col1+1).value = int(d['equip_no'])
                    sheet2.cell(data_row1, data_col1+2).value = int(d['wait_time'])
                    sheet2.cell(data_row1, data_col1+3).value = d['plan_trains']
                    data_col1 += 4
                data_row1 += 1
            data_row += 1

        for j in left_plans:
            if j['status'] == 'COMMITED' or release_date <= 0:
                res_date = 0
            else:
                res_date = release_date
            if j['status'] == 'COMMITED':
                d_time = int(j['time_consume'] + j['begin_time'])
            else:
                d_time = int(j['delivery_time']) + release_date
            # 写入project list
            sheet1.cell(data_row, 1).value = data_row - 1  # 序号
            sheet1.cell(data_row, 2).value = '-'.join(j['recipe_name'].split('-')[-2:])  # 规格名称（带版本号）
            sheet1.cell(data_row, 3).value = res_date  # 胶料代码开始时间(暂时无用)
            sheet1.cell(data_row, 4).value = 0  # 关键路径持续时间（暂时无用）
            sheet1.cell(data_row, 5).value = d_time
            sheet1.cell(data_row, 6).value = 1  # job_list_size(总共需要打待段次数量)
            # 写入job list
            sheet2.cell(data_row1, 1).value = data_row - 1
            sheet2.cell(data_row1, 2).value = '-'.join(j['recipe_name'].split('-')[-2:])
            sheet2.cell(data_row1, 3).value = 1
            sheet2.cell(data_row1, 4).value = j['recipe_name']
            sheet2.cell(data_row1, 5).value = j['status']
            sheet2.cell(data_row1, 6).value = round(j['weight'], 2)
            sheet2.cell(data_row1, 7).value = 1
            sheet2.cell(data_row1, 8).value = int(j['time_consume'])
            sheet2.cell(data_row1, 9).value = int(j['equip_no'][-2:])
            sheet2.cell(data_row1, 10).value = int(j.get('begin_time', 0)) if j['status'] == 'COMMITED' else int(j['time_consume']/j['plan_trains'] * sps.scheduling_interval_trains)
            sheet2.cell(data_row1, 11).value = int(j['plan_trains'])
            data_row1 += 1
            data_row += 1
        # machine stop plan sheet
        sheet3 = wb.worksheets[3]
        data_row2 = 2
        for j in equip_stop_plan:
            sheet3.cell(data_row2, 1).value = int(j.equip_no[-2:])
            sheet3.cell(data_row2, 2).value = j.begin_time.strftime('%Y-%m-%d %H:%M:%S')
            sheet3.cell(data_row2, 3).value = j.duration * 60
            data_row2 += 1

        output = BytesIO()
        wb.save(output)
        # 重新定位到开始
        output.seek(0)
        response = HttpResponse(content_type='application/vnd.ms-excel')
        filename = 'aps_import'
        response['Content-Disposition'] = u'attachment;filename= ' + filename.encode('gbk').decode(
            'ISO-8859-1') + '.xlsx'
        response.write(output.getvalue())
        return response


@method_decorator([api_recorder], name='dispatch')
class APSGanttView(APIView):

    def get(self, request):
        data_type = self.request.query_params.get('data_type')  # product:规格别 equip:机台别
        schedule_no = self.request.query_params.get('schedule_no')
        if not schedule_no:
            raise ValidationError('请输入排程单号！')
        ret = {}
        if data_type == 'product':
            queryset = SchedulingResult.objects.filter(
                schedule_no=schedule_no).order_by('equip_no', 'sn').values(
                'id', 'equip_no', 'recipe_name', 'time_consume', 'start_time', 'end_time', 'plan_trains', 'is_locked')
            st_idx = len(ret) + 1111
            for item in queryset:
                product_no = '-'.join(item['recipe_name'].split('-')[-2:])
                if product_no not in ret:
                    ret[product_no] = [{'id': len(ret) + 1, 'render': 'split', 'owner': product_no}]
                ret[product_no].append({'id': st_idx + 1,
                                        'parent': ret[product_no][0]['id'],
                                        'text': '{}/{}/{}'.format(item['equip_no'], item['recipe_name'], str(item['plan_trains'])),
                                        'start_date': item['start_time'].strftime('%Y-%m-%d %H:%M:%S'),
                                        'end_date': item['end_time'].strftime('%Y-%m-%d %H:%M:%S'),
                                        'equip_no': item['equip_no'],
                                        'is_locked': item['is_locked']
                                        })
                st_idx += 1
        else:
            for idx, equip_no in enumerate(list(Equip.objects.filter(
                    category__equip_type__global_name='密炼设备'
            ).order_by('equip_no').values_list('equip_no', flat=True))):
                ret[equip_no] = [{'id': idx+1, 'render': 'split', 'owner': equip_no}]
            queryset = SchedulingResult.objects.filter(
                schedule_no=schedule_no).order_by('equip_no', 'sn').values(
                'id', 'equip_no', 'recipe_name', 'time_consume', 'start_time', 'end_time', 'plan_trains', 'is_locked')
            st_idx = len(ret) + 1
            for item in queryset:
                equip_no = item['equip_no']
                ret[equip_no].append({'id': st_idx + 1,
                                      'parent': ret[equip_no][0]['id'],
                                      'text': item['recipe_name'] + '/' + str(item['plan_trains']),
                                      'start_date': item['start_time'].strftime('%Y-%m-%d %H:%M:%S'),
                                      'end_date': item['end_time'].strftime('%Y-%m-%d %H:%M:%S'),
                                      'equip_no': equip_no,
                                      'is_locked': item['is_locked']
                                      })
                st_idx += 1
        results = []
        for i in list(ret.values()):
            if len(i) == 1:
                continue
            results += i
        return Response(results)


@method_decorator([api_recorder], name='dispatch')
class APSPlanImport(APIView):

    def post(self, request):
        factory_date = self.request.data.get('factory_date')
        if not factory_date:
            raise ValidationError('请选择日期！')
        date_splits = factory_date.split('-')
        excel_file = request.FILES.get('file', None)
        if not excel_file:
            raise ValidationError('文件不可为空！')
        if not excel_file.name.split('.')[-1] in ['xls', 'xlsx', 'xlsm']:
            raise ValidationError('文件格式错误,仅支持 xls、xlsx、xlsm文件')
        try:
            data = xlrd.open_workbook(filename=None, file_contents=excel_file.read())
        except Exception as e:
            raise ValidationError('打开文件失败，请用文档另存为xlsx文件后导入！'.format(e))
        try:
            cur_sheet = data.sheet_by_name('排程结果列表')
        except Exception:
            raise ValidationError('未找到排程结果sheet表格，请检查后重试！！')
        schedule_no = cur_sheet.cell(0, 0).value
        if not schedule_no:
            schedule_no = 'APS1{}'.format(datetime.datetime.now().strftime('%Y%m%d%H%M%S'))
        SchedulingResult.objects.filter(schedule_no=schedule_no).delete()
        data = get_sheet_data(cur_sheet, start_row=3)
        i = 0
        ret = []
        aps_st = datetime.datetime(year=int(date_splits[0]), month=int(date_splits[1]), day=int(date_splits[2]), hour=8)
        eq_st = 'Z01'
        for idx, item in enumerate(data):
            if item[0] == 'Z09':
                eq_st = 'Z09'
                continue
            if '规格' in item[0]:
                continue
            if eq_st == 'Z01':
                equip_nos = ['Z01', 'Z02', 'Z03', 'Z04', 'Z05', 'Z06', 'Z07', 'Z08']
            # elif 21 < idx < 42:
            #     equip_nos = ['Z09', 'Z10', 'Z11', 'Z12', 'Z13', 'Z14', 'Z15']
            else:
                equip_nos = ['Z09', 'Z10', 'Z11', 'Z12', 'Z13', 'Z14', 'Z15']
            for j, equip_no in enumerate(equip_nos):
                product_no = item[j * 6]
                if not product_no:
                    continue
                try:
                    plan_trains = int(item[j * 6 + 1])
                    time_consume = round(item[j * 6 + 2]/60, 1)
                    desc = str(item[j * 6 + 3])
                    st = int(item[j * 6 + 4])
                    et = int(item[j * 6 + 5])
                except Exception:
                    raise ValidationError('数据错误，请检查后重试！')
                pb = ProductBatching.objects.using('SFJ').filter(
                    stage_product_batch_no=product_no,
                    used_type=4,
                    equip__equip_no=equip_no
                ).first()
                if not pb:
                    raise ValidationError('{}机台未找到此启用配方：{}'.format(equip_no, product_no))
                try:
                    ret.append(SchedulingResult(**{'factory_date': factory_date,
                                                    'schedule_no': schedule_no,
                                                    'equip_no': equip_no,
                                                    'sn': i + 1,
                                                    'recipe_name': product_no,
                                                    'plan_trains': plan_trains,
                                                    'time_consume': time_consume,
                                                    'start_time': aps_st + datetime.timedelta(minutes=st),
                                                    'end_time': aps_st + datetime.timedelta(minutes=et),
                                                    'is_locked': True if '锁定' in desc else False
                                                   }))
                except Exception:
                    raise ValidationError('导入数据有误，请检查后重试!')
            i += 1
        SchedulingResult.objects.bulk_create(ret)
        return Response('导入排程结果成功!')
