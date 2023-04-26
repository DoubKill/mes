import calendar
import datetime
import json
import re
from decimal import Decimal
from io import BytesIO

from django.http import HttpResponse
from openpyxl import load_workbook
from suds.client import Client
from django.db import connection
from django.forms import model_to_dict
from django.utils import timezone
from datetime import timedelta
from django.db.transaction import atomic
from django.utils.decorators import method_decorator
from django.views.decorators.cache import cache_page
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework import mixins, status, viewsets
from rest_framework.decorators import action
from rest_framework.exceptions import ValidationError
from rest_framework.generics import ListAPIView
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from rest_framework.viewsets import GenericViewSet, ModelViewSet, ReadOnlyModelViewSet

from basics.models import GlobalCodeType, Equip, GlobalCode, WorkSchedulePlan
from basics.serializers import GlobalCodeSerializer
import uuid

from equipment.utils import gen_template_response
from inventory.models import WmsNucleinManagement, ProductInventoryLocked, BzFinalMixingRubberInventory, \
    BzFinalMixingRubberInventoryLB
from mes import settings
from mes.common_code import CommonDeleteMixin, date_range, get_template_response, GroupConcat
from mes.conf import WMS_URL
from mes.paginations import SinglePageNumberPagination
from mes.derorators import api_recorder
from mes.permissions import PermissionClass
from plan.models import ProductClassesPlan
from production.models import PalletFeedbacks, TrainsFeedbacks
from quality.deal_result import receive_deal_result
from quality.filters import TestMethodFilter, DataPointFilter, \
    MaterialTestMethodFilter, MaterialDataPointIndicatorFilter, MaterialTestOrderFilter, MaterialDealResulFilter, \
    DealSuggestionFilter, PalletFeedbacksTestFilter, UnqualifiedDealOrderFilter, MaterialExamineTypeFilter, \
    ExamineMaterialFilter, MaterialEquipFilter, MaterialExamineResultFilter, MaterialReportEquipFilter, \
    MaterialReportValueFilter, ProductReportEquipFilter, ProductReportValueFilter, ProductTestResumeFilter, \
    MaterialTestPlanFilter, MaterialInspectionRegistrationFilter, UnqualifiedPalletFeedBackListFilter, ScorchTimeFilter
from quality.models import TestIndicator, MaterialDataPointIndicator, TestMethod, MaterialTestOrder, \
    MaterialTestMethod, TestType, DataPoint, DealSuggestion, MaterialDealResult, LevelResult, MaterialTestResult, \
    LabelPrint, UnqualifiedDealOrder, \
    MaterialExamineResult, MaterialExamineType, MaterialExamineRatingStandard, ExamineValueUnit, ExamineMaterial, \
    DataPointStandardError, MaterialSingleTypeExamineResult, MaterialEquipType, MaterialEquip, \
    QualifiedRangeDisplay, IgnoredProductInfo, MaterialReportEquip, MaterialReportValue, \
    ProductReportEquip, ProductReportValue, ProductTestPlan, ProductTestPlanDetail, RubberMaxStretchTestResult, \
    LabelPrintLog, MaterialTestPlan, MaterialTestPlanDetail, MaterialDataPointIndicatorHistory, \
    MaterialInspectionRegistration, WMSMooneyLevel, UnqualifiedDealOrderDetail, ScorchTime

from quality.serializers import MaterialDataPointIndicatorSerializer, \
    MaterialTestOrderSerializer, MaterialTestOrderListSerializer, \
    MaterialTestMethodSerializer, TestMethodSerializer, TestTypeSerializer, DataPointSerializer, \
    DealSuggestionSerializer, DealResultDealSerializer, MaterialDealResultListSerializer, LevelResultSerializer, \
    TestIndicatorSerializer, LabelPrintSerializer, \
    UnqualifiedDealOrderCreateSerializer, UnqualifiedDealOrderSerializer, UnqualifiedDealOrderUpdateSerializer, \
    MaterialDealResultListSerializer1, ExamineMaterialSerializer, MaterialExamineTypeSerializer, \
    ExamineValueUnitSerializer, MaterialExamineResultMainSerializer, DataPointStandardErrorSerializer, \
    MaterialEquipTypeSerializer, MaterialEquipSerializer, MaterialEquipTypeUpdateSerializer, \
    ExamineMaterialCreateSerializer, IgnoredProductInfoSerializer, \
    MaterialExamineResultMainCreateSerializer, MaterialReportEquipSerializer, MaterialReportValueSerializer, \
    MaterialReportValueCreateSerializer, ProductReportEquipSerializer, ProductReportValueViewSerializer, \
    ProductTestPlanSerializer, ProductTEstResumeSerializer, ReportValueSerializer, RubberMaxStretchTestResultSerializer, \
    UnqualifiedPalletFeedBackSerializer, LabelPrintLogSerializer, ProductTestPlanDetailSerializer, \
    ProductTestPlanDetailBulkCreateSerializer, MaterialTestPlanSerializer, MaterialTestPlanCreateSerializer, \
    MaterialTestPlanDetailSerializer, MaterialTestOrderExportSerializer, MaterialInspectionRegistrationSerializer, \
    MaterialDataPointIndicatorHistorySerializer, WMSMooneyLevelSerializer, ERPMESMaterialRelationSerializer, \
    ScorchTimeSerializer

from django.db.models import Prefetch, F, StdDev
from django.db.models import Q
from quality.utils import get_cur_sheet, get_sheet_data, export_mto, gen_pallet_test_result
from recipe.models import Material, ProductBatching, ERPMESMaterialRelation, ZCMaterial, ProductInfo
from django.db.models import Max, Sum, Avg, Count, Min


@method_decorator([api_recorder], name="dispatch")
class TestIndicatorViewSet(ModelViewSet):
    """试验指标列表"""
    queryset = TestIndicator.objects.filter(delete_flag=False)
    serializer_class = TestIndicatorSerializer

    def list(self, request, *args, **kwargs):
        data = self.queryset.values('id', 'name')
        return Response(data)


@method_decorator([api_recorder], name="dispatch")
class DataPointListView(APIView):
    """数据点列表"""

    def get(self, request):
        ret = []
        indicator_names = {'门尼': 1, '硬度': 2, '比重': 3, '流变': 4, '钢拔': 5, '物性': 6}
        for indicator_name in indicator_names:
            data_points = DataPoint.objects.filter(
                test_type__test_indicator__name=indicator_name).order_by('name').values_list('name', flat=True)
            for data_point in data_points:
                if data_point not in ret:
                    ret.append(data_point)
        return Response(ret)


@method_decorator([api_recorder], name="dispatch")
class TestTypeViewSet(ModelViewSet):
    """试验类型管理"""
    queryset = TestType.objects.filter(delete_flag=False)
    serializer_class = TestTypeSerializer
    filter_backends = (DjangoFilterBackend,)
    filter_fields = ('test_indicator',)

    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        if self.request.query_params.get('all'):
            data = queryset.values('id', 'name')
            return Response({'results': data})
        return super().list(self, request, *args, **kwargs)


@method_decorator([api_recorder], name="dispatch")
class DataPointViewSet(ModelViewSet):
    """试验类型数据点管理"""
    queryset = DataPoint.objects.filter(delete_flag=False)
    serializer_class = DataPointSerializer
    filter_backends = (DjangoFilterBackend,)
    filter_class = DataPointFilter
    pagination_class = None

    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        if self.request.query_params.get('all'):
            data = queryset.values('id', 'name', 'unit')
            return Response({'results': data})
        return super().list(self, request, *args, **kwargs)


@method_decorator([api_recorder], name="dispatch")
class DataPointStandardErrorViewSet(ModelViewSet):
    """数据点误差(不合格pass指标管理)"""
    queryset = DataPointStandardError.objects.filter(delete_flag=False)
    serializer_class = DataPointStandardErrorSerializer
    filter_backends = (DjangoFilterBackend,)
    filter_fields = ('data_point_id',)
    pagination_class = None


@method_decorator([api_recorder], name="dispatch")
class DataPointLabelHistoryView(APIView):
    """标记历史记录"""

    def get(self, request):
        return Response(set(DataPointStandardError.objects.values_list('label', flat=True)))


@method_decorator([api_recorder], name="dispatch")
class TestMethodViewSet(ModelViewSet):
    """试验方法管理"""
    queryset = TestMethod.objects.filter(delete_flag=False)
    serializer_class = TestMethodSerializer
    filter_backends = (DjangoFilterBackend,)
    filter_class = TestMethodFilter

    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        if self.request.query_params.get('all'):
            data = queryset.values('id', 'name')
            return Response({'results': data})
        return super().list(self, request, *args, **kwargs)


@method_decorator([api_recorder], name="dispatch")
class TestIndicatorDataPointListView(ListAPIView):
    """获取试验指标及其所有的试验方法数据点"""
    queryset = TestIndicator.objects.filter(delete_flag=False)

    def list(self, request, *args, **kwargs):
        ret = []
        test_indicators_names = ['门尼', '比重', '硬度', '流变', '钢拔', '物性']
        for name in test_indicators_names:
            test_indicator = TestIndicator.objects.filter(name__icontains=name).first()
            if test_indicator:
                data_indicator_detail = []
                data_names = DataPoint.objects.filter(
                    test_type__test_indicator=test_indicator).order_by('name').values_list('name', flat=True)
                for data_name in data_names:
                    if data_name not in data_indicator_detail:
                        data_indicator_detail.append(data_name)
                data = {'test_type_id': test_indicator.id,
                        'test_type_name': test_indicator.name,
                        'data_indicator_detail': data_indicator_detail
                        }
                ret.append(data)
        return Response(ret)


@method_decorator([api_recorder], name="dispatch")
class MaterialTestIndicatorMethods(APIView):
    """获取原材料指标试验方法"""

    def get(self, request):
        material_no = self.request.query_params.get('material_no')
        ret = {}
        for method in MaterialTestMethod.objects.filter(material__material_no=material_no):
            indicator_name = method.test_method.test_type.test_indicator.name
            data_points = method.mat_indicators.filter(
                level=1, delete_flag=False).values(name=F('data_point__name'),
                                                   judge_upper_limit=F('upper_limit'),
                                                   judge_lower_limit=F('lower_limit')).order_by('name')
            if not data_points:
                continue
            if indicator_name not in ret:
                data = {
                    'test_indicator': indicator_name,
                    'methods': [
                        {'id': method.test_method.id,
                         'name': method.test_method.name,
                         'allowed': True,
                         'data_points': data_points}
                    ]
                }
                ret[indicator_name] = data
            else:
                ret[indicator_name]['methods'].append(
                    {'id': method.test_method.id, 'name': method.test_method.name, 'allowed': True, 'data_points': data_points})
        return Response(ret.values())


@method_decorator([api_recorder], name="dispatch")
class MaterialTestOrderViewSet(mixins.CreateModelMixin,
                               mixins.ListModelMixin,
                               GenericViewSet):
    """
    list:
        列表展示
    create:
        手工录入数据
    """
    queryset = MaterialTestOrder.objects.filter(
        delete_flag=False).prefetch_related(
        Prefetch('order_results',
                 queryset=MaterialTestResult.objects.order_by('id'))
    ).order_by('-production_factory_date',
               '-production_class',
               'production_equip_no',
               'product_no',
               'actual_trains')
    serializer_class = MaterialTestOrderSerializer
    filter_backends = (DjangoFilterBackend,)
    permission_classes = (IsAuthenticated,)
    filter_class = MaterialTestOrderFilter

    def get_serializer_class(self):
        if self.action == 'create':
            return MaterialTestOrderSerializer
        else:
            return MaterialTestOrderListSerializer

    def export_xls(self, result, lot_deal_result_dict, template, data_row):
        if not result:
            return Response('暂无数据！')
        wb = load_workbook(template)
        ws = wb.worksheets[0]
        sheet0 = wb.copy_worksheet(ws)
        sheet0.title = '汇总'
        row_num = data_row
        sheet0.delete_rows(2, data_row)
        sheet = wb.copy_worksheet(ws)
        data_row0 = 2
        product_no = result[0]['product_no']
        sheet.title = product_no
        indicators_data = MaterialDataPointIndicator.objects.filter(
            material_test_method__material__material_no=product_no,
            level=1,
            delete_flag=False
        ).values('data_point__name', 'upper_limit', 'lower_limit')
        indicators_data_dict = {i['data_point__name']: i for i in indicators_data}
        sheet.cell(2, 1).value = product_no
        sheet.cell(5, 10).value = indicators_data_dict.get('ML(1+4)', {}).get('upper_limit')
        sheet.cell(6, 10).value = indicators_data_dict.get('ML(1+4)', {}).get('lower_limit')
        sheet.cell(5, 11).value = indicators_data_dict.get('比重值', {}).get('upper_limit')
        sheet.cell(6, 11).value = indicators_data_dict.get('比重值', {}).get('lower_limit')
        sheet.cell(5, 12).value = indicators_data_dict.get('硬度值', {}).get('upper_limit')
        sheet.cell(6, 12).value = indicators_data_dict.get('硬度值', {}).get('lower_limit')
        sheet.cell(5, 13).value = indicators_data_dict.get('MH', {}).get('upper_limit')
        sheet.cell(6, 13).value = indicators_data_dict.get('MH', {}).get('lower_limit')
        sheet.cell(5, 14).value = indicators_data_dict.get('ML', {}).get('upper_limit')
        sheet.cell(6, 14).value = indicators_data_dict.get('ML', {}).get('lower_limit')
        sheet.cell(5, 15).value = indicators_data_dict.get('TC10', {}).get('upper_limit')
        sheet.cell(6, 15).value = indicators_data_dict.get('TC10', {}).get('lower_limit')
        sheet.cell(5, 16).value = indicators_data_dict.get('TC50', {}).get('upper_limit')
        sheet.cell(6, 16).value = indicators_data_dict.get('TC50', {}).get('lower_limit')
        sheet.cell(5, 17).value = indicators_data_dict.get('TC90', {}).get('upper_limit')
        sheet.cell(6, 17).value = indicators_data_dict.get('TC90', {}).get('lower_limit')
        sheet.cell(5, 18).value = indicators_data_dict.get('T5', {}).get('upper_limit')
        sheet.cell(6, 18).value = indicators_data_dict.get('T5', {}).get('lower_limit')
        sheet.cell(5, 19).value = indicators_data_dict.get('钢拔', {}).get('upper_limit')
        sheet.cell(6, 19).value = indicators_data_dict.get('钢拔', {}).get('lower_limit')
        sheet.cell(5, 20).value = indicators_data_dict.get('100%', {}).get('upper_limit')
        sheet.cell(6, 20).value = indicators_data_dict.get('100%', {}).get('lower_limit')
        sheet.cell(5, 21).value = indicators_data_dict.get('M300', {}).get('upper_limit')
        sheet.cell(6, 21).value = indicators_data_dict.get('M300', {}).get('lower_limit')
        sheet.cell(5, 22).value = indicators_data_dict.get('伸长率%', {}).get('upper_limit')
        sheet.cell(6, 22).value = indicators_data_dict.get('伸长率%', {}).get('lower_limit')
        sheet.cell(5, 23).value = indicators_data_dict.get('扯断强度', {}).get('upper_limit')
        sheet.cell(6, 23).value = indicators_data_dict.get('扯断强度', {}).get('lower_limit')
        sheet.cell(5, 24).value = indicators_data_dict.get('永久变形', {}).get('upper_limit')
        sheet.cell(6, 24).value = indicators_data_dict.get('永久变形', {}).get('lower_limit')
        for i in result:
            deal_suggestion = ''
            result_data = lot_deal_result_dict.get(i['lot_no'])
            if result_data:
                if result_data['deal_user']:
                    deal_suggestion = result_data['deal_suggestion']
                else:
                    deal_suggestion = 'PASS' if result_data['test_result'] == 'PASS' else None
            if i['product_no'] != product_no:
                row_num = data_row
                product_no = i['product_no']
                sheet = wb.copy_worksheet(ws)
                sheet.title = product_no
                indicators_data = MaterialDataPointIndicator.objects.filter(
                    material_test_method__material__material_no=product_no,
                    level=1,
                    delete_flag=False
                ).values('data_point__name', 'upper_limit', 'lower_limit')
                indicators_data_dict = {i['data_point__name']: i for i in indicators_data}
                sheet.cell(2, 1).value = product_no
                sheet.cell(5, 10).value = indicators_data_dict.get('ML(1+4)', {}).get('upper_limit')
                sheet.cell(6, 10).value = indicators_data_dict.get('ML(1+4)', {}).get('lower_limit')
                sheet.cell(5, 11).value = indicators_data_dict.get('比重值', {}).get('upper_limit')
                sheet.cell(6, 11).value = indicators_data_dict.get('比重值', {}).get('lower_limit')
                sheet.cell(5, 12).value = indicators_data_dict.get('硬度值', {}).get('upper_limit')
                sheet.cell(6, 12).value = indicators_data_dict.get('硬度值', {}).get('lower_limit')
                sheet.cell(5, 13).value = indicators_data_dict.get('MH', {}).get('upper_limit')
                sheet.cell(6, 13).value = indicators_data_dict.get('MH', {}).get('lower_limit')
                sheet.cell(5, 14).value = indicators_data_dict.get('ML', {}).get('upper_limit')
                sheet.cell(6, 14).value = indicators_data_dict.get('ML', {}).get('lower_limit')
                sheet.cell(5, 15).value = indicators_data_dict.get('TC10', {}).get('upper_limit')
                sheet.cell(6, 15).value = indicators_data_dict.get('TC10', {}).get('lower_limit')
                sheet.cell(5, 16).value = indicators_data_dict.get('TC50', {}).get('upper_limit')
                sheet.cell(6, 16).value = indicators_data_dict.get('TC50', {}).get('lower_limit')
                sheet.cell(5, 17).value = indicators_data_dict.get('TC90', {}).get('upper_limit')
                sheet.cell(6, 17).value = indicators_data_dict.get('TC90', {}).get('lower_limit')
                sheet.cell(5, 18).value = indicators_data_dict.get('T5', {}).get('upper_limit')
                sheet.cell(6, 18).value = indicators_data_dict.get('T5', {}).get('lower_limit')
                sheet.cell(5, 19).value = indicators_data_dict.get('钢拔', {}).get('upper_limit')
                sheet.cell(6, 19).value = indicators_data_dict.get('钢拔', {}).get('lower_limit')
                sheet.cell(5, 20).value = indicators_data_dict.get('100%', {}).get('upper_limit')
                sheet.cell(6, 20).value = indicators_data_dict.get('100%', {}).get('lower_limit')
                sheet.cell(5, 21).value = indicators_data_dict.get('M300', {}).get('upper_limit')
                sheet.cell(6, 21).value = indicators_data_dict.get('M300', {}).get('lower_limit')
                sheet.cell(5, 22).value = indicators_data_dict.get('伸长率%', {}).get('upper_limit')
                sheet.cell(6, 22).value = indicators_data_dict.get('伸长率%', {}).get('lower_limit')
                sheet.cell(5, 23).value = indicators_data_dict.get('扯断强度', {}).get('upper_limit')
                sheet.cell(6, 23).value = indicators_data_dict.get('扯断强度', {}).get('lower_limit')
                sheet.cell(5, 24).value = indicators_data_dict.get('永久变形', {}).get('upper_limit')
                sheet.cell(6, 24).value = indicators_data_dict.get('永久变形', {}).get('lower_limit')
            order_results = i['order_results']
            mn_machine_name = ''
            lb_machine_name = ''
            ret = {}
            for j in order_results:
                ret[j['data_point_name']] = j['value']
                machine_name = j['machine_name']
                if machine_name:
                    if j['data_point_name'] == 'ML(1+4)':
                        mn_machine_name = machine_name
                    elif j['data_point_name'] in ('MH', 'ML', 'TC10', 'TC50', 'TC90'):
                        lb_machine_name = machine_name

            ret = {i['data_point_name']: i['value'] for i in order_results}
            sheet.cell(row_num, 1).value = i['product_no']
            sheet.cell(row_num, 2).value = i['production_factory_date']
            sheet.cell(row_num, 3).value = i['production_class']
            sheet.cell(row_num, 4).value = i['production_group']
            sheet.cell(row_num, 5).value = i['production_equip_no']
            sheet.cell(row_num, 6).value = mn_machine_name
            sheet.cell(row_num, 7).value = lb_machine_name
            sheet.cell(row_num, 8).value = i['actual_trains']
            sheet.cell(row_num, 9).value = '复检' if i['is_recheck'] else '正常'
            sheet.cell(row_num, 10).value = ret.get('ML(1+4)')
            sheet.cell(row_num, 11).value = ret.get('比重值')
            sheet.cell(row_num, 12).value = ret.get('硬度值')
            sheet.cell(row_num, 13).value = ret.get('MH')
            sheet.cell(row_num, 14).value = ret.get('ML')
            sheet.cell(row_num, 15).value = ret.get('TC10')
            sheet.cell(row_num, 16).value = ret.get('TC50')
            sheet.cell(row_num, 17).value = ret.get('TC90')
            sheet.cell(row_num, 18).value = ret.get('T5')
            sheet.cell(row_num, 19).value = ret.get('钢拔')
            sheet.cell(row_num, 20).value = ret.get('100%')
            sheet.cell(row_num, 21).value = ret.get('M300')
            sheet.cell(row_num, 22).value = ret.get('伸长率%')
            sheet.cell(row_num, 23).value = ret.get('扯断强度')
            sheet.cell(row_num, 24).value = ret.get('永久变形')
            sheet.cell(row_num, 25).value = i['state']
            sheet.cell(row_num, 26).value = deal_suggestion

            # 写入汇总
            sheet0.cell(data_row0, 1).value = i['product_no']
            sheet0.cell(data_row0, 2).value = i['production_factory_date']
            sheet0.cell(data_row0, 3).value = i['production_class']
            sheet0.cell(data_row0, 4).value = i['production_group']
            sheet0.cell(data_row0, 5).value = i['production_equip_no']
            sheet0.cell(data_row0, 6).value = mn_machine_name
            sheet0.cell(data_row0, 7).value = lb_machine_name
            sheet0.cell(data_row0, 8).value = i['actual_trains']
            sheet0.cell(data_row0, 9).value = '复检' if i['is_recheck'] else '正常'
            sheet0.cell(data_row0, 10).value = ret.get('ML(1+4)')
            sheet0.cell(data_row0, 11).value = ret.get('比重值')
            sheet0.cell(data_row0, 12).value = ret.get('硬度值')
            sheet0.cell(data_row0, 13).value = ret.get('MH')
            sheet0.cell(data_row0, 14).value = ret.get('ML')
            sheet0.cell(data_row0, 15).value = ret.get('TC10')
            sheet0.cell(data_row0, 16).value = ret.get('TC50')
            sheet0.cell(data_row0, 17).value = ret.get('TC90')
            sheet0.cell(data_row0, 18).value = ret.get('T5')
            sheet0.cell(data_row0, 19).value = ret.get('钢拔')
            sheet0.cell(data_row0, 20).value = ret.get('100%')
            sheet0.cell(data_row0, 21).value = ret.get('M300')
            sheet0.cell(data_row0, 22).value = ret.get('伸长率%')
            sheet0.cell(data_row0, 23).value = ret.get('扯断强度')
            sheet0.cell(data_row0, 24).value = ret.get('永久变形')
            sheet0.cell(data_row0, 25).value = i['state']
            sheet0.cell(data_row0, 26).value = deal_suggestion

            row_num += 1
            data_row0 += 1
        wb.remove_sheet(ws)
        output = BytesIO()
        wb.save(output)
        # 重新定位到开始
        output.seek(0)
        response = HttpResponse(content_type='application/vnd.ms-excel')
        filename = '快检详细信息'
        response['Content-Disposition'] = u'attachment;filename= ' + filename.encode('gbk').decode(
            'ISO-8859-1') + '.xls'
        response.write(output.getvalue())
        return response

    def list(self, request, *args, **kwargs):
        state = self.request.query_params.get('state')
        export = self.request.query_params.get('export')
        queryset = self.filter_queryset(self.get_queryset())
        sum_project = self.request.query_params.get('sum_project')
        recipe_type = self.request.query_params.get('recipe_type')
        stage = self.request.query_params.get('stage')
        if stage:
            queryset = queryset.filter(product_no__icontains='-{}-'.format(stage))
        if state:
            if state == '检测中':
                queryset = queryset.filter(is_finished=False)
            elif state == '合格':
                queryset = queryset.filter(is_finished=True, is_qualified=True)
            elif state == '不合格':
                queryset = queryset.filter(is_finished=True, is_qualified=False)
        if recipe_type:
            if recipe_type:
                stage_prefix = re.split(r'[,|，]', recipe_type)
                filter_str = ''
                for i in stage_prefix:
                    filter_str += ('' if not filter_str else '|') + f"Q(product_info__product_name__startswith='{i.strip()}')"
                product_qs = ProductBatching.objects.filter(eval(filter_str))
                if 'C' in stage_prefix or 'TC' in stage_prefix:  # 车胎类别(C)与半钢类别(CJ)需要区分
                    product_qs = product_qs.filter(~Q(product_info__product_name__startswith='CJ'),
                                                   ~Q(product_info__product_name__startswith='TCJ'))
                if 'U' in stage_prefix or 'TU' in stage_prefix:  # 车胎类别(UC)与斜胶类别(U)需要区分
                    product_qs = product_qs.filter(~Q(product_info__product_name__startswith='UC'),
                                                   ~Q(product_info__product_name__startswith='TUC'))
                product_nos = product_qs.values_list('stage_product_batch_no', flat=True)
                queryset = queryset.filter(product_no__in=product_nos)

        if export:
            st = self.request.query_params.get('st')
            et = self.request.query_params.get('et')
            if not all([st, et]):
                raise ValidationError('请选择导出的时间范围！')
            diff = datetime.datetime.strptime(et, '%Y-%m-%d') - \
                   datetime.datetime.strptime(st, '%Y-%m-%d')
            if self.request.query_params.get('product_no'):
                if diff.days > 30:
                    raise ValidationError('导出数据的日期跨度不得超过一个月！')
            else:
                if diff.days > 6:
                    raise ValidationError('导出数据的日期跨度不得超过一个周！')
            queryset = queryset.order_by('product_no', 'production_factory_date', '-production_class', 'production_equip_no', 'actual_trains')
            lot_nos = set(queryset.values_list('lot_no', flat=True))
            deal_results = MaterialDealResult.objects.filter(
                lot_no__in=lot_nos,
                test_result__in=('PASS', '三等品')).values('lot_no', 'deal_suggestion', 'deal_user', 'test_result')
            lot_deal_result_dict = {i['lot_no']: i for i in deal_results}
            data = MaterialTestOrderExportSerializer(queryset, many=True).data
            if sum_project:
                return self.export_xls(data, lot_deal_result_dict, 'xlsx_template/product_test_result2.xlsx', 13)
            return self.export_xls(data, lot_deal_result_dict, 'xlsx_template/product_test_result.xlsx', 7)
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            s_data = serializer.data
            lot_nos = [i['lot_no'] for i in s_data]
            result = MaterialDealResult.objects.filter(
                lot_no__in=lot_nos,
                test_result__in=('PASS', '三等品')).values('lot_no', 'deal_suggestion', 'deal_user', 'test_result')
            lot_dict = {i['lot_no']: i for i in result}
            for item in s_data:
                result_data = lot_dict.get(item['lot_no'])
                if result_data:
                    if result_data['deal_user']:
                        item['deal_suggestion'] = result_data['deal_suggestion']
                    else:
                        item['deal_suggestion'] = 'PASS' if result_data['test_result'] == 'PASS' else None
            return self.get_paginated_response(s_data)

        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)

    @atomic()
    def create(self, request, *args, **kwargs):
        data = request.data
        if not isinstance(data, list):
            raise ValidationError('参数错误')
        s = MaterialTestOrderSerializer(data=data, many=True)
        if not s.is_valid():
            raise ValidationError(s.errors)
        lot_nos = []
        factory_date = data[0]['production_factory_date']
        classes = data[0]['production_class']
        product_no = data[0]['product_no']
        equip_no = data[0]['production_equip_no']
        ws = WorkSchedulePlan.objects.filter(plan_schedule__day_time=factory_date,
                                             classes__global_name=classes,
                                             plan_schedule__work_schedule__work_procedure__global_name='密炼').first()
        if not ws:
            production_group = '班'
        else:
            production_group = ws.group.global_name
        data_point_method_map = {}

        for validated_data in s.validated_data:
            order_results = validated_data.pop('order_results', None)
            if not order_results:
                continue
            pallets = PalletFeedbacks.objects.filter(
                equip_no=equip_no,
                product_no=product_no,
                classes=classes,
                factory_date=factory_date,
                begin_trains__lte=validated_data['actual_trains'],
                end_trains__gte=validated_data['actual_trains']
            )
            for pallet in pallets:
                lot_nos.append(pallet.lot_no)
                validated_data['lot_no'] = pallet.lot_no
                validated_data['material_test_order_uid'] = uuid.uuid1()
                validated_data['production_group'] = production_group
                while 1:
                    try:
                        instance, created = MaterialTestOrder.objects.get_or_create(
                            defaults=validated_data, **{'lot_no': validated_data['lot_no'],
                                                        'actual_trains': validated_data['actual_trains']})
                        break
                    except Exception:
                        pass
                # is_recheck = False
                for item in order_results:
                    test_value = item['value']
                    if not test_value:
                        continue
                    data_point_name = item['data_point_name']
                    item['material_test_order'] = instance
                    item['test_factory_date'] = datetime.datetime.now()
                    item['created_user'] = self.request.user
                    item['test_class'] = classes
                    item['test_group'] = production_group
                    method = data_point_method_map.get(data_point_name)
                    if method:
                        if method['qualified_range'][0] <= test_value <= method['qualified_range'][1]:
                            item['mes_result'] = '一等品'
                            item['level'] = 1
                        else:
                            item['mes_result'] = '三等品'
                            item['level'] = 2
                        item['is_judged'] = method['is_judged']
                        item['judged_upper_limit'] = method['qualified_range'][1]
                        item['judged_lower_limit'] = method['qualified_range'][0]
                    else:
                        material_test_method = MaterialTestMethod.objects.filter(
                            material__material_no=product_no,
                            test_method__name=item['test_method_name'],
                            test_method__test_type__test_indicator__name=item['test_indicator_name'],
                            data_point__name=data_point_name,
                            data_point__test_type__test_indicator__name=item['test_indicator_name']).first()
                        if material_test_method:
                            item['is_judged'] = material_test_method.is_judged
                            indicator = MaterialDataPointIndicator.objects.filter(
                                material_test_method=material_test_method,
                                data_point__name=data_point_name,
                                data_point__test_type__test_indicator__name=item['test_indicator_name'],
                                level=1).first()
                            if indicator:
                                data_point_method_map[data_point_name] = {
                                    'qualified_range': [indicator.lower_limit,  indicator.upper_limit],
                                    'is_judged': material_test_method.is_judged}
                                if indicator.lower_limit <= item['value'] <= indicator.upper_limit:
                                    item['mes_result'] = '一等品'
                                    item['level'] = 1
                                else:
                                    item['mes_result'] = '三等品'
                                    item['level'] = 2
                                item['judged_upper_limit'] = indicator.upper_limit
                                item['judged_lower_limit'] = indicator.lower_limit
                            else:
                                continue
                        else:
                            raise ValidationError('该胶料实验方法不存在！')
                    if not created:
                        dp_instances = instance.order_results.filter(data_point_name=item['data_point_name'])
                        if dp_instances:
                            v = dp_instances.first()
                            item['value0'] = v.value
                            item['judged_upper_limit0'] = v.judged_upper_limit
                            item['judged_lower_limit0'] = v.judged_lower_limit
                            dp_instances.delete()
                            instance.is_recheck = True
                            instance.save()
                    MaterialTestResult.objects.create(**item)
                # if is_recheck:
                #     instance.is_recheck = True
                #     instance.save()
        gen_pallet_test_result(lot_nos)
        return Response('新建成功')


@method_decorator([api_recorder], name="dispatch")
class TestedMaterials(APIView):

    def get(self, request):
        stage = self.request.query_params.get('stage')
        product_type = self.request.query_params.get('product_type')
        test_indicator_id = self.request.query_params.get('test_indicator_id')
        test_type_id = self.request.query_params.get('test_type_id')
        queryset = MaterialTestMethod.objects.filter(delete_flag=False)
        if stage:
            queryset = queryset.filter(material__material_name__icontains='-{}-'.format(stage))
        if product_type:
            queryset = queryset.filter(material__material_name__icontains='{}-'.format(product_type))
        if test_indicator_id:
            queryset = queryset.filter(test_method__test_type__test_indicator_id=test_indicator_id)
        if test_type_id:
            queryset = queryset.filter(test_method__test_type_id=test_type_id)
        return Response(set(queryset.values_list('material__material_no', flat=True)))


@method_decorator([api_recorder], name="dispatch")
class MaterialTestMethodViewSet(ModelViewSet):
    """物料试验方法"""
    queryset = MaterialTestMethod.objects.filter(delete_flag=False)
    serializer_class = MaterialTestMethodSerializer
    filter_backends = (DjangoFilterBackend,)
    permission_classes = (IsAuthenticated,)
    filter_class = MaterialTestMethodFilter

    def list(self, request, *args, **kwargs):
        stage = self.request.query_params.get('stage')
        product_type = self.request.query_params.get('product_type')
        queryset = self.filter_queryset(self.get_queryset())
        if stage:
            queryset = queryset.filter(material__material_name__icontains='-{}-'.format(stage))
        if product_type:
            queryset = queryset.filter(material__material_name__icontains='{}-'.format(product_type))
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)

        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)

    @action(methods=['post'], detail=False, permission_classes=[IsAuthenticated], url_path='batch-set',
            url_name='batch-set')
    def batch_set(self, request):
        data = self.request.data
        method_ids = data.get('test_method_ids')
        indicator_data = data.get('indicator_data')  # [{'data_point': 1, upper_limit:12, lower_limit:1, level:1, result:1}]
        for item in indicator_data:
            qs = MaterialDataPointIndicator.objects.filter(delete_flag=0,
                                                           data_point_id=item['data_point'],
                                                           material_test_method__id__in=method_ids,
                                                           level=item['level'],
                                                           )
            for instance in qs:
                if not instance.last_updated_user:
                    username = self.request.user.username
                else:
                    username = instance.last_updated_user.username
                MaterialDataPointIndicatorHistory.objects.create(
                    product_no=instance.material_test_method.material.material_no,
                    test_method=instance.material_test_method.test_method,
                    data_point=instance.data_point,
                    level=instance.level,
                    result=instance.result,
                    upper_limit=instance.upper_limit,
                    lower_limit=instance.lower_limit,
                    created_username=username,
                    created_date=instance.last_updated_date
                )
            qs.update(upper_limit=item['upper_limit'],
                      lower_limit=item['lower_limit'],
                      last_updated_date=datetime.datetime.now(),
                      last_updated_user=self.request.user
                      )
        return Response('ok')


@method_decorator([api_recorder], name="dispatch")
class MaterialDataPointIndicatorViewSet(ModelViewSet):
    """物料数据点评判指标"""
    queryset = MaterialDataPointIndicator.objects.filter(delete_flag=False)
    serializer_class = MaterialDataPointIndicatorSerializer
    filter_backends = (DjangoFilterBackend,)
    permission_classes = (IsAuthenticated,)
    filter_class = MaterialDataPointIndicatorFilter
    pagination_class = None


@method_decorator([api_recorder], name="dispatch")
class MaterialDataPointIndicatorHistoryView(ListAPIView):
    """物料数据点评判指标历史修改数据"""
    queryset = MaterialDataPointIndicatorHistory.objects.order_by('-id')
    serializer_class = MaterialDataPointIndicatorHistorySerializer
    filter_backends = (DjangoFilterBackend,)
    permission_classes = (IsAuthenticated,)
    filter_fields = ('product_no', 'data_point_id', 'level', 'test_method_id')


@method_decorator([api_recorder], name="dispatch")
class ProductBatchingMaterialListView(ListAPIView):
    """胶料原材料列表，（可根据生产信息过滤）"""
    queryset = Material.objects.filter(delete_flag=False)

    def list(self, request, *args, **kwargs):
        m_type = self.request.query_params.get('type', '1')  # 1胶料  2原材料
        factory_date = self.request.query_params.get('factory_date')  # 工厂日期
        equip_no = self.request.query_params.get('equip_no')  # 设备编号
        classes = self.request.query_params.get('classes')  # 班次
        used_type = self.request.query_params.get('used_type')  # 启用状态
        stages = self.request.query_params.get('stage')  # 段次
        recipe_type = self.request.query_params.get('recipe_type')  # 配方类别

        pbs = ProductBatching.objects.all()
        if used_type:
            pbs = pbs.filter(used_type=used_type)
        if stages:
            pbs = pbs.filter(stage__global_name__in=stages.split(','))
        if recipe_type:
            stage_prefix = re.split(r'[,|，]', recipe_type)
            product_nos = list(ProductInfo.objects.values_list('product_no', flat=True))
            filter_product_nos = list(filter(lambda x: ''.join(re.findall(r'[A-Za-z]', x)) in stage_prefix, product_nos))
            pbs = pbs.filter(product_info__product_no__in=filter_product_nos)
        batching_no = set(pbs.values_list('stage_product_batch_no', flat=True))
        if m_type == '1':
            kwargs = {}
            if factory_date:
                kwargs['factory_date'] = factory_date
            if equip_no:
                kwargs['equip_no'] = equip_no
            if classes:
                kwargs['classes'] = classes
            if kwargs:
                batching_no = TrainsFeedbacks.objects.filter(**kwargs).values_list('product_no', flat=True)
            material_data = self.queryset.filter(
                material_no__in=batching_no).values('id', 'material_no', 'material_name')
        elif m_type == '2':
            material_data = self.queryset.exclude(
                material_no__in=batching_no).values('id', 'material_no', 'material_name')
        else:
            raise ValidationError('参数错误')
        return Response(material_data)


@method_decorator([api_recorder], name="dispatch")
class DealSuggestionViewSet(CommonDeleteMixin, ModelViewSet):
    """处理意见
        list: 查询处理意见列表
        retrive: 查询处理意见详情
        post: 新增处理意见
        put: 修改处理意见
    """
    queryset = DealSuggestion.objects.filter(delete_flag=False)
    serializer_class = DealSuggestionSerializer
    filter_backends = (DjangoFilterBackend,)
    filter_class = DealSuggestionFilter
    pagination_class = SinglePageNumberPagination


@method_decorator([api_recorder], name="dispatch")
class MaterialDealResultViewSet(CommonDeleteMixin, ModelViewSet):
    """胶料处理结果
    list: 查询胶料处理结果列表
    post: 创建胶料处理结果
    put: 创建胶料处理结果
    """
    queryset = MaterialDealResult.objects.filter(~Q(deal_result="一等品")).filter(~Q(status="复测")).filter(
        delete_flag=False)
    serializer_class = DealResultDealSerializer
    filter_backends = (DjangoFilterBackend,)
    filter_class = MaterialDealResulFilter


@method_decorator([api_recorder], name="dispatch")
class MaterialDealStatusListView(APIView):
    """胶料状态列表"""

    def get(self, request):
        filter_set = MaterialDealResult.objects.filter(delete_flag=False).values("status").annotate()
        return Response(filter_set)


@method_decorator([api_recorder], name="dispatch")
class DealTypeView(APIView):
    # 创建处理类型
    def post(self, request):
        data = request.data
        gct = GlobalCodeType.objects.filter(type_name="处理类型").first()
        if not gct:
            raise ValidationError("请先在基础信息管理下的公用代码管理内启用/创建'处理类型'")
        data.update(global_type=gct.id)
        serializer = GlobalCodeSerializer(data=data, context={"request": request})
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response({"message": "ok"}, status=status.HTTP_201_CREATED)


@method_decorator([api_recorder], name="dispatch")
class MaterialDealResultUpdateValidTime(APIView):
    # 快检信息综合管理修改有效时间
    @atomic()
    def post(self, request):
        id = self.request.data.get('id', None)
        valid_time = self.request.data.get('valid_time', None)
        if not id or not valid_time:
            raise ValidationError('id或有效时间必传')
        MaterialDealResult.objects.filter(id=id).update(valid_time=valid_time)
        return Response('修改成功')


@method_decorator([api_recorder], name="dispatch")
class PalletFeedbacksTestListView(ModelViewSet):
    # 快检信息综合管里
    queryset = MaterialDealResult.objects.filter(delete_flag=False).order_by('factory_date', 'classes',
                                                                             'equip_no', 'product_no',
                                                                             'begin_trains')
    serializer_class = MaterialDealResultListSerializer
    filter_backends = (DjangoFilterBackend,)
    filter_class = PalletFeedbacksTestFilter
    permission_classes = (IsAuthenticated, )

    def get_serializer_class(self):
        if self.action == "retrieve":
            return MaterialDealResultListSerializer
        else:
            return MaterialDealResultListSerializer1

    def list(self, request, *args, **kwargs):
        locked_status = self.request.query_params.get('locked_status')
        is_instock = self.request.query_params.get('is_instock')
        queryset = self.filter_queryset(self.get_queryset())
        if locked_status:
            # if locked_status == '0':  # 空白
            #     locked_lot_nos = list(
            #         ProductInventoryLocked.objects.filter(is_locked=True).values_list('lot_no', flat=True))
            #     queryset = queryset.exclude(lot_no__in=locked_lot_nos)
            if locked_status == '1':  # 工艺锁定
                locked_lot_nos = list(
                    ProductInventoryLocked.objects.filter(is_locked=True, locked_status__in=(1, 3)).values_list(
                        'lot_no', flat=True))
                queryset = queryset.filter(lot_no__in=locked_lot_nos)
            elif locked_status == '2':  # 快检锁定
                locked_lot_nos = list(
                    ProductInventoryLocked.objects.filter(is_locked=True, locked_status__in=(2, 3)).values_list(
                        'lot_no',
                        flat=True))
                queryset = queryset.filter(lot_no__in=locked_lot_nos)
        if is_instock:
            stock_lot_nos = list(BzFinalMixingRubberInventory.objects.using('bz').values_list('lot_no', flat=True)) + \
                            list(BzFinalMixingRubberInventoryLB.objects.using('lb').filter(
                                store_name='炼胶库').values_list('lot_no', flat=True))
            if is_instock == 'Y':
                queryset = queryset.filter(lot_no__in=stock_lot_nos)
            else:
                stock_ids = queryset.filter(lot_no__in=stock_lot_nos).values_list('id', flat=True)
                queryset = queryset.exclude(id__in=list(stock_ids))
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            s_data = serializer.data
            lot_nos = [i['lot_no'] for i in s_data]
            pallet_weight_dict = dict(PalletFeedbacks.objects.filter(
                lot_no__in=lot_nos).values_list('lot_no', 'actual_weight'))
            pallet_group_dict = dict(MaterialTestOrder.objects.filter(
                lot_no__in=lot_nos).values_list('lot_no', 'production_group'))
            max_test_ids = list(MaterialTestResult.objects.filter(
                material_test_order__lot_no__in=lot_nos
            ).values('material_test_order__lot_no').annotate(mid=Max('id')).values_list('mid', flat=True))
            last_test_data = MaterialTestResult.objects.filter(
                id__in=max_test_ids).values('material_test_order__lot_no',
                                            'test_factory_date',
                                            'created_user__username')
            last_test_dict = {i['material_test_order__lot_no']: i for i in last_test_data}
            locked_dict = dict(ProductInventoryLocked.objects.filter(lot_no__in=lot_nos).values_list('lot_no', 'locked_status'))
            stock_lot_nos = list(BzFinalMixingRubberInventory.objects.using('bz').filter(lot_no__in=lot_nos).values_list('lot_no', flat=True)) + \
                            list(BzFinalMixingRubberInventoryLB.objects.using('lb').filter(
                                store_name='炼胶库', lot_no__in=lot_nos).values_list('lot_no', flat=True))
            for item in s_data:
                lot_no = item['lot_no']
                item['actual_weight'] = pallet_weight_dict.get(lot_no)
                item['classes_group'] = item['classes'] + '/' + pallet_group_dict.get(lot_no)  # 班次班组
                item['test'] = {'test_status': '正常',
                                'test_factory_date': last_test_dict.get(lot_no, {}).get('test_factory_date'),
                                'test_class': item['classes'],
                                'test_user': last_test_dict.get(lot_no, {}).get('created_user__username')}
                item['residual_weight'] = None
                item['day_time'] = item['factory_date']
                item["trains"] = ",".join([str(x) for x in range(item['begin_trains'], item['end_trains'] + 1)])
                item['locked_status'] = locked_dict.get(lot_no, 0)
                item['is_instock'] = lot_no in stock_lot_nos
            return self.get_paginated_response(s_data)
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)

    def get_queryset(self):
        equip_no = self.request.query_params.get('equip_no', None)
        product_no = self.request.query_params.get('product_no', None)
        day_time = self.request.query_params.get('day_time', None)
        classes = self.request.query_params.get('classes', None)
        # schedule_name = self.request.query_params.get('schedule_name', None)
        is_print = self.request.query_params.get('is_print', None)
        filter_dict = {'delete_flag': False}
        # pfb_filter = {}
        if day_time:
            filter_dict['factory_date'] = day_time
        if equip_no:
            filter_dict['equip_no'] = equip_no
        if product_no:
            filter_dict['product_no'] = product_no
        if classes:
            filter_dict['classes'] = classes
        # if pfb_filter:
        #     pfb_product_list = MaterialTestOrder.objects.filter(**pfb_filter).values_list('lot_no', flat=True)
        #     filter_dict['lot_no__in'] = list(pfb_product_list)
        if is_print == "已打印":
            filter_dict['print_time__isnull'] = False
        elif is_print == "未打印":
            filter_dict['print_time__isnull'] = True
        pfb_queryset = MaterialDealResult.objects.filter(
            **filter_dict).exclude(status='复测').order_by('factory_date', 'classes', 'equip_no',
                                                          'product_no', 'begin_trains')
        return pfb_queryset


@method_decorator([api_recorder], name="dispatch")
class LevelResultViewSet(ModelViewSet):
    """等级和结果"""
    queryset = LevelResult.objects.filter(delete_flag=False)
    serializer_class = LevelResultSerializer
    filter_backends = (DjangoFilterBackend,)
    permission_classes = (IsAuthenticated,)

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        mdp_set = MaterialDataPointIndicator.objects.filter(level=instance.level, result=instance.deal_result,
                                                            delete_flag=False)
        if mdp_set:
            raise ValidationError('该等级已被使用，不能删除')
        instance.delete_flag = True
        instance.last_updated_user = request.user
        instance.save()
        return Response(status=status.HTTP_204_NO_CONTENT)

    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        if self.request.query_params.get('all'):
            data = queryset.values('id', 'deal_result', 'level')
            return Response({'results': data})
        return super().list(self, request, *args, **kwargs)

    def create(self, request, *args, **kwargs):
        deal_result = self.request.data.get('deal_result', None)
        level = self.request.data.get('level', None)
        if not deal_result or not level:
            raise ValidationError('等级和检测结果必传')
        lr_obj = LevelResult.objects.filter(deal_result=deal_result, level=level, delete_flag=False).first()
        if lr_obj:
            raise ValidationError('不可重复新建')
        lr_obj = LevelResult.objects.filter(deal_result=deal_result, level=level, delete_flag=True).first()
        if lr_obj:
            lr_obj.delete_flag = False
            lr_obj.save()
            return Response('新建成功')
        else:
            serializer = self.get_serializer(data=request.data)
            serializer.is_valid(raise_exception=True)
            self.perform_create(serializer)
            headers = self.get_success_headers(serializer.data)
            return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)


@method_decorator([api_recorder], name="dispatch")
class ProductDayStatistics(APIView):
    """胶料日合格率统计"""

    def get(self, request, *args, **kwargs):
        params = request.query_params
        month_time = params.get('ym_time', datetime.datetime.now()).month
        year_time = params.get('ym_time', datetime.datetime.now()).year
        pass_type = params.get('pass_type', '1')  # 1:综合合格率  2：一次合格率  3：流变合格率
        pass_dict = {'1': ['门尼', '比重', '硬度', '流变'], '2': ['门尼', '比重', '硬度'], '3': ['流变']}
        test_indicator_name_dict = pass_dict[pass_type]
        product_no_list = MaterialTestOrder.objects.filter(delete_flag=False,
                                                           production_factory_date__year=year_time,
                                                           production_factory_date__month=month_time).values(
            'product_no').annotate().distinct()
        ruturn_pass = []
        for product_no_dict in product_no_list:
            return_dict = {}
            return_dict['product_no'] = product_no_dict['product_no']
            for day_time in range(1, int(datetime.datetime.now().day) + 1):
                lot_no_list = MaterialTestOrder.objects.filter(delete_flag=False,
                                                               production_factory_date__year=year_time,
                                                               production_factory_date__month=month_time,
                                                               production_factory_date__day=day_time,
                                                               **product_no_dict).values('lot_no').annotate().distinct()

                # mto_count = lot_no_list.count()
                mto_count = 0  # 粒度比等级综合判定更细，是基于每一车的。而不是每一托的
                if lot_no_list.count() == 0:
                    continue
                pass_count = 0

                for lot_no_dict in lot_no_list:
                    mto_set = MaterialTestOrder.objects.filter(delete_flag=False,
                                                               production_factory_date__year=year_time,
                                                               production_factory_date__month=month_time,
                                                               production_factory_date__day=day_time,
                                                               **product_no_dict, **lot_no_dict).all()
                    if not mto_set:
                        continue
                    mto_count += mto_set.count()
                    for mto_obj in mto_set:
                        level_list = []

                        mrt_list = mto_obj.order_results.filter(
                            test_indicator_name__in=test_indicator_name_dict).all().values('data_point_name').annotate(
                            max_test_time=Max('test_times'))
                        for mrt_dict in mrt_list:
                            mrt_dict_obj = MaterialTestResult.objects.filter(material_test_order=mto_obj,
                                                                             data_point_name=mrt_dict[
                                                                                 'data_point_name'],
                                                                             test_times=mrt_dict[
                                                                                 'max_test_time']).last()
                            level_list.append(mrt_dict_obj)
                        quality_sign = True
                        for mtr_obj in level_list:
                            if not mtr_obj.mes_result:  # mes没有数据
                                if not mtr_obj.result:  # 快检也没有数据
                                    quality_sign = False
                                elif mtr_obj.result != '一等品':
                                    quality_sign = False

                            elif mtr_obj.mes_result == '一等品':
                                if mtr_obj.result not in ['一等品', None]:
                                    quality_sign = False

                            elif mtr_obj.mes_result != '一等品':
                                quality_sign = False

                        if quality_sign:
                            pass_count += 1
                percent_of_pass = str((pass_count / mto_count) * 100) + '%'
                return_dict[f'{month_time}-{day_time}'] = percent_of_pass
            ruturn_pass.append(return_dict)
        return Response(ruturn_pass)


@method_decorator([api_recorder], name="dispatch")
class LabelPrintViewSet(mixins.CreateModelMixin,
                        mixins.UpdateModelMixin,
                        GenericViewSet):
    """
    list: 获取一条打印标签
    create: 存储一条打印标签
    """
    queryset = LabelPrint.objects.all()
    serializer_class = LabelPrintSerializer
    # permission_classes = (IsAuthenticated, )

    def create(self, request, *args, **kwargs):
        lot_no_list = request.data.get('lot_no')
        if not isinstance(lot_no_list, list):
            raise ValidationError('数据格式错误！')
        for lot_no in lot_no_list:
            data = receive_deal_result(lot_no)
            LabelPrint.objects.create(label_type=2, lot_no=lot_no, status=0, data=data)
            try:
                LabelPrintLog.objects.create(result=MaterialDealResult.objects.filter(lot_no=lot_no).first(),
                                             created_user=self.request.user.username,
                                             location='快检')
            except Exception:
                pass
        return Response('打印任务已下发')

    def list(self, request, *args, **kwargs):
        station_dict = {
            "收皮": 1,
            "快检": 2,
            "一层前端": 3,
            "一层后端": 4,
            "二层前端": 5,
            "二层后端": 6,
            "炼胶#出库口#1": 7,
            "炼胶#出库口#2": 8,
            "炼胶#出库口#3": 9,
            "帘布#出库口#0": 10
        }
        station = request.query_params.get("station")
        ip_address = request.query_params.get("ip_address")
        if ip_address:
            instance = self.get_queryset().filter(label_type=station_dict.get(station), status=0, ip_address=ip_address).order_by('id').first()
        else:
            instance = self.get_queryset().filter(label_type=station_dict.get(station), status=0, ip_address__isnull=True).order_by('id').first()
        if instance:
            instance.status = 2
            instance.save()
            serializer = self.get_serializer(instance)
            data = serializer.data
        else:
            data = {}
        if data:
            data["data"] = json.loads(data.get("data"))
        return Response(data)

    @atomic()
    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        data = request.data if request.data else {"status": 1}
        serializer = self.get_serializer(instance, data=data, partial=partial)
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)

        if getattr(instance, '_prefetched_objects_cache', None):
            # If 'prefetch_related' has been applied to a queryset, we need to
            # forcibly invalidate the prefetch cache on the instance.
            instance._prefetched_objects_cache = {}
        MaterialDealResult.objects.filter(lot_no=instance.lot_no).update(print_time=datetime.datetime.now())
        st_ct = instance.created_date - datetime.timedelta(minutes=1)
        et_ct = instance.created_date + datetime.timedelta(minutes=1)
        LabelPrint.objects.filter(
            lot_no=instance.lot_no,
            label_type=instance.label_type
        ).filter(Q(created_date__gt=st_ct) | Q(created_date__lt=et_ct)).update(status=1)
        return Response("打印完成")


@method_decorator([api_recorder], name="dispatch")
class LabelPrintLogView(ListAPIView):
    """打印履历"""
    queryset = LabelPrintLog.objects.all()
    serializer_class = LabelPrintLogSerializer
    permission_classes = (IsAuthenticated,)
    filter_backends = (DjangoFilterBackend,)
    filter_fields = ('result_id',)
    pagination_class = None


@method_decorator([api_recorder], name="dispatch")
class DealSuggestionView(APIView):
    """处理意见展示"""

    def get(self, request, *args, **kwargs):
        queryset = DealSuggestion.objects.filter(delete_flag=False).values('suggestion_desc').annotate().distinct()
        return Response(queryset.values_list('suggestion_desc', flat=True))


@method_decorator([api_recorder], name="dispatch")
class MaterialTestResultHistoryView(APIView):
    """试验结果数据展开列表， 参数：?test_order_id=检测单id"""

    def get(self, request):
        test_order_id = self.request.query_params.get('test_order_id')
        try:
            test_order = MaterialTestOrder.objects.get(id=test_order_id)
        except Exception:
            raise ValidationError('参数错误')
        data = MaterialTestResult.objects.filter(material_test_order=test_order).all()
        max_test_times = MaterialTestResult.objects.filter(material_test_order=test_order
                                                           ).aggregate(max_time=Max('test_times'))['max_time']
        ret = {i: {} for i in range(1, max_test_times + 1)}

        for item in data:
            indicator_name = item.test_indicator_name
            data_point_name = item.data_point_name
            test_times = item.test_times
            test_result = {
                'value': item.value,
                'result': item.result,
                'mes_result': item.mes_result,
                'machine_name': item.machine_name,
                'level': item.level,
                'test_times': item.test_times
            }
            if indicator_name not in ret[test_times]:
                ret[test_times][indicator_name] = {data_point_name: test_result}
            else:
                ret[test_times][indicator_name][data_point_name] = test_result
        return Response(ret)


# @method_decorator([api_recorder], name="dispatch")
# class PrintMaterialDealResult(APIView):
#     """不合格品打印功能"""
#
#     def get(self, request, *args, **kwargs):
#         day = self.request.query_params.get('day', None)
#         status = self.request.query_params.get('status', None)
#         filter_dict = {}
#         if day:
#             filter_dict['production_factory_date__icontains'] = day
#         if status:
#             filter_dict['status'] = status
#         MaterialDealResult.objects.filter()
#         mdr_set = MaterialDealResult.objects.filter(~Q(deal_result="一等品")).filter(~Q(status="复测")).filter(**filter_dict,
#                                                                                                           delete_flag=False)
#         return print_mdr("results", mdr_set)


# 1天
CACHE_TIME_TIMEOUT = 60 * 60 * 24


class AllMixin:

    # @cache_response(timeout=60 * 10, cache='default')
    def list(self, request, *args, **kwargs):
        if 'all' in self.request.query_params:
            queryset = self.filter_queryset(self.get_queryset())
            serializer = self.get_serializer(queryset, many=True)
            return Response(serializer.data)
        return super().list(request, *args, **kwargs)


def get_statics_query_dates(query_params):
    start_time = query_params.get('start_time')
    end_time = query_params.get('end_time')
    try:
        start_time = datetime.datetime.strptime(start_time, '%Y-%m') \
            if start_time else timezone.now() - timedelta(days=365)
        end_time = datetime.datetime.strptime(end_time, '%Y-%m') \
            if end_time else timezone.now()
    except ValueError:
        raise ValidationError('日期格式:yyyy-mm')
    return start_time, end_time


def get_statics_query_date(query_params):
    date = query_params.get('date')
    try:
        date = datetime.datetime.strptime(date, '%Y-%m') if date else timezone.now()
    except ValueError:
        raise ValidationError('日期格式:yyyy-mm')
    return date


@method_decorator([api_recorder], name="dispatch")
class UnqualifiedOrderTrains(APIView):
    """不合格车次汇总列表"""

    def get(self, request, *args, **kwargs):
        if settings.DATABASES['default']['ENGINE'] == 'django.db.backends.oracle':
            engine = 1
        else:
            engine = 2
        if engine == 2:
            where_str = 'where mtr.level>1'
        else:
            where_str = 'where mtr."LEVEL">1'
        st = self.request.query_params.get('st')
        if st:
            if engine == 2:
                where_str += " and date(mto.production_factory_date)>='{}'".format(st)
            else:
                where_str += " and to_char(mto.PRODUCTION_FACTORY_DATE, 'yyyy-mm-dd') >= '{}'".format(st)
        et = self.request.query_params.get('et')
        if et:
            if engine == 2:
                where_str += " and date(mto.production_factory_date)<='{}'".format(et)
            else:
                where_str += " and to_char(mto.PRODUCTION_FACTORY_DATE, 'yyyy-mm-dd') <= '{}'".format(st)

        classes = self.request.query_params.get('classes')
        if classes:
            where_str += " and mto.production_class='{}'".format(classes)
        product_no = self.request.query_params.get('product_no')
        if product_no:
            where_str += " and mto.product_no='{}'".format(product_no)
        if engine == 2:
            sql = """
            select
                   mto.production_factory_date,
                   mto.production_class,
                   mto.production_equip_no,
                   mto.product_no,
                   mto.actual_trains,
                   mtr.test_indicator_name,
                   mtr.data_point_name,
                   mtr.value,
                   mtr.material_test_order_id,
                   udod.id
            from material_test_result mtr
            inner join (select
                   material_test_order_id,
                   test_indicator_name,
                   data_point_name,
                    max(test_times) max_times
                from material_test_result
                group by test_indicator_name, data_point_name, material_test_order_id
                ) tmp on tmp.material_test_order_id=mtr.material_test_order_id
                             and tmp.data_point_name=mtr.data_point_name
                             and tmp.test_indicator_name=mtr.test_indicator_name
                             and tmp.max_times=mtr.test_times
            inner join material_test_order mto on mtr.material_test_order_id = mto.id
            left join unqualified_deal_order_detail udod on mto.id = udod.material_test_order_id
            {};""".format(where_str)
        else:
            sql = """
            select
                   mto.PRODUCTION_FACTORY_DATE,
                   mto.PRODUCTION_CLASS,
                   mto.PRODUCTION_EQUIP_NO,
                   mto.PRODUCT_NO,
                   mto.ACTUAL_TRAINS,
                   mtr.TEST_INDICATOR_NAME,
                   mtr.DATA_POINT_NAME,
                   mtr.VALUE,
                   mtr.MATERIAL_TEST_ORDER_ID,
                   udod.ID
            from MATERIAL_TEST_RESULT mtr
            inner join (
                select
                    MATERIAL_TEST_ORDER_ID,
                    TEST_INDICATOR_NAME,
                    DATA_POINT_NAME,
                    max(TEST_TIMES) as max_times
                from
                     MATERIAL_TEST_RESULT
                group by TEST_INDICATOR_NAME, DATA_POINT_NAME, MATERIAL_TEST_ORDER_ID
                ) tmp on tmp.MATERIAL_TEST_ORDER_ID=mtr.MATERIAL_TEST_ORDER_ID
                             and tmp.DATA_POINT_NAME=mtr.DATA_POINT_NAME
                             and tmp.TEST_INDICATOR_NAME=mtr.TEST_INDICATOR_NAME
                             and tmp.max_times=mtr.TEST_TIMES
            inner join MATERIAL_TEST_ORDER mto on mtr.MATERIAL_TEST_ORDER_ID = mto.ID
            left join UNQUALIFIED_DEAL_ORDER_DETAIL udod on mto.ID = udod.MATERIAL_TEST_ORDER_ID
            {};""".format(where_str)
        cursor = connection.cursor()
        cursor.execute(sql)
        data = cursor.fetchall()
        ret = {}
        form_head_data = set()
        for item in data:
            if item[-1]:
                continue
            item_key = datetime.datetime.strftime(item[0], '%Y-%m-%d') + '-' + item[1] + '-' + item[2] + '-' + item[3]
            data_point_key = item[6] if item[5] == '流变' else item[5]
            form_head_data.add(data_point_key)
            if item_key not in ret:
                ret[item_key] = {
                    'date': datetime.datetime.strftime(item[0], '%Y-%m-%d'),
                    'classes': item[1],
                    'equip_no': item[2],
                    'product_no': item[3],
                    'actual_trains': {item[4]},
                    'indicator_data': {data_point_key: [item[7]]},
                    'order_ids': {item[8]}
                }
            else:
                ret[item_key]['actual_trains'].add(item[4])
                ret[item_key]['order_ids'].add(item[8])
                if data_point_key not in ret[item_key]['indicator_data']:
                    ret[item_key]['indicator_data'][data_point_key] = [item[7]]
                else:
                    ret[item_key]['indicator_data'][data_point_key].append(item[7])
        return Response({'form_head_data': form_head_data,
                         'ret': ret.values()})


@method_decorator([api_recorder], name="dispatch")
class UnqualifiedDealOrderViewSet(ModelViewSet):
    """不合格处置"""
    queryset = UnqualifiedDealOrder.objects.all()
    filter_backends = (DjangoFilterBackend,)
    filter_class = UnqualifiedDealOrderFilter
    permission_classes = (IsAuthenticated, )

    def get_queryset(self):
        queryset = UnqualifiedDealOrder.objects.order_by('-id')
        t_deal = self.request.query_params.get('t_solved')
        c_deal = self.request.query_params.get('c_solved')
        factory_date = self.request.query_params.get('factory_date')
        equip_no = self.request.query_params.get('equip_no')
        classes = self.request.query_params.get('classes')
        product_no = self.request.query_params.get('product_no')
        if t_deal == 'Y':  # 技术部门已处理
            queryset = queryset.filter(state=1)
        elif t_deal == 'N':  # 技术部门未处理
            queryset = queryset.filter(state=0)
        if c_deal == 'Y':  # 检查部门已处理
            queryset = queryset.filter(state=2)
        elif c_deal == 'N':  # 检查部门未处理
            queryset = queryset.filter(state=1)
        if factory_date:
            ids1 = UnqualifiedDealOrderDetail.objects.filter(
                factory_date=factory_date).values_list('unqualified_deal_order_id', flat=True)
            queryset = queryset.filter(id__in=ids1)
        if equip_no:
            ids2 = UnqualifiedDealOrderDetail.objects.filter(
                equip_no=equip_no).values_list('unqualified_deal_order_id', flat=True)
            queryset = queryset.filter(id__in=ids2)
        if classes:
            ids3 = UnqualifiedDealOrderDetail.objects.filter(
                classes=classes).values_list('unqualified_deal_order_id', flat=True)
            queryset = queryset.filter(id__in=ids3)
        if product_no:
            ids4 = UnqualifiedDealOrderDetail.objects.filter(
                product_no=product_no).values_list('unqualified_deal_order_id', flat=True)
            queryset = queryset.filter(id__in=ids4)
        return queryset.distinct()

    def get_serializer_class(self):
        if self.action == 'create':
            return UnqualifiedDealOrderCreateSerializer
        elif self.action == 'retrieve':
            return UnqualifiedDealOrderCreateSerializer
        elif self.action in ('update', 'partial_update'):
            return UnqualifiedDealOrderUpdateSerializer
        else:
            return UnqualifiedDealOrderSerializer


@method_decorator([api_recorder], name="dispatch")
class DealMethodHistoryView(APIView):

    def get(self, request):
        return Response(set(UnqualifiedDealOrder.objects.filter(
            deal_method__isnull=False).values_list('deal_method', flat=True)))


@method_decorator([api_recorder], name="dispatch")
class TestDataPointCurveView(APIView):
    """胶料数据点检测值曲线"""
    permission_classes = (IsAuthenticated, PermissionClass({'view': 'curve_result_info'}))

    def get(self, request):
        st = self.request.query_params.get('st')
        et = self.request.query_params.get('et')
        product_no = self.request.query_params.get('product_no')
        equip_no = self.request.query_params.get('equip_no')
        if not all([st, et, product_no]):
            raise ValidationError('参数缺失')
        # try:
        #     days = date_range(datetime.datetime.strptime(st, '%Y-%m-%d'),
        #                       datetime.datetime.strptime(et, '%Y-%m-%d'))
        # except Exception:
        #     raise ValidationError('参数错误')
        query_set = MaterialTestResult.objects.filter(material_test_order__production_factory_date__gte=st,
                                                      material_test_order__production_factory_date__lte=et,
                                                      material_test_order__product_no=product_no)
        if equip_no:
            query_set = query_set.filter(material_test_order__production_equip_no=equip_no)
            equip_nos = [equip_no]
        else:
            equip_nos = list(query_set.values(
                'material_test_order__production_equip_no'
            ).annotate(a=Count('id')).values_list('material_test_order__production_equip_no', flat=True))

        indicators = MaterialDataPointIndicator.objects.filter(
            material_test_method__material__material_no=product_no,
            level=1).values('data_point__name', 'upper_limit', 'lower_limit')
        sorted_rules = {'MH': 7, 'ML': 8, 'TC10': 9, 'TC50': 10, 'TC90': 11,
                        '比重值': 12, 'ML(1+4)': 13, '硬度值': 14,
                        'M300': 15, '扯断强度': 16, '伸长率%': 17, '焦烧': 18, '钢拔': 19}
        indicators = sorted(indicators, key=lambda d: sorted_rules.get(d['data_point__name'], 999))
        data_point_names = []
        ret = []
        indicators_dict = {}
        for i in indicators:
            data_point_name = i['data_point__name']
            indicators_dict[data_point_name] = {'upper_limit': i['upper_limit'], 'lower_limit': i['lower_limit']}
            data_point_names.append(data_point_name)
            test_data = query_set.filter(
                data_point_name=data_point_name
            ).values(date=F('material_test_order__production_factory_date'), v=F('value')).order_by('date')
            if not test_data:
                continue
            ret.append({'name': data_point_name, 'data': test_data})
        return Response(
            {'indicators': indicators_dict, 'data': ret, 'equip_nos': equip_nos}
        )


@method_decorator([api_recorder], name="dispatch")
class ImportAndExportView(APIView):
    """快检数据导入，一次只能导入同一批生产数据"""
    permission_classes = (IsAuthenticated, )

    def get(self, request, *args, **kwargs):
        """快检数据导入模板"""
        return export_mto()

    @atomic()
    def post(self, request, *args, **kwargs):
        """快检数据导入"""
        file = request.FILES.get('file')
        cur_sheet = get_cur_sheet(file)
        data = get_sheet_data(cur_sheet, start_row=2)
        by_dict = {'MH': 7,
                   'ML': 8,
                   'TC10': 9,
                   'TC50': 10,
                   'TC90': 11,
                   '比重值': 12,
                   'ML(1+4)': 13,
                   '硬度值': 14,
                   'M300': 15,
                   '扯断强度': 16,
                   '伸长率%': 17,
                   '焦烧': 18,
                   '钢拔': 19}
        # 取第一行数据
        try:
            production_data = data[0]
        except Exception:
            raise ValidationError('请填入检测数据后再导入！')
        # 胶料
        product_no = production_data[0].strip()
        # 密炼日期
        try:
            delta = datetime.timedelta(days=production_data[2])
            date_1 = datetime.datetime.strptime('1899-12-30', '%Y-%m-%d') + delta
            factory_date = datetime.datetime.strftime(date_1, '%Y-%m-%d')
        except Exception:
            raise ValidationError('密炼日期格式错误！')
        # 班次
        classes = production_data[3].strip() + '班'
        # 机台
        equip_no = production_data[4].strip()
        # 班组
        group = production_data[5].strip() + '班'

        if not group:
            raise ValidationError('请输入班组！')
        ws = WorkSchedulePlan.objects.filter(plan_schedule__day_time=factory_date,
                                             classes__global_name=classes,
                                             plan_schedule__work_schedule__work_procedure__global_name='密炼').first()
        if ws:
            group = ws.group.global_name
        # 取数据点
        reverse_dict = {value: key for key, value in by_dict.items()}
        data_points = [reverse_dict.get(i) for i in range(7, 20) if production_data[i]]
        if not data_points:
            raise ValidationError('请填入检测数据后再导入！')
        data_point_method_map = {}
        for data_point_name in data_points:
            mtm = MaterialTestMethod.objects.filter(material__material_no=product_no,
                                                    data_point__name=data_point_name)
            if not mtm:
                raise ValidationError('该胶料{}不存在数据点{}试验方法！'.format(product_no, data_point_name))
            if mtm.count() > 1:
                raise ValidationError('该胶料{}存在多种数据点{}试验方法，请联系管理员！'.format(product_no, data_point_name))
            else:
                data_point_method_map[data_point_name] = mtm.values('id',
                                                                    'test_method__name',
                                                                    'test_method__test_type__test_indicator__name',
                                                                    'is_judged')[0]
                indicator = MaterialDataPointIndicator.objects.filter(
                    material_test_method=mtm.first(),
                    data_point__name=data_point_name,
                    data_point__test_type__test_indicator__name=mtm.first().test_method.test_type.test_indicator.name,
                    level=1
                ).first()
                if indicator:
                    data_point_method_map[data_point_name]['qualified_range'] = [indicator.lower_limit,
                                                                                 indicator.upper_limit]
        pallet_data = PalletFeedbacks.objects.filter(equip_no=equip_no,
                                                     factory_date=factory_date,
                                                     classes=classes,
                                                     product_no=product_no
                                                     ).values('lot_no', 'begin_trains',
                                                              'end_trains', 'plan_classes_uid')
        if not pallet_data:
            raise ValidationError('未找到该批次生产数据：【{}】-【{}】-【{}】-【{}】！！'.format(factory_date, classes, equip_no, product_no))
        pallet_trains_map = {}  # 车次与收皮条码map数据
        for pallet in pallet_data:
            for j in range(pallet['begin_trains'], pallet['end_trains']+1):
                if j not in pallet_trains_map:
                    pallet_trains_map[j] = {'lot_no': [pallet['lot_no']],
                                            'plan_classes_uid': pallet['plan_classes_uid']}
                else:
                    pallet_trains_map[j]['lot_no'].append(pallet['lot_no'])

        del j, data_points, pallet_data, production_data, reverse_dict

        lot_nos = []
        for item in data:
            try:
                actual_trains = int(item[6])
            except Exception:
                raise ValidationError('车次数据错误！')
            if not pallet_trains_map.get(actual_trains):  # 未找到收皮条码
                continue
            for lot_no in pallet_trains_map[actual_trains]['lot_no']:
                lot_nos.append(lot_no)
                plan_classes_uid = pallet_trains_map[actual_trains]['plan_classes_uid']
                validated_data = dict()
                validated_data['material_test_order_uid'] = uuid.uuid1()
                validated_data['actual_trains'] = actual_trains
                validated_data['lot_no'] = lot_no
                validated_data['product_no'] = product_no
                validated_data['plan_classes_uid'] = plan_classes_uid
                validated_data['production_class'] = classes
                validated_data['production_equip_no'] = equip_no
                validated_data['production_factory_date'] = factory_date
                validated_data['production_group'] = group
                while 1:
                    try:
                        instance, created = MaterialTestOrder.objects.get_or_create(
                            defaults=validated_data, **{'lot_no': lot_no,
                                                        'actual_trains': actual_trains})
                        break
                    except Exception:
                        pass
                # is_recheck = False
                for data_point_name, method in data_point_method_map.items():
                    test_method_name = method['test_method__name']
                    test_indicator_name = method['test_method__test_type__test_indicator__name']
                    is_judged = method['is_judged']
                    try:
                        point_value = Decimal(item[by_dict[data_point_name]]).quantize(Decimal('0.000'))
                    except Exception:
                        raise ValidationError('检测值{}数据错误'.format(item[by_dict[data_point_name]]))
                    if not point_value:
                        continue
                    result_data = {'material_test_order': instance,
                                   'test_factory_date': datetime.datetime.now(),
                                   'value': point_value,
                                   'data_point_name': data_point_name,
                                   'test_method_name': test_method_name,
                                   'test_indicator_name': test_indicator_name,
                                   'mes_result': '三等品',
                                   'result': '三等品',
                                   'level': 2,
                                   'is_judged': is_judged,
                                   'created_user': self.request.user,
                                   'test_class': classes,
                                   'test_group': group
                                   }
                    if method.get('qualified_range'):
                        if method['qualified_range'][0] <= point_value <= method['qualified_range'][1]:
                            result_data['mes_result'] = '一等品'
                            result_data['result'] = '一等品'
                            result_data['level'] = 1
                        result_data['judged_upper_limit'] = method['qualified_range'][1]
                        result_data['judged_lower_limit'] = method['qualified_range'][0]
                    else:
                        continue
                    if not created:
                        dp_instances = instance.order_results.filter(data_point_name=data_point_name)
                        if dp_instances:
                            v = dp_instances.first()
                            result_data['value0'] = v.value
                            result_data['judged_upper_limit0'] = v.judged_upper_limit
                            result_data['judged_lower_limit0'] = v.judged_lower_limit
                            dp_instances.delete()
                            instance.is_recheck = True
                            instance.save()
                    MaterialTestResult.objects.create(**result_data)
        gen_pallet_test_result(lot_nos)
        return Response('导入成功')


@method_decorator([api_recorder], name="dispatch")
class BarCodePreview(APIView):
    # 条码追溯中的条码预览接口
    def get(self, request):
        lot_no = request.query_params.get("lot_no")
        # try:
        instance = MaterialDealResult.objects.filter(lot_no=lot_no).first()
        if not instance:
            return Response({})
        serializer = MaterialDealResultListSerializer(instance)
        return Response(serializer.data)
        # except Exception as e:
        #     raise ValidationError(f"该条码无快检结果:{e}")


@method_decorator([api_recorder], name="dispatch")
class ShowQualifiedRange(APIView):

    def get(self, request):
        instance = QualifiedRangeDisplay.objects.first()
        if instance:
            return Response({'is_showed': instance.is_showed})
        return Response({'is_showed': False})

    def post(self, request):
        is_showed = self.request.data.get('is_showed')
        if not isinstance(is_showed, bool):
            raise ValidationError('参数错误')
        instance = QualifiedRangeDisplay.objects.first()
        if not instance:
            QualifiedRangeDisplay.objects.create(is_showed=is_showed)
        else:
            instance.is_showed = is_showed
            instance.save()
        return Response('设置成功！')


@method_decorator([api_recorder], name="dispatch")
class IgnoredProductInfoViewSet(viewsets.GenericViewSet,
                                mixins.ListModelMixin,
                                mixins.CreateModelMixin,
                                mixins.DestroyModelMixin):
    """不做pass章的判定胶种"""
    queryset = IgnoredProductInfo.objects.all()
    serializer_class = IgnoredProductInfoSerializer
    permission_classes = (IsAuthenticated,)


@method_decorator([api_recorder], name="dispatch")
class ProductReportEquipViewSet(mixins.CreateModelMixin,
                                mixins.UpdateModelMixin,
                                mixins.ListModelMixin,
                                viewsets.GenericViewSet):
    """胶料上报设备管理"""
    queryset = ProductReportEquip.objects.order_by('no')
    serializer_class = ProductReportEquipSerializer
    # permission_classes = (IsAuthenticated,)
    filter_backends = (DjangoFilterBackend,)
    filter_class = ProductReportEquipFilter

    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        if request.query_params.get('all'):
            serializer = self.get_serializer(queryset, many=True)
            return Response(serializer.data)

        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)

        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)

    def create(self, request, *args, **kwargs):
        # 更新设备的状态
        if request.data.get('equip'):
            data = request.data.get('data')
            for item in data:
                equip_obj = ProductReportEquip.objects.filter(ip=item['machine']).first()
                if equip_obj:
                    equip_obj.status = 1 if item['status'] else 2
                    equip_obj.save()
            return Response('ok')
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        headers = self.get_success_headers(serializer.data)
        return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)


@method_decorator([api_recorder], name="dispatch")
class ProductReportValueViewSet(mixins.CreateModelMixin,
                                mixins.ListModelMixin,
                                viewsets.GenericViewSet):
    """胶料设备数据上报"""
    permission_classes = (IsAuthenticated,)
    queryset = ProductReportValue.objects.filter(is_binding=False)
    serializer_class = ProductReportValueViewSerializer
    filter_backends = (DjangoFilterBackend,)
    filter_class = ProductReportValueFilter

    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        value_data = queryset.values('id', 'ip', 'value', 'created_date')
        equip_data = ProductReportEquip.objects.values('ip', 'data_point__test_type', 'no',
                                                       'data_point', 'data_point__name')
        equip_data_dict = {item['ip']: item for item in equip_data}
        ret = []
        for item in value_data:
            item['created_date'] = datetime.datetime.strftime(item['created_date'], '%Y-%m-%d %H:%M:%S')
            if item['ip'] in equip_data_dict:
                item['data_point_name'] = equip_data_dict[item['ip']]['data_point__name']
                item['test_type'] = equip_data_dict[item['ip']]['data_point__test_type']
                item['data_point'] = equip_data_dict[item['ip']]['data_point']
                item['report_equip_no'] = equip_data_dict[item['ip']]['no']
            ret.append(item)
        return Response(ret)

    def create(self, request, *args, **kwargs):
        data = request.data
        if not isinstance(data, list):
            raise ValidationError('参数错误')
        for item in data:
            s = ProductReportValueViewSerializer(data=item, context={'request': request})
            if not s.is_valid():
                raise ValidationError(s.errors)
            s.save()
        return Response('新建成功！')


"""新原材料快检"""


@method_decorator([api_recorder], name="dispatch")
class ExamineValueUnitViewSet(viewsets.GenericViewSet,
                              mixins.ListModelMixin,
                              mixins.CreateModelMixin,
                              mixins.RetrieveModelMixin):
    """
    list:
        检测值单位列表
    create:
        新建检测值单位
    """
    queryset = ExamineValueUnit.objects.all()
    serializer_class = ExamineValueUnitSerializer
    pagination_class = SinglePageNumberPagination


class MaterialEquipTypeViewSet(viewsets.GenericViewSet,
                               mixins.ListModelMixin,
                               mixins.CreateModelMixin,
                               mixins.RetrieveModelMixin,
                               mixins.UpdateModelMixin):
    """
    list:
        检测设备类型列表
    create:
        创建检测设备类型
    update:
        修改检测设备类型
    """
    queryset = MaterialEquipType.objects.all()
    serializer_class = MaterialEquipTypeSerializer
    permission_classes = (IsAuthenticated,)
    filter_backends = (DjangoFilterBackend,)
    pagination_class = None

    def get_serializer_class(self):
        if self.action == 'update':
            return MaterialEquipTypeUpdateSerializer
        return MaterialEquipTypeSerializer


class MaterialEquipViewSet(viewsets.GenericViewSet,
                           mixins.ListModelMixin,
                           mixins.CreateModelMixin,
                           mixins.RetrieveModelMixin,
                           mixins.UpdateModelMixin):
    """
    list:
        检测设备列表
    create:
        创建检测设备
    update:
        修改检测设备
    """
    queryset = MaterialEquip.objects.all()
    serializer_class = MaterialEquipSerializer
    permission_classes = (IsAuthenticated,)
    filter_backends = (DjangoFilterBackend,)
    filter_class = MaterialEquipFilter

    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        examine_type_id = self.request.query_params.get('examine_type')
        if examine_type_id:
            try:
                examine_type = MaterialExamineType.objects.get(id=examine_type_id)
            except Exception:
                raise ValidationError('参数错误')
            data = queryset.filter(equip_type__examine_type=examine_type).values('id', 'equip_name')
            return Response({'results': data})
        return super().list(request, *args, **kwargs)


@method_decorator([api_recorder], name="dispatch")
class MaterialExamineTypeViewSet(viewsets.GenericViewSet,
                                 mixins.ListModelMixin,
                                 mixins.CreateModelMixin,
                                 mixins.RetrieveModelMixin,
                                 mixins.UpdateModelMixin):
    """
    list:
        原材料检测类型列表
    create:
        创建原材料检测类型
    update:
        修改原材料检测类型
    """
    queryset = MaterialExamineType.objects.all().select_related("unit").prefetch_related('standards')
    serializer_class = MaterialExamineTypeSerializer
    filter_backends = (DjangoFilterBackend,)
    filter_class = MaterialExamineTypeFilter
    titles = ["比值类型", "检测类型", "边界值", "单位", "上限值", "下限值", "级别"]
    description = "比值类型填 上下限, <=, >=, 外观确认"
    permission_classes = ()

    def get_permissions(self):
        if self.request.query_params.get('all'):  # 检测类型列表
            return ()
        elif self.request.query_params.get('types'):  # 比值类型列表
            return ()
        else:
            return (IsAuthenticated(),)

    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        if self.request.query_params.get('all'):
            data = queryset.values("id", "name", 'interval_type', 'limit_value')
            for item in data:
                if item['interval_type'] == 1:
                    standard = MaterialExamineRatingStandard.objects.filter(level=1, examine_type=item['id']).first()
                    if standard:
                        item['qualified_range'] = [standard.lower_limiting_value, standard.upper_limit_value]
                elif item['interval_type'] == 2:
                    item['qualified_range'] = [None, item['limit_value']]
                elif item['interval_type'] == 3:
                    item['qualified_range'] = [item['limit_value'], None]
            return Response({'results': data})
        elif self.request.query_params.get('types'):
            qs = [{"id": x[0], "value": x[1]} for x in MaterialExamineType.INTERVAL_TYPES]
            return Response({'results': qs})
        else:
            return super().list(request, *args, **kwargs)

    @action(methods=['get'], detail=False, permission_classes=[IsAuthenticated], url_path='export-template',
            url_name='export-template')
    def export_template(self, request):
        """资产导入模板"""

        filename = '原材料检测指标导入模板'
        return get_template_response(self.titles, filename=filename, description="比值类型填: 上下限, <=, >=, 外观确认")

    @action(methods=['post'], detail=False, permission_classes=[IsAuthenticated], url_path='import-data',
            url_name='import-data')
    def import_data(self, request):
        file = request.FILES.get('file')
        cur_sheet = get_cur_sheet(file)
        data = get_sheet_data(cur_sheet, start_row=2)
        interval_dict = {x[1]: x[0] for x in MaterialExamineType.INTERVAL_TYPES}
        for x in data:
            if x[3]:
                unit, tag = ExamineValueUnit.objects.get_or_create(name=x[3])
            else:
                unit = None
            temp = {
                "interval_type": interval_dict[x[0]],
                "limit_value": float(x[2]) if x[2] else None,
                "unit": unit
            }
            instance = MaterialExamineType.objects.get_or_create(defaults=temp, **{"name": x[1]})[0]
            if str(x[0]) == "上下限":
                MaterialExamineRatingStandard.objects.create(examine_type=instance, upper_limit_value=float(x[4]),
                                                             lower_limiting_value=float(x[5]), level=x[6])
        return Response("导入成功")


@method_decorator([api_recorder], name="dispatch")
class ExamineMaterialViewSet(viewsets.GenericViewSet,
                             mixins.CreateModelMixin,
                             mixins.ListModelMixin,
                             mixins.RetrieveModelMixin):
    """
    list:
        原材料列表
    create:
        创建原材料
    """
    queryset = ExamineMaterial.objects \
        .prefetch_related(Prefetch('examine_results',
                                   queryset=MaterialExamineResult.objects.order_by(
                                       '-examine_date', '-create_time'))
                          ).distinct().order_by('-create_time')
    permission_classes = (IsAuthenticated,)
    filter_backends = (DjangoFilterBackend,)
    filter_class = ExamineMaterialFilter

    def get_serializer_class(self):
        if self.action == 'create':
            return ExamineMaterialCreateSerializer
        return ExamineMaterialSerializer

    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        deal_status = self.request.query_params.get('deal_status')
        if deal_status == '未处理':
            queryset = queryset.filter(qualified=False)
        if self.request.query_params.get('all'):
            data = queryset.values("id", "name", 'sample_name', 'batch', 'supplier')
            return Response({'results': data})
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)

    @action(methods=['post'], detail=False)
    def disqualification(self, request):
        material_ids = self.request.data.get('material_ids')
        desc = self.request.data.get('desc')
        deal_result = self.request.data.get('deal_result')
        materials = ExamineMaterial.objects.filter(id__in=material_ids)
        batch_nos = list(materials.values_list('batch', flat=True))
        if WmsNucleinManagement.objects.filter(
                locked_status='已锁定',
                batch_no__in=batch_nos).exists():
            raise ValidationError('该批次物料已锁定核酸管控，无法处理！')
        materials.update(
            deal_status='已处理',
            deal_result=deal_result,
            desc=desc,
            deal_username=self.request.user.username,
            deal_time=datetime.datetime.now(),
            status=1
        )
        # if deal_result == '放行':
        #     url = WMS_URL + '/MESApi/UpdateTestingResult'
        #     for m in materials:
        #         data = {
        #             "TestingType": 2,
        #             "SpotCheckDetailList": [{
        #                 "BatchNo": m.batch,
        #                 "MaterialCode": m.wlxxid,
        #                 "CheckResult": 1
        #             }]
        #         }
        #         headers = {"Content-Type": "application/json ;charset=utf-8"}
        #         try:
        #             r = requests.post(url, json=data, headers=headers, timeout=5)
        #             r = r.json()
        #         except Exception:
        #             continue
        #         resp_status = r.get('state')
        #         m.status = 2 if resp_status == 1 else 3
        #         m.save()
        return Response('成功')


class WMSMaterialSearchView(APIView):
    """根据条码号搜索中策总厂wms物料信息，参数:?tmh=BHZ12105311651140001"""

    def get(self, request):
        # ret = [{"ZCDBH": "DDRK2211150241", "KFID": "4", "KFMC": "中策安吉", "TOFAC": "AJ1", "TOFACNM": "AJ1",
        #         "TMH": "BAJ12211151454000002", "WLXXID": "0F049162-3479-4C5E-8031-A8373289BE03", "WLMC": "溴化丁基2828(京博)",
        #         "WLDWXXID": "186A2FBC-A8C6-4706-8576-5CA19A86A0C8", "WLDWMC": "山东京博中聚新材料有限公司", "CD": "京博",
        #         "SCRQ": "2022-11-14", "SYQX": "2023-11-13", "PH": "202211151453", "PDM": "B2828", "SL": 36.0,
        #         "BZDW": "块（34千克）", "ZL": 1224.0, "ZLDW": "千克kg", "SLDW": "块", "SM_USERID": "5008",
        #         "SM_CREATE": "2022-11-15 17:06:25", "DDH": "CGDD2211100021", "ZSL": 864.0}]
        tmh = self.request.query_params.get('tmh')
        if not tmh:
            raise ValidationError('请输入条码号')
        url = 'http://10.1.10.157:9091/WebService.asmx?wsdl'
        try:
            client = Client(url)
            json_data = {"tofac": "AJ1", "tmh": tmh}
            resp = client.service.FindZcdtmList(json.dumps(json_data))
        except Exception:
            raise ValidationError('网络异常！')
        try:
            data = json.loads(resp)
        except Exception:
            raise ValidationError(resp)
        ret = data.get('Table')
        if not ret:
            raise ValidationError('未找到该条码对应物料信息！')
        return Response(ret)


@method_decorator([api_recorder], name="dispatch")
class MaterialInspectionRegistrationViewSet(viewsets.GenericViewSet,
                                            mixins.ListModelMixin,
                                            mixins.RetrieveModelMixin,
                                            mixins.CreateModelMixin,
                                            mixins.DestroyModelMixin):
    queryset = MaterialInspectionRegistration.objects.order_by('-id')
    serializer_class = MaterialInspectionRegistrationSerializer
    filter_backends = (DjangoFilterBackend,)
    filter_class = MaterialInspectionRegistrationFilter
    permission_classes = (IsAuthenticated, PermissionClass({'view': 'view_material_sjdj',
                                                            'add': 'add_material_sjdj',
                                                            'delete': 'delete_material_sjdj'}))
    FILE_NAME = '原材料总部送检条码登记'
    EXPORT_FIELDS_DICT = {"质检条码": "tracking_num", "总部品质状态": "quality_status", "物料名称": "material_name",
                          "物料编码": "material_no", "批次号": "batch", "安吉质检状态": "mes_quality_status",
                          "核酸状态": "locked_status", "送检日期": "created_date", "送检人": "created_username"}

    def list(self, request, *args, **kwargs):
        export = self.request.query_params.get('export')
        queryset = self.filter_queryset(self.get_queryset())
        page = self.paginate_queryset(queryset)
        if export:
            data = self.get_serializer(queryset, many=True).data
            return gen_template_response(self.EXPORT_FIELDS_DICT, data, self.FILE_NAME)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)


@method_decorator([api_recorder], name="dispatch")
class MaterialExamineResultViewSet(viewsets.GenericViewSet,
                                   mixins.CreateModelMixin,
                                   mixins.ListModelMixin,
                                   mixins.RetrieveModelMixin,
                                   mixins.UpdateModelMixin):
    """
    list:
        原材料检测结果列表
    create:
        创建原材料检测结果
    update:
        修改原材料检测结果
    """
    queryset = MaterialExamineResult.objects.all().select_related(
        "material", "recorder", "sampling_user").prefetch_related("single_examine_results").order_by('-id')
    serializer_class = MaterialExamineResultMainSerializer
    permission_classes = (IsAuthenticated,)
    filter_backends = (DjangoFilterBackend,)
    filter_class = MaterialExamineResultFilter

    def get_serializer_class(self):
        if self.action == 'create':
            return MaterialExamineResultMainCreateSerializer
        return MaterialExamineResultMainSerializer

    def get_queryset(self):
        type_id = self.request.query_params.get('type_id')
        if type_id:
            result_ids = MaterialSingleTypeExamineResult.objects.filter(
                type_id=type_id).values_list('material_examine_result_id', flat=True)
            return self.queryset.filter(id__in=result_ids)
        return self.queryset


@method_decorator([api_recorder], name="dispatch")
class MaterialSingleTypeExamineResultView(APIView):
    """批次原材料不合格项，参数：material=原材料id"""

    def get(self, request):
        ret = {}
        material_id = self.request.query_params.get('material')
        if not material_id:
            raise ValidationError('参数缺失！')
        try:
            material = ExamineMaterial.objects.get(id=material_id)
        except Exception:
            raise ValidationError('该原材料不存在！')
        last_examine_result = MaterialExamineResult.objects.filter(material=material, qualified=False).last()
        material_data = model_to_dict(material)
        ret['material_data'] = material_data
        if last_examine_result:
            material_data['re_examine'] = last_examine_result.re_examine
            material_data['recorder_username'] = last_examine_result.recorder.username
            material_data['sampling_username'] = last_examine_result.sampling_user.username
            material_data['examine_date'] = last_examine_result.examine_date
            material_data['transport_date'] = last_examine_result.transport_date
            ret['unqualified_type_data'] = last_examine_result.single_examine_results.filter(
                mes_decide_qualified=False).values('value', 'type__name')
        ret['mode'] = {'mode': material.desc,
                       'created_username': material.deal_username,
                       'create_time': material.deal_time.strftime('%Y-%m-%d %H:%M:%S') if material.deal_time else None}
        return Response(ret)


@method_decorator([api_recorder], name="dispatch")
class ExamineResultCurveView(APIView):
    """原材料历史检测类型值记录"""

    def get(self, request):
        material_id = self.request.query_params.get('material')
        st = self.request.query_params.get('st')
        et = self.request.query_params.get('et')
        if not all([st, et, material_id]):
            raise ValidationError('参数缺失')

        if not material_id:
            raise ValidationError('参数缺失！')
        try:
            days = date_range(datetime.datetime.strptime(st, '%Y-%m-%d'),
                              datetime.datetime.strptime(et, '%Y-%m-%d'))
            material = ExamineMaterial.objects.get(id=material_id)
        except Exception:
            raise ValidationError('参数错误！')
        last_type_results = MaterialSingleTypeExamineResult.objects.filter(
            material_examine_result__material=material,
            material_examine_result__examine_date__in=days).values(
            'material_examine_result__examine_date',
            'type__name').annotate(max_id=Max('id')).values_list('max_id', flat=True)
        examine_results = MaterialSingleTypeExamineResult.objects.filter(
            id__in=last_type_results)
        type_names = set(examine_results.values_list('type__name', flat=True))
        y_axis = {
            type_name: {
                'name': type_name,
                'type': 'line',
                'data': [0] * len(days)}
            for type_name in type_names
        }
        for item in examine_results:
            date = datetime.datetime.strftime(item.material_examine_result.examine_date, '%Y-%m-%d')
            y_axis[item.type.name]['data'][days.index(date)] = item.value
        return Response({'x_axis': days, 'y_axis': y_axis.values()})


@method_decorator([api_recorder], name='dispatch')
class MaterialReportEquipViewSet(mixins.CreateModelMixin,
                                 mixins.ListModelMixin,
                                 mixins.UpdateModelMixin,
                                 mixins.RetrieveModelMixin,
                                 viewsets.GenericViewSet):
    """
    list:
        原材料上报设备列表
    create:
        新建原材料上报设备
    update：
        修改原材料上报设备
    """
    queryset = MaterialReportEquip.objects.all()
    permission_classes = (IsAuthenticated,)
    filter_backends = (DjangoFilterBackend,)
    filter_class = MaterialReportEquipFilter
    serializer_class = MaterialReportEquipSerializer


@method_decorator([api_recorder], name='dispatch')
class MaterialReportValueViewSet(mixins.CreateModelMixin,
                                 mixins.ListModelMixin,
                                 viewsets.GenericViewSet):
    queryset = MaterialReportValue.objects.filter(is_binding=False)
    permission_classes = (IsAuthenticated,)
    pagination_class = SinglePageNumberPagination
    filter_backends = (DjangoFilterBackend,)
    filter_class = MaterialReportValueFilter

    def get_serializer_class(self):
        if self.action == 'create':
            return MaterialReportValueCreateSerializer
        return MaterialReportValueSerializer

    def create(self, request, *args, **kwargs):
        data = self.request.data
        for item in data:
            s = MaterialReportValueCreateSerializer(data=item, context={'request': self.request})
            s.is_valid(raise_exception=True)
            self.perform_create(s)
        return Response("提交成功")

    def list(self, request, *args, **kwargs):
        # 获取类型判断值
        data = MaterialExamineType.objects.all().select_related("unit").prefetch_related('standards').values("id",
                                                                                                             "name",
                                                                                                             'interval_type',
                                                                                                             'limit_value')
        for item in data:
            if item['interval_type'] == 1:
                standard = MaterialExamineRatingStandard.objects.filter(level=1, examine_type=item['id']).first()
                if standard:
                    item['qualified_range'] = [standard.lower_limiting_value, standard.upper_limit_value]
            elif item['interval_type'] == 2:
                item['qualified_range'] = [None, item['limit_value']]
            elif item['interval_type'] == 3:
                item['qualified_range'] = [item['limit_value'], None]

        # 返回未绑定数据
        queryset = self.filter_queryset(self.get_queryset())
        prepare_data = list(queryset.values())
        for i in prepare_data:
            row = list(MaterialReportEquip.objects.filter(ip=i['ip']).values('no', 'type'))[0]
            for j in data:
                j['type'] = j.pop('id') if 'id' in j else j['type']
                if j['type'] == row['type']:
                    row.update(j)
            i.update(row)
            if None in i['qualified_range']:
                if i['qualified_range'].index(None) == 0:
                    i['qualified'] = 1 if i['value'] <= i['qualified_range'][1] else 0
                else:
                    i['qualified'] = 1 if i['qualified_range'][0] <= i['value'] else 0
            else:
                i['qualified'] = 1 if i['qualified_range'][0] <= i['value'] <= i['qualified_range'][1] else 0
        return Response({'results': prepare_data})


@method_decorator([api_recorder], name="dispatch")
class ProductTestPlanViewSet(ModelViewSet):

    """门尼检测计划"""
    queryset = ProductTestPlan.objects.prefetch_related(
        Prefetch('product_test_plan_detail', queryset=ProductTestPlanDetail.objects.order_by('id')),
    ).order_by('-id')
    serializer_class = ProductTestPlanSerializer
    permission_classes = (IsAuthenticated, )
    filter_backends = (DjangoFilterBackend,)
    filter_fields = ('id', 'test_equip', 'status')

    def perform_destroy(self, instance):
        """结束检测"""
        instance.status = 4
        ProductTestPlanDetail.objects.filter(test_plan=instance, status=1).update(status=4)
        instance.save()

    @action(methods=['get'], detail=False, permission_classes=[IsAuthenticated], url_path='underway-plan',
            url_name='underway-plan')
    def underway_plan(self, request):
        """查看机台正在检测或者上一次检测完成但是未合格的计划"""
        test_equip = self.request.query_params.get('test_equip')
        try:
            test_equip = int(test_equip)
        except Exception:
            raise ValidationError('参数错误！')
        current_plan = self.queryset.filter(status=1, test_equip=test_equip).first()
        if current_plan:
            data = self.serializer_class(instance=current_plan).data
            return Response(data)
        else:
            current_plan = self.queryset.filter(test_equip=test_equip).order_by('id').last()
            if current_plan:
                product_test_plan_detail = list(
                    current_plan.product_test_plan_detail.filter(is_qualified=False).order_by('id').values(
                        'equip_no', 'product_no', 'lot_no', 'factory_date', 'production_classes',
                        'production_group', 'actual_trains', 'is_qualified'))
                if not product_test_plan_detail:
                    return Response({})
                for item in product_test_plan_detail:
                    item['is_recheck'] = True
                data = {'count': current_plan.count,
                        "test_classes": current_plan.test_classes,
                        "test_group": current_plan.test_group,
                        "test_indicator_name": current_plan.test_indicator_name,
                        "test_method_name": current_plan.test_method_name,
                        "test_times": current_plan.test_times,
                        "test_interval": current_plan.test_interval,
                        "test_equip": current_plan.test_equip_id,
                        "product_test_plan_detail": product_test_plan_detail
                        }
                return Response(data)
            else:
                return Response({})


@method_decorator([api_recorder], name="dispatch")
class ProductTestPlanDetailViewSet(ModelViewSet):
    """门尼检测计划详情"""
    queryset = ProductTestPlanDetail.objects.all()
    serializer_class = ProductTestPlanDetailSerializer
    permission_classes = (IsAuthenticated,)
    filter_backends = (DjangoFilterBackend,)

    def perform_destroy(self, instance):
        if instance.value:
            raise ValidationError('该数据已检测，无法删除！')
        return super().perform_destroy(instance)

    @action(methods=['post'], detail=False)
    def bulk_create(self, request):
        serializer = ProductTestPlanDetailBulkCreateSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        return Response('ok')


@method_decorator([api_recorder], name="dispatch")
class ProductTestResumeViewSet(mixins.ListModelMixin, GenericViewSet):
    """门尼检测履历"""
    queryset = ProductTestPlanDetail.objects.order_by('-id')
    serializer_class = ProductTEstResumeSerializer
    permission_classes = (IsAuthenticated,)
    filter_backends = (DjangoFilterBackend,)
    filter_class = ProductTestResumeFilter


@method_decorator([api_recorder], name="dispatch")
class TestDataView(APIView):
    """设备监控"""
    def get(self, request):
        data = ProductTestPlanDetail.objects.filter(test_plan__test_time__year=datetime.datetime.now().year,
                                                    test_plan__test_time__month=datetime.datetime.now().month,
                                                    test_plan__test_time__day=datetime.datetime.now().day
                                                    ).values('raw_value')
        return Response(data)


@method_decorator([api_recorder], name='dispatch')
class ReportValueView(APIView):
    """
    原材料、胶料检测数据上报，
    {"report_type": 上报类型  1原材料  2胶料，
    "ip": IP地址，
    "value": 检测值{"l_4: 12"},
    "raw_value": ""}
    """

    @atomic()
    def post(self, request):
        # 原材料：{"report_type": 1, "ip": "IP地址", "value": {"l_4": 12}, "raw_value": "机台检测完整数据"}
        # 胶料门尼：{"report_type": 2, "ip": "IP地址", "value": {"l_4: 12"}, "raw_value": "机台检测完整数据"}

        test_type = self.request.data.get('type')
        if not test_type:
            # 门尼数据上报
            s = ReportValueSerializer(data=self.request.data)
            s.is_valid(raise_exception=True)
            data = s.validated_data
            data['value'] = {'ML(1+4)': data['value']['l_4']}
            report_type = data['report_type']
            if report_type == 1:
                # 原材料数据上报
                MaterialReportValue.objects.create(ip=data['ip'],
                                                   created_date=datetime.datetime.now(),
                                                   value=data['value']['ML(1+4)'])
                # 1. 将添加到检测计划详情
                material_test_plan_detail = MaterialTestPlanDetail.objects.filter(
                    material_test_plan__material_report_equip__ip=data['ip'],
                    material_test_plan__status=1,
                    value__isnull=True).first()
                if not material_test_plan_detail:
                    raise ValidationError('当前机台没有进行中的检测计划')
                material_test_plan_detail.value = data['value']['ML(1+4)']
                material_test_plan_detail.status = 2
                # 2. 判断计划详情的检测结果是否是合格
                material = material_test_plan_detail.material
                test_method = material_test_plan_detail.material_test_plan.test_method
                if test_method.interval_type == 1:
                    standard = MaterialExamineRatingStandard.objects.filter(level=1, examine_type=test_method).first()
                    if standard:
                        flat = True if standard.lower_limiting_value <= data['value']['ML(1+4)']  <= standard.upper_limit_value else False
                elif test_method.interval_type == 2:
                    flat = True if data['value']['ML(1+4)'] <= test_method.limit_value else False
                elif test_method.interval_type == 3:
                    flat = True if data['value']['ML(1+4)'] >= test_method.limit_value else False
                material_test_plan_detail.flat = flat
                material_test_plan_detail.save()

                # 4. 添加到 MaterialExamineResult 在添加到 MaterialSingleTypeExamineResult

                if MaterialExamineResult.objects.filter(material=material,
                                                        examine_date=material_test_plan_detail.material_test_plan.test_time).exists():
                    material_examine_result = MaterialExamineResult.objects.filter(material=material,
                                                                                   examine_date=material_test_plan_detail.material_test_plan.test_time).first()
                    if not material_test_plan_detail.flat:
                        material_examine_result.qualified = False
                        material_examine_result.save()
                else:
                    material_examine_result = MaterialExamineResult.objects.create(
                        material=material,
                        examine_date=material_test_plan_detail.material_test_plan.test_time,
                        transport_date=material_test_plan_detail.transport_date,
                        qualified=material_test_plan_detail.flat,
                        re_examine=False,
                        recorder=material_test_plan_detail.recorder,
                        sampling_user=material_test_plan_detail.sampling_user
                    )
                MaterialSingleTypeExamineResult.objects.create(
                    material_examine_result=material_examine_result,
                    type=material_test_plan_detail.material_test_plan.test_method,
                    mes_decide_qualified=material_test_plan_detail.flat,
                    value=data['value']['ML(1+4)'],
                )
                return Response({'msg': '上报成功', 'success': True})
            else:
                # 胶料数据上报
                equip_test_plan = ProductTestPlan.objects.filter(test_equip__ip=data['ip'], status=1).last()
        else:
            # 钢拔/物性数据上报
            data = self.request.data
            equip_test_plan = ProductTestPlan.objects.filter(test_equip__no=data['test_equip_no'], status=1).last()

        # 获取当前机台正在进行中的检测计划
        if not equip_test_plan:
            return Response({'mes': '未找到该机台正在进行中的计划', 'success': False})

        # 获取当前检测任务
        current_test_detail = ProductTestPlanDetail.objects.filter(test_plan=equip_test_plan,
                                                                   status=1
                                                                   ).order_by('id').first()
        if not current_test_detail:
            return Response({'msg': '全部检测完成', 'success': True})

        # 如果是钢拔应检测四/五次
        if test_type:
            if equip_test_plan.test_indicator_name == '钢拔' and test_type == '钢拔':
                # 判断有没有
                ordering = RubberMaxStretchTestResult.objects.filter(
                    product_test_plan_detail=current_test_detail).count() + 1
                RubberMaxStretchTestResult.objects.create(product_test_plan_detail=current_test_detail,
                                                          ordering=ordering,
                                                          speed=data['Speed'],
                                                          max_strength=data['MaxF'],
                                                          max_length=data['MaxL'],
                                                          end_strength=data['BF'],
                                                          end_length=data['BL'],
                                                          yield_strength=data['YieldF'],
                                                          yield_length=data['YieldL'],
                                                          test_time=data['DateTime'],
                                                          test_method=data['TestMethod'],
                                                          ds1=data['DS1'],
                                                          ds2=data['DS2'],
                                                          ds3=data['DS3'],
                                                          result=data['Result'])
                if ordering == equip_test_plan.count:
                    current_test_detail.status = 2  # 完成
                    values = RubberMaxStretchTestResult.objects.filter(product_test_plan_detail=current_test_detail).aggregate(钢拔=Avg('max_strength'))
                    values.update({'钢拔': round(values['钢拔'], 3)})
                    current_test_detail.value = json.dumps(values, ensure_ascii=False)
                    current_test_detail.save()
                else:
                    # current_test_detail.status = 3  # 检测中
                    # current_test_detail.save()
                    return Response('ok')
            # 如果是物性应检测三次
            elif equip_test_plan.test_indicator_name == '物性' and test_type == '物性':
                ordering = RubberMaxStretchTestResult.objects.filter(
                    product_test_plan_detail=current_test_detail).count() + 1
                RubberMaxStretchTestResult.objects.create(product_test_plan_detail=current_test_detail,
                                                          ordering=ordering,
                                                          speed=data['Speed'],
                                                          thickness=data['Thickness'],
                                                          width=data['Width'],
                                                          ds1=data['DS1'],
                                                          ds2=data['DS2'],
                                                          ds3=data['DS3'],
                                                          ds4=data['DS4'],
                                                          max_strength=data['MStrength'],
                                                          max_length=data['MLength'],
                                                          break_strength=data['BSrength'],
                                                          break_length=data['BLength'],
                                                          n1=data['N1'],
                                                          n2=data['N2'],
                                                          n3=data['N3'],
                                                          test_time=data['DateTime'],
                                                          test_method=data['TestMethod'],
                                                          result=data['Result'])
                if ordering == 3:
                    values = RubberMaxStretchTestResult.objects.filter(
                        product_test_plan_detail=current_test_detail).aggregate(扯断强度=Avg('break_strength'),
                                                                                伸长率=Avg('max_length'),
                                                                                M300=Avg('ds2'))
                    values['伸长率%'] = values['伸长率']
                    del values['伸长率']
                    values.update({'扯断强度': round(values['扯断强度'], 3)})
                    values.update({'伸长率%': round(values['伸长率%'], 3)})
                    values.update({'M300': round(values['M300'], 3)})
                    current_test_detail.status = 2  # 完成
                    current_test_detail.value = json.dumps(values, ensure_ascii=False)
                    current_test_detail.save()
                else:
                    # current_test_detail.status = 3  # 检测中
                    # current_test_detail.save()
                    return Response('ok')
        else:
            current_test_detail.value = json.dumps(data['value'])
            current_test_detail.raw_value = data['raw_value']
            current_test_detail.status = 2
            current_test_detail.save()

        # if equip_test_plan.product_test_plan_detail.filter(
        #         value__isnull=False).count() == equip_test_plan.product_test_plan_detail.count():
        #     equip_test_plan.status = 2
        #     equip_test_plan.save()

        product_no = current_test_detail.product_no  # 胶料编码
        production_class = current_test_detail.production_classes  # 班次
        group = current_test_detail.production_group  # 班组
        equip_no = current_test_detail.equip_no  # 机台
        product_date = current_test_detail.factory_date  # 工厂日期
        method_name = equip_test_plan.test_method_name  # 实验方法名称
        indicator_name = equip_test_plan.test_indicator_name  # 实验指标名称
        test_times = equip_test_plan.test_times  # 检测次数

        if equip_test_plan.test_indicator_name == '门尼':
            data_point_list = ['ML(1+4)']
        elif equip_test_plan.test_indicator_name == '钢拔':
            data_point_list = ['钢拔']
        elif equip_test_plan.test_indicator_name == '物性':
            data_point_list = ['扯断强度', '伸长率%', 'M300']
        else:
            data_point_list = []
        is_qualified = True
        test_results = {}
        for k, v in json.loads(current_test_detail.value).items():
            test_results[k] = {"name": k, "value": v, "flag": ""}
        lot_nos = []
        # 根据检测间隔，补充车次相关test_order和test_result表数据
        for train in range(current_test_detail.actual_trains,
                           current_test_detail.actual_trains + equip_test_plan.test_interval):
            pallets = PalletFeedbacks.objects.filter(
                equip_no=equip_no,
                product_no=product_no,
                classes=production_class,
                factory_date=product_date,
                begin_trains__lte=train,
                end_trains__gte=train
            )
            if not pallets:
                continue
            for pallet in pallets:
                lot_no = pallet.lot_no
                lot_nos.append(lot_no)
                test_order_data = {
                    "lot_no": lot_no,
                    'material_test_order_uid': uuid.uuid1(),
                    'actual_trains': train,
                    'product_no': product_no,
                    'plan_classes_uid': pallet.plan_classes_uid,
                    'production_class': production_class,
                    'production_group': group,
                    'production_equip_no': equip_no,
                    'production_factory_date': product_date}
                while 1:
                    try:
                        test_order, created = MaterialTestOrder.objects.get_or_create(
                            defaults=test_order_data, **{'lot_no': lot_no,
                                                         'actual_trains': train})
                        break
                    except Exception:
                        pass

                # 由MES判断检测结果
                material_test_method = MaterialTestMethod.objects.filter(
                    material__material_no=product_no,
                    test_method__name=method_name).first()
                if not material_test_method:
                    continue

                # is_recheck = False
                for data_point in data_point_list:
                    data_point_name = data_point
                    try:
                        if equip_test_plan.test_indicator_name == '门尼':
                            test_value = Decimal(list(json.loads(current_test_detail.value).values())[0]).quantize(Decimal('0.000'))
                        else:
                            test_value = json.loads(current_test_detail.value)[data_point_name]
                    except Exception:
                        raise ValidationError('检测值数据错误')

                    indicator = MaterialDataPointIndicator.objects.filter(
                        material_test_method=material_test_method,
                        data_point__name=data_point_name,
                        data_point__test_type__test_indicator__name=indicator_name,
                        level=1).first()
                    if indicator:
                        if indicator.lower_limit <= test_value <= indicator.upper_limit:
                            mes_result = '一等品'
                            level = 1
                        else:
                            mes_result = '三等品'
                            level = 2
                            try:
                                if test_value > indicator.upper_limit:
                                    test_results[data_point_name]['flag'] = 'H'
                                elif test_value < indicator.lower_limit:
                                    test_results[data_point_name]['flag'] = 'L'
                            except Exception:
                                pass
                            is_qualified = False
                    else:
                        # mes_result = '三等品'
                        # level = 2
                        continue
                    value0 = None
                    judged_upper_limit0 = None
                    judged_lower_limit0 = None
                    if not created:
                        dp_instances = test_order.order_results.filter(data_point_name=data_point_name)
                        if dp_instances:
                            v = dp_instances.first()
                            value0 = v.value
                            judged_upper_limit0 = v.judged_upper_limit
                            judged_lower_limit0 = v.judged_lower_limit
                            dp_instances.delete()
                            test_order.is_recheck = True
                            test_order.save()
                        # try:
                        #     if a[0]:
                        #         is_recheck = True
                        # except Exception:
                        #     pass
                    MaterialTestResult.objects.create(
                        material_test_order=test_order,
                        test_factory_date=datetime.datetime.now(),
                        value=test_value,
                        test_times=test_times,
                        data_point_name=data_point_name,
                        test_method_name=method_name,
                        test_indicator_name=indicator_name,
                        result=mes_result,
                        mes_result=mes_result,
                        machine_name=equip_test_plan.test_equip.no,
                        test_group=group,
                        level=level,
                        test_class=production_class,
                        is_judged=material_test_method.is_judged,
                        created_user=equip_test_plan.created_user,
                        judged_upper_limit=indicator.upper_limit,
                        judged_lower_limit=indicator.lower_limit,
                        value0=value0,
                        judged_upper_limit0=judged_upper_limit0,
                        judged_lower_limit0=judged_lower_limit0
                    )

                # if is_recheck:
                #     test_order.is_recheck = True
                #     test_order.save()

        # test_indicator_name = current_test_detail.test_plan.test_indicator_name
        # mto = MaterialTestOrder.objects.filter(lot_no=current_test_detail.lot_no,
        #                                        actual_trains=current_test_detail.actual_trains).first()
        # if mto:
        #     max_result_ids = list(mto.order_results.filter(
        #         test_indicator_name=test_indicator_name
        #     ).values('data_point_name').annotate(max_id=Max('id')).values_list('max_id', flat=True))
        #     if mto.order_results.filter(id__in=max_result_ids, level__gt=1).exists():
        #         current_test_detail.is_qualified = False
        #     else:
        #         current_test_detail.is_qualified = True
        #     current_test_detail.save()
        current_test_detail.value = json.dumps(list(test_results.values()))
        current_test_detail.is_qualified = is_qualified
        current_test_detail.status = 2
        current_test_detail.save()
        gen_pallet_test_result(lot_nos)
        return Response({'msg': '检测完成', 'success': True})


class CheckEquip(APIView):
    """检测设备状态"""
    def get(self, request):
        equip = ProductReportEquip.objects.order_by('-last_updated_date').first()
        last_date = datetime.datetime.timestamp(equip.last_updated_date)
        now_date = datetime.datetime.timestamp(datetime.datetime.now())
        return Response({'status': False} if now_date - last_date > 10 else {'status': True})


class RubberMaxStretchTestResultViewSet(GenericViewSet, mixins.ListModelMixin, mixins.UpdateModelMixin):
    """物性/钢拔检测数据查看"""
    queryset = RubberMaxStretchTestResult.objects.order_by('ordering')
    serializer_class = RubberMaxStretchTestResultSerializer
    filter_backends = (DjangoFilterBackend,)
    filter_fields = ('product_test_plan_detail_id',)

    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        serializer = self.get_serializer(queryset, many=True)
        avg_value = queryset.aggregate(最大力=Avg('max_strength'),
                                       结束力=Avg('end_strength'),
                                       厚度=Avg('thickness'),
                                       百分之百=Avg('ds1'),
                                       百分之三百=Avg('ds2'),
                                       断裂强力=Avg('break_strength'),
                                       断裂伸长=Avg('break_length'),
                                       )
        for k, v in avg_value.items():
            if avg_value[k]:
                avg_value[k] = round(v, 3)
        return Response({'results': serializer.data, 'avg_value': avg_value})

    def update(self, request, *args, **kwargs):
        if kwargs.get('pk'):
            pk = kwargs.get('pk')
            test_plan_detail_obj = ProductTestPlanDetail.objects.get(test_results__id=pk)
            RubberMaxStretchTestResult.objects.filter(pk=pk).update(**request.data)
            test_plan_obj = ProductTestPlan.objects.filter(product_test_plan_detail=test_plan_detail_obj).first()
            if test_plan_obj.test_indicator_name == '钢拔':
                data_point_list = ['钢拔']
                values = RubberMaxStretchTestResult.objects.filter(product_test_plan_detail=test_plan_detail_obj).aggregate(
                    钢拔=Avg('max_strength'))
                values = {'钢拔': round(values['钢拔'], 3)}
                test_plan_detail_obj.value = values
                test_plan_detail_obj.save()
            elif test_plan_obj.test_indicator_name == '物性':
                data_point_list = ['扯断强度', '伸长率%', 'M300']
                values = RubberMaxStretchTestResult.objects.filter(product_test_plan_detail=test_plan_detail_obj).aggregate(
                                                                            扯断强度=Avg('break_strength'),
                                                                            伸长率=Avg('max_length'),
                                                                            M300=Avg('ds2'))
                values = {'扯断强度': round(values['扯断强度'], 3),
                          '伸长率%': round(values['伸长率'], 3),
                          'M300': round(values['M300'], 3)}
                test_plan_detail_obj.value = values
                test_plan_detail_obj.save()

            for data_point in data_point_list:
                data_point_name = data_point
                test_value = values[data_point_name]
                material_test_method = MaterialTestMethod.objects.filter(
                    material__material_no=test_plan_detail_obj.product_no,
                    test_method__name=test_plan_obj.test_method_name).first()
                if not material_test_method:
                    continue
                indicator = MaterialDataPointIndicator.objects.filter(
                    material_test_method=material_test_method,
                    data_point__name=data_point_name,
                    data_point__test_type__test_indicator__name=test_plan_obj.test_indicator_name,
                    upper_limit__gte=test_value,
                    lower_limit__lte=test_value).first()
                if indicator:
                    mes_result = indicator.result
                    level = indicator.level
                else:
                    mes_result = '三等品'
                    level = 2

                material_test_order_list = MaterialTestOrder.objects.filter(lot_no=test_plan_detail_obj.lot_no,
                                                                            actual_trains=test_plan_detail_obj.actual_trains)
                for material_test_order in material_test_order_list:
                    MaterialTestResult.objects.filter(material_test_order=material_test_order,
                                                      data_point_name=data_point).update(value=test_value,
                                                                                                      result=mes_result,
                                                                                                      mes_result=mes_result,
                                                                                                      level=level)
            return Response('ok')


@method_decorator([api_recorder], name='dispatch')
class UnqualifiedPalletFeedBackListView(ListAPIView):
    """不合格收皮数据列表"""
    queryset = MaterialDealResult.objects.filter(test_result='三等品').order_by('factory_date', 'classes',
                                                                               'equip_no', 'product_no',
                                                                               'begin_trains')
    serializer_class = UnqualifiedPalletFeedBackSerializer
    permission_classes = (IsAuthenticated, )
    filter_backends = (DjangoFilterBackend, )
    filter_class = UnqualifiedPalletFeedBackListFilter

    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        sulfur_flag = self.request.query_params.get('sulfur_flag')
        if sulfur_flag:
            if sulfur_flag == '加硫':
                queryset = queryset.filter(Q(product_no__icontains='-FM') |
                                           Q(product_no__icontains='-RFM') |
                                           Q(product_no__icontains='-RE'))
            else:
                queryset = queryset.exclude(Q(product_no__icontains='-FM') |
                                            Q(product_no__icontains='-RFM') |
                                            Q(product_no__icontains='-RE'))
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)

        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)


@method_decorator([api_recorder], name='dispatch')
class ProductTestStaticsView(APIView):
    """胶料别不合格率统计"""
    permission_classes = (IsAuthenticated,)

    def get(self, request):
        stage = self.request.query_params.get('station', '')  # 段次
        product_standard = self.request.query_params.get('product_type', '')  # 胶料规格
        production_equip_no = self.request.query_params.get('equip_no', '')  # 机台
        production_class = self.request.query_params.get('classes', '')  # 班次
        st = self.request.query_params.get('s_time')  # 开始时间
        et = self.request.query_params.get('e_time')  # 结束时间
        sy_flag = self.request.query_params.get('sy_flag', 'N')
        if not all([st, et]):
            raise ValidationError('请选择日期范围查询！')
        diff = datetime.datetime.strptime(et, '%Y-%m-%d') - datetime.datetime.strptime(st, '%Y-%m-%d')
        if diff.days > 31:
            raise ValidationError('搜索日期跨度不得超过一个月！')
        filter_kwargs = {
            'production_factory_date__gte': st,
            'production_factory_date__lte': et,
        }
        where_str = """ mtr."LEVEL" = 2 AND to_char(mto."PRODUCTION_FACTORY_DATE", 'yyyy-mm-dd') >= '{}' 
                      AND to_char(mto."PRODUCTION_FACTORY_DATE", 'yyyy-mm-dd') <= '{}'""".format(st, et)
        product_filter_str = ""
        if stage:
            product_filter_str += '-{}-'.format(stage)
            filter_kwargs['product_no__icontains'] = product_filter_str
            where_str += " AND mto.PRODUCT_NO like '%-{}-%'".format(stage)
        if product_standard:
            if product_filter_str:
                product_filter_str += '{}-'.format(product_standard)
            else:
                product_filter_str += '-{}-'.format(product_standard)
            filter_kwargs['product_no__icontains'] = product_filter_str
            where_str += " AND mto.PRODUCT_NO like '%-{}-%'".format(product_standard)
        if production_equip_no:
            filter_kwargs['production_equip_no'] = production_equip_no
            where_str += " AND mto.PRODUCTION_EQUIP_NO='{}'".format(production_equip_no)
        if production_class:
            filter_kwargs['production_class'] = production_class
            where_str += " AND mto.PRODUCTION_CLASS='{}'".format(production_class)
        if sy_flag == 'N':
            filter_kwargs['is_experiment'] = False
            where_str += " AND mto.IS_EXPERIMENT={}".format(0)
        mto_data = MaterialTestOrder.objects.filter(**filter_kwargs).values('product_no', 'is_qualified').annotate(qty=Count('id'))

        # 流变总检查车次
        lb_check_cnt = MaterialTestOrder.objects.filter(**filter_kwargs).filter(
            order_results__test_indicator_name='流变').distinct().values('id').count()
        # 流变总检查不合格车次
        lb_qualified_cnt = MaterialTestOrder.objects.filter(**filter_kwargs).filter(
            order_results__test_indicator_name='流变', order_results__level=2).values('id').distinct().count()

        # 一次总检查车次
        yc_check_cnt = MaterialTestOrder.objects.filter(**filter_kwargs).filter(
            order_results__test_indicator_name__in=('门尼', '比重', '硬度')).values('id').distinct().count()
        # 一次总检查不合格车次
        yc_qualified_cnt = MaterialTestOrder.objects.filter(**filter_kwargs).filter(
            order_results__test_indicator_name__in=('门尼', '比重', '硬度'),
            order_results__level=2).values('id').distinct().count()

        # 按照胶料规格分组，每个规格流变检测多少车次
        lb_product_check_data = MaterialTestOrder.objects.filter(**filter_kwargs).filter(
            order_results__test_indicator_name='流变'
        ).values('product_no').annotate(cnt=Count('id', distinct=True))
        # 按照胶料规格分组，每个规格流变合格多少车次
        lb_product_qualified_data = dict(MaterialTestOrder.objects.filter(**filter_kwargs).filter(
            order_results__test_indicator_name='流变',
            order_results__level=2).values('product_no').annotate(cnt=Count('id', distinct=True)).values_list('product_no', 'cnt'))

        # 各胶料流变总检车次、不合格车次
        datas = {}  # {"C590": [2, 1]}
        for item in lb_product_check_data:
            s_pn = item['product_no'].split('-')[2]
            s_total_qty = item['cnt']
            s_unq_qty = lb_product_qualified_data.get(item['product_no'], 0)
            if s_pn in datas:
                datas[s_pn][0] += s_total_qty
                datas[s_pn][1] += s_unq_qty
            else:
                datas[s_pn] = [s_total_qty, s_unq_qty]
        # 按照胶料规格分组，每个规格一次检测多少车次
        yc_product_check_data = MaterialTestOrder.objects.filter(**filter_kwargs).filter(
            order_results__test_indicator_name__in=('门尼', '比重', '硬度')
        ).values('product_no').annotate(cnt=Count('id', distinct=True))
        # 按照胶料规格分组，每个规格一次合格多少车次
        yc_product_qualified_data = dict(MaterialTestOrder.objects.filter(**filter_kwargs).filter(
            order_results__test_indicator_name__in=('门尼', '比重', '硬度'),
            order_results__level=2).values('product_no').annotate(cnt=Count('id', distinct=True)).values_list('product_no', 'cnt'))

        # 各胶料一次总检车次、不合格车次
        data1 = {}  # {"C590": [2, 1]}
        for item in yc_product_check_data:
            pn_1 = item['product_no'].split('-')[2]
            total_qty_1 = item['cnt']
            unq_qty_1 = yc_product_qualified_data.get(item['product_no'], 0)
            if pn_1 in data1:
                data1[pn_1][0] += total_qty_1
                data1[pn_1][1] += unq_qty_1
            else:
                data1[pn_1] = [total_qty_1, unq_qty_1]
        if not mto_data:
            return Response({})

        total_check_count = 0
        total_qualified_count = 0
        # 各胶料总检车、合格、不合格车次
        mto_data_dict = {}  # {"C590": {"total_qty": 12, "qualified_qty": 10, "unqualified_qty": 2}}
        for item in mto_data:
            mto_product_no = item['product_no'].split('-')[2]
            mto_qty = item['qty']
            mto_qualified_qty = 0
            mto_unqualified_qty = 0
            is_qualified = item['is_qualified']
            total_check_count += mto_qty
            if is_qualified:
                mto_qualified_qty += mto_qty
                total_qualified_count += mto_qty
            else:
                mto_unqualified_qty += mto_qty
            if mto_product_no not in mto_data_dict:
                mto_data_dict[mto_product_no] = {
                    'total_qty': mto_qty,
                    'qualified_qty': mto_qualified_qty,
                    'unqualified_qty': mto_unqualified_qty
                }
            else:
                mto_data_dict[mto_product_no]['total_qty'] += mto_qty
                mto_data_dict[mto_product_no]['qualified_qty'] += mto_qualified_qty
                mto_data_dict[mto_product_no]['unqualified_qty'] += mto_unqualified_qty

        sql = """
        select
    temp.PRODUCT_NO product_no,
    temp.DATA_POINT_NAME,
    temp.TEST_INDICATOR_NAME,
    count(*) uqd,
    temp.flag
from (SELECT
       mtr.DATA_POINT_NAME,
       (CASE
           WHEN mtr.VALUE < mtr.JUDGED_LOWER_LIMIT THEN '-'
           WHEN mtr.VALUE > mtr.JUDGED_UPPER_LIMIT THEN '+'
           ELSE '+' END) as flag,
       REGEXP_SUBSTR(PRODUCT_NO, '[^-]+', 1, 3, 'i') as PRODUCT_NO,
       mtr.TEST_INDICATOR_NAME
FROM "MATERIAL_TEST_RESULT" mtr
INNER JOIN "MATERIAL_TEST_ORDER" mto ON (mtr."MATERIAL_TEST_ORDER_ID" = mto."ID")
WHERE {}
    ) temp
group by temp.DATA_POINT_NAME, temp.PRODUCT_NO, temp.TEST_INDICATOR_NAME, temp.flag
order by temp.PRODUCT_NO, temp.TEST_INDICATOR_NAME;""".format(where_str)
        cursor = connection.cursor()
        cursor.execute(sql)
        query_data = cursor.fetchall()
        ret = {}
        for item in query_data:
            product_type = item[0]
            data_point_name = item[1]
            test_indicator_name = item[2]
            qty = item[3]
            flag = item[4]
            mn_unqualified_count, mn_lower_count, mn_upper_count = 0, 0, 0
            bz_unqualified_count, bz_lower_count, bz_upper_count = 0, 0, 0
            yd_unqualified_count, yd_lower_count, yd_upper_count = 0, 0, 0
            ml_unqualified_count, ml_lower_count, ml_upper_count = 0, 0, 0
            mh_unqualified_count, mh_lower_count, mh_upper_count = 0, 0, 0
            tc10_unqualified_count, tc10_lower_count, tc10_upper_count = 0, 0, 0
            tc50_unqualified_count, tc50_lower_count, tc50_upper_count = 0, 0, 0
            tc90_unqualified_count, tc90_lower_count, tc90_upper_count = 0, 0, 0

            if test_indicator_name == '门尼':
                if data_point_name == 'ML(1+4)':
                    mn_unqualified_count = qty
                    if flag == '-':
                        mn_lower_count = qty
                    else:
                        mn_upper_count = qty
            elif test_indicator_name == '硬度':
                yd_unqualified_count = qty
                if flag == '-':
                    yd_lower_count = qty
                else:
                    yd_upper_count = qty
            elif test_indicator_name == '比重':
                bz_unqualified_count = qty
                if flag == '-':
                    bz_lower_count = qty
                else:
                    bz_upper_count = qty
            if test_indicator_name == '流变':
                if data_point_name == 'ML':
                    ml_unqualified_count = qty
                    if flag == '-':
                        ml_lower_count = qty
                    else:
                        ml_upper_count = qty
                elif data_point_name == 'MH':
                    mh_unqualified_count = qty
                    if flag == '-':
                        mh_lower_count = qty
                    else:
                        mh_upper_count = qty
                elif data_point_name == 'TC10':
                    tc10_unqualified_count = qty
                    if flag == '-':
                        tc10_lower_count = qty
                    else:
                        tc10_upper_count = qty
                elif data_point_name == 'TC50':
                    tc50_unqualified_count = qty
                    if flag == '-':
                        tc50_lower_count = qty
                    else:
                        tc50_upper_count = qty
                elif data_point_name == 'TC90':
                    tc90_unqualified_count = qty
                    if flag == '-':
                        tc90_lower_count = qty
                    else:
                        tc90_upper_count = qty
            if product_type not in ret:
                qt_data = mto_data_dict.get(product_type)
                rate_date1 = data1.get(product_type)
                rate_dates = datas.get(product_type)
                if qt_data:
                    check_qty = qt_data.get('total_qty', 0)
                    qualified_qty = qt_data.get('qualified_qty', 0)
                    unqualified_qty = qt_data.get('unqualified_qty', 0)
                    rate = round(qualified_qty / check_qty * 100, 2) if check_qty else ''
                else:
                    check_qty = qualified_qty = unqualified_qty = ''
                    rate = ''
                if rate_date1:
                    rate1 = round((rate_date1[0] - rate_date1[1]) / rate_date1[0] * 100, 2) if rate_date1[0] else ''
                else:
                    rate1 = ''
                if rate_dates:
                    rates = round((rate_dates[0] - rate_dates[1]) / rate_dates[0] * 100, 2) if rate_dates[0] else ''
                    sum_s = rate_dates[1]
                else:
                    rates = ''
                    sum_s = ''
                ret[product_type] = {"product_type": product_type,  # 胶料
                                     "JC": check_qty,  # 检查数量
                                     "HG": qualified_qty,  # 合格数量
                                     "cp_all": unqualified_qty,  # 次品合计
                                     "MN": mn_unqualified_count,  # 门尼不合格数量
                                     "YD": yd_unqualified_count,  # 硬度不合格数量
                                     "BZ": bz_unqualified_count,  # 比重不合格数量
                                     "MH": mh_unqualified_count,  # 流变MH不合格数量
                                     "ML": ml_unqualified_count,  # 流变ML不合格数量
                                     "TC10": tc10_unqualified_count,  # 流变TC10不合格数量
                                     "TC50": tc50_unqualified_count,  # 流变TC50不合格数量
                                     "TC90": tc90_unqualified_count,  # 流变TC90不合格数量
                                     "sum_s": sum_s,  # 流变合计
                                     "mn_lower": mn_lower_count,  # 门尼低于下限不合格数量
                                     "mn_upper": mn_upper_count,  # 门尼高于上限不合格数量
                                     "bz_lower": bz_lower_count,  # 比重低于下限不合格数量
                                     "bz_upper": bz_upper_count,  # 门尼高于上限不合格数量
                                     "yd_lower": yd_lower_count,  # 硬度低于下限不合格数量
                                     "yd_upper": yd_upper_count,  # 硬度高于上限不合格数量
                                     "MH_lower": mh_lower_count,  # MH低于下限不合格数量
                                     "MH_upper": mh_upper_count,  # MH高于上限不合格数量
                                     "ML_lower": ml_lower_count,  # ML低于下限不合格数量
                                     "ML_upper": ml_upper_count,  # ML高于上限不合格数量
                                     "TC10_lower": tc10_lower_count,  # ML低于下限不合格数量
                                     "TC10_upper": tc10_upper_count,  # ML高于上限不合格数量
                                     "TC50_lower": tc50_lower_count,  # ML低于下限不合格数量
                                     "TC50_upper": tc50_upper_count,  # ML高于上限不合格数量
                                     "TC90_lower": tc90_lower_count,  # ML低于下限不合格数量
                                     "TC90_upper": tc90_upper_count,  # ML高于上限不合格数量
                                     'RATE_1_PASS': rate1,
                                     'RATE_S_PASS': rates,
                                     "rate": rate
                                     }
            else:
                ret[product_type]['MN'] += mn_unqualified_count
                ret[product_type]['YD'] += yd_unqualified_count
                ret[product_type]['BZ'] += bz_unqualified_count
                ret[product_type]['MH'] += mh_unqualified_count
                ret[product_type]['ML'] += ml_unqualified_count
                ret[product_type]['TC10'] += tc10_unqualified_count
                ret[product_type]['TC50'] += tc50_unqualified_count
                ret[product_type]['TC90'] += tc90_unqualified_count
                ret[product_type]['mn_lower'] += mn_lower_count
                ret[product_type]['mn_upper'] += mn_upper_count
                ret[product_type]['bz_lower'] += bz_lower_count
                ret[product_type]['bz_upper'] += bz_upper_count
                ret[product_type]['yd_lower'] += yd_lower_count
                ret[product_type]['yd_upper'] += yd_upper_count
                ret[product_type]['MH_lower'] += mh_lower_count
                ret[product_type]['MH_upper'] += mh_upper_count
                ret[product_type]['ML_lower'] += ml_lower_count
                ret[product_type]['ML_upper'] += ml_upper_count
                ret[product_type]['TC10_lower'] += tc10_lower_count
                ret[product_type]['TC10_upper'] += tc10_upper_count
                ret[product_type]['TC50_lower'] += tc50_lower_count
                ret[product_type]['TC50_upper'] += tc50_upper_count
                ret[product_type]['TC90_lower'] += tc90_lower_count
                ret[product_type]['TC90_upper'] += tc90_upper_count

        # 补充全部合格的胶料规格数据
        for k, v in mto_data_dict.items():
            product_type = k
            unqualified_qty = v['unqualified_qty']
            check_qty = v['total_qty']
            qualified_qty = v['qualified_qty']
            if unqualified_qty == 0:
                if product_type not in ret:
                    rate = round(qualified_qty / check_qty * 100, 2) if check_qty else ''
                    rate_date1 = data1.get(product_type)
                    rate_dates = datas.get(product_type)
                    if rate_date1:
                        rate1 = round((rate_date1[0] - rate_date1[1]) / rate_date1[0] * 100, 2) if rate_date1[0] else ''
                    else:
                        rate1 = ''
                    if rate_dates:
                        rates = round((rate_dates[0] - rate_dates[1]) / rate_dates[0] * 100, 2) if rate_dates[0] else ''
                    else:
                        rates = ''
                    ret[product_type] = {
                        "product_type": product_type,  # 胶料
                        "JC": check_qty,  # 检查数量
                        "HG": qualified_qty,  # 合格数量
                        "cp_all": 0,  # 次品合计
                        "MN": '',  # 门尼不合格数量
                        "YD": '',  # 硬度不合格数量
                        "BZ": '',  # 比重不合格数量
                        "MH": '',  # 流变MH不合格数量
                        "ML": '',  # 流变ML不合格数量
                        "TC10": '',  # 流变TC10不合格数量
                        "TC50": '',  # 流变TC50不合格数量
                        "TC90": '',  # 流变TC90不合格数量
                        "sum_s": '',  # 流变合计
                        "mn_lower": '',  # 门尼低于下限不合格数量
                        "mn_upper": '',  # 门尼高于上限不合格数量
                        "bz_lower": '',  # 比重低于下限不合格数量
                        "bz_upper": '',  # 门尼高于上限不合格数量
                        "yd_lower": '',  # 硬度低于下限不合格数量
                        "yd_upper": '',  # 硬度高于上限不合格数量
                        "MH_lower": '',  # MH低于下限不合格数量
                        "MH_upper": '',  # MH高于上限不合格数量
                        "ML_lower": '',  # ML低于下限不合格数量
                        "ML_upper": '',  # ML高于上限不合格数量
                        "TC10_lower": '',  # ML低于下限不合格数量
                        "TC10_upper": '',  # ML高于上限不合格数量
                        "TC50_lower": '',  # ML低于下限不合格数量
                        "TC50_upper": '',  # ML高于上限不合格数量
                        "TC90_lower": '',  # ML低于下限不合格数量
                        "TC90_upper": '',  # ML高于上限不合格数量
                        'RATE_1_PASS': rate1,
                        'RATE_S_PASS': rates,
                        "rate": rate
                    }

        resp_data = ret.values()
        for dt in resp_data:
            for k, v in dt.items():
                if k not in ('rate', 'HG', 'RATE_1_PASS', 'RATE_S_PASS'):
                    if v == 0:
                        dt[k] = ''
        summary_data = {'rate': round(total_qualified_count/total_check_count*100, 2) if total_check_count else '',
                        'rate_1': round((yc_check_cnt-yc_qualified_cnt)/yc_check_cnt*100, 2) if yc_check_cnt else '',
                        'rate_lb': round((lb_check_cnt-lb_qualified_cnt)/lb_check_cnt*100, 2) if lb_check_cnt else ''}
        return Response({'result': resp_data, 'all': summary_data})


@method_decorator([api_recorder], name='dispatch')
class DailyProductTestStaticsView(APIView):
    """胶料别不合格率 规格每日合格率汇总"""
    permission_classes = (IsAuthenticated,)

    def get(self, request):
        stage = self.request.query_params.get('station', '')  # 段次
        product_standard = self.request.query_params.get('product_type', '')  # 胶料规格
        production_equip_no = self.request.query_params.get('equip_no', '')  # 机台
        production_class = self.request.query_params.get('classes', '')  # 班次
        st = self.request.query_params.get('s_time')  # 开始时间
        et = self.request.query_params.get('e_time')  # 结束时间
        sy_flag = self.request.query_params.get('sy_flag', 'N')
        if not all([st, et]):
            raise ValidationError('请选择日期范围查询！')
        diff = datetime.datetime.strptime(et, '%Y-%m-%d') - datetime.datetime.strptime(st, '%Y-%m-%d')
        if diff.days > 31:
            raise ValidationError('搜索日期跨度不得超过一个月！')
        filter_kwargs = {
            'production_factory_date__gte': st,
            'production_factory_date__lte': et,
        }
        where_str = """ mtr."LEVEL" = 2 AND to_char(mto."PRODUCTION_FACTORY_DATE", 'yyyy-mm-dd') >= '{}' 
                      AND to_char(mto."PRODUCTION_FACTORY_DATE", 'yyyy-mm-dd') <= '{}'""".format(st, et)
        product_filter_str = ""
        if stage:
            product_filter_str += '-{}-'.format(stage)
            filter_kwargs['product_no__icontains'] = product_filter_str
            where_str += " AND mto.PRODUCT_NO like '%-{}-%'".format(stage)
        if product_standard:
            if product_filter_str:
                product_filter_str += '{}-'.format(product_standard)
            else:
                product_filter_str += '-{}-'.format(product_standard)
            filter_kwargs['product_no__icontains'] = product_filter_str
            where_str += " AND mto.PRODUCT_NO like '%-{}-%'".format(product_standard)
        if production_equip_no:
            filter_kwargs['production_equip_no'] = production_equip_no
            where_str += " AND mto.PRODUCTION_EQUIP_NO='{}'".format(production_equip_no)
        if production_class:
            filter_kwargs['production_class'] = production_class
            where_str += " AND mto.PRODUCTION_CLASS='{}'".format(production_class)
        if sy_flag == 'N':
            filter_kwargs['is_experiment'] = False
            where_str += " AND mto.IS_EXPERIMENT={}".format(0)
        results = []
        mto_data = MaterialTestOrder.objects.filter(**filter_kwargs).values('production_factory_date', 'product_no', 'is_qualified').annotate(qty=Count('id'))
        if mto_data:
            total_check_count = {}
            total_qualified_count = {}
            # 各胶料总检车、合格、不合格车次
            mto_data_dict = {}  # {'2023-04-26_C590': {"total_qty": 12, "qualified_qty": 10, "unqualified_qty": 2}}
            for item in mto_data:
                mo_date = item['production_factory_date'].strftime('%Y-%m-%d')
                mto_product_no = item['product_no'].split('-')[2]
                mto_qty = item['qty']
                mto_qualified_qty = 0
                mto_unqualified_qty = 0
                is_qualified = item['is_qualified']
                keyword = f"{mo_date}_{mto_product_no}"
                total_check_count[mo_date] = total_check_count.get(mo_date, 0) + mto_qty
                if is_qualified:
                    mto_qualified_qty += mto_qty
                    total_qualified_count[mo_date] = total_qualified_count.get(mo_date, 0) + mto_qty
                else:
                    mto_unqualified_qty += mto_qty
                if keyword not in mto_data_dict:
                    mto_data_dict[keyword] = {'total_qty': mto_qty, 'qualified_qty': mto_qualified_qty, 'unqualified_qty': mto_unqualified_qty}
                else:
                    mto_data_dict[keyword]['total_qty'] += mto_qty
                    mto_data_dict[keyword]['qualified_qty'] += mto_qualified_qty
                    mto_data_dict[keyword]['unqualified_qty'] += mto_unqualified_qty

            sql = """
                            select
                            temp.PRODUCT_NO product_no,
                            temp.DATA_POINT_NAME,
                            temp.TEST_INDICATOR_NAME,
                            count(*) uqd,
                            temp.flag,
                            temp.PRODUCTION_FACTORY_DATE
                        from (SELECT
                               mtr.DATA_POINT_NAME,
                               (CASE
                                   WHEN mtr.VALUE < mtr.JUDGED_LOWER_LIMIT THEN '-'
                                   WHEN mtr.VALUE > mtr.JUDGED_UPPER_LIMIT THEN '+'
                                   ELSE '+' END) as flag,
                               REGEXP_SUBSTR(PRODUCT_NO, '[^-]+', 1, 3, 'i') as PRODUCT_NO,
                               mtr.TEST_INDICATOR_NAME,
                               mto.PRODUCTION_FACTORY_DATE
                        FROM "MATERIAL_TEST_RESULT" mtr
                        INNER JOIN "MATERIAL_TEST_ORDER" mto ON (mtr."MATERIAL_TEST_ORDER_ID" = mto."ID")
                        WHERE {}
                            ) temp
                        group by temp.PRODUCTION_FACTORY_DATE, temp.DATA_POINT_NAME, temp.PRODUCT_NO, temp.TEST_INDICATOR_NAME, temp.flag
                        order by temp.PRODUCTION_FACTORY_DATE, temp.PRODUCT_NO, temp.TEST_INDICATOR_NAME;""".format(where_str)
            cursor = connection.cursor()
            cursor.execute(sql)
            query_data = cursor.fetchall()
            ret = {}
            for item in query_data:
                _date = item[5].strftime('%Y-%m-%d')
                product_type = item[0]
                data_point_name = item[1]
                test_indicator_name = item[2]
                qty = item[3]
                flag = item[4]
                keyword = f"{_date}_{product_type}"
                if keyword not in ret:
                    qt_data = mto_data_dict.get(keyword, {})
                    if qt_data:
                        check_qty = qt_data.get('total_qty', 0)
                        qualified_qty = qt_data.get('qualified_qty', 0)
                        unqualified_qty = qt_data.get('unqualified_qty', 0)
                        rate = round(qualified_qty / check_qty * 100, 2) if check_qty else ''
                    else:
                        check_qty = qualified_qty = unqualified_qty = ''
                        rate = ''
                    ret[keyword] = {"product_type": product_type,  # 胶料
                                    "JC": check_qty,  # 检查数量
                                    "HG": qualified_qty,  # 合格数量
                                    "rate": rate}

            # 补充全部合格的胶料规格数据
            for k, v in mto_data_dict.items():
                date, product_type = k.split('_')
                unqualified_qty = v['unqualified_qty']
                check_qty = v['total_qty']
                qualified_qty = v['qualified_qty']
                if unqualified_qty == 0:
                    if k not in ret:
                        rate = round(qualified_qty / check_qty * 100, 2) if check_qty else ''
                        ret[k] = {
                            "product_type": product_type,  # 胶料
                            "JC": check_qty,  # 检查数量
                            "HG": qualified_qty,  # 合格数量
                            "rate": rate
                        }
            results = [{'product_date': k.split('_')[0], 'product_type': v.get('product_type'), 'ratio': v.get('rate')} for k, v in ret.items()]
        return Response({'results': results})


@method_decorator([api_recorder], name='dispatch')
class ClassTestStaticsView(APIView):
    """班次别不合格率统计"""
    permission_classes = (IsAuthenticated,)

    def get(self, request):
        product_segment = self.request.query_params.get('station', '')
        product_standard = self.request.query_params.get('product_type', '')
        production_equip_no = self.request.query_params.get('equip_no', '')
        production_class = self.request.query_params.get('classes', '')
        start_time = self.request.query_params.get('s_time')
        end_time = self.request.query_params.get('e_time')
        diff = datetime.datetime.strptime(end_time, '%Y-%m-%d') - datetime.datetime.strptime(start_time, '%Y-%m-%d')
        if diff.days > 31:
            raise ValidationError('搜索日期跨度不得超过一个月！')
        product_str = ''
        if product_segment:
            product_str += '-{}'.format(product_segment)
        if product_standard:
            product_str += '-{}'.format(product_standard)
        queryset = MaterialTestResult.objects.filter(
            material_test_order__product_no__icontains=product_str,
            material_test_order__production_factory_date__gte=start_time,
            material_test_order__production_factory_date__lte=end_time,
            material_test_order__production_equip_no__icontains=production_equip_no,
            material_test_order__production_class__icontains=production_class
        )
        # 统计不合格中超过和小于标准的数量
        # ---------------- begin ---------------
        dic = {}
        data_point_dic = {}
        data_point_query = MaterialDataPointIndicator.objects.filter(level=1).values(
            'material_test_method__material__material_no', 'data_point__name',
            'upper_limit', 'lower_limit')
        for i in data_point_query:
            if data_point_dic.get(i['material_test_method__material__material_no']):
                data_point_dic[i['material_test_method__material__material_no']][i['data_point__name']] = [
                    i['lower_limit'], i['upper_limit']]
            else:
                data_point_dic[i['material_test_method__material__material_no']] = {
                    i['data_point__name']: [i['lower_limit'], i['upper_limit']]}

        # res = queryset.values('material_test_order__production_factory_date',
        #                       'material_test_order__product_no',
        #                       'material_test_order__production_class',
        #                       'data_point_name', 'value')
        sql = """
            SELECT
         * 
        FROM
         (
         SELECT
         a.data_point_name,
          a.value,
          b.product_no,
          b.production_factory_date,
          b.production_class,
          a.test_times,
          b.actual_trains,
          b.production_equip_no 
         FROM
          material_test_result a
          LEFT JOIN material_test_order b ON a.material_test_order_id = b.id 
         WHERE
          b.PRODUCTION_EQUIP_NO LIKE '%{}%'
                    AND b.PRODUCT_NO LIKE '%{}%'
                    AND b.PRODUCTION_CLASS LIKE '%{}%'
                    AND b.PRODUCTION_FACTORY_DATE >= TO_DATE('{}', 'YYYY-MM-DD')
                    AND b.PRODUCTION_FACTORY_DATE <= TO_DATE('{}', 'YYYY-MM-DD')

         ) a 
        WHERE
         a.test_times = (
         SELECT
          max( x.test_times ) 
         FROM
          (
          SELECT
           a.value,
           a.test_class,
           a.test_times,
           a.data_point_name,
           b.production_factory_date,
           b.actual_trains,
           b.product_no,
           b.production_equip_no 
          FROM
           material_test_result a
           LEFT JOIN material_test_order b ON a.material_test_order_id = b.id 
          WHERE
           b.PRODUCTION_EQUIP_NO LIKE '%{}%'
                    AND b.PRODUCT_NO LIKE '%{}%'
                    AND b.PRODUCTION_CLASS LIKE '%{}%'
                    AND b.PRODUCTION_FACTORY_DATE >= TO_DATE('{}', 'YYYY-MM-DD')
                    AND b.PRODUCTION_FACTORY_DATE <= TO_DATE('{}', 'YYYY-MM-DD')

          ) x 
         WHERE
          x.data_point_name = a.data_point_name 
          AND x.actual_trains = a.actual_trains 
          AND x.product_no = a.product_no 
         AND x.production_equip_no = a.production_equip_no 
         )""".format(production_equip_no, product_str, production_class, start_time, end_time,
                     production_equip_no, product_str, production_class, start_time, end_time)
        cursor = connection.cursor()
        cursor.execute(sql)
        data = cursor.fetchall()
        for item in data:
            if data_point_dic.get(item[2]):
                data_point_list = data_point_dic[item[2]].get(item[0])
                if data_point_list:
                    MH_lower = MH_upper = ML_lower = ML_upper = TC10_lower = TC10_upper = TC50_lower = TC50_upper = TC90_lower = TC90_upper = 0
                    bz_lower = bz_upper = mn_lower = mn_upper = yd_lower = yd_upper = 0
                    if 'ML(1+4)' == item[0]:
                        mn_lower = 1 if item[1] < data_point_list[0] else 0
                        mn_upper = 1 if item[1] > data_point_list[1] else 0
                    if '比重值' == item[0]:
                        bz_lower = 1 if item[1] < data_point_list[0] else 0
                        bz_upper = 1 if item[1] > data_point_list[1] else 0
                    if '硬度值' == item[0]:
                        yd_lower = 1 if item[1] < data_point_list[0] else 0
                        yd_upper = 1 if item[1] > data_point_list[1] else 0
                    if 'MH' in item[0]:
                        MH_lower = 1 if item[1] < data_point_list[0] else 0
                        MH_upper = 1 if item[1] > data_point_list[1] else 0
                    if 'ML' == item[0]:
                        ML_lower = 1 if item[1] < data_point_list[0] else 0
                        ML_upper = 1 if item[1] > data_point_list[1] else 0
                    if 'TC10' in item[0]:
                        TC10_lower = 1 if item[1] < data_point_list[0] else 0
                        TC10_upper = 1 if item[1] > data_point_list[1] else 0
                    if 'TC50' in item[0]:
                        TC50_lower = 1 if item[1] < data_point_list[0] else 0
                        TC50_upper = 1 if item[1] > data_point_list[1] else 0
                    if 'TC90' in item[0]:
                        TC90_lower = 1 if item[1] < data_point_list[0] else 0
                        TC90_upper = 1 if item[1] > data_point_list[1] else 0
                    date = item[3].strftime('%Y-%m-%d')
                    spe = date + '_' + item[4]
                    if dic.get(spe):
                        data = dic.get(spe)
                        dic[spe].update({
                            'mn_lower': data['mn_lower'] + mn_lower,
                            'mn_upper': data['mn_upper'] + mn_upper,
                            'bz_lower': data['bz_lower'] + bz_lower,
                            'bz_upper': data['bz_upper'] + bz_upper,
                            'yd_lower': data['yd_lower'] + yd_lower,
                            'yd_upper': data['yd_upper'] + yd_upper,
                            'MH_lower': data['MH_lower'] + MH_lower,
                            'MH_upper': data['MH_upper'] + MH_upper,
                            'ML_lower': data['ML_lower'] + ML_lower,
                            'ML_upper': data['ML_upper'] + ML_upper,
                            'TC10_lower': data['TC10_lower'] + TC10_lower,
                            'TC10_upper': data['TC10_upper'] + TC10_upper,
                            'TC50_lower': data['TC50_lower'] + TC50_lower,
                            'TC50_upper': data['TC50_upper'] + TC50_upper,
                            'TC90_lower': data['TC90_lower'] + TC90_lower,
                            'TC90_upper': data['TC90_upper'] + TC90_upper,
                        })
                    else:
                        dic[spe] = {
                            'mn_lower': mn_lower,
                            'mn_upper': mn_upper,
                            'bz_lower': bz_lower,
                            'bz_upper': bz_upper,
                            'yd_lower': yd_lower,
                            'yd_upper': yd_upper,
                            'MH_lower': MH_lower,
                            'MH_upper': MH_upper,
                            'ML_lower': ML_lower,
                            'ML_upper': ML_upper,
                            'TC10_lower': TC10_lower,
                            'TC10_upper': TC10_upper,
                            'TC50_lower': TC50_lower,
                            'TC50_upper': TC50_upper,
                            'TC90_lower': TC90_lower,
                            'TC90_upper': TC90_upper,
                        }
        # --------------- end -----------------
        # 检查数与合格数
        records = queryset.values('material_test_order__production_factory_date',
                                  'material_test_order__production_class').annotate(
            JC=Count('material_test_order_id', distinct=True),
            HG=Count('material_test_order_id', distinct=True, filter=Q(material_test_order__is_qualified=True)))
        if not records:
            return Response([])
        result = {}
        for j in records:
            factory_date = str(j['material_test_order__production_factory_date'])
            production_class = j['material_test_order__production_class']
            result.update({
                factory_date + '_' + production_class: {
                    'date': factory_date,
                    'class': production_class,
                    'JC': j['JC'], 'HG': j['HG'], 'MN': 0, 'YD': 0, 'BZ': 0, 'RATE_1': [], 'MH': 0, 'ML': 0, 'TC10': 0,
                    'TC50': 0, 'TC90': 0, 'RATE_S': [], 'sum_s': 0, 'rate': '%.2f' % (j['HG'] / j['JC'] * 100),
                    'sort_class': 1 if production_class == '早班' else 2
                }
            })
        # ---------------- begin ---------------
        for i in result.keys():
            if dic.get(i):
                result[i].update(dic[i])
            else:
                result[i].update({
                    'mn_lower': 0,
                    'mn_upper': 0,
                    'bz_lower': 0,
                    'bz_upper': 0,
                    'yd_lower': 0,
                    'yd_upper': 0,
                    'MH_lower': 0,
                    'MH_upper': 0,
                    'ML_lower': 0,
                    'ML_upper': 0,
                    'TC10_lower': 0,
                    'TC10_upper': 0,
                    'TC50_lower': 0,
                    'TC50_upper': 0,
                    'TC90_lower': 0,
                    'TC90_upper': 0,
                })
        # --------------- end -----------------
        pre_data = queryset.values('material_test_order__production_factory_date',
                                   'material_test_order__production_class', 'test_indicator_name', 'data_point_name') \
            .annotate(num=Count('id', distinct=True, filter=Q(~Q(level=1)))) \
            .values('material_test_order__production_factory_date', 'material_test_order__production_class',
                    'material_test_order_id', 'test_indicator_name', 'data_point_name', 'num', 'test_times') \
            .order_by('test_times')
        # 处理数据
        data = {(str(i['material_test_order_id']) + '_' + str(i['material_test_order__production_factory_date']) + '_'
                 + i['material_test_order__production_class'] + '_' + i['test_indicator_name'] + '_' + i[
                     'data_point_name']): [i['num'], i['test_times']] for i in pre_data}
        for k, v in data.items():
            single_data = result.get(k.split('_')[1] + '_' + k.split('_')[2])
            order_id = k.split('_')[0]
            if 'ML(1+4)' in k:
                single_data['MN'] += v[0]
                if order_id not in single_data['RATE_1'] and v[0] != 0:
                    single_data['RATE_1'].append(order_id)
            elif '硬度值' in k:
                single_data['YD'] += v[0]
                if order_id not in single_data['RATE_1'] and v[0] != 0:
                    single_data['RATE_1'].append(order_id)
            elif '比重值' in k:
                single_data['BZ'] += v[0]
                if order_id not in single_data['RATE_1'] and v[0] != 0:
                    single_data['RATE_1'].append(order_id)
            elif 'MH' in k:
                single_data['MH'] += v[0]
                if order_id not in single_data['RATE_S'] and v[0] != 0:
                    single_data['RATE_S'].append(order_id)
            elif 'ML' in k:
                single_data['ML'] += v[0]
                if order_id not in single_data['RATE_S'] and v[0] != 0:
                    single_data['RATE_S'].append(order_id)
            elif 'TC10' in k:
                single_data['TC10'] += v[0]
                if order_id not in single_data['RATE_S'] and v[0] != 0:
                    single_data['RATE_S'].append(order_id)
            elif 'TC50' in k:
                single_data['TC50'] += v[0]
                if order_id not in single_data['RATE_S'] and v[0] != 0:
                    single_data['RATE_S'].append(order_id)
            elif 'TC90' in k:
                single_data['TC90'] += v[0]
                if order_id not in single_data['RATE_S'] and v[0] != 0:
                    single_data['RATE_S'].append(order_id)
            else:
                continue
        res_data = result.values()
        all = {}
        rate_1, rate_lb, rate_test_all, rate_pass_sum = 0, 0, 0, 0
        for v in res_data:
            v['sum_s'] = len(v.pop('RATE_S'))
            rate_1_pass_sum = v['JC'] - len(v.pop('RATE_1'))
            rate_s_pass_sum = v['JC'] - v['sum_s']
            v['RATE_1_PASS'] = '%.2f' % (rate_1_pass_sum / v['JC'] * 100)
            v['cp_all'] = v['JC'] - v['HG']
            v['RATE_S_PASS'] = '%.2f' % (rate_s_pass_sum / v['JC'] * 100)
            rate_1 += rate_1_pass_sum
            rate_lb += rate_s_pass_sum
            rate_test_all += v['JC']
            rate_pass_sum += v['HG']
        all.update(rate_1='%.2f' % (rate_1 / rate_test_all * 100), rate_lb='%.2f' % (rate_lb / rate_test_all * 100),
                   rate='%.2f' % (rate_pass_sum / rate_test_all * 100))
        for dic in res_data:
            for key, value in dic.items():
                dic[key] = None if not dic[key] else dic[key]
        return Response({'result': sorted(res_data, key=lambda x: (x['date'], x['sort_class'])), 'all': all})


@method_decorator([api_recorder], name="dispatch")
class UnqialifiedEquipView(APIView):
    """机台别不合格率统计"""
    permission_classes = (IsAuthenticated,)

    def get(self, request):
        station = self.request.query_params.get("station", '')
        product_type = self.request.query_params.get("product_type", '')
        s_time = self.request.query_params.get('s_time')
        e_time = self.request.query_params.get('e_time')
        equip_no = self.request.query_params.get('equip_no', '')
        classes = self.request.query_params.get('classes', '')
        product_str = ''
        if station:
            product_str += '-{}'.format(station)
        if product_type:
            product_str += '-{}'.format(product_type)
        e = datetime.datetime.strptime(e_time, '%Y-%m-%d')
        s = datetime.datetime.strptime(s_time, '%Y-%m-%d')
        delta = e - s
        if delta.days > 31:
            raise ValidationError('搜索日期跨度不得超过一个月！')
        if not s_time and not e_time:
            raise ValidationError('请输入检测时间！')

        queryset = MaterialTestOrder.objects.filter(product_no__icontains=product_str,
                                                    production_factory_date__gte=s_time,
                                                    production_factory_date__lte=e_time,
                                                    production_equip_no__icontains=equip_no,
                                                    production_class__icontains=classes
                                                    )
        material_test_result = MaterialTestResult.objects.filter(material_test_order__product_no__icontains=product_str,
                                                                 material_test_order__production_factory_date__gte=s_time,
                                                                 material_test_order__production_factory_date__lte=e_time,
                                                                 material_test_order__production_equip_no__icontains=equip_no,
                                                                 material_test_order__production_class__icontains=classes
                                                                 )
        # 统计不合格中超过和小于标准的数量
        # ---------------- begin ---------------
        dic_ = {}
        data_point_dic = {}
        data_point_query = MaterialDataPointIndicator.objects.filter(level=1).values(
            'material_test_method__material__material_no', 'data_point__name',
            'upper_limit', 'lower_limit')
        for i in data_point_query:
            if data_point_dic.get(i['material_test_method__material__material_no']):
                data_point_dic[i['material_test_method__material__material_no']][i['data_point__name']] = [
                    i['lower_limit'], i['upper_limit']]
            else:
                data_point_dic[i['material_test_method__material__material_no']] = {
                    i['data_point__name']: [i['lower_limit'], i['upper_limit']]}

        # res = material_test_result.values('material_test_order__production_equip_no',
        #                                   'material_test_order__product_no',
        #                                   'data_point_name', 'value')
        sql = """
            SELECT
         * 
        FROM
         (
         SELECT
         a.data_point_name,
          a.value,
          b.product_no,
          b.production_factory_date,
          b.production_class,
          a.test_times,
          b.actual_trains,
          b.production_equip_no 
         FROM
          material_test_result a
          LEFT JOIN material_test_order b ON a.material_test_order_id = b.id 
         WHERE
          b.PRODUCTION_EQUIP_NO LIKE '%{}%'
                    AND b.PRODUCT_NO LIKE '%{}%'
                    AND b.PRODUCTION_CLASS LIKE '%{}%'
                    AND b.PRODUCTION_FACTORY_DATE >= TO_DATE('{}', 'YYYY-MM-DD')
                    AND b.PRODUCTION_FACTORY_DATE <= TO_DATE('{}', 'YYYY-MM-DD')

         ) a 
        WHERE
         a.test_times = (
         SELECT
          max( x.test_times ) 
         FROM
          (
          SELECT
           a.value,
           a.test_class,
           a.test_times,
           a.data_point_name,
           b.production_factory_date,
           b.actual_trains,
           b.product_no,
           b.production_equip_no 
          FROM
           material_test_result a
           LEFT JOIN material_test_order b ON a.material_test_order_id = b.id 
          WHERE
           b.PRODUCTION_EQUIP_NO LIKE '%{}%'
                    AND b.PRODUCT_NO LIKE '%{}%'
                    AND b.PRODUCTION_CLASS LIKE '%{}%'
                    AND b.PRODUCTION_FACTORY_DATE >= TO_DATE('{}', 'YYYY-MM-DD')
                    AND b.PRODUCTION_FACTORY_DATE <= TO_DATE('{}', 'YYYY-MM-DD')

          ) x 
         WHERE
          x.data_point_name = a.data_point_name 
          AND x.actual_trains = a.actual_trains 
          AND x.product_no = a.product_no 
         AND x.production_equip_no = a.production_equip_no 
         )""".format(equip_no, product_str, classes, s_time, e_time,
                     equip_no, product_str, classes, s_time, e_time)
        cursor = connection.cursor()
        cursor.execute(sql)
        data = cursor.fetchall()
        for item in data:
            if data_point_dic.get(item[2]):
                data_point_list = data_point_dic[item[2]].get(item[0])
                if data_point_list:
                    MH_lower = MH_upper = ML_lower = ML_upper = TC10_lower = TC10_upper = TC50_lower = TC50_upper = TC90_lower = TC90_upper = 0
                    bz_lower = bz_upper = mn_lower = mn_upper = yd_lower = yd_upper = 0
                    if 'ML(1+4)' == item[0]:
                        mn_lower = 1 if item[1] < data_point_list[0] else 0
                        mn_upper = 1 if item[1] > data_point_list[1] else 0
                    if '比重值' == item[0]:
                        bz_lower = 1 if item[1] < data_point_list[0] else 0
                        bz_upper = 1 if item[1] > data_point_list[1] else 0
                    if '硬度值' == item[0]:
                        yd_lower = 1 if item[1] < data_point_list[0] else 0
                        yd_upper = 1 if item[1] > data_point_list[1] else 0
                    if 'MH' in item[0]:
                        MH_lower = 1 if item[1] < data_point_list[0] else 0
                        MH_upper = 1 if item[1] > data_point_list[1] else 0
                    if 'ML' == item[0]:
                        ML_lower = 1 if item[1] < data_point_list[0] else 0
                        ML_upper = 1 if item[1] > data_point_list[1] else 0
                    if 'TC10' in item[0]:
                        TC10_lower = 1 if item[1] < data_point_list[0] else 0
                        TC10_upper = 1 if item[1] > data_point_list[1] else 0
                    if 'TC50' in item[0]:
                        TC50_lower = 1 if item[1] < data_point_list[0] else 0
                        TC50_upper = 1 if item[1] > data_point_list[1] else 0
                    if 'TC90' in item[0]:
                        TC90_lower = 1 if item[1] < data_point_list[0] else 0
                        TC90_upper = 1 if item[1] > data_point_list[1] else 0

                    spe = item[7]
                    if dic_.get(spe):
                        data = dic_.get(spe)
                        dic_[spe].update({
                            'mn_lower': data['mn_lower'] + mn_lower,
                            'mn_upper': data['mn_upper'] + mn_upper,
                            'bz_lower': data['bz_lower'] + bz_lower,
                            'bz_upper': data['bz_upper'] + bz_upper,
                            'yd_lower': data['yd_lower'] + yd_lower,
                            'yd_upper': data['yd_upper'] + yd_upper,
                            'MH_lower': data['MH_lower'] + MH_lower,
                            'MH_upper': data['MH_upper'] + MH_upper,
                            'ML_lower': data['ML_lower'] + ML_lower,
                            'ML_upper': data['ML_upper'] + ML_upper,
                            'TC10_lower': data['TC10_lower'] + TC10_lower,
                            'TC10_upper': data['TC10_upper'] + TC10_upper,
                            'TC50_lower': data['TC50_lower'] + TC50_lower,
                            'TC50_upper': data['TC50_upper'] + TC50_upper,
                            'TC90_lower': data['TC90_lower'] + TC90_lower,
                            'TC90_upper': data['TC90_upper'] + TC90_upper,
                        })
                    else:
                        dic_[spe] = {
                            'mn_lower': mn_lower,
                            'mn_upper': mn_upper,
                            'bz_lower': bz_lower,
                            'bz_upper': bz_upper,
                            'yd_lower': yd_lower,
                            'yd_upper': yd_upper,
                            'MH_lower': MH_lower,
                            'MH_upper': MH_upper,
                            'ML_lower': ML_lower,
                            'ML_upper': ML_upper,
                            'TC10_lower': TC10_lower,
                            'TC10_upper': TC10_upper,
                            'TC50_lower': TC50_lower,
                            'TC50_upper': TC50_upper,
                            'TC90_lower': TC90_lower,
                            'TC90_upper': TC90_upper,
                        }
        # --------------- end -----------------

        # 检查数
        test_all = queryset.values('production_equip_no').annotate(count=Count('product_no')).values(
            'production_equip_no', 'count')
        # 合格数
        test_right = queryset.filter(is_qualified=True).values('production_equip_no').annotate(
            count=Count('product_no'))

        result = material_test_result.values('material_test_order_id', 'data_point_name',
                                             'test_indicator_name',
                                             'material_test_order__production_equip_no'
                                             ).annotate(count=Count('id')).values(
            'material_test_order_id', 'data_point_name', 'test_indicator_name', 'level',
            'material_test_order__production_equip_no', 'test_times').order_by('test_times')
        equip_queryset = MaterialTestOrder.objects.filter(production_equip_no__icontains=equip_no).values(
            'production_equip_no').annotate(sum=Count('production_equip_no')).values('production_equip_no')
        equip_list = [equip['production_equip_no'] for equip in equip_queryset]
        if not equip_no:
            dic = {'Z01': {}, 'Z02': {}, 'Z03': {}, 'Z04': {}, 'Z05': {}, 'Z06': {}, 'Z07': {}, 'Z08': {}, 'Z09': {},
                   'Z10': {}, 'Z11': {}, 'Z12': {}, 'Z13': {}, 'Z14': {}, 'Z15': {}}
        else:
            dic = {}
            for equip in equip_list:
                dic.update({equip: {}})

        if len(test_all) > 0:
            for i in result:
                if dic[i['material_test_order__production_equip_no']].get(
                        f"{i['material_test_order_id']}_{i['test_indicator_name']}_{i['data_point_name']}"):
                    if i['level'] == 1:
                        del dic[i['material_test_order__production_equip_no']][
                            f"{i['material_test_order_id']}_{i['test_indicator_name']}_{i['data_point_name']}"]
                else:
                    if i['level'] == 2:
                        dic[i['material_test_order__production_equip_no']].update({
                            f"{i['material_test_order_id']}_{i['test_indicator_name']}_{i['data_point_name']}": {
                                'data_point_name': i[
                                    'data_point_name'],
                                'test_indicator_name': i[
                                    'test_indicator_name']}})

            results = []
            if not equip_no:
                equip_list = ['Z01', 'Z02', 'Z03', 'Z04', 'Z05', 'Z06', 'Z07', 'Z08', 'Z09', 'Z10', 'Z11', 'Z12', 'Z13',
                              'Z14', 'Z15']

            for equip in equip_list:
                MN = YD = BZ = MH = ML = TC10 = TC50 = TC90 = 0
                RATE_1 = []
                RATE_LB = []

                for i in test_all:
                    if equip == i['production_equip_no']:
                        TEST_ALL = i['count']
                        break
                    else:
                        TEST_ALL = 0
                if len(test_right) > 0:
                    for i in test_right:
                        if equip == i['production_equip_no']:
                            TEST_RIGHT = i['count']
                            break
                        else:
                            TEST_RIGHT = 0
                else:
                    TEST_RIGHT = 0
                for i in dic[equip].keys():
                    if i.split('_')[2] == 'ML(1+4)':
                        MN += 1
                    elif i.split('_')[2] == '硬度值':
                        YD += 1
                    elif i.split('_')[2] == '比重值':
                        BZ += 1
                    elif i.split('_')[2] == 'MH':
                        MH += 1
                    elif i.split('_')[2] == 'ML':
                        ML += 1
                    elif i.split('_')[2] == 'TC10':
                        TC10 += 1
                    elif i.split('_')[2] == 'TC50':
                        TC50 += 1
                    elif i.split('_')[2] == 'TC90':
                        TC90 += 1

                    if i.split('_')[2] == 'ML(1+4)' or i.split('_')[2] == '硬度值' or i.split('_')[2] == '比重值':
                        RATE_1.append(i.split('_')[0])
                    if i.split('_')[2] == 'MH' or i.split('_')[2] == 'ML' or i.split('_')[2] == 'TC10' or i.split('_')[
                        2] == 'TC50' or i.split('_')[2] == 'TC90':
                        RATE_LB.append(i.split('_')[0])
                RATE_1 = len(set(RATE_1))
                RATE_LB = len(set(RATE_LB))
                data = {
                    'equip': equip,
                    'test_all': TEST_ALL,
                    'test_right': TEST_RIGHT,
                    'mn': MN,
                    'yd': YD,
                    'bz': BZ,
                    'rate_1': '%.2f' % (((TEST_ALL - RATE_1) / TEST_ALL) * 100) if TEST_ALL else 0,
                    'MH': MH,
                    'ML': ML,
                    'TC10': TC10,
                    'TC50': TC50,
                    'TC90': TC90,
                    'lb_all': RATE_LB,
                    'rate_lb': '%.2f' % (((TEST_ALL - RATE_LB) / TEST_ALL) * 100) if TEST_ALL else 0,
                    'cp_all': TEST_ALL - TEST_RIGHT,
                    'rate': '%.2f' % ((TEST_RIGHT / TEST_ALL) * 100) if TEST_ALL else 0,
                    'rate_1_sum': TEST_ALL - RATE_1,
                    'rate_s_sum': TEST_ALL - RATE_LB
                }
                if dic_.get(equip):
                    data.update(dic_[equip])
                else:
                    data.update({
                        'mn_lower': 0,
                        'mn_upper': 0,
                        'bz_lower': 0,
                        'bz_upper': 0,
                        'yd_lower': 0,
                        'yd_upper': 0,
                        'MH_lower': 0,
                        'MH_upper': 0,
                        'ML_lower': 0,
                        'ML_upper': 0,
                        'TC10_lower': 0,
                        'TC10_upper': 0,
                        'TC50_lower': 0,
                        'TC50_upper': 0,
                        'TC90_lower': 0,
                        'TC90_upper': 0,
                    })
                results.append(
                    data
                )
            all = {}
            num = rate_1 = rate_lb = test_all = rate_pass_sum = 0
            for i in results:
                if i['test_all'] == 0:
                    pass
                else:
                    num += 1
                    rate_1 += i['rate_1_sum']
                    rate_lb += i['rate_s_sum']
                    test_all += i['test_all']
                    rate_pass_sum += i['test_right']
            if num != 0:
                all.update(rate_1='%.2f' % (rate_1 / test_all * 100), rate_lb='%.2f' % (rate_lb / test_all * 100),
                           rate='%.2f' % (rate_pass_sum / test_all * 100))
        else:
            results = []
            all = {}
        for dic in results:
            for key, value in dic.items():
                dic[key] = None if not dic[key] else dic[key]
        return Response({'results': results, 'all': all})


@method_decorator([api_recorder], name="dispatch")
class MaterialTestPlanViewSet(ModelViewSet):
    queryset = MaterialTestPlan.objects.filter(status=1)
    serializer_class = MaterialTestPlanSerializer
    filter_backends = (DjangoFilterBackend,)
    filter_class = MaterialTestPlanFilter

    def get_serializer_class(self):
        if self.action == 'create':
            return MaterialTestPlanCreateSerializer
        return MaterialTestPlanSerializer

    @atomic
    def update(self, request, *args, **kwargs):  # 检测中添加
        instance = self.get_object()
        material_dic = self.request.data
        material_name = material_dic.get('material_name')
        material_sample_name = material_dic.get('material_sample_name', None)
        material_batch = material_dic.get('material_batch')
        material_supplier = material_dic.get('material_supplier', None)
        material_tmh = material_dic.get('material_tmh')
        material_wlxxid = material_dic.get('material_wlxxid')
        material = ExamineMaterial.objects.filter(batch=material_batch,
                                                  wlxxid=material_wlxxid).first()
        if not material:
            material = ExamineMaterial.objects.create(batch=material_batch,
                                                      name=material_name,
                                                      sample_name=material_sample_name,
                                                      supplier=material_supplier,
                                                      tmh=material_tmh,
                                                      wlxxid=material_wlxxid)
        MaterialTestPlanDetail.objects.create(material_test_plan=instance,
                                              material=material, recorder=self.request.user,
                                              sampling_user=self.request.user)
        return Response('添加成功')

    def perform_destroy(self, instance):
        """结束检测"""
        instance.status = 4
        MaterialTestPlanDetail.objects.filter(material_test_plan=instance, status=1).update(status=4)
        instance.save()


@method_decorator([api_recorder], name="dispatch")
class MaterialTestPlanDetailViewSet(ModelViewSet):
    queryset = MaterialTestPlanDetail.objects.all()
    serializer_class = MaterialTestPlanDetailSerializer

    def perform_destroy(self, instance):
        if instance.value:
            raise ValidationError('该数据已检测，无法删除！')
        return super().perform_destroy(instance)


@method_decorator([api_recorder], name="dispatch")
class WMSMooneyLevelView(ModelViewSet):
    queryset = WMSMooneyLevel.objects.all()
    serializer_class = WMSMooneyLevelSerializer
    permission_classes = (IsAuthenticated, PermissionClass({'view': 'view_wms_mooney_level',
                                                            'change': 'change_wms_mooney_level',
                                                            }))

    def get_serializer_class(self):
        if self.action == 'list':
            return ERPMESMaterialRelationSerializer
        return WMSMooneyLevelSerializer

    @action(methods=['post'], detail=False, permission_classes=[IsAuthenticated], url_path='clear-level',
            url_name='clear-level')
    def clear_level(self, request):
        """清除等级"""
        material_no = self.request.data.get('material_no')
        WMSMooneyLevel.objects.filter(material_no=material_no).delete()
        return Response('success')

    @action(methods=['post'], detail=False, permission_classes=[IsAuthenticated], url_path='batch-set',
            url_name='batch-set')
    def batch_set(self, request):
        """批量设置"""
        req_data = self.request.data
        if not isinstance(req_data, dict):
            raise ValidationError('参数错误！')
        material_nos = req_data.pop('material_no', [])
        if not isinstance(material_nos, list):
            raise ValidationError('参数错误！')
        for material_no in material_nos:
            req_data.update({'material_no': material_no})
            s = WMSMooneyLevelSerializer(data=req_data, context={'request': request})
            s.is_valid(raise_exception=True)
            s.save()
        return Response('ok')

    def list(self, request, *args, **kwargs):
        material_type = self.request.query_params.get('material_type')
        material_no = self.request.query_params.get('material_no')
        material_name = self.request.query_params.get('material_name')
        filter_kwargs = {'material__material_type__global_name__icontains': '胶'}
        if material_type:
            filter_kwargs['material__material_type__global_name'] = material_type
        if material_no:
            filter_kwargs['zc_material__wlxxid__icontains'] = material_no
        if material_name:
            filter_kwargs['zc_material__material_name__icontains'] = material_name
        erp_ids = set(ERPMESMaterialRelation.objects.filter(**filter_kwargs).values_list('zc_material_id', flat=True))
        queryset = ZCMaterial.objects.filter(id__in=erp_ids)
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)

        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)


@method_decorator([api_recorder], name="dispatch")
class ProductSynthesisRate(APIView):
    # [{"name": "综合合格率%", "rate": 22.2, "1": "", "2": "", "3": ""},
    #  {"name": "一次合格率%", "rate": 22.2, "1": "", "2": "", "3": ""},
    #  {"name": "流变合格率%", "rate": 22.2, "1": "", "2": "", "3": ""}]
    permission_classes = (IsAuthenticated, PermissionClass({'view': 'view_product_daily_rate'}))

    def get(self, request):
        date = self.request.query_params.get('date')
        sy_flag = self.request.query_params.get('sy_flag', 'N')
        if not date:
            raise ValidationError('请选择月份！')
        year = int(date.split('-')[0])
        month = int(date.split('-')[1])
        filter_kwargs = {'production_factory_date__year': year,
                         'production_factory_date__month': month,
                         'product_no__icontains': '-FM-'}
        if sy_flag == 'N':
            filter_kwargs['is_experiment'] = False
        # 每日检查量
        mto_data = dict(MaterialTestOrder.objects.filter(
            **filter_kwargs).values('production_factory_date__day').annotate(qty=Count('id')).values_list('production_factory_date__day', 'qty'))
        if not mto_data:
            return Response({})
        total_mto_qty = sum(mto_data.values())

        # 每日合格量
        mto_qualified_data = dict(MaterialTestOrder.objects.filter(
            **filter_kwargs).filter(is_qualified=True).values('production_factory_date__day').annotate(qty=Count('id')).values_list('production_factory_date__day', 'qty'))
        total_mto_qualified_qty = sum(mto_qualified_data.values())

        # 按照胶料规格分组，每个规格流变检测多少车次
        lb_product_check_data = dict(MaterialTestOrder.objects.filter(**filter_kwargs).filter(
            order_results__test_indicator_name='流变'
        ).values('production_factory_date__day').annotate(cnt=Count('id', distinct=True)).values_list('production_factory_date__day', 'cnt'))
        total_lb_product_check_qty = sum(lb_product_check_data.values())

        # 按照胶料规格分组，每个规格流变不合格多少车次
        lb_product_unqualified_data = dict(MaterialTestOrder.objects.filter(**filter_kwargs).filter(
            order_results__test_indicator_name='流变',
            order_results__level=2).values('production_factory_date__day').annotate(cnt=Count('id', distinct=True)).values_list('production_factory_date__day', 'cnt'))
        total_lb_product_unqualified_qty = sum(lb_product_unqualified_data.values())

        # 按照胶料规格分组，每个规格一次检测多少车次
        yc_product_check_data = dict(MaterialTestOrder.objects.filter(**filter_kwargs).filter(
            order_results__test_indicator_name__in=('门尼', '比重', '硬度')
        ).values('production_factory_date__day').annotate(cnt=Count('id', distinct=True)).values_list('production_factory_date__day', 'cnt'))
        total_yc_product_check_qty = sum(yc_product_check_data.values())

        # 按照胶料规格分组，每个规格一次不合格多少车次
        yc_product_unqualified_data = dict(MaterialTestOrder.objects.filter(**filter_kwargs).filter(
            order_results__test_indicator_name__in=('门尼', '比重', '硬度'),
            order_results__level=2).values('production_factory_date__day').annotate(cnt=Count('id', distinct=True)).values_list(
            'production_factory_date__day', 'cnt'))
        total_yc_product_unqualified_qty = sum(yc_product_unqualified_data.values())

        days = calendar.monthrange(year, month)[1]

        data1 = [{"name": "综合合格率%", "rate": round(total_mto_qualified_qty/total_mto_qty*100, 1)},
                 {"name": "一次合格率%", "rate": "" if not total_yc_product_check_qty else round((1-total_yc_product_unqualified_qty/total_yc_product_check_qty)*100, 1)},
                 {"name": "流变合格率%", "rate": "" if not total_lb_product_check_qty else round((1-total_lb_product_unqualified_qty/total_lb_product_check_qty)*100, 1)}]

        for i in range(1, days+1):
            try:
                zh_rate = round(mto_qualified_data.get(i) / mto_data.get(i) * 100, 1)
            except Exception:
                zh_rate = ''
            try:
                lb_rate = round((1-lb_product_unqualified_data.get(i) / lb_product_check_data.get(i)) * 100, 1)
            except Exception:
                lb_rate = ''
            try:
                yc_rate = round((1-yc_product_unqualified_data.get(i) / yc_product_check_data.get(i)) * 100, 1)
            except Exception:
                yc_rate = ''
            data1[0][str(i)] = zh_rate
            data1[1][str(i)] = yc_rate
            data1[2][str(i)] = lb_rate
        return Response({'data': data1})


@method_decorator([api_recorder], name="dispatch")
class ProductSynthesisMonthRate(APIView):
    # [{"name": "综合合格率%", "rate": 22.2, "1": "", "2": "", "3": ""},
    #  {"name": "一次合格率%", "rate": 22.2, "1": "", "2": "", "3": ""},
    #  {"name": "流变合格率%", "rate": 22.2, "1": "", "2": "", "3": ""}]
    permission_classes = (IsAuthenticated, PermissionClass({'view': 'view_product_daily_rate'}))

    def get(self, request):
        year = self.request.query_params.get('year')
        sy_flag = self.request.query_params.get('sy_flag', 'N')
        if not year:
            raise ValidationError('请选择年份！')
        year = int(year)
        filter_kwargs = {'production_factory_date__year': year, 'product_no__icontains': '-FM-'}
        if sy_flag == 'N':
            filter_kwargs['is_experiment'] = False
        # 每月检查量
        mto_data = dict(MaterialTestOrder.objects.filter(
            **filter_kwargs).values('production_factory_date__month').annotate(qty=Count('id')).values_list('production_factory_date__month', 'qty'))
        if not mto_data:
            return Response({})
        total_mto_qty = sum(mto_data.values())

        # 每月合格量
        mto_qualified_data = dict(MaterialTestOrder.objects.filter(
            **filter_kwargs).filter(is_qualified=True).values('production_factory_date__month').annotate(qty=Count('id')).values_list('production_factory_date__month', 'qty'))
        total_mto_qualified_qty = sum(mto_qualified_data.values())

        # 按照胶料规格分组，每个规格流变检测多少车次
        lb_product_check_data = dict(MaterialTestOrder.objects.filter(**filter_kwargs).filter(
            order_results__test_indicator_name='流变'
        ).values('production_factory_date__month').annotate(cnt=Count('id', distinct=True)).values_list('production_factory_date__month', 'cnt'))
        total_lb_product_check_qty = sum(lb_product_check_data.values())

        # 按照胶料规格分组，每个规格流变不合格多少车次
        lb_product_unqualified_data = dict(MaterialTestOrder.objects.filter(**filter_kwargs).filter(
            order_results__test_indicator_name='流变',
            order_results__level=2).values('production_factory_date__month').annotate(cnt=Count('id', distinct=True)).values_list('production_factory_date__month', 'cnt'))
        total_lb_product_unqualified_qty = sum(lb_product_unqualified_data.values())

        # 按照胶料规格分组，每个规格一次检测多少车次
        yc_product_check_data = dict(MaterialTestOrder.objects.filter(**filter_kwargs).filter(
            order_results__test_indicator_name__in=('门尼', '比重', '硬度')
        ).values('production_factory_date__month').annotate(cnt=Count('id', distinct=True)).values_list('production_factory_date__month', 'cnt'))
        total_yc_product_check_qty = sum(yc_product_check_data.values())

        # 按照胶料规格分组，每个规格一次不合格多少车次
        yc_product_unqualified_data = dict(MaterialTestOrder.objects.filter(**filter_kwargs).filter(
            order_results__test_indicator_name__in=('门尼', '比重', '硬度'),
            order_results__level=2).values('production_factory_date__month').annotate(cnt=Count('id', distinct=True)).values_list(
            'production_factory_date__month', 'cnt'))
        total_yc_product_unqualified_qty = sum(yc_product_unqualified_data.values())

        data1 = [{"name": "综合合格率%", "rate": round(total_mto_qualified_qty/total_mto_qty*100, 1)},
                 {"name": "一次合格率%", "rate": "" if not total_yc_product_check_qty else round((1-total_yc_product_unqualified_qty/total_yc_product_check_qty)*100, 1)},
                 {"name": "流变合格率%", "rate": "" if not total_lb_product_check_qty else round((1-total_lb_product_unqualified_qty/total_lb_product_check_qty)*100, 1)}]

        for i in range(1, 13):
            try:
                zh_rate = round(mto_qualified_data.get(i) / mto_data.get(i) * 100, 1)
            except Exception:
                zh_rate = ''
            try:
                lb_rate = round((1-lb_product_unqualified_data.get(i) / lb_product_check_data.get(i)) * 100, 1)
            except Exception:
                lb_rate = ''
            try:
                yc_rate = round((1-yc_product_unqualified_data.get(i) / yc_product_check_data.get(i)) * 100, 1)
            except Exception:
                yc_rate = ''
            data1[0][str(i)] = zh_rate
            data1[1][str(i)] = yc_rate
            data1[2][str(i)] = lb_rate
        return Response({'data': data1})


@method_decorator([api_recorder], name="dispatch")
class ProductSynthesisEquipRate(APIView):
    # [{"name": "Z01", "rate": 22.2, "1": "", "2": "", "3": ""},
    #  {"name": "Z02", "rate": 22.2, "1": "", "2": "", "3": ""},
    #  {"name": "Z03", "rate": 22.2, "1": "", "2": "", "3": ""}]
    permission_classes = (IsAuthenticated, PermissionClass({'view': 'view_product_daily_rate'}))

    def get(self, request):
        date = self.request.query_params.get('date')
        sy_flag = self.request.query_params.get('sy_flag', 'N')
        if not date:
            raise ValidationError('请选择月份！')
        year = int(date.split('-')[0])
        month = int(date.split('-')[1])
        days = calendar.monthrange(year, month)[1]
        filter_kwargs = {'production_factory_date__year': year,
                         'production_factory_date__month': month}
        if sy_flag == 'N':
            filter_kwargs['is_experiment'] = False
        # 机台每日总检查量
        mto_equip_data = MaterialTestOrder.objects.filter(
            **filter_kwargs).values('production_factory_date__day', 'production_equip_no').annotate(qty=Count('id'))
        mto_equip_data_dict = {}
        for item in mto_equip_data:
            equip_no = item['production_equip_no']
            day = item['production_factory_date__day']
            qty = item['qty']
            if equip_no in mto_equip_data_dict:
                mto_equip_data_dict[equip_no][day] = qty
                mto_equip_data_dict[equip_no]['total'] += qty
            else:
                mto_equip_data_dict[equip_no] = {day: qty, 'total': qty}

        # 机台每日合格量
        mto_equip_qualified_data = MaterialTestOrder.objects.filter(
            **filter_kwargs).filter(is_qualified=True).values(
            'production_factory_date__day', 'production_equip_no').annotate(qty=Count('id'))
        mto_equip_qualified_data_dict = {}
        for item in mto_equip_qualified_data:
            equip_no = item['production_equip_no']
            day = item['production_factory_date__day']
            qty = item['qty']
            if equip_no in mto_equip_qualified_data_dict:
                mto_equip_qualified_data_dict[equip_no][day] = qty
                mto_equip_qualified_data_dict[equip_no]['total'] += qty
            else:
                mto_equip_qualified_data_dict[equip_no] = {day: qty, 'total': qty}

        equip_nos = list(Equip.objects.filter(
                category__equip_type__global_name="密炼设备"
            ).order_by('equip_no').values_list("equip_no", flat=True))

        data2 = [{"name": equip_no, "rate": "" if not mto_equip_data_dict.get(equip_no) else
                    round(mto_equip_qualified_data_dict.get(equip_no, {}).get('total', 0) / mto_equip_data_dict.get(equip_no, {}).get('total') * 100, 1)}
                 for equip_no in equip_nos]

        # 机台合格率
        for i in range(1, days+1):
            for idx, equip_no in enumerate(equip_nos):
                q_qty = mto_equip_qualified_data_dict.get(equip_no, {}).get(i, 0)
                t_qty = mto_equip_data_dict.get(equip_no, {}).get(i, 0)
                data2[idx][str(i)] = '' if not t_qty else round(q_qty/t_qty*100, 1)
        return Response({'data': data2})


@method_decorator([api_recorder], name="dispatch")
class ProductSynthesisGroupRate(APIView):
    # [{"name": "A班", "rate": 22.2, "1": "", "2": "", "3": ""},
    #  {"name": "B班", "rate": 22.2, "1": "", "2": "", "3": ""},
    #  {"name": "C班", "rate": 22.2, "1": "", "2": "", "3": ""}]
    permission_classes = (IsAuthenticated, PermissionClass({'view': 'view_product_daily_rate'}))

    def get(self, request):
        date = self.request.query_params.get('date')
        sy_flag = self.request.query_params.get('sy_flag', 'N')
        if not date:
            raise ValidationError('请选择月份！')
        year = int(date.split('-')[0])
        month = int(date.split('-')[1])
        days = calendar.monthrange(year, month)[1]
        filter_kwargs = {'production_factory_date__year': year,
                         'production_factory_date__month': month,
                         'product_no__icontains': '-FM-'}
        if sy_flag == 'N':
            filter_kwargs['is_experiment'] = False
        # 班组每日总检查量
        mto_group_data = MaterialTestOrder.objects.filter(
            **filter_kwargs).values('production_factory_date__day', 'production_group').annotate(qty=Count('id'))
        mto_group_data_dict = {}
        group_names = set()
        for item in mto_group_data:
            group = item['production_group']
            day = item['production_factory_date__day']
            qty = item['qty']
            group_names.add(group)
            if group in mto_group_data_dict:
                mto_group_data_dict[group][day] = qty
                mto_group_data_dict[group]['total'] += qty
            else:
                mto_group_data_dict[group] = {day: qty, 'total': qty}

        # 班组每日合格量
        mto_group_qualified_data = MaterialTestOrder.objects.filter(
            **filter_kwargs).filter(is_qualified=True).values(
            'production_factory_date__day', 'production_group').annotate(qty=Count('id'))
        mto_group_qualified_data_dict = {}
        for item in mto_group_qualified_data:
            group = item['production_group']
            day = item['production_factory_date__day']
            qty = item['qty']
            if group in mto_group_qualified_data_dict:
                mto_group_qualified_data_dict[group][day] = qty
                mto_group_qualified_data_dict[group]['total'] += qty
            else:
                mto_group_qualified_data_dict[group] = {day: qty, 'total': qty}

        group_names = sorted(list(group_names))
        data3 = [{"name": group_name, "rate": "" if not mto_group_data_dict.get(group_name) else
                    round(mto_group_qualified_data_dict.get(group_name, {}).get('total', 0) / mto_group_data_dict.get(group_name, {}).get('total') * 100, 1)}
                 for group_name in group_names]

        # 班组合格率
        for i in range(1, days+1):
            for idx, group in enumerate(group_names):
                q_qty = mto_group_qualified_data_dict.get(group, {}).get(i, 0)
                t_qty = mto_group_data_dict.get(group, {}).get(i, 0)
                data3[idx][str(i)] = '' if not t_qty else round(q_qty / t_qty * 100, 1)
        return Response({'data': data3})


@method_decorator([api_recorder], name="dispatch")
class ProductSynthesisProductRate(APIView):
    permission_classes = (IsAuthenticated, PermissionClass({'view': 'view_product_monthly_rate'}))

    def get(self, request):
        date = self.request.query_params.get('date')
        sy_flag = self.request.query_params.get('sy_flag', 'N')
        if not date:
            raise ValidationError('请选择月份！')
        year = int(date.split('-')[0])
        month = int(date.split('-')[1])
        days = calendar.monthrange(year, month)[1]
        filter_kwargs = {'production_factory_date__year': year,
                         'production_factory_date__month': month,
                         'product_no__icontains': '-FM-'}
        if sy_flag == 'N':
            filter_kwargs['is_experiment'] = False
        pt_dict = {}
        pt_list = []
        pt = GlobalCode.objects.filter(global_type__type_name='配方类别').order_by('id').values('global_no', 'global_name')
        for item in pt:
            type_name = item['global_no']
            if type_name == '斜胶胎':
                type_name = '斜胶胎（不含实心胎）'
            pt_list.append(type_name)
            for i in item['global_name'].split(','):
                pt_dict[i] = type_name
        pt_list.insert(1, '斜胶胎（含实心胎）')
        # 规格每日总检查量
        mto_product_data = MaterialTestOrder.objects.filter(
            **filter_kwargs).values('production_factory_date__day', 'product_no').annotate(qty=Count('id'))
        mto_product_dict = {}
        for item in mto_product_data:
            product_no = item['product_no'].split('-')[2]
            day = item['production_factory_date__day']
            qty = item['qty']
            re_result = re.match(r'[A-Z]+', product_no)
            if not re_result:
                continue
            pb_type = pt_dict.get(re_result.group(), '未知')
            if pb_type in mto_product_dict:
                if day not in mto_product_dict[pb_type]:
                    mto_product_dict[pb_type][day] = qty
                else:
                    mto_product_dict[pb_type][day] += qty
                mto_product_dict[pb_type]['total'] += qty
            else:
                mto_product_dict[pb_type] = {day: qty, 'total': qty}

        # 规格每日合格量
        mto_product_qualified_data = MaterialTestOrder.objects.filter(
            **filter_kwargs).filter(is_qualified=True).values(
            'production_factory_date__day', 'product_no').annotate(qty=Count('id'))
        mto_product_qualified_data_dict = {}
        for item in mto_product_qualified_data:
            product_no = item['product_no'].split('-')[2]
            day = item['production_factory_date__day']
            qty = item['qty']
            re_result = re.match(r'[A-Z]+', product_no)
            if not re_result:
                continue
            pb_type = pt_dict.get(re_result.group(), '未知')
            if pb_type in mto_product_qualified_data_dict:
                if day not in mto_product_qualified_data_dict[pb_type]:
                    mto_product_qualified_data_dict[pb_type][day] = qty
                else:
                    mto_product_qualified_data_dict[pb_type][day] += qty
                mto_product_qualified_data_dict[pb_type]['total'] += qty
            else:
                mto_product_qualified_data_dict[pb_type] = {day: qty, 'total': qty}

        # 规格每日流变检测多少车次
        lb_product_daily_check_data = MaterialTestOrder.objects.filter(**filter_kwargs).filter(
            order_results__test_indicator_name='流变'
        ).values('production_factory_date__day', 'product_no').annotate(cnt=Count('id', distinct=True))
        lb_product_daily_check_dict = {}
        for item in lb_product_daily_check_data:
            day = item['production_factory_date__day']
            product_no = item['product_no'].split('-')[2]
            qty = item['cnt']
            re_result = re.match(r'[A-Z]+', product_no)
            if not re_result:
                continue
            pb_type = pt_dict.get(re_result.group(), '未知')
            if pb_type in lb_product_daily_check_dict:
                if day not in lb_product_daily_check_dict[pb_type]:
                    lb_product_daily_check_dict[pb_type][day] = qty
                else:
                    lb_product_daily_check_dict[pb_type][day] += qty
                lb_product_daily_check_dict[pb_type]['total'] += qty
            else:
                lb_product_daily_check_dict[pb_type] = {day: qty, 'total': qty}

        # 规格每日流变不合格多少车次
        lb_product_daily_unqualified_data = MaterialTestOrder.objects.filter(**filter_kwargs).filter(
            order_results__test_indicator_name='流变',
            order_results__level=2).values('production_factory_date__day', 'product_no').annotate(cnt=Count('id', distinct=True))
        lb_product_daily_unqualified_dict = {}
        for item in lb_product_daily_unqualified_data:
            day = item['production_factory_date__day']
            product_no = item['product_no'].split('-')[2]
            qty = item['cnt']
            re_result = re.match(r'[A-Z]+', product_no)
            if not re_result:
                continue
            pb_type = pt_dict.get(re_result.group(), '未知')
            if pb_type in lb_product_daily_unqualified_dict:
                if day not in lb_product_daily_unqualified_dict[pb_type]:
                    lb_product_daily_unqualified_dict[pb_type][day] = qty
                else:
                    lb_product_daily_unqualified_dict[pb_type][day] += qty
                lb_product_daily_unqualified_dict[pb_type]['total'] += qty
            else:
                lb_product_daily_unqualified_dict[pb_type] = {day: qty, 'total': qty}

        # 规格一次每日检测多少车次
        yc_product_daily_check_data = MaterialTestOrder.objects.filter(**filter_kwargs).filter(
            order_results__test_indicator_name__in=('门尼', '比重', '硬度')
        ).values('production_factory_date__day', 'product_no').annotate(cnt=Count('id', distinct=True))
        yc_product_daily_check_dict = {}
        for item in yc_product_daily_check_data:
            day = item['production_factory_date__day']
            product_no = item['product_no'].split('-')[2]
            qty = item['cnt']
            re_result = re.match(r'[A-Z]+', product_no)
            if not re_result:
                continue
            pb_type = pt_dict.get(re_result.group(), '未知')
            if pb_type in yc_product_daily_check_dict:
                if day not in yc_product_daily_check_dict[pb_type]:
                    yc_product_daily_check_dict[pb_type][day] = qty
                else:
                    yc_product_daily_check_dict[pb_type][day] += qty
                yc_product_daily_check_dict[pb_type]['total'] += qty
            else:
                yc_product_daily_check_dict[pb_type] = {day: qty, 'total': qty}

        # 规格一次每日不合格多少车次
        yc_product_daily_unqualified_data = MaterialTestOrder.objects.filter(**filter_kwargs).filter(
            order_results__test_indicator_name__in=('门尼', '比重', '硬度'),
            order_results__level=2).values('production_factory_date__day', 'product_no').annotate(cnt=Count('id', distinct=True))
        yc_product_daily_unqualified_dict = {}
        for item in yc_product_daily_unqualified_data:
            day = item['production_factory_date__day']
            product_no = item['product_no'].split('-')[2]
            qty = item['cnt']
            re_result = re.match(r'[A-Z]+', product_no)
            if not re_result:
                continue
            pb_type = pt_dict.get(re_result.group(), '未知')
            if pb_type in yc_product_daily_unqualified_dict:
                if day not in yc_product_daily_unqualified_dict[pb_type]:
                    yc_product_daily_unqualified_dict[pb_type][day] = qty
                else:
                    yc_product_daily_unqualified_dict[pb_type][day] += qty
                yc_product_daily_unqualified_dict[pb_type]['total'] += qty
            else:
                yc_product_daily_unqualified_dict[pb_type] = {day: qty, 'total': qty}

        data4 = {}
        # {'半钢-综合合格率': {'name': '半钢', 'type': '综合合格率', 1:1, 2:2, 3:3},
        #  '半钢-流变合格率': {'name': '半钢', 'type': '流变合格率', 1:1, 2:2, 3:3},
        #  '半钢-一次合格率': {'name': '半钢', 'type': '一次合格率', 1:1, 2:2, 3:3},
        #  '车胎-综合合格率': {'name': '车胎', 'type': '综合合格率', 1: 1, 2: 2, 3: 3},
        #  '车胎-流变合格率': {'name': '车胎', 'type': '一次合格率', 1: 1, 2: 2, 3: 3},
        #  '车胎-一次合格率': {'name': '车胎', 'type': '', 1: 1, 2: 2, 3: 3},
        #  '斜胶胎-综合合格率': {'name': '斜胶胎', 'type'综合合格率: '', 1: 1, 2: 2, 3: 3},
        #  '斜胶胎-流变合格率': {'name': '斜胶胎', 'type': '一次合格率', 1: 1, 2: 2, 3: 3},
        #  '斜胶胎-一次合格率': {'name': '斜胶胎', 'type': '', 1: 1, 2: 2, 3: 3},
        #  }

        # 规格合格率
        for i in range(1, days+1):
            for product_type in pt_list:
                if product_type == '斜胶胎（含实心胎）':
                    zh_qualified_qty = mto_product_qualified_data_dict.get('斜胶胎（不含实心胎）', {}).get(i, 0) +\
                                       mto_product_qualified_data_dict.get('实心胎', {}).get(i, 0)
                    zh_total_qty = mto_product_dict.get('斜胶胎（不含实心胎）', {}).get(i, 0) + \
                                   mto_product_dict.get('实心胎', {}).get(i, 0)
                    zh_rate = "" if not zh_total_qty else round(zh_qualified_qty / zh_total_qty * 100, 1)
                    k = product_type + '-综合合格率'
                    if k not in data4:
                        q_qty1 = mto_product_qualified_data_dict.get('斜胶胎（不含实心胎）', {}).get('total', 0) + \
                                 mto_product_qualified_data_dict.get('实心胎', {}).get('total', 0)
                        t_qty1 = mto_product_dict.get('斜胶胎（不含实心胎）', {}).get('total', 0) + \
                                 mto_product_dict.get('实心胎', {}).get('total', 0)
                        rate1 = "" if not t_qty1 else round(q_qty1 / t_qty1 * 100, 1)
                        data4[k] = {"name": product_type, "type": "综合合格率%", "rate": rate1, i: zh_rate}
                    else:
                        data4[k][i] = zh_rate

                    lb_unqualified_qty = lb_product_daily_unqualified_dict.get('斜胶胎（不含实心胎）', {}).get(i, 0) +\
                                         lb_product_daily_unqualified_dict.get('实心胎', {}).get(i, 0)
                    lb_total_qty = lb_product_daily_check_dict.get('斜胶胎（不含实心胎）', {}).get(i, 0) +\
                                   lb_product_daily_check_dict.get('实心胎', {}).get(i, 0)
                    lb_rate = "" if not lb_total_qty else round((1-lb_unqualified_qty / lb_total_qty) * 100, 1)
                    k = product_type + '-流变合格率'
                    if k not in data4:
                        q_qty2 = lb_product_daily_unqualified_dict.get('斜胶胎（不含实心胎）', {}).get('total', 0) +\
                                 lb_product_daily_unqualified_dict.get('实心胎', {}).get('total', 0)
                        t_qty2 = lb_product_daily_check_dict.get('斜胶胎（不含实心胎）', {}).get('total', 0) +\
                                 lb_product_daily_check_dict.get('实心胎', {}).get('total', 0)
                        rate2 = "" if not t_qty2 else round((1-q_qty2 / t_qty2) * 100, 1)
                        data4[k] = {"name": product_type, "type": "流变合格率%", "rate": rate2, i: lb_rate}
                    else:
                        data4[k][i] = lb_rate

                    yc_unqualified_qty = yc_product_daily_unqualified_dict.get('斜胶胎（不含实心胎）', {}).get(i, 0) +\
                                         yc_product_daily_unqualified_dict.get('实心胎', {}).get(i, 0)
                    yc_total_qty = yc_product_daily_check_dict.get('斜胶胎（不含实心胎）', {}).get(i, 0) +\
                                   yc_product_daily_check_dict.get('实心胎', {}).get(i, 0)
                    yc_rate = "" if not yc_total_qty else round((1-yc_unqualified_qty / yc_total_qty) * 100, 1)
                    k = product_type + '-一次合格率'
                    if k not in data4:
                        q_qty3 = yc_product_daily_unqualified_dict.get('斜胶胎（不含实心胎）', {}).get('total', 0) +\
                                 yc_product_daily_unqualified_dict.get('实心胎', {}).get('total', 0)
                        t_qty3 = yc_product_daily_check_dict.get('斜胶胎（不含实心胎）', {}).get('total', 0) +\
                                 yc_product_daily_check_dict.get('实心胎', {}).get('total', 0)
                        rate3 = "" if not t_qty3 else round((1-q_qty3 / t_qty3) * 100, 1)
                        data4[k] = {"name": product_type, "type": "一次合格率%", "rate": rate3, i: yc_rate}
                    else:
                        data4[k][i] = yc_rate
                else:
                    zh_qualified_qty = mto_product_qualified_data_dict.get(product_type, {}).get(i, 0)
                    zh_total_qty = mto_product_dict.get(product_type, {}).get(i)
                    zh_rate = "" if not zh_total_qty else round(zh_qualified_qty/zh_total_qty*100, 1)
                    k = product_type + '-综合合格率'
                    if k not in data4:
                        q_qty1 = mto_product_qualified_data_dict.get(product_type, {}).get('total', 0)
                        t_qty1 = mto_product_dict.get(product_type, {}).get('total')
                        rate1 = "" if not t_qty1 else round(q_qty1/t_qty1*100, 1)
                        data4[k] = {"name": product_type, "type": "综合合格率%", "rate": rate1, i: zh_rate}
                    else:
                        data4[k][i] = zh_rate

                    lb_unqualified_qty = lb_product_daily_unqualified_dict.get(product_type, {}).get(i, 0)
                    lb_total_qty = lb_product_daily_check_dict.get(product_type, {}).get(i)
                    lb_rate = "" if not lb_total_qty else round((1-lb_unqualified_qty / lb_total_qty) * 100, 1)
                    k = product_type + '-流变合格率'
                    if k not in data4:
                        q_qty2 = lb_product_daily_unqualified_dict.get(product_type, {}).get('total', 0)
                        t_qty2 = lb_product_daily_check_dict.get(product_type, {}).get('total')
                        rate2 = "" if not t_qty2 else round((1-q_qty2 / t_qty2) * 100, 1)
                        data4[k] = {"name": product_type, "type": "流变合格率%", "rate": rate2, i: lb_rate}
                    else:
                        data4[k][i] = lb_rate

                    yc_unqualified_qty = yc_product_daily_unqualified_dict.get(product_type, {}).get(i, 0)
                    yc_total_qty = yc_product_daily_check_dict.get(product_type, {}).get(i)
                    yc_rate = "" if not yc_total_qty else round((1-yc_unqualified_qty / yc_total_qty) * 100, 1)
                    k = product_type + '-一次合格率'
                    if k not in data4:
                        q_qty3 = yc_product_daily_unqualified_dict.get(product_type, {}).get('total', 0)
                        t_qty3 = yc_product_daily_check_dict.get(product_type, {}).get('total')
                        rate3 = "" if not t_qty3 else round((1-q_qty3 / t_qty3) * 100, 1)
                        data4[k] = {"name": product_type, "type": "一次合格率%", "rate": rate3, i: yc_rate}
                    else:
                        data4[k][i] = yc_rate
        return Response({'data': data4.values()})


@method_decorator([api_recorder], name="dispatch")
class ProductTestValueHistoryView(APIView):

    def get(self, request):
        factory_date = self.request.query_params.get('factory_date')
        classes = self.request.query_params.get('classes')
        product_no = self.request.query_params.get('product_no')
        equip_no = self.request.query_params.get('equip_no')
        data_point = self.request.query_params.get('data_point')
        test_order_ids = list(MaterialTestOrder.objects.filter(
            product_no=product_no,
            production_class=classes,
            production_equip_no=equip_no,
            production_factory_date=factory_date).values_list('id', flat=True))
        test_result = MaterialTestResult.objects.filter(
            material_test_order_id__in=test_order_ids,
            data_point_name=data_point
        ).aggregate(min_trains=Min('material_test_order__actual_trains'),
                    max_trains=Max('material_test_order__actual_trains'))
        return Response(test_result)


@method_decorator([api_recorder], name="dispatch")
class ProductIndicatorStandard(APIView):

    def get(self, request):
        product_no = self.request.query_params.get('product_no')
        st = self.request.query_params.get('st')
        et = self.request.query_params.get('et')
        equip_no = self.request.query_params.get('equip_no')
        classes = self.request.query_params.get('classes')
        stage = self.request.query_params.get('stage')
        production_group = self.request.query_params.get('production_group')
        state = self.request.query_params.get('state')
        sum_project = self.request.query_params.get('sum_project')
        indicators_data = MaterialDataPointIndicator.objects.filter(
            material_test_method__material__material_no=product_no,
            level=1,
            delete_flag=False
        ).values('data_point__name', 'upper_limit', 'lower_limit', 'data_point__unit')
        if sum_project:
            queryset = MaterialTestResult.objects.all()
            if st:
                queryset = queryset.filter(material_test_order__production_factory_date__gte=st)
            if et:
                queryset = queryset.filter(material_test_order__production_factory_date__lte=et)
            if equip_no:
                queryset = queryset.filter(material_test_order__production_equip_no=equip_no)
            if product_no:
                queryset = queryset.filter(material_test_order__product_no=product_no)
            if classes:
                queryset = queryset.filter(material_test_order__production_class=classes)
            if stage:
                queryset = queryset.filter(material_test_order__product_no__icontains='-{}-'.format(stage))
            if production_group:
                queryset = queryset.filter(material_test_order__production_group=production_group)
            if state:
                if state == '检测中':
                    queryset = queryset.filter(material_test_order__is_finished=False)
                elif state == '合格':
                    queryset = queryset.filter(material_test_order__is_finished=True,
                                               material_test_order__is_qualified=True)
                elif state == '不合格':
                    queryset = queryset.filter(material_test_order__is_finished=True,
                                               material_test_order__is_qualified=False)
            summary_data = queryset.values('data_point_name').annotate(avg=Avg('value'), std=StdDev('value', sample=True))
            indicators_data_dict = {i['data_point__name']: i for i in indicators_data}
            for item in summary_data:
                data_point_name = item.pop('data_point_name')
                item['data_point__name'] = data_point_name
                id_data = indicators_data_dict.get(data_point_name)
                if id_data:
                    item['upper_limit'] = id_data['upper_limit']
                    item['lower_limit'] = id_data['lower_limit']
                    item['data_point__unit'] = id_data['data_point__unit']
            return Response(summary_data)
        return Response(indicators_data)


@method_decorator([api_recorder], name="dispatch")
class ProductMaterials(APIView):

    def get(self, request):
        all_product_nos = set(ProductBatching.objects.exclude(stage_product_batch_no__endswith='_NEW').values_list('stage_product_batch_no', flat=True))
        # 群控启用
        sfj_used_recipes = set(ProductBatching.objects.using('SFJ').filter(used_type=4, batching_type=1)
                               .order_by('-used_time').values_list('stage_product_batch_no', flat=True))
        # mes启用
        mes_used_recipes = set(ProductBatching.objects.filter(~Q(stage_product_batch_no__in=sfj_used_recipes),
                                                              ~Q(stage_product_batch_no__endswith='_NEW'),
                                                              used_type=4, batching_type=2)
                               .order_by('-used_time').values_list('stage_product_batch_no', flat=True))
        used_recipes = sfj_used_recipes | mes_used_recipes
        unused_products = all_product_nos - used_recipes
        ret = [{'product_no': j, 'used': True} for j in used_recipes] + [{'product_no': i, 'used': False} for i in unused_products]
        return Response(ret)


@method_decorator([api_recorder], name="dispatch")
class ProductTestedTrains(APIView):

    def get(self, request):
        factory_date = self.request.query_params.get('factory_date')
        classes = self.request.query_params.get('classes')
        product_no = self.request.query_params.get('product_no')
        equip_no = self.request.query_params.get('equip_no')
        test_indicator_name = self.request.query_params.get('test_indicator_name')
        test_plan_data = list(ProductTestPlanDetail.objects.filter(
            factory_date=factory_date,
            product_no=product_no,
            equip_no=equip_no,
            production_classes=classes,
            test_plan__test_indicator_name=test_indicator_name
        ).order_by('actual_trains').values('actual_trains', 'test_plan__test_interval'))
        if not test_plan_data:
            return Response({})
        last_test_plan = test_plan_data[-1]
        first_test_plan = test_plan_data[0]
        return Response({'max_trains': last_test_plan['actual_trains'] + last_test_plan['test_plan__test_interval'] - 1,
                        'min_trains': first_test_plan['actual_trains']})


@method_decorator([api_recorder], name="dispatch")
class ScorchTimeView(ModelViewSet):
    queryset = ScorchTime.objects.order_by('input_date')
    serializer_class = ScorchTimeSerializer
    permission_classes = (IsAuthenticated, PermissionClass({'view': 'view_scorch_time',
                                                            'add': 'add_scorch_time',
                                                            'change': 'change_scorch_time',
                                                            'delete': 'delete_scorch_time'}))
    filter_backends = (DjangoFilterBackend,)
    filter_class = ScorchTimeFilter
    pagination_class = None

    def list(self, request, *args, **kwargs):
        combine_flag = self.request.query_params.get('combine_flag')
        if combine_flag:
            queryset = self.filter_queryset(self.get_queryset())
            data = self.get_serializer(queryset, many=True).data
            ret = {}
            # {"product_no": "A",
            #  "2022-01-01": [{"equip_no": "Z01", "method_name": "test1"}, {"equip_no": "Z02", "method_name": "test2"}],
            #  "2022-02-02": [{"equip_no": "Z01", "method_name": "test1"}, {"equip_no": "Z02", "method_name": "test2"}]}
            for item in data:
                product_no = item['product_no']
                if product_no in ret:
                    if item['input_date'] in ret[product_no]:
                        ret[product_no][item['input_date']].append({"equip_no": item['equip_no'],
                                                                    "classes": item['classes'],
                                                                    "test_method_name": item['test_method_name'],
                                                                    "test_time": item['test_time'],
                                                                    "id": item['id']})
                    else:
                        ret[product_no][item['input_date']] = [{"equip_no": item['equip_no'],
                                                                "classes": item['classes'],
                                                                "test_method_name": item['test_method_name'],
                                                                "test_time": item['test_time'],
                                                                "id": item['id']}]
                else:
                    ret[product_no] = {"product_no": item['product_no'],
                                       "recipe_type": item['recipe_type'],
                                       item['input_date']: [{"equip_no": item['equip_no'],
                                                             "classes": item['classes'],
                                                             "test_method_name": item['test_method_name'],
                                                             "test_time": item['test_time'],
                                                             "id": item['id']}]}
            dates = queryset.values('input_date').annotate(a=Count('id')).values_list('input_date', flat=True)

            return Response({'data': ret.values(), 'dates': dates})
        else:
            return super().list(self, request, *args, **kwargs)


@method_decorator([api_recorder], name="dispatch")
class ProductTestPlanInterval(APIView):

    def get(self, request):
        product_no = self.request.query_params.get('product_no')
        details = ProductTestPlanDetail.objects.filter(product_no=product_no).order_by('id').last()
        if details:
            interval = details.test_plan.test_interval
        else:
            interval = None
        return Response({'interval': interval})