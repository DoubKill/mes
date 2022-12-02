import calendar
from io import BytesIO

from django.db import connection
from django.db.models import Q, Max, Sum, Count, F, DecimalField, Min
from django.http import HttpResponse
from django.utils.decorators import method_decorator
from django_filters.rest_framework import DjangoFilterBackend
from openpyxl import load_workbook
from rest_framework.exceptions import ValidationError
from rest_framework.generics import ListAPIView
from rest_framework.permissions import IsAuthenticatedOrReadOnly, IsAuthenticated
from rest_framework.views import APIView

from basics.models import WorkSchedulePlan, Equip, GlobalCode
from equipment.utils import gen_template_response
from inventory.models import InventoryLog, DispatchLog, FinalGumInInventoryLog, MixGumInInventoryLog, \
    FinalGumOutInventoryLog, MixGumOutInventoryLog
from mes import settings
from mes.common_code import get_weekdays
from mes.derorators import api_recorder
from plan.models import ProductClassesPlan, SchedulingEquipCapacity
from mes.permissions import PermissionClass
from production.filters import CollectTrainsFeedbacksFilter
from production.models import TrainsFeedbacks, EquipDownDetails
from production.serializers import CollectTrainsFeedbacksSerializer
import datetime
from rest_framework.response import Response

from quality.models import MaterialTestOrder


@method_decorator([api_recorder], name="dispatch")
class ClassesBanBurySummaryView(ListAPIView):
    """班次密炼统计"""
    queryset = TrainsFeedbacks.objects.all()
    pagination_class = None

    @staticmethod
    def get_class_dimension_page_data(page):
        factory_dates = set([str(item['date']) for item in page])
        classes = set([item['classes'] for item in page])
        schedule_plans = WorkSchedulePlan.objects.filter(
            plan_schedule__work_schedule__work_procedure__global_name='密炼').filter(
            Q(plan_schedule__day_time__in=factory_dates) | Q(classes__global_name__in=classes)).values(
            'plan_schedule__day_time', 'classes__global_name', 'start_time', 'end_time'
        )
        schedule_plans_dict = {
            str(schedule_plan['plan_schedule__day_time']) + schedule_plan['classes__global_name']:
                schedule_plan for schedule_plan in schedule_plans}
        """如果是按照工厂时间并且是按照班次分组则需要找出该班次的总时间"""
        for value in page:
            key = str(value['date']) + value['classes']
            if key in schedule_plans_dict:
                value['classes_time'] = (schedule_plans_dict[key]['end_time']
                                         - schedule_plans_dict[key]['start_time']).seconds
        return page

    def list(self, request, *args, **kwargs):
        st = self.request.query_params.get('st')  # 开始时间
        et = self.request.query_params.get('et')  # 结束时间
        dimension = self.request.query_params.get('dimension', '2')  # 维度 1：班次  2：日 3：月
        day_type = self.request.query_params.get('day_type', '2')  # 日期类型 1：自然日  2：工厂日
        equip_no = self.request.query_params.get('equip_no')  # 设备编号
        product_no = self.request.query_params.get('product_no')  # 胶料编码

        where_str = """not(CLASSES=' ') """
        if not all([st, et]):
            raise ValidationError('请选择开始和结束日期查询！')
        try:
            e_time = datetime.datetime.strptime(et, '%Y-%m-%d')
            s_time = datetime.datetime.strptime(st, '%Y-%m-%d')
        except Exception:
            raise ValidationError('日期错误！')
        diff = e_time - s_time
        if diff.days > 7:
            raise ValidationError('搜索日期跨度不得超过一周！')

        if day_type == '1':  # 按照自然日
            group_date_field = 'end_time'
        else:  # 按照工厂日期
            group_date_field = 'factory_date'
            where_str += 'and factory_date is not null '

        if dimension == '3':  # 按照月份分组
            date_format = 'yyyy-mm'
        else:
            date_format = 'yyyy-mm-dd'

        select_str = """plan_classes_uid,
                    equip_no,
                    product_no,
                   to_char({}, '{}') AS "date",
                   MAX(actual_trains) AS max_trains,
                   MIN(actual_trains) AS min_trains,
                   max(ceil(
                    (To_date(to_char(END_TIME,'yyyy-mm-dd hh24:mi:ss') , 'yyyy-mm-dd hh24-mi-ss')
                     - To_date(to_char(BEGIN_TIME,'yyyy-mm-dd hh24:mi:ss'), 'yyyy-mm-dd hh24-mi-ss')
                    ) * 24 * 60 * 60 )) as max_train_time,
                    min(ceil(
                    (To_date(to_char(END_TIME,'yyyy-mm-dd hh24:mi:ss') , 'yyyy-mm-dd hh24-mi-ss')
                     - To_date(to_char(BEGIN_TIME,'yyyy-mm-dd hh24:mi:ss'), 'yyyy-mm-dd hh24-mi-ss')
                    ) * 24 * 60 * 60 )) as min_train_time,
                   sum(ceil(
                    (To_date(to_char(END_TIME,'yyyy-mm-dd hh24:mi:ss') , 'yyyy-mm-dd hh24-mi-ss')
                     - To_date(to_char(BEGIN_TIME,'yyyy-mm-dd hh24:mi:ss'), 'yyyy-mm-dd hh24-mi-ss')
                    ) * 24 * 60 * 60 )) as total_time""".format(group_date_field, date_format)

        group_by_str = """plan_classes_uid,
                   equip_no,
                   product_no,
                   to_char({}, '{}')""".format(group_date_field, date_format)
        classes_order = ''

        if dimension == '1':
            group_by_str += ' ,classes'
            select_str += ' ,classes'
            classes_order = ',classes desc'
        order_by_str = """ "date",product_no {}, equip_no""".format(classes_order)

        if equip_no:
            where_str += """and equip_no like '%{}%' """.format(equip_no)
        if product_no:
            where_str += """and product_no like '%{}%' """.format(product_no)

        if st:
            try:
                datetime.datetime.strptime(st, "%Y-%m-%d")
            except Exception:
                raise ValidationError("开始日期格式错误")
            if day_type == '1':
                where_str += """and to_char(end_time, 'yyyy-mm-dd') >= '{}' """.format(st)
            else:
                where_str += """and to_char(factory_date, 'yyyy-mm-dd') >= '{}' """.format(st)

        if et:
            try:
                datetime.datetime.strptime(et, "%Y-%m-%d")
            except Exception:
                raise ValidationError("结束日期格式错误")
            if day_type == '1':
                where_str += """and to_char(end_time, 'yyyy-mm-dd') <= '{}' """.format(et)
            else:
                where_str += """and to_char(factory_date, 'yyyy-mm-dd') <= '{}' """.format(et)

        sql = """
            SELECT {}
           FROM
           trains_feedbacks
           where {}
           GROUP BY {}
           order by {};""".format(select_str, where_str, group_by_str, order_by_str)
        ret = {}
        cursor = connection.cursor()
        cursor.execute(sql)
        data = cursor.fetchall()
        for item in data:
            query_group_by_index = [3]
            item_dict = {
                'equip_no': item[1],
                'product_no': item[2],
                'max_trains': item[4],
                'min_trains': item[5],
                'max_train_time': item[6],
                'min_train_time': item[7],
                'total_time': item[8] if item[1] != 'Z04' else int(item[8]/2),
                'date': item[3]
            }
            if dimension == '1':
                query_group_by_index.append(-1)
                item_dict['classes'] = item[-1]
            diff_trains = item[4] - item[5] + 1
            item_key = ''.join([str(item[i]) for i in query_group_by_index]) + item[1] + item[2]
            if item_key not in ret:
                item_dict['total_trains'] = diff_trains
                ret[item_key] = item_dict
            else:
                ret[item_key]['total_trains'] += diff_trains
                ret[item_key]['total_time'] += item_dict['total_time']
                if ret[item_key]['max_train_time'] < item[6]:
                    ret[item_key]['max_train_time'] = item[6]
                if ret[item_key]['min_train_time'] > item[7]:
                    ret[item_key]['min_train_time'] = item[7]

        data = ret.values()
        if day_type == '2' and dimension == '1':
            data = self.get_class_dimension_page_data(ret.values())

        return Response(data)


@method_decorator([api_recorder], name="dispatch")
class EquipBanBurySummaryView(ClassesBanBurySummaryView):
    """机台密炼统计"""
    queryset = TrainsFeedbacks.objects.all()

    def list(self, request, *args, **kwargs):
        st = self.request.query_params.get('st')  # 开始时间
        et = self.request.query_params.get('et')  # 结束时间
        dimension = self.request.query_params.get('dimension', '2')  # 维度 1：班次  2：日 3：月
        day_type = self.request.query_params.get('day_type', '2')  # 日期类型 1：自然日  2：工厂日
        equip_no = self.request.query_params.get('equip_no')  # 设备编号

        if not all([st, et]):
            raise ValidationError('请选择开始和结束日期查询！')
        try:
            e_time = datetime.datetime.strptime(et, '%Y-%m-%d')
            s_time = datetime.datetime.strptime(st, '%Y-%m-%d')
        except Exception:
            raise ValidationError('日期错误！')
        diff = e_time - s_time
        if diff.days > 7:
            raise ValidationError('搜索日期跨度不得超过一周！')

        where_str = """not(CLASSES=' ') """

        if day_type == '1':  # 按照自然日
            group_date_field = 'end_time'
        else:  # 按照工厂日期
            group_date_field = 'factory_date'
            where_str += 'and factory_date is not null '

        if dimension == '3':  # 按照月份分组
            date_format = 'yyyy-mm'
        else:
            date_format = 'yyyy-mm-dd'

        select_str = """plan_classes_uid,
                            equip_no,
                           to_char({}, '{}') AS "date",
                           MAX(actual_trains) AS max_trains,
                           MIN(actual_trains) AS min_trains,
                           sum(ceil(
                            (To_date(to_char(END_TIME,'yyyy-mm-dd hh24:mi:ss') , 'yyyy-mm-dd hh24-mi-ss')
                             - To_date(to_char(BEGIN_TIME,'yyyy-mm-dd hh24:mi:ss'), 'yyyy-mm-dd hh24-mi-ss')
                            ) * 24 * 60 * 60 )) as total_time""".format(group_date_field, date_format)

        group_by_str = """plan_classes_uid,
                           equip_no,
                           to_char({}, '{}')""".format(group_date_field, date_format)

        if dimension == '1':
            group_by_str += ' ,classes'
            select_str += ' ,classes'

        if equip_no:
            where_str += """and equip_no like '%{}%' """.format(equip_no)

        if st:
            try:
                datetime.datetime.strptime(st, "%Y-%m-%d")
            except Exception:
                raise ValidationError("开始日期格式错误")
            if day_type == '1':
                where_str += """and to_char(end_time, 'yyyy-mm-dd') >= '{}' """.format(st)
            else:
                where_str += """and to_char(factory_date, 'yyyy-mm-dd') >= '{}' """.format(st)

        if et:
            try:
                datetime.datetime.strptime(et, "%Y-%m-%d")
            except Exception:
                raise ValidationError("结束日期格式错误")
            if day_type == '1':
                where_str += """and to_char(end_time, 'yyyy-mm-dd') <= '{}' """.format(et)
            else:
                where_str += """and to_char(factory_date, 'yyyy-mm-dd') <= '{}' """.format(et)

        sql = """
                    SELECT {}
                   FROM
                   trains_feedbacks
                   where {}
                   GROUP BY {}
                   order by "date";""".format(select_str, where_str, group_by_str)
        ret = {}
        cursor = connection.cursor()
        cursor.execute(sql)
        data = cursor.fetchall()
        for item in data:
            query_group_by_index = [2]
            item_dict = {
                'equip_no': item[1],
                'max_trains': item[3],
                'min_trains': item[4],
                'total_time': item[5] if item[1] != 'Z04' else int(item[5]/2),
                'date': item[2]
            }
            if dimension == '1':
                query_group_by_index.append(-1)
                item_dict['classes'] = item[-1]
            diff_trains = item[3] - item[4] + 1
            item_key = ''.join([str(item[i]) for i in query_group_by_index]) + item[1] + item[2]
            if item_key not in ret:
                item_dict['total_trains'] = diff_trains
                ret[item_key] = item_dict
            else:
                ret[item_key]['total_trains'] += diff_trains
                ret[item_key]['total_time'] += item_dict['total_time']

        data = ret.values()
        if day_type == '2' and dimension == '1':
            data = self.get_class_dimension_page_data(ret.values())

        return Response(data)


@method_decorator([api_recorder], name="dispatch")
class CollectTrainsFeedbacksList(ListAPIView):
    """胶料单车次时间汇总"""
    queryset = TrainsFeedbacks.objects.filter(delete_flag=False).order_by('product_time', 'actual_trains')
    serializer_class = CollectTrainsFeedbacksSerializer
    filter_backends = (DjangoFilterBackend,)
    filter_class = CollectTrainsFeedbacksFilter
    permission_classes = (IsAuthenticatedOrReadOnly,)
    EXPORT_FIELDS_DICT = {'班次': "classes", '设备编码': "equip_no", '胶料编码': "product_no", '车次': "actual_trains", '耗时/s': "time_consuming", '间隔时间/s': "interval_time"}
    FILE_NAME = '胶料单车次时间汇总'

    def list(self, request, *args, **kwargs):
        export = self.request.query_params.get('export')
        queryset = self.filter_queryset(self.get_queryset())
        if export:
            data = self.get_serializer(queryset, many=True).data
            return gen_template_response(self.EXPORT_FIELDS_DICT, data, self.FILE_NAME)
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)

        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)


@method_decorator([api_recorder], name="dispatch")
class SumCollectTrains(APIView):
    """胶料单车次时间汇总最大最小平均时间"""

    def get(self, request, *args, **kwargs):
        params = request.query_params
        st = params.get("st", None)  # 开始时间
        classes = params.get("classes", None)  # 班次
        equip_no = params.get("equip_no", None)  # 设备编号
        product_no = params.get("product_no", None)  # 胶料编码
        dict_filter = {'delete_flag': False}
        if st:
            dict_filter['factory_date'] = st
        if classes:
            dict_filter['classes'] = classes
        if equip_no:
            dict_filter['equip_no'] = equip_no
        if product_no:
            dict_filter['product_no__icontains'] = product_no
        tfb_set = TrainsFeedbacks.objects.filter(**dict_filter).values()
        for tfb_obj in tfb_set:
            if not tfb_obj['end_time'] or not tfb_obj['begin_time']:
                tfb_obj['time_consuming'] = None
            else:
                tfb_obj['time_consuming'] = tfb_obj['end_time'] - tfb_obj['begin_time']
        if tfb_set:
            sum_time = datetime.timedelta(days=0, seconds=0, microseconds=0, milliseconds=0, minutes=0, hours=0,
                                          weeks=0)
            max_time = tfb_set[0]['time_consuming']
            min_time = tfb_set[0]['time_consuming']
            for tfb_obj in tfb_set:
                if tfb_obj['time_consuming'] > max_time:
                    max_time = tfb_obj['time_consuming']
                if tfb_obj['time_consuming'] < min_time:
                    min_time = tfb_obj['time_consuming']
                sum_time += tfb_obj['time_consuming']
            avg_time = sum_time / len(tfb_set)
            return Response(
                {'results': {'sum_time': sum_time, 'max_time': max_time, 'min_time': min_time, 'avg_time': avg_time}})
        else:
            return Response({'results': {'sum_time': None, 'max_time': None, 'min_time': None, 'avg_time': None}})


@method_decorator([api_recorder], name="dispatch")
class CutTimeCollect(APIView):
    """规格切换时间汇总"""
    EXPORT_FIELDS_DICT = {'时间': "time", '设备编码': "equip_no",
                          '切换前计划号': "plan_classes_uid_age",
                          '切换后计划号': "plan_classes_uid_later",
                          '切换前胶料编码': "cut_ago_product_no",
                          '切换后胶料编码': "cut_later_product_no",
                          '切换规格时间标准': 'standard_time',
                          '切换规格耗时（秒）': "normal_cut_time_consumer",
                          '异常时间(秒）': 'err_cut_time_consumer',
                          '切换规格时间完成率%': 'rate'}
    FILE_NAME = '规格切换时间汇总'
    permission_classes = (IsAuthenticated, PermissionClass({'view': 'view_product_exchange_consume'}))

    def export_xlsx(self, export_fields_dict, data, file_name, columns, result, date_month):
        export_fields = list(export_fields_dict.values())
        sheet_heads = list(export_fields_dict.keys())
        wb = load_workbook('xlsx_template/example.xlsx')
        sheet = wb.worksheets[0]
        sheet.title = '规格切换时间明细'
        sheet1 = wb.create_sheet(title='规格切换次数统计')

        sheet1.cell(1, 1).value = date_month

        for idx, sheet_head in enumerate(sheet_heads):
            sheet.cell(1, idx + 1).value = sheet_head

        data_row = 2
        for i in data:
            for col_num, data_key in enumerate(export_fields):
                set_value = i[data_key]
                sheet.cell(data_row, col_num + 1).value = set_value
            data_row += 1

        columns.insert(0, '机台')
        columns.insert(1, '总次数')
        columns.insert(2, '平均次数')
        # 写入文件标题
        for idx, sheet_head in enumerate(columns):
            sheet1.cell(2, idx + 1).value = sheet_head
        data_row1 = 3
        for i in result:
            for col_num, data_key in enumerate(columns):
                sheet1.cell(data_row1, col_num + 1).value = i.get(data_key)
            data_row1 += 1
        # wb.remove_sheet(ws)
        output = BytesIO()
        wb.save(output)
        # 重新定位到开始
        output.seek(0)
        response = HttpResponse(content_type='application/vnd.ms-excel')
        filename = file_name
        response['Content-Disposition'] = u'attachment;filename= ' + filename.encode('gbk').decode(
            'ISO-8859-1') + '.xls'
        response.write(output.getvalue())
        return response

    def get(self, request, *args, **kwargs):
        # 筛选工厂
        equip_no = self.request.query_params.get('equip_no')  # 设备编号
        st = self.request.query_params.get('st')
        et = self.request.query_params.get('et')
        classes = self.request.query_params.get('classes')
        group = self.request.query_params.get('group')
        page = int(self.request.query_params.get('page', 1))
        page_size = int(self.request.query_params.get('page_size', 10))
        try:
            s_time = datetime.datetime.strptime(st, '%Y-%m-%d')
            e_time = datetime.datetime.strptime(et, '%Y-%m-%d')
        except Exception:
            raise ValidationError('日期错误！')
        filter_kwargs = {'classes__isnull': False}
        if st:
            filter_kwargs['factory_date__gte'] = s_time
        if et:
            filter_kwargs['factory_date__lte'] = e_time
        if equip_no:
            filter_kwargs['equip_no'] = equip_no
        query_set = list(TrainsFeedbacks.objects.filter(**filter_kwargs).values(
            'plan_classes_uid', 'factory_date', 'classes', 'equip_no', 'product_no'
        ).annotate(st_time=Min('begin_time'), et_time=Max('end_time')
                   ).values(
            'plan_classes_uid', 'factory_date', 'classes', 'equip_no', 'st_time', 'et_time', 'product_no'
        ).order_by('factory_date', 'equip_no', 'st_time'))
        gc = GlobalCode.objects.filter(global_type__type_name='规格切换时间标准', global_name=equip_no).first()
        if gc:
            standard_time = int(gc.description)
        else:
            standard_time = None

        factory_classes_group_map = WorkSchedulePlan.objects.filter(
            plan_schedule__day_time__gte=s_time,
            plan_schedule__day_time__lte=e_time,
            plan_schedule__work_schedule__work_procedure__global_name='密炼'
        ).values('plan_schedule__day_time', 'classes__global_name', 'group__global_name')
        factory_classes_group_map_dict = {
            i['plan_schedule__day_time'].strftime('%Y-%m-%d') + '-' + i['classes__global_name']: i for i in
            factory_classes_group_map}
        return_list = []
        sum_normal = 0
        sum_error = 0
        cnt_normal = 0
        cnt_error = 0
        sum_rate = 0
        for idx, item in enumerate(query_set[:-1]):
            key = item['factory_date'].strftime('%Y-%m-%d') + '-' + item['classes']
            next_st_time = query_set[idx + 1]['st_time']
            if item['equip_no'] != query_set[idx + 1]['equip_no']:
                continue
            time_consuming = int((next_st_time - item['et_time']).total_seconds())
            data = dict()
            data['group'] = factory_classes_group_map_dict[key]['group__global_name']
            if abs(time_consuming) >= 1000 or time_consuming <= 0:
                data['err_cut_time_consumer'] = time_consuming
                data['normal_cut_time_consumer'] = None
                sum_error += time_consuming
                cnt_error += 1
            else:
                data['err_cut_time_consumer'] = None
                data['normal_cut_time_consumer'] = time_consuming
                sum_normal += time_consuming
                cnt_normal += 1
            data['classes'] = item['classes']
            data['time'] = item['factory_date'].strftime('%Y-%m-%d')
            data['cut_ago_product_no'] = item['product_no']
            data['equip_no'] = item['equip_no']
            data['cut_later_product_no'] = query_set[idx + 1]['product_no']
            data['plan_classes_uid_age'] = item['plan_classes_uid']
            data['plan_classes_uid_later'] = query_set[idx + 1]['plan_classes_uid']
            data['standard_time'] = standard_time
            rate = None if not time_consuming else round((standard_time / time_consuming) * 100, 2)
            data['rate'] = None if not standard_time else '{}%'.format(rate)
            sum_rate += 0 if not rate else rate
            return_list.append(data)
        if group:
            return_list = list(filter(lambda x: x['group'] == group, return_list))
        if classes:
            return_list = list(filter(lambda x: x['classes'] == classes, return_list))
        if self.request.query_params.get('export'):
            year = s_time.year
            month = s_time.month
            product_info_shift_queryset = TrainsFeedbacks.objects.filter(
                factory_date__year=year, factory_date__month=month).values(
                'factory_date', 'equip_no').annotate(cnt=Count('product_no', distinct=True))
            product_info_shift_data = {i: {} for i in Equip.objects.filter(
                category__equip_type__global_name="密炼设备"
            ).order_by('equip_no').values_list('equip_no', flat=True)}
            days_list = [str(i) + '日' for i in range(1, calendar.monthrange(year, month)[1] + 1)]
            for i in product_info_shift_queryset:
                equip_no = i['equip_no']
                factory_date = str(i['factory_date'].day) + '日'
                cnt = i['cnt']
                product_info_shift_data[equip_no][factory_date] = cnt
            for k, v in product_info_shift_data.items():
                total_cnt = sum(v.values())
                avg_cnt = '' if not v else round(total_cnt / len(v), 1)
                v['机台'] = k
                v['总次数'] = total_cnt
                v['平均次数'] = avg_cnt
            date_month = '月份：{}年{}月'.format(year, month)
            return self.export_xlsx(self.EXPORT_FIELDS_DICT, return_list, self.FILE_NAME, days_list, product_info_shift_data.values(), date_month)
        avg_normal = None if not cnt_normal else round(sum_normal / cnt_normal)
        avg_error = None if not cnt_error else round(sum_error / cnt_error)
        avg_rate = None if cnt_normal + cnt_error == 0 else round(sum_rate / (cnt_normal + cnt_error), 2)
        # 分页
        counts = len(return_list)
        return_list = return_list[(page - 1) * page_size:page_size * page]
        # return_list.append(
        #     {'sum_time': sum_time, 'max_time': max_time, 'min_time': min_time, 'avg_time': avg_time})
        return Response({'count': counts, 'results': return_list,
                         'avg_normal': avg_normal, 'avg_error': avg_error, 'avg_rate': avg_rate})


@method_decorator([api_recorder], name="dispatch")
class IndexOverview(APIView):
    """首页-今日概况"""
    authentication_classes = ()

    @staticmethod
    def get_current_factory_date():
        # 获取当前时间的工厂日期，开始、结束时间
        now = datetime.datetime.now()
        current_work_schedule_plan = WorkSchedulePlan.objects.filter(
            start_time__lte=now,
            end_time__gte=now,
            plan_schedule__work_schedule__work_procedure__global_name='密炼'
        ).first()
        date_now = str(now.date())
        if current_work_schedule_plan:
            date_now = str(current_work_schedule_plan.plan_schedule.day_time)
            st = current_work_schedule_plan.plan_schedule.work_schedule_plan.filter(
                classes__global_name='早班').first()
            et = current_work_schedule_plan.plan_schedule.work_schedule_plan.filter(
                classes__global_name='夜班').first()
            if st:
                begin_time = str(st.start_time)
            else:
                begin_time = date_now + ' 00:00:01'
            if et:
                end_time = str(et.end_time)
            else:
                end_time = date_now + ' 23:59:59'
        else:
            begin_time = date_now + ' 00:00:01'
            end_time = date_now + ' 23:59:59'

        return date_now, begin_time, end_time

    def get(self, request):
        ret = {}
        factory_date, begin_time, _ = self.get_current_factory_date()

        # 日计划量
        plan_data = ProductClassesPlan.objects.filter(
            work_schedule_plan__plan_schedule__day_time=factory_date,
            delete_flag=False
        ).aggregate(total_trains=Sum('plan_trains'),
                    total_weight=Sum(F('plan_trains') * F('product_batching__batching_weight'),
                                     output_field=DecimalField())/1000)

        # 日总产量
        actual_data = TrainsFeedbacks.objects.filter(
            factory_date=factory_date
        ).exclude(operation_user='Mixer2'
                  ).aggregate(total_trains=Count('id'),
                              total_weight=Sum('plan_weight')/1000)

        # 日入库量
        try:
            final_gum_data = FinalGumInInventoryLog.objects.using('lb').filter(
                start_time__gte=begin_time).filter(Q(location__startswith='1') |
                                                   Q(location__startswith='2') |
                                                   Q(location__startswith='3') |
                                                   Q(location__startswith='4')
                                                   ).aggregate(
                total_trains=Sum('qty'),
                total_weight=Sum('weight')/1000)
            final_gum_qyt = final_gum_data['total_trains'] if final_gum_data['total_trains'] else 0
            final_gum_weight = final_gum_data['total_weight'] if final_gum_data['total_weight'] else 0

            mix_gum_data = MixGumInInventoryLog.objects.using('bz').filter(
                start_time__gte=begin_time).aggregate(
                total_trains=Sum('qty'),
                total_weight=Sum('weight') / 1000)
            mix_gum_qyt = mix_gum_data['total_trains'] if mix_gum_data['total_trains'] else 0
            mix_gum_weight = mix_gum_data['total_weight'] if mix_gum_data['total_weight'] else 0
        except Exception:
            final_gum_qyt = final_gum_weight = 0
            mix_gum_qyt = mix_gum_weight = 0
        inbound_data = {
                        'total_trains': int(final_gum_qyt + mix_gum_qyt),
                        'total_weight': final_gum_weight + mix_gum_weight
        }

        # 日出库量
        try:
            final_outbound_data = FinalGumOutInventoryLog.objects.using('lb').filter(
                start_time__gte=begin_time).filter(Q(location__startswith='1') |
                                                   Q(location__startswith='2') |
                                                   Q(location__startswith='3') |
                                                   Q(location__startswith='4')
                                                   ).aggregate(
                total_trains=Sum('qty'),
                total_weight=Sum('weight')/1000)
            final_outbound_qyt = final_outbound_data['total_trains'] if final_outbound_data['total_trains'] else 0
            final_outbound_weight = final_outbound_data['total_weight'] if final_outbound_data['total_weight'] else 0

            mix_outbound_data = MixGumOutInventoryLog.objects.using('bz').filter(
                start_time__gte=begin_time
            ).aggregate(total_trains=Sum('qty'),
                        total_weight=Sum('weight') / 1000)
            mix_outbound_qyt = mix_outbound_data['total_trains'] if mix_outbound_data['total_trains'] else 0
            mix_outbound_weight = mix_outbound_data['total_weight'] if mix_outbound_data['total_weight'] else 0
        except Exception:
            final_outbound_qyt = final_outbound_weight = 0
            mix_outbound_qyt = mix_outbound_weight = 0
        outbound_data = {
                        'total_trains': int(final_outbound_qyt + mix_outbound_qyt),
                        'total_weight': final_outbound_weight + mix_outbound_weight
        }

        # 日发货量
        dispatch_data = DispatchLog.objects.filter(
            order_created_time=factory_date
        ).aggregate(total_trains=Sum('qty'),
                    total_weight=Sum('weight')/1000)

        # 日合格率
        qualified_count = MaterialTestOrder.objects.filter(production_factory_date=factory_date,
                                                           product_no__icontains='-FM',
                                                           is_qualified=True,
                                                           is_experiment=False
                                                           ).count()
        total_test_count = MaterialTestOrder.objects.filter(production_factory_date=factory_date,
                                                            product_no__icontains='-FM',
                                                            is_experiment=False
                                                            ).count()
        try:
            qualified_rate = round(qualified_count / total_test_count * 100, 2)
        except ZeroDivisionError:
            qualified_rate = 0
        except Exception:
            raise

        ret['plan_data'] = {key: round(value, 2) if value else 0 for key, value in plan_data.items()}
        ret['actual_data'] = {key: round(value, 2) if value else 0 for key, value in actual_data.items()}
        ret['qualified_rate'] = '{}%'.format(qualified_rate)
        ret['outbound_data'] = {key: round(value, 2) if value else 0 for key, value in outbound_data.items()}
        ret['dispatch_data'] = {key: round(value, 2) if value else 0 for key, value in dispatch_data.items()}
        ret['inbound_data'] = {key: round(value, 2) if value else 0 for key, value in inbound_data.items()}
        return Response(ret)


@method_decorator([api_recorder], name="dispatch")
class IndexProductionAnalyze(APIView):
    """首页-产量分析/合格率分析，参数?dimension=xxx   1:周；2:月"""
    authentication_classes = ()

    def get(self, request):
        dimension = self.request.query_params.get('dimension', '1')
        if dimension == '1':  # 周
            date_range = get_weekdays(7)
        else:  # 月
            date_range = get_weekdays(datetime.datetime.now().day)
        plan_data = ProductClassesPlan.objects.filter(
            work_schedule_plan__plan_schedule__day_time__gte=date_range[0],
            work_schedule_plan__plan_schedule__day_time__lte=date_range[-1],
            delete_flag=False
        ).values('work_schedule_plan__plan_schedule__day_time',
                 'product_batching__stage_product_batch_no'
                 ).annotate(total_trains=Sum('plan_trains'))

        # 日总产量
        actual_data = TrainsFeedbacks.objects.filter(
            factory_date__gte=date_range[0],
            factory_date__lte=date_range[-1]
        ).exclude(operation_user='Mixer2'
                  ).values('factory_date',
                           'product_no'
                           ).annotate(total_trains=Count('id'))

        plan_data_dict = {}
        for item in plan_data:
            plan_date = item['work_schedule_plan__plan_schedule__day_time'].strftime("%Y-%m-%d")
            product_no = item['product_batching__stage_product_batch_no']
            trains = item['total_trains']
            try:
                stage = product_no.split('-')[1]
                if stage in ('FM', 'RFM', 'RE'):
                    flag = '加硫'
                else:
                    flag = '无硫'
            except Exception:
                flag = '无硫'
            if plan_date in plan_data_dict:
                if flag in plan_data_dict[plan_date]:
                    plan_data_dict[plan_date][flag] += trains
                else:
                    plan_data_dict[plan_date][flag] = trains
            else:
                plan_data_dict[plan_date] = {flag: trains}

        actual_data_dict = {}
        for item in actual_data:
            actual_date = item['factory_date'].strftime("%Y-%m-%d")
            product_no = item['product_no']
            trains = item['total_trains']
            try:
                stage = product_no.split('-')[1]
                if stage in ('FM', 'RFM', 'RE'):
                    flag = '加硫'
                else:
                    flag = '无硫'
            except Exception:
                flag = '无硫'
            if actual_date in actual_data_dict:
                if flag in actual_data_dict[actual_date]:
                    actual_data_dict[actual_date][flag] += trains
                else:
                    actual_data_dict[actual_date][flag] = trains
            else:
                actual_data_dict[actual_date] = {flag: trains}

        # 日合格率
        test_date_qualify_group = MaterialTestOrder.objects.filter(
            production_factory_date__in=date_range,
            product_no__icontains='-FM',
            is_experiment=False
        ).values('is_qualified', 'production_factory_date').annotate(count=Count('id'))
        test_date_group = MaterialTestOrder.objects.filter(
            production_factory_date__in=date_range,
            product_no__icontains='-FM',
            is_experiment=False
        ).values('production_factory_date').annotate(count=Count('id'))
        test_date_group_dict = {str(item['production_factory_date']): item for item in test_date_group}
        test_date_qualify_group_dict = {}
        for item in test_date_qualify_group:
            factory_date = str(item['production_factory_date'])
            if item['is_qualified']:  # 合格
                test_date_qualify_group_dict[factory_date] = {
                    'rate': round(item['count'] / test_date_group_dict[factory_date]['count'] * 100, 2),
                    'total': test_date_group_dict[factory_date]['count'],
                    'qualified_count': item['count']
                }
            else:
                if factory_date in test_date_qualify_group_dict:
                    continue
                test_date_qualify_group_dict[factory_date] = {
                    'rate': 0,
                    'total': test_date_group_dict[factory_date]['count'],
                    'qualified_count': 0
                }

        plan_actual_data = {}
        qualified_rate_data = {}
        for date in date_range:
            qualified_rate_data[date] = {'rate': 0, 'total': 0, 'qualified_count': 0}
            plan_actual_data[date] = {
                                            'plan_trains': 0,
                                            'actual_trains': 0,
                                            'plan_add_sulfur_trains': 0,
                                            'plan_without_sulfur_trains': 0,
                                            'actual_add_sulfur_trains': 0,
                                            'actual_without_sulfur_trains': 0,
                                            'diff_trains': 0
                                        }
            if date in plan_data_dict:
                if '无硫' in plan_data_dict[date]:
                    without_sulfur_num = plan_data_dict[date]['无硫']
                    plan_actual_data[date]['plan_without_sulfur_trains'] = without_sulfur_num
                if '加硫' in plan_data_dict[date]:
                    add_sulfur_num = plan_data_dict[date]['加硫']
                    plan_actual_data[date]['plan_add_sulfur_trains'] = add_sulfur_num
                plan_actual_data[date]['plan_trains'] = plan_actual_data[date]['plan_add_sulfur_trains'] + plan_actual_data[date]['plan_without_sulfur_trains']

            if date in actual_data_dict:
                if '无硫' in actual_data_dict[date]:
                    without_sulfur_num = actual_data_dict[date]['无硫']
                    plan_actual_data[date]['actual_without_sulfur_trains'] = without_sulfur_num
                if '加硫' in actual_data_dict[date]:
                    add_sulfur_num = actual_data_dict[date]['加硫']
                    plan_actual_data[date]['actual_add_sulfur_trains'] = add_sulfur_num
                plan_actual_data[date]['actual_trains'] = plan_actual_data[date]['actual_without_sulfur_trains'] + plan_actual_data[date]['actual_add_sulfur_trains']

            plan_actual_data[date]['diff_trains'] = plan_actual_data[date]['plan_trains'] - \
                                                        plan_actual_data[date]['actual_trains']

            if date in test_date_qualify_group_dict:
                qualified_rate_data[date] = test_date_qualify_group_dict[date]
        return Response({'plan_actual_data': plan_actual_data,
                         'qualified_rate_data': qualified_rate_data,
                         'date_range': date_range})


@method_decorator([api_recorder], name="dispatch")
class IndexEquipProductionAnalyze(IndexOverview):
    """首页-机台产量分析, 参数?dimension=1:今天；2:昨天&st=开始日期&et=结束日期"""
    authentication_classes = ()

    def get(self, request):
        dimension = self.request.query_params.get('dimension')
        st = self.request.query_params.get('st')
        et = self.request.query_params.get('et')

        factory_date, _, _ = self.get_current_factory_date()
        plan_query_params = {}
        actual_query_params = {}

        if dimension:
            if dimension == '1':  # 今天
                plan_query_params['work_schedule_plan__plan_schedule__day_time'] = factory_date
                actual_query_params['factory_date'] = factory_date
            elif dimension == '2':  # 昨天
                yesterday = datetime.datetime.strptime(factory_date, '%Y-%m-%d') - datetime.timedelta(days=1)
                plan_query_params['work_schedule_plan__plan_schedule__day_time'] = yesterday.date()
                actual_query_params['factory_date'] = yesterday.date()

        if st and et:
            if et > factory_date:
                raise ValidationError('结束日期不得大于当前工厂日期！')
            plan_query_params = {'work_schedule_plan__plan_schedule__day_time__gte': st,
                                 'work_schedule_plan__plan_schedule__day_time__lte': et}
            actual_query_params = {'factory_date__gte': st, 'factory_date__lte': et}

        equip_data = [equip.equip_no for equip in Equip.objects.filter(
            category__equip_type__global_name='密炼设备').order_by('equip_no')]

        # 当日计划与实际数据
        plan_data = ProductClassesPlan.objects.filter(**plan_query_params).filter(delete_flag=False).values(
            'equip__equip_no').annotate(plan_trains=Sum('plan_trains'))
        actual_data = TrainsFeedbacks.objects.filter(
            **actual_query_params).exclude(operation_user='Mixer2').values('equip_no').annotate(actual_trains=Count('id'))

        plan_data_dict = {item['equip__equip_no']: item for item in plan_data}
        actual_data_dict = {item['equip_no']: item for item in actual_data}

        plan_actual_data = {}
        for equip_no in equip_data:
            plan_actual_data[equip_no] = {'plan_trains': 0, 'actual_trains': 0}
            if equip_no in plan_data_dict:
                plan_actual_data[equip_no]['plan_trains'] = plan_data_dict[equip_no]['plan_trains']
            if equip_no in actual_data_dict:
                plan_actual_data[equip_no]['actual_trains'] = actual_data_dict[equip_no]['actual_trains']
            plan_actual_data[equip_no]['diff_trains'] = plan_actual_data[equip_no]['plan_trains'] - \
                                                        plan_actual_data[equip_no]['actual_trains'] if \
                plan_actual_data[equip_no]['plan_trains'] > plan_actual_data[equip_no]['actual_trains'] else 0
        return Response({'plan_actual_data': plan_actual_data,
                         'equip_data': equip_data})


@method_decorator([api_recorder], name="dispatch")
class IndexEquipMaintenanceAnalyze(IndexOverview):
    """首页-机台停机时间分析,参数?dimension=1:今天；2:昨天&st=开始日期&et=结束日期"""
    authentication_classes = ()

    def get(self, request):
        dimension = self.request.query_params.get('dimension')
        st = self.request.query_params.get('st')
        et = self.request.query_params.get('et')
        factory_date, factory_begin_time, _ = self.get_current_factory_date()
        filter_kwargs = {}

        if dimension:
            if dimension == '1':  # 今天
                filter_kwargs['factory_date'] = factory_date
            elif dimension == '2':  # 昨天
                yesterday = datetime.datetime.strptime(factory_date, '%Y-%m-%d') - datetime.timedelta(days=1)
                filter_kwargs['factory_date'] = yesterday.date()
            else:  # 本月
                st = datetime.datetime.strptime(factory_date[:-2] + '01', '%Y-%m-%d')
                filter_kwargs.update({'factory_date__gte': st, 'factory_date__lte': factory_date})

        if st and et:
            filter_kwargs = {'factory_date__gte': st, 'factory_date__lte': et}

        equip_data = [equip.equip_no for equip in Equip.objects.filter(category__equip_type__global_name='密炼设备').order_by('equip_no')]
        init_data = EquipDownDetails.objects.filter(**filter_kwargs).values('equip_no').annotate(total_time=Sum('times')).values('equip_no', 'total_time')
        time_data = {i['equip_no']: round(i['total_time'], 2) for i in init_data}

        maintenance_data = {}
        # 当日超过平均工作时间、平均间隔时间数据
        for equip_no in equip_data:
            s_minute = time_data.get(equip_no, 0)
            maintenance_data.update({equip_no: {'minutes': s_minute}})

        return Response({'maintenance_data': maintenance_data, 'equip_data': equip_data})


@method_decorator([api_recorder], name="dispatch")
class CutTimeCollectSummary(APIView):
    permission_classes = (IsAuthenticated, PermissionClass({'view': 'view_product_exchange_consume'}))

    def get(self, request):
        st = self.request.query_params.get('st')
        et = self.request.query_params.get('et')
        classes = self.request.query_params.get('classes')
        group = self.request.query_params.get('group')
        try:
            s_time = datetime.datetime.strptime(st, '%Y-%m-%d')
            e_time = datetime.datetime.strptime(et, '%Y-%m-%d')
        except Exception:
            raise ValidationError('日期错误！')
        filter_kwargs = {'classes__isnull': False}
        if st:
            filter_kwargs['factory_date__gte'] = s_time
        if et:
            filter_kwargs['factory_date__lte'] = e_time
        query_set = list(TrainsFeedbacks.objects.filter(**filter_kwargs).values(
            'plan_classes_uid', 'factory_date', 'classes', 'equip_no'
        ).annotate(st_time=Min('begin_time'), et_time=Max('end_time')
                   ).values(
            'plan_classes_uid', 'factory_date', 'classes', 'equip_no', 'st_time', 'et_time'
        ).order_by('factory_date', 'equip_no', 'st_time'))
        factory_classes_group_map = WorkSchedulePlan.objects.filter(
            plan_schedule__day_time__gte=s_time,
            plan_schedule__day_time__lte=e_time,
            plan_schedule__work_schedule__work_procedure__global_name='密炼'
        ).values('plan_schedule__day_time', 'classes__global_name', 'group__global_name')
        factory_classes_group_map_dict = {
            i['plan_schedule__day_time'].strftime('%Y-%m-%d') + '-' + i['classes__global_name']: i for i in factory_classes_group_map}
        ret = {}
        for idx, item in enumerate(query_set[:-1]):
            key = item['factory_date'].strftime('%Y-%m-%d') + '-' + item['classes']
            next_st_time = query_set[idx + 1]['st_time']
            if not query_set[idx + 1]['equip_no'] == item['equip_no']:
                continue
            cut_time_consume = int((next_st_time - item['et_time']).total_seconds())
            if key in ret:
                if item['equip_no'] in ret[key]:
                    ret[key][item['equip_no']] += cut_time_consume
                    ret[key][item['equip_no'] + 'cnt'] += 1
                else:
                    ret[key][item['equip_no']] = cut_time_consume
                    ret[key][item['equip_no'] + 'cnt'] = 1
            else:
                ret[key] = {'factory_date': item['factory_date'].strftime('%Y-%m-%d'),
                            'classes': item['classes'],
                            item['equip_no']: cut_time_consume,
                            item['equip_no'] + 'cnt': 1,
                            'group': factory_classes_group_map_dict[key]['group__global_name'],
                            }
            avg_cut_time_consume = int(ret[key][item['equip_no']] / ret[key][item['equip_no'] + 'cnt'])
            if abs(avg_cut_time_consume) >= 1000 or avg_cut_time_consume <= 0:
                err_cut_time_consumer = avg_cut_time_consume
                normal_cut_time_consumer = None
            else:
                err_cut_time_consumer = None
                normal_cut_time_consumer = avg_cut_time_consume
            ret[key][item['equip_no'] + 'err_cut_time_consumer'] = err_cut_time_consumer
            ret[key][item['equip_no'] + 'normal_cut_time_consumer'] = normal_cut_time_consumer
        result = ret.values()
        if group:
            result = list(filter(lambda x: x['group'] == group, result))
        if classes:
            result = list(filter(lambda x: x['classes'] == classes, result))
        return Response(result)
