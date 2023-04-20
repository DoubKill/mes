import calendar
import decimal
import json
import datetime
import math
import re
import logging
from bisect import bisect
from copy import deepcopy

from io import BytesIO
from itertools import groupby
from operator import itemgetter

import requests
import xlwt
from django.db import connection
from django.http import HttpResponse, FileResponse
from django_pandas.io import read_frame
import pandas as pd
from openpyxl import load_workbook
from multiprocessing.dummy import Pool as ThreadPool

from equipment.utils import gen_template_response
from django.db.models.functions import TruncMonth
from django.db.models import Max, Sum, Count, Min, F, Q, DecimalField, Avg
from django.db.transaction import atomic
from django.utils.decorators import method_decorator
from django_filters.rest_framework import DjangoFilterBackend

from rest_framework import mixins, status
from rest_framework.decorators import action
from rest_framework.exceptions import ValidationError
from rest_framework.filters import OrderingFilter
from rest_framework.permissions import IsAuthenticatedOrReadOnly, IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.viewsets import GenericViewSet, ModelViewSet, ViewSet
from basics.models import GlobalCode, WorkSchedulePlan
from equipment.models import EquipMaintenanceOrder
from inventory.models import ProductInventoryLocked, BzFinalMixingRubberInventory, BzFinalMixingRubberInventoryLB
from mes import settings
from mes.common_code import OSum, date_range, days_cur_month_dates, get_virtual_time, OAvg
from mes.conf import EQUIP_LIST, JZ_EQUIP_NO
from mes.derorators import api_recorder
from mes.paginations import SinglePageNumberPagination
from mes.permissions import PermissionClass
from plan.filters import ProductClassesPlanFilter
from plan.models import ProductClassesPlan, SchedulingEquipShutDownPlan, SchedulingResult, SchedulingRecipeMachineSetting
from basics.models import Equip
from production.filters import TrainsFeedbacksFilter, PalletFeedbacksFilter, QualityControlFilter, EquipStatusFilter, \
    PlanStatusFilter, ExpendMaterialFilter, UnReachedCapacityCause, \
    ProductInfoDingJiFilter, SubsidyInfoFilter, PerformanceJobLadderFilter, AttendanceGroupSetupFilter, Equip190EFilter, \
    AttendanceClockDetailFilter, EmployeeAttendanceRecordsLogFilter
from production.models import TrainsFeedbacks, PalletFeedbacks, EquipStatus, PlanStatus, ExpendMaterial, OperationLog, \
    QualityControl, ProcessFeedback, AlarmLog, MaterialTankStatus, ProductionDailyRecords, ProductionPersonnelRecords, \
    RubberCannotPutinReason, MachineTargetYieldSettings, EmployeeAttendanceRecords, PerformanceJobLadder, \
    PerformanceUnitPrice, ProductInfoDingJi, SetThePrice, SubsidyInfo, IndependentPostTemplate, AttendanceGroupSetup, \
    FillCardApply, ApplyForExtraWork, EquipMaxValueCache, Equip190EWeight, OuterMaterial, Equip190E, \
    AttendanceClockDetail, AttendanceResultAudit, ManualInputTrains, ActualWorkingDay, EmployeeAttendanceRecordsLog, \
    RubberFrameRepair, ToolManageAccount, ActualWorkingEquip, ActualWorkingDay190E, WeightClassPlan, \
    WeightClassPlanDetail, EquipDownDetails, FinishRatio, RubberLog, RubberLogSummary, HistoryProductionGroup
from production.serializers import QualityControlSerializer, OperationLogSerializer, ExpendMaterialSerializer, \
    PlanStatusSerializer, EquipStatusSerializer, PalletFeedbacksSerializer, TrainsFeedbacksSerializer, \
    ProductionRecordSerializer, TrainsFeedbacksBatchSerializer, \
    ProductionPlanRealityAnalysisSerializer, UnReachedCapacityCauseSerializer, TrainsFeedbacksSerializer2, \
    CurveInformationSerializer, MixerInformationSerializer2, WeighInformationSerializer2, AlarmLogSerializer, \
    ProcessFeedbackSerializer, TrainsFixSerializer, PalletFeedbacksBatchModifySerializer, ProductPlanRealViewSerializer, \
    RubberCannotPutinReasonSerializer, PerformanceJobLadderSerializer, ProductInfoDingJiSerializer, \
    SetThePriceSerializer, SubsidyInfoSerializer, AttendanceGroupSetupSerializer, EmployeeAttendanceRecordsSerializer, \
    FillCardApplySerializer, ApplyForExtraWorkSerializer, Equip190EWeightSerializer, \
    Equip190ESerializer, EquipStatusBatchSerializer, AttendanceClockDetailSerializer, \
    EmployeeAttendanceRecordsLogSerializer, WeightClassPlanSerializer, WeightClassPlanUpdateSerializer, \
    EquipDownDetailsImportSerializer
from rest_framework.generics import ListAPIView, UpdateAPIView, \
    get_object_or_404
from datetime import timedelta

from production.utils import get_standard_time, get_classes_plan, get_user_group, get_user_weight_flag, get_work_time, actual_clock_data, get_user_level
from quality.models import MaterialTestOrder, MaterialDealResult, MaterialTestResult, MaterialDataPointIndicator
from quality.utils import get_cur_sheet, get_sheet_data
from system.models import Section
from recipe.models import Material, MaterialAttribute, ProductBatching
from system.models import User
from terminal.models import Plan, JZPlan, JZReportBasic, ReportBasic
from equipment.utils import DinDinAPI
from terminal.serializers import WeightPackagePlanSerializer
from terminal.utils import get_current_factory_date

logger = logging.getLogger('error_log')


@method_decorator([api_recorder], name="dispatch")
class TrainsFeedbacksViewSet(mixins.CreateModelMixin,
                             mixins.RetrieveModelMixin,
                             GenericViewSet):
    """
    list:
        车次/批次产出反馈列表
    retrieve:
        车次/批次产出反馈详情
    create:
        创建车次/批次产出反馈
    """
    queryset = TrainsFeedbacks.objects.filter(delete_flag=False)
    permission_classes = (IsAuthenticatedOrReadOnly,)
    serializer_class = TrainsFeedbacksSerializer
    filter_backends = [DjangoFilterBackend, OrderingFilter]
    ordering_fields = ('id',)
    filter_class = TrainsFeedbacksFilter

    def list(self, request, *args, **kwargs):
        actual_trains = request.query_params.get("actual_trains", '')
        if "," in actual_trains:
            train_list = actual_trains.split(",")
            try:
                queryset = self.filter_queryset(self.get_queryset().filter(actual_trains__gte=train_list[0],
                                                                           actual_trains__lte=train_list[-1],
                                                                           ))
            except:
                return Response({"actual_trains": "请输入: <开始车次>,<结束车次>。这类格式"})
        else:
            queryset = self.filter_queryset(self.get_queryset())
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)

        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)


@method_decorator([api_recorder], name="dispatch")
class PalletFeedbacksViewSet(mixins.CreateModelMixin,
                             mixins.ListModelMixin,
                             GenericViewSet):
    """
        list:
            托盘产出反馈列表
        retrieve:
            托盘产出反馈详情
        create:
            托盘产出反馈反馈
    """
    queryset = PalletFeedbacks.objects.filter(delete_flag=False).order_by("-plan_classes_uid")
    permission_classes = (IsAuthenticatedOrReadOnly,)
    serializer_class = PalletFeedbacksSerializer
    filter_backends = [DjangoFilterBackend, OrderingFilter]
    ordering_fields = ('id', 'product_time')
    filter_class = PalletFeedbacksFilter

    @action(methods=['post'], detail=False, permission_classes=[], url_path='bind-rfid',
            url_name='bind-rfid')
    def bind_rfid(self, request):
        # {"plan_classes_uid": "2021102520011276Z11", "bath_no": 367, "equip_no": "Z11", "product_no": "C-FM-F978-04",
        #  "plan_weight": "669.66", "actual_weight": "670", "begin_time": "2021-10-26 00:18:20.436",
        #  "end_time": "2021-10-26 00:18:20.436", "operation_user": "人工", "begin_trains": 81, "end_trains": 82,
        #  "pallet_no": "20102905", "classes": "夜班", "lot_no": "AAJ1Z112021102630367",
        #  "product_time": "2021-10-26 00:18:20.436", "factory_date": "2021-10-26"}
        if request.data:
            data = dict(request.data)
        else:
            data = dict(request.query_params)
        validated_data = {}
        for key, value in data.items():
            if isinstance(value, str):
                value = value.strip()
            validated_data[key] = value
            if key in ('begin_trains', 'end_trains', 'actual_weight'):
                if not value:
                    return Response('补充成功')
                if float(value) < 1:
                    return Response('补充成功')
            if key == 'pallet_no':
                if not value:
                    return Response('补充成功')
                if all(i == '0' for i in str(value)):
                    return Response('补充成功')
        lot_no = validated_data.pop("lot_no", None)
        validated_data.pop("factory_date", None)
        plan_classes_uid = validated_data.get('plan_classes_uid')
        if plan_classes_uid:
            plan = ProductClassesPlan.objects.filter(plan_classes_uid=plan_classes_uid).first()
            if plan:
                validated_data['factory_date'] = plan.work_schedule_plan.plan_schedule.day_time
        if not lot_no:
            raise ValidationError("请传入lot_no")
        if MaterialTestOrder.objects.filter(lot_no=lot_no).exists():
            # raise ValidationError("该批次数据已绑定快检数据，不可修改！")
            # 根据条码修正重量
            fix_weight = validated_data.get('actual_weight')
            PalletFeedbacks.objects.filter(lot_no=lot_no).update(actual_weight=fix_weight)
            return Response("补充成功")
        instance, flag = PalletFeedbacks.objects.update_or_create(defaults=validated_data, **{"lot_no": lot_no})
        if flag:
            message = "补充成功"
        else:
            message = "重新绑定成功"
        return Response(message)

    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        auto = request.query_params.get("auto")
        manual = request.query_params.get("manual")
        target_equip = request.query_params.get("equip_no", '')
        product_no = request.query_params.get("product_no", '')
        day_time = request.query_params.get("day_time", '')
        if auto or manual:
            try:
                auto, manual = int(auto), int(manual)
                data = []
                if auto and manual:
                    queryset = list(queryset.values()) + list(
                        ManualInputTrains.objects.filter(equip_no=target_equip, product_no=product_no, factory_date=day_time).order_by('-id').values())
                elif not auto and manual:
                    queryset = ManualInputTrains.objects.filter(equip_no=target_equip, product_no=product_no, factory_date=day_time).order_by('-id').values()
                elif auto and not manual:
                    queryset = queryset.values()
                else:
                    queryset = []
                for item in queryset:
                    weight = item.get('weight')
                    if weight is None:  # 自动
                        data.append(item)
                    else:  # 手动
                        created_time = item['created_time'].strftime('%Y-%m-%d %H:%M:%S')
                        data.append({
                            "plan_classes_uid": '', "bath_no": '', "equip_no": item['equip_no'], "product_no": item['product_no'], "id": item['id'],
                            "plan_weight": weight, "actual_weight": weight, "begin_time": item['created_time'], "end_time": created_time,
                            "operation_user": item['created_username'], "begin_trains": 1, "end_trains": item['actual_trains'], "pallet_no": '',
                            "classes": item['classes'], "lot_no": '', "product_time": created_time, "factory_date": item['factory_date']
                        })
                page = self.paginate_queryset(data)
                if page is not None:
                    return self.get_paginated_response(page)
                return Response(data)
            except Exception as e:
                logger.error(f"密炼实绩异常: {e.args[0]}")
                raise ValidationError('解析异常')

        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)

        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)


@method_decorator([api_recorder], name="dispatch")
class EquipStatusViewSet(mixins.CreateModelMixin,
                         mixins.ListModelMixin,
                         GenericViewSet):
    """
    list:
        机台状况反馈列表
    retrieve:
        机台状况反馈详情
    create:
        创建机台状况反馈
    """
    queryset = EquipStatus.objects.filter(delete_flag=False)
    pagination_class = SinglePageNumberPagination
    permission_classes = (IsAuthenticatedOrReadOnly,)
    serializer_class = EquipStatusSerializer
    filter_backends = [DjangoFilterBackend, OrderingFilter]
    ordering_fields = ('id',)
    filter_class = EquipStatusFilter

    def list(self, request, *args, **kwargs):
        actual_trains = request.query_params.get("actual_trains", '')
        if "," in actual_trains:
            train_list = actual_trains.split(",")
            try:
                queryset = self.filter_queryset(self.get_queryset().filter(current_trains__gte=train_list[0],
                                                                           current_trains__lte=train_list[-1]))
            except:
                return Response({"actual_trains": "请输入: <开始车次>,<结束车次>。这类格式"})
        else:
            queryset = self.filter_queryset(self.get_queryset())
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)

        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)


@method_decorator([api_recorder], name="dispatch")
class PlanStatusViewSet(mixins.CreateModelMixin,
                        mixins.RetrieveModelMixin,
                        mixins.ListModelMixin,
                        GenericViewSet):
    """
    list:
        计划状态变更列表
    retrieve:
        计划状态变更详情
    create:
        创建计划状态变更
    """
    queryset = PlanStatus.objects.filter(delete_flag=False)
    permission_classes = (IsAuthenticatedOrReadOnly,)
    serializer_class = PlanStatusSerializer
    filter_backends = [DjangoFilterBackend, OrderingFilter]
    ordering_fields = ('id',)
    filter_class = PlanStatusFilter


@method_decorator([api_recorder], name="dispatch")
class ExpendMaterialViewSet(mixins.CreateModelMixin,
                            mixins.RetrieveModelMixin,
                            mixins.ListModelMixin,
                            GenericViewSet):
    """
    list:
        原材料消耗列表
    retrieve:
        原材料消耗详情
    create:
        创建原材料消耗
    """
    queryset = ExpendMaterial.objects.filter(delete_flag=False)
    permission_classes = (IsAuthenticatedOrReadOnly,)
    serializer_class = ExpendMaterialSerializer
    filter_backends = [OrderingFilter]
    ordering_fields = ('id',)
    filter_class = ExpendMaterialFilter


@method_decorator([api_recorder], name="dispatch")
class OperationLogViewSet(mixins.CreateModelMixin,
                          mixins.RetrieveModelMixin,
                          mixins.ListModelMixin,
                          GenericViewSet):
    """
    list:
        操作日志列表
    retrieve:
        操作日志详情
    create:
        创建操作日志
    """
    queryset = OperationLog.objects.filter(delete_flag=False)
    permission_classes = (IsAuthenticatedOrReadOnly,)
    serializer_class = OperationLogSerializer


@method_decorator([api_recorder], name="dispatch")
class QualityControlViewSet(mixins.CreateModelMixin,
                            mixins.RetrieveModelMixin,
                            mixins.ListModelMixin,
                            GenericViewSet):
    """
    list:
        质检结果列表
    retrieve:
        质检结果详情
    create:
        创建质检结果
    """
    queryset = QualityControl.objects.filter(delete_flag=False)
    permission_classes = (IsAuthenticatedOrReadOnly,)
    serializer_class = QualityControlSerializer
    filter_backends = [DjangoFilterBackend, OrderingFilter]
    ordering_fields = ('id',)
    filter_class = QualityControlFilter


@method_decorator([api_recorder], name="dispatch")
class PlanRealityViewSet(mixins.ListModelMixin,
                         GenericViewSet):

    def list(self, request, *args, **kwargs):
        params = request.query_params
        search_time_str = params.get("search_time")
        target_equip_no = params.get('equip_no')
        time_now = datetime.datetime.now()
        if search_time_str:
            if not re.search(r"[0-9]{4}\-[0-9]{1,2}\-[0-9]{1,2}", search_time_str):
                raise ValidationError("查询时间格式异常")
        else:
            search_time_str = str(datetime.date.today())
        if target_equip_no:
            pcp_set = ProductClassesPlan.objects.filter(work_schedule_plan__plan_schedule__day_time=search_time_str,
                                                        work_schedule_plan__plan_schedule__work_schedule__work_procedure__global_name="密炼",
                                                        equip__equip_no=target_equip_no,
                                                        delete_flag=False).select_related(
                'equip__equip_no',
                'product_batching__stage_product_batch_no',
                'product_day_plan_id', 'time', 'product_batching__stage__global_name')
        else:
            pcp_set = ProductClassesPlan.objects.filter(work_schedule_plan__plan_schedule__day_time=search_time_str,
                                                        work_schedule_plan__plan_schedule__work_schedule__work_procedure__global_name="密炼",
                                                        delete_flag=False).select_related(
                'equip__equip_no',
                'product_batching__stage_product_batch_no',
                'product_day_plan_id', 'time', 'product_batching__stage__global_name')
        # 班次计划号列表
        uid_list = pcp_set.values_list("plan_classes_uid", flat=True)
        # 日计划号对比
        day_plan_list_temp = pcp_set.values_list("product_batching__stage_product_batch_no", "equip__equip_no")
        day_plan_list = list(set([x[0] + x[1] for x in day_plan_list_temp]))
        tf_set = TrainsFeedbacks.objects.values('plan_classes_uid').filter(plan_classes_uid__in=uid_list).annotate(
            actual_trains=Max('actual_trains'), actual_weight=Sum('actual_weight'), begin_time=Min('begin_time'),
            actual_time=Max('product_time'))
        tf_dict = {x.get("plan_classes_uid"): [x.get("actual_trains"), x.get("actual_weight"),
                                               x.get("begin_time", time_now), x.get("actual_time", time_now),
                                               (x.get("actual_time", time_now) - x.get("begin_time",
                                                                                       time_now)).total_seconds()] for x
                   in tf_set}
        day_plan_dict = {x: {"plan_weight": 0, "plan_trains": 0, "actual_trains": 0, "actual_weight": 0, "plan_time": 0,
                             "start_rate": None, "all_time": 0}
                         for x in day_plan_list}
        pcp_data = pcp_set.values("plan_classes_uid", "weight", "plan_trains", 'equip__equip_no',
                                  'product_batching__stage_product_batch_no',
                                  'product_day_plan_id', 'time', 'product_batching__stage__global_name')
        for pcp in pcp_data:
            day_plan_id = pcp.get("product_batching__stage_product_batch_no") + pcp.get("equip__equip_no")
            plan_classes_uid = pcp.get('plan_classes_uid')
            day_plan_dict[day_plan_id].update(
                equip_no=pcp.get('equip__equip_no'),
                product_no=pcp.get('product_batching__stage_product_batch_no'),
                stage=pcp.get('product_batching__stage__global_name'))
            day_plan_dict[day_plan_id]["plan_weight"] = pcp.get('weight', 0)
            day_plan_dict[day_plan_id]["plan_trains"] += pcp.get('plan_trains', 0)
            day_plan_dict[day_plan_id]["plan_time"] += pcp.get('time', 0)
            if not tf_dict.get(plan_classes_uid):
                day_plan_dict[day_plan_id]["actual_trains"] += 0
                day_plan_dict[day_plan_id]["actual_weight"] += 0
                continue
            day_plan_dict[day_plan_id]["actual_trains"] += tf_dict[plan_classes_uid][0]
            day_plan_dict[day_plan_id]["actual_weight"] += round(tf_dict[plan_classes_uid][1] / 100, 2)
            day_plan_dict[day_plan_id]["begin_time"] = tf_dict[plan_classes_uid][2].strftime('%Y-%m-%d %H:%M:%S') if \
                tf_dict[plan_classes_uid][2] else ""
            day_plan_dict[day_plan_id]["actual_time"] = tf_dict[plan_classes_uid][3].strftime('%Y-%m-%d %H:%M:%S')
            day_plan_dict[day_plan_id]["plan_time"] += pcp.get("time", 0)
            day_plan_dict[day_plan_id]["all_time"] += tf_dict[plan_classes_uid][4]
        temp_data = {}
        for equip_no in EQUIP_LIST:
            temp_data[equip_no] = []
            for temp in day_plan_dict.values():
                if temp.get("equip_no") == equip_no:
                    temp_data[equip_no].append(temp)
        datas = []
        for equip_data in temp_data.values():
            equip_data.sort(key=lambda x: (x.get("equip_no", ""), x.get("begin_time", "")))
            new_equip_data = []
            for _ in equip_data:
                _.update(sn=equip_data.index(_) + 1)
                _.update(ach_rate=round(_.get('actual_trains') / _.get('plan_trains') * 100, 2))
                new_equip_data.append(_)
            datas += new_equip_data
        return Response({"data": datas})


@method_decorator([api_recorder], name="dispatch")
class ProductActualViewSet(mixins.ListModelMixin,
                           GenericViewSet):
    """密炼实绩"""

    def list(self, request, *args, **kwargs):
        params = request.query_params
        day_time = params.get("search_time", datetime.datetime.now().date())
        target_equip_no = params.get('equip_no')
        auto = int(params.get('auto', 1))
        manual = int(params.get('manual', 1))
        plan_queryset = ProductClassesPlan.objects.filter(delete_flag=False,
                                                          work_schedule_plan__plan_schedule__day_time=day_time
                                                          ).order_by('equip__equip_no')
        if target_equip_no:
            plan_queryset = plan_queryset.filter(equip__equip_no=target_equip_no)
        plan_data = plan_queryset.values(
            'equip__equip_no',
            'product_batching__stage_product_batch_no',
            'work_schedule_plan__classes__global_name'
        ).annotate(plan_trains=Sum('plan_trains')).values(equip_no=F('equip__equip_no'),
                                                           product_no=F('product_batching__stage_product_batch_no'),
                                                           plan_trains=F('plan_trains'),
                                                           classes=F('work_schedule_plan__classes__global_name'))
        plan_data_dict = {item['equip_no']+item['product_no']+item['classes']: item for item in plan_data}
        manual_plan_data = ManualInputTrains.objects.filter(factory_date=day_time).values('equip_no', 'product_no', 'classes')\
            .annotate(plan_trains=Sum('actual_trains'), actual_weight=F('weight'), plan_weight=F('weight'))\
            .values('equip_no', 'product_no', 'plan_trains', 'classes', 'actual_trains', 'actual_weight', 'plan_weight').order_by('equip_no')
        if target_equip_no:
            manual_plan_data = manual_plan_data.filter(equip_no=target_equip_no)
        manual_plan_data_dict = {item['equip_no']+item['product_no']+item['classes']: item for item in manual_plan_data}
        real_plan_data_dict = {}
        if auto and manual:
            keys = set(list(plan_data_dict.keys()) + list(manual_plan_data_dict.keys()))
            sorted_keys = sorted(keys)
            for k in sorted_keys:
                if k in plan_data_dict and k in manual_plan_data_dict:
                    real_plan_data_dict[k] = plan_data_dict[k]
                    real_plan_data_dict[k]['plan_trains'] += manual_plan_data_dict[k]['plan_trains']
                elif k in plan_data_dict and k not in manual_plan_data_dict:
                    real_plan_data_dict[k] = plan_data_dict[k]
                elif k not in plan_data_dict and k in manual_plan_data_dict:
                    real_plan_data_dict[k] = manual_plan_data_dict[k]
                else:
                    continue
        elif auto and not manual:
            real_plan_data_dict = plan_data_dict
        elif not auto and manual:
            real_plan_data_dict = manual_plan_data_dict
        else:
            real_plan_data_dict = real_plan_data_dict

        plan_classes_uid_list = plan_queryset.values_list('plan_classes_uid', flat=True)
        tf_set = TrainsFeedbacks.objects.filter(
            plan_classes_uid__in=plan_classes_uid_list
        ).values('equip_no',
                 'product_no',
                 'classes'
                 ).annotate(
            actual_trains=Count('actual_trains'),
            actual_weight=Max('actual_weight'),
            plan_weight=Max('plan_weight')
        ).values('equip_no', 'product_no', 'classes', 'actual_trains', 'actual_weight', 'plan_weight')
        actual_data_dict = {item['equip_no']+item['product_no']+item['classes']: item for item in tf_set}
        real_data_dict = {}
        if auto and manual:
            keys = set(list(actual_data_dict.keys()) + list(manual_plan_data_dict.keys()))
            sorted_keys = sorted(keys)
            for k in sorted_keys:
                if k in actual_data_dict and k in manual_plan_data_dict:
                    real_data_dict[k] = actual_data_dict[k]
                    real_data_dict[k]['actual_trains'] += manual_plan_data_dict[k]['actual_trains']
                elif k in actual_data_dict and k not in manual_plan_data_dict:
                    real_data_dict[k] = actual_data_dict[k]
                elif k not in actual_data_dict and k in manual_plan_data_dict:
                    real_data_dict[k] = manual_plan_data_dict[k]
                else:
                    continue
        elif auto and not manual:
            real_data_dict = actual_data_dict
        elif not auto and manual:
            real_data_dict = manual_plan_data_dict
        else:
            real_data_dict = real_data_dict
        ret = {}
        aps_result = SchedulingResult.objects.filter(factory_date=day_time).order_by('id').last()
        if aps_result:
            aps_data = SchedulingResult.objects.filter(
                factory_date=day_time,
                schedule_no=aps_result.schedule_no
            ).values('equip_no', 'recipe_name').annotate(plan_trains=Sum('plan_trains'))
            aps_data_dict = {item['equip_no']+item['recipe_name']: item['plan_trains'] for item in aps_data}
        else:
            aps_data_dict = {}

        for key, value in real_plan_data_dict.items():
            if key in real_data_dict:
                value['actual_trains'] = real_data_dict[key]['actual_trains']
                value['actual_weight'] = real_data_dict[key]['actual_weight']
                value['plan_weight'] = real_data_dict[key]['plan_weight']
            else:
                value['actual_trains'] = 0
                value['actual_weight'] = 0
                value['plan_weight'] = 0
            equip_product_key = value['equip_no']+value['product_no']
            if equip_product_key not in ret:
                classes_data = {'早班': {'plan_trains': 0, 'actual_trains': 0, 'classes': '早班'},
                                '中班': {'plan_trains': 0, 'actual_trains': 0, 'classes': '中班'},
                                '夜班': {'plan_trains': 0, 'actual_trains': 0, 'classes': '夜班'}}
                classes_data[value['classes']] = {'plan_trains': value['plan_trains'],
                                                  'actual_trains': value['actual_trains'],
                                                  'classes': value['classes']}
                ret[equip_product_key] = {
                               'equip_no': value['equip_no'],
                               'product_no': value['product_no'],
                               'plan_trains': aps_data_dict.get(equip_product_key),
                               'actual_trains': value['actual_trains'],
                               'plan_weight': value['plan_weight'],
                               'actual_weight': value['actual_weight'],
                               'classes_data': classes_data,
                           }
            else:
                # ret[equip_product_key]['plan_trains'] += value['plan_trains']
                ret[equip_product_key]['actual_trains'] += value['actual_trains']
                ret[equip_product_key]['classes_data'][value['classes']] = {
                    'plan_trains': value['plan_trains'],
                    'actual_trains': value['actual_trains'],
                    'classes': value['classes']}

        for item in ret.values():
            item['classes_data'] = item['classes_data'].values()
        ret = {"data": ret.values()}
        return Response(ret)


@method_decorator([api_recorder], name="dispatch")
class ProductionRecordViewSet(mixins.ListModelMixin,
                              GenericViewSet):
    queryset = PalletFeedbacks.objects.filter(delete_flag=False).order_by("-factory_date", 'equip_no', 'classes', '-plan_classes_uid',
                                                                          'product_no', '-begin_trains')
    permission_classes = (IsAuthenticatedOrReadOnly,)
    serializer_class = ProductionRecordSerializer
    filter_backends = [DjangoFilterBackend, ]
    filter_class = PalletFeedbacksFilter

    def list(self, request, *args, **kwargs):
        locked_status = self.request.query_params.get('locked_status')
        is_instock = self.request.query_params.get('is_instock')
        queryset = self.filter_queryset(self.get_queryset())
        if locked_status:
            # if locked_status == '0':  # 空白
            #     locked_lot_nos = list(
            #         ProductInventoryLocked.objects.filter(is_locked=True).values_list('lot_no', flat=True))
            #     queryset = queryset.exclude(lot_no__in=locked_lot_nos)
            if locked_status == '1':  # 工艺锁定
                locked_lot_nos = list(
                    ProductInventoryLocked.objects.filter(is_locked=True, locked_status__in=(1, 3)).values_list('lot_no', flat=True))
                queryset = queryset.filter(lot_no__in=locked_lot_nos)
            elif locked_status == '2':  # 快检锁定
                locked_lot_nos = list(
                    ProductInventoryLocked.objects.filter(is_locked=True, locked_status__in=(2, 3)).values_list('lot_no',
                                                                                                       flat=True))
                queryset = queryset.filter(lot_no__in=locked_lot_nos)
        if is_instock:
            stock_lot_nos = list(BzFinalMixingRubberInventory.objects.using('bz').values_list('lot_no', flat=True)) + \
                            list(BzFinalMixingRubberInventoryLB.objects.using('lb').filter(
                                store_name='炼胶库').values_list('lot_no', flat=True))
            if is_instock == 'Y':
                queryset = queryset.filter(lot_no__in=stock_lot_nos)
            else:
                stock_ids = queryset.filter(lot_no__in=stock_lot_nos).values_list('id', flat=True)
                queryset = queryset.exclude(id__in=list(stock_ids))
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            s_data = serializer.data
            lot_nos = [i['lot_no'] for i in s_data]
            plan_classes_uids = [i['plan_classes_uid'] for i in s_data]
            locked_dict = dict(
                ProductInventoryLocked.objects.filter(lot_no__in=lot_nos).values_list('lot_no', 'locked_status'))
            stock_lot_nos = list(
                BzFinalMixingRubberInventory.objects.using('bz').filter(lot_no__in=lot_nos).values_list('lot_no',
                                                                                                        flat=True)) + \
                            list(BzFinalMixingRubberInventoryLB.objects.using('lb').filter(
                                store_name='炼胶库', lot_no__in=lot_nos).values_list('lot_no', flat=True))
            product_validity_dict = dict(MaterialAttribute.objects.filter(
                period_of_validity__isnull=False
            ).values_list('material__material_no', 'period_of_validity'))
            group_dict = dict(
                ProductClassesPlan.objects.filter(
                    plan_classes_uid__in=plan_classes_uids
                ).values_list('plan_classes_uid', 'work_schedule_plan__group__global_name'))
            for item in s_data:
                lot_no = item['lot_no']
                item['locked_status'] = locked_dict.get(lot_no, 0)
                item['is_instock'] = lot_no in stock_lot_nos
                item['class_group'] = group_dict.get(item['plan_classes_uid'])
                if product_validity_dict.get(item['product_no']):
                    item['validtime'] = (datetime.datetime.strptime(item['end_time'], '%Y-%m-%d %H:%M:%S') + datetime.timedelta(days=product_validity_dict.get(item['product_no']))).strftime('%Y-%m-%d %H:%M:%S')
            return self.get_paginated_response(s_data)
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)

    def get_queryset(self):
        queryset = self.queryset
        st = self.request.query_params.get('st')
        et = self.request.query_params.get('et')
        if st:
            queryset = queryset.filter(factory_date__gte=st[:10])
        if et:
            queryset = queryset.filter(factory_date__lte=et[:10])
        return queryset


@method_decorator([api_recorder], name="dispatch")
class MaterialInventory(GenericViewSet,
                        mixins.ListModelMixin, ):

    def get_queryset(self):
        return

    def list(self, request, *args, **kwargs):
        ret = requests.get("http://49.235.45.128:8169/storageSpace/GetInventoryCount")
        ret_json = json.loads(ret.text)
        results = []
        for i in ret_json.get("datas"):
            results = [{
                "sn": 1,
                "material_no": i.get('material_no'),
                "material_name": i.get('materialName'),
                "material_type": "1MB",
                "qty": 11,
                "unit": "吨",
                "unit_weight": 1,
                "total_weight": 1,
                "need_weight": 1,
                "standard_flag": True,
                "site": "立库",
            }]
        return Response({'results': results})


@method_decorator([api_recorder], name="dispatch")
class ProductInventory(GenericViewSet,
                       mixins.ListModelMixin, ):

    def get_queryset(self):
        return

    def list(self, request, *args, **kwargs):
        results = [{"sn": 1,
                    "id": 1,
                    "material_id": 1,
                    "material_no": "C9001",
                    "material_name": "C9001",
                    "material_type_id": 1,
                    "material_type": "天然胶",
                    "qty": 11,
                    "unit": "吨",
                    "unit_weight": 1,
                    "total_weight": 1,
                    "need_weight": 1,
                    "site": "安吉仓库",
                    "standard_flag": True,
                    }, {"sn": 2,
                        "id": 2,
                        "material_id": 2,
                        "material_no": "C9002",
                        "material_name": "C9002",
                        "material_type_id": 1,
                        "material_type": "天然胶",
                        "qty": 11,
                        "unit": "吨",
                        "unit_weight": 1,
                        "total_weight": 1,
                        "need_weight": 1,
                        "site": "安吉仓库",
                        "standard_flag": False,
                        }]
        return Response({'results': results})


@method_decorator([api_recorder], name="dispatch")
class TrainsFeedbacksBatch(APIView):
    """批量同步车次生产数据接口"""

    @atomic
    def post(self, request):
        serializer = TrainsFeedbacksBatchSerializer(data=request.data, many=True, context={'request': request})
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response("sync success", status=201)


@method_decorator([api_recorder], name="dispatch")
class PalletFeedbacksBatch(APIView):
    """批量同步托次生产数据接口"""

    @atomic
    def post(self, request):
        for item in request.data:
            p = item['pallet_no']
            b = re.sub(u'\u0000', "", p)
            item['pallet_no'] = b
        serializer = PalletFeedbacksSerializer(data=request.data, many=True, context={'request': request})
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response("sync success", status=201)


@method_decorator([api_recorder], name="dispatch")
class EquipStatusBatch(APIView):
    """批量同步设备生产数据接口"""

    @atomic
    def post(self, request):
        serializer = EquipStatusBatchSerializer(data=request.data, many=True, context={'request': request})
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response("sync success", status=201)


@method_decorator([api_recorder], name="dispatch")
class PlanStatusBatch(APIView):
    """批量同步计划状态数据接口"""

    @atomic
    def post(self, request):
        serializer = PlanStatusSerializer(data=request.data, many=True, context={'request': request})
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response("sync success", status=201)


@method_decorator([api_recorder], name="dispatch")
class ExpendMaterialBatch(APIView):
    """批量同步原材料消耗数据接口"""

    @atomic
    def post(self, request):
        serializer = ExpendMaterialSerializer(data=request.data, many=True, context={'request': request})
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response("sync success", status=201)


@method_decorator([api_recorder], name="dispatch")
class ProcessFeedbackBatch(APIView):
    """批量同步原材料消耗数据接口"""

    @atomic
    def post(self, request):
        serializer = ProcessFeedbackSerializer(data=request.data, many=True, context={'request': request})
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response("sync success", status=201)


@method_decorator([api_recorder], name="dispatch")
class AlarmLogBatch(APIView):
    """批量同步原材料消耗数据接口"""

    @atomic
    def post(self, request):
        serializer = AlarmLogSerializer(data=request.data, many=True, context={'request': request})
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response("sync success", status=201)


@method_decorator([api_recorder], name="dispatch")
class PalletTrainFeedback(APIView):
    """获取托盘开始车次-结束车次的数据，过滤字段：equip_no=设备编码&factory_date=工厂日期
        &classes=班次&product_no=胶料编码&stage=胶料段次&is_tested=是否已检测（Y已检测 N未检测）"""

    def get(self, request):
        equip_no = self.request.query_params.get('equip_no')
        factory_date = self.request.query_params.get('factory_date')
        classes = self.request.query_params.get('classes')
        product_no = self.request.query_params.get('product_no')
        stage = self.request.query_params.get('stage')
        if not all([equip_no, factory_date, classes, product_no]):
            raise ValidationError('缺少参数')
        pallet_feed_backs = PalletFeedbacks.objects.filter(
            equip_no=equip_no,
            factory_date=factory_date,
            classes=classes,
            product_no=product_no
        )
        if stage:
            pallet_feed_backs = pallet_feed_backs.filter(product_no__icontains=stage)
        ret = []
        trains = []
        for pallet_feed_back in pallet_feed_backs:
            begin_trains = pallet_feed_back.begin_trains
            end_trains = pallet_feed_back.end_trains
            for i in range(begin_trains, end_trains + 1):
                if i not in trains:
                    trains.append(i)
                else:
                    continue
                # test_order = MaterialTestOrder.objects.filter(product_no=product_no,
                #                                               production_equip_no=equip_no,
                #                                               production_factory_date=factory_date,
                #                                               actual_trains=i,
                #                                               production_class=classes).first()
                data = {
                    'product_no': pallet_feed_back.product_no,
                    'lot_no': pallet_feed_back.lot_no,
                    'classes': pallet_feed_back.classes,
                    'equip_no': pallet_feed_back.equip_no,
                    'actual_trains': i,
                    'plan_classes_uid': pallet_feed_back.plan_classes_uid,
                    # 'factory_date': pallet_feed_back.end_time,
                    'factory_date': pallet_feed_back.factory_date,
                    'is_tested': 'N'
                    # 'is_tested': 'Y' if test_order else 'N'
                }
                ret.append(data)
        ret.sort(key=lambda x: x.get('actual_trains'))
        return Response(ret)


class UpdateUnReachedCapacityCauseView(UpdateAPIView):
    queryset = UnReachedCapacityCause.objects.all()
    serializer_class = UnReachedCapacityCauseSerializer

    def get_object(self):
        data = dict(self.request.data)
        filter_kwargs = {
            'factory_date': data.get('factory_date'),
            'classes': data.get('classes'),
            'equip_no': data.get('equip_no')
        }
        obj = get_object_or_404(self.get_queryset(), **filter_kwargs)
        return obj


def get_trains_feed_backs_query_params(view):
    query_params = view.request.query_params
    hour_step = int(query_params.get('hour_step', 2))
    classes = query_params.getlist('classes[]')
    factory_date = query_params.get('factory_date', datetime.datetime.now().strftime('%Y-%m-%d'))
    return hour_step, classes, factory_date


def zno_(obj):
    return int(obj['equip_no'].lower()[1:])


# 产量计划实际分析
@method_decorator([api_recorder], name="dispatch")
class ProductionPlanRealityAnalysisView(ListAPIView):
    queryset = TrainsFeedbacks.objects.all()
    serializer_class = ProductionPlanRealityAnalysisSerializer

    def get_serializer_context(self):
        hour_step = get_trains_feed_backs_query_params(self)[0]
        context = super().get_serializer_context()
        context.update({'hour_step': hour_step})
        return context

    def list(self, request, *args, **kwargs):
        hour_step, classes, factory_date = get_trains_feed_backs_query_params(self)
        trains_feed_backs = TrainsFeedbacks.objects \
            .filter(factory_date=factory_date) \
            .values('factory_date',
                    'classes',
                    'equip_no',
                    'plan_classes_uid'
                    ) \
            .order_by('-factory_date') \
            .annotate(plan_train_sum=Max('plan_trains'),
                      finished_train_count=Max('actual_trains'))
        data = {
            'headers': list(range(hour_step, 13, hour_step))
        }
        for class_ in classes:
            feed_backs = trains_feed_backs.filter(classes=class_)
            feed_backs = sorted(feed_backs, key=zno_)
            train_count_group = {}
            for feed_back in feed_backs:
                train_counts = train_count_group.setdefault((feed_back['factory_date'],
                                                             feed_back['classes'],
                                                             feed_back['equip_no']), {
                                                                'plan_train_sum': 0,
                                                                'finished_train_count': 0
                                                            })
                train_counts['plan_train_sum'] += feed_back['plan_train_sum']
                train_counts['finished_train_count'] += feed_back['finished_train_count']
            group_feed_backs = []
            for key, value in train_count_group.items():
                feed_back = {
                    'factory_date': key[0],
                    'classes': key[1],
                    'equip_no': key[2],
                }
                feed_back.update(**value)
                group_feed_backs.append(feed_back)
            serializer = self.get_serializer(group_feed_backs, many=True)
            data[class_] = serializer.data
        return Response(data)


# 区间产量统计
@method_decorator([api_recorder], name="dispatch")
class IntervalOutputStatisticsView(APIView):

    def get(self, request, *args, **kwargs):
        hour_step, classes, factory_date = get_trains_feed_backs_query_params(self)

        if not TrainsFeedbacks.objects.filter(factory_date=factory_date).exists():
            return Response({})

        ws = WorkSchedulePlan.objects.filter(
            plan_schedule__work_schedule__work_procedure__global_name='密炼',
            plan_schedule__day_time=factory_date
        ).values('classes__global_name', 'start_time', 'end_time')
        if not ws:
            raise ValidationError('未找到当日排班信息！')
        ws_dict = {i['classes__global_name']: i for i in ws}
        day_start_time = datetime.datetime.strptime(factory_date + ' 08:00:00', "%Y-%m-%d %H:%M:%S")
        factory_end_time = day_start_time + datetime.timedelta(days=1)

        day_end_time = (datetime.datetime.now() + datetime.timedelta(hours=1)).strftime("%Y-%m-%d %H:%M:%S")
        day_end_time = day_end_time[:14] + '00:00'
        day_end_time = datetime.datetime.strptime(day_end_time, "%Y-%m-%d %H:%M:%S")
        if day_end_time > factory_end_time:
            day_end_time = factory_end_time

        time_spans = []
        end_time = day_start_time
        while end_time < day_end_time:
            time_spans.append(end_time)
            end_time = end_time + datetime.timedelta(hours=hour_step)
        time_spans.append(day_end_time)

        data = {
            'equips': list(Equip.objects.filter(
                category__equip_type__global_name='密炼设备'
            ).order_by('equip_no').values_list('equip_no', flat=True))
        }
        for class_ in classes:
            data[class_] = []
            for i in range(len(time_spans) - 1):
                time_span_data = {}
                interval_trains_feed_backs = None
                total_trains_feed_backs = None
                if class_ == '整日':
                    interval_trains_feed_backs = TrainsFeedbacks.objects \
                        .filter(factory_date=factory_date,
                                end_time__gte=time_spans[i],
                                end_time__lte=time_spans[i + 1]) \
                        .values('equip_no') \
                        .annotate(interval_finished_train_count=Count('id', distinct=True))

                    total_trains_feed_backs = TrainsFeedbacks.objects \
                        .filter(factory_date=factory_date,
                                end_time__gte=day_start_time,
                                end_time__lte=time_spans[i + 1]) \
                        .values('equip_no') \
                        .annotate(total_finished_train_count=Count('id', distinct=True))
                else:
                    ws_data = ws_dict.get(class_)
                    if not ws_data:
                        continue
                    c_st = ws_data['start_time']
                    c_et = ws_data['end_time']
                    if not c_st <= time_spans[i] < c_et:
                        continue
                    interval_trains_feed_backs = TrainsFeedbacks.objects \
                        .filter(classes=class_,
                                factory_date=factory_date,
                                end_time__gte=time_spans[i],
                                end_time__lte=time_spans[i + 1]) \
                        .values('equip_no') \
                        .annotate(interval_finished_train_count=Count('id', distinct=True))

                    total_trains_feed_backs = TrainsFeedbacks.objects \
                        .filter(classes=class_,
                                factory_date=factory_date,
                                end_time__gte=day_start_time,
                                end_time__lte=time_spans[i + 1]) \
                        .values('equip_no') \
                        .annotate(total_finished_train_count=Count('id', distinct=True))

                time_span_data['time_span'] = "{0}:00-{1}:00".format(time_spans[i].hour,
                                                                     time_spans[i + 1].hour
                                                                     if time_spans[i + 1].hour != time_spans[i].hour
                                                                     else time_spans[i + 1].hour + hour_step)

                for total_trains_feed_back in total_trains_feed_backs:
                    equip_no = total_trains_feed_back.get('equip_no')
                    same_equip_no_interval_trains_feed_back = None
                    for interval_trains_feed_back in interval_trains_feed_backs:
                        if interval_trains_feed_back.get('equip_no') == equip_no:
                            same_equip_no_interval_trains_feed_back = interval_trains_feed_back
                            break
                    time_span_data[equip_no] = (total_trains_feed_back
                                                .get('total_finished_train_count'),
                                                same_equip_no_interval_trains_feed_back
                                                .get('interval_finished_train_count')
                                                if same_equip_no_interval_trains_feed_back else 0)
                if time_span_data:
                    data[class_].append(time_span_data)
        return Response(data)


# 将群控的车次报表移植过来 （中间表就吗，没有用了 关于中间表的代码直接删除了）
@method_decorator([api_recorder], name="dispatch")
class TrainsFeedbacksAPIView(mixins.ListModelMixin,
                             GenericViewSet):
    """车次报表展示接口"""
    queryset = TrainsFeedbacks.objects.order_by('factory_date', 'equip_no', '-plan_classes_uid', 'product_no', '-actual_trains')
    permission_classes = (IsAuthenticatedOrReadOnly,)
    serializer_class = TrainsFeedbacksSerializer2
    filter_backends = [DjangoFilterBackend, OrderingFilter]
    filter_class = TrainsFeedbacksFilter

    @staticmethod
    def calculate_energy(df):
        try:
            if df['equip_no'] == 'Z01':
                df['evacuation_energy'] = df['evacuation_energy'] / 10
            elif df['equip_no'] == 'Z02':
                df['evacuation_energy'] = df['evacuation_energy'] / 0.6
            elif df['equip_no'] == 'Z04':
                df['evacuation_energy'] = df['evacuation_energy'] * 0.28 * df['plan_weight'] / 1000
            elif df['equip_no'] == 'Z12':
                df['evacuation_energy'] = df['evacuation_energy'] / 5.3
            elif df['equip_no'] == 'Z01':
                df['evacuation_energy'] = df['evacuation_energy'] / 31.7
        except Exception:
            pass
        return df

    def list(self, request, *args, **kwargs):
        params = request.query_params
        trains = params.get("trains")
        queryset = self.filter_queryset(self.get_queryset())
        st = params.get('begin_time')
        et = params.get('end_time')
        export = params.get('export')
        if st:
            queryset = queryset.filter(factory_date__gte=st[:10])
        if et:
            queryset = queryset.filter(factory_date__lte=et[:10])
        if trains:
            try:
                train_range = trains.split(",")
            except:
                raise ValidationError("trains参数错误,参考: trains=5,10")
            queryset = queryset.filter(actual_trains__range=train_range)
        if export:
            if not all([st, et]):
                raise ValidationError('请选择时间范围进行导出！')
            qs_df = read_frame(qs=queryset,
                               fieldnames=['equip_no', 'factory_date', 'product_no', 'classes', 'plan_classes_uid',
                                           'begin_time', 'end_time', 'plan_trains', 'actual_trains', 'control_mode',
                                           'operating_type', 'plan_weight', 'actual_weight', 'evacuation_time',
                                           'evacuation_temperature', 'evacuation_energy',  'operation_user',
                                           'interval_time', 'ai_power'])
            bio = BytesIO()
            writer = pd.ExcelWriter(bio, engine='xlsxwriter')  # 注意安装这个包 pip install xlsxwriter
            qs_df["mixer_time"] = round((qs_df["end_time"] - qs_df["begin_time"]).dt.seconds)
            # qs_df['factory_date'] = qs_df['factory_date'].apply(lambda x: x.strftime('%Y-%m-%d'))
            qs_df['begin_time'] = qs_df['begin_time'].apply(lambda x: x.strftime('%Y-%m-%d %H:%M:%S'))
            qs_df['end_time'] = qs_df['end_time'].apply(lambda x: x.strftime('%Y-%m-%d %H:%M:%S'))
            try:
                qs_df['actual_weight'] = qs_df['actual_weight'].astype(float)
                qs_df['plan_weight'] = qs_df['plan_weight'].astype(float)
                qs_df['ai_power'] = qs_df['ai_power'].astype(float)
            except:
                pass
            qs_df['actual_weight'] = qs_df['actual_weight'].apply(lambda x: x/100)
            qs_df = qs_df.apply(self.calculate_energy, axis=1)
            order = ['equip_no', 'factory_date', 'product_no', 'classes', 'plan_classes_uid', 'begin_time', 'end_time',
                     'plan_trains', 'actual_trains', 'control_mode', 'operating_type', 'plan_weight', 'actual_weight',
                     'evacuation_time', 'evacuation_temperature', 'evacuation_energy',  'operation_user',
                     'mixer_time', 'interval_time', 'ai_power']
            qs_df = qs_df[order]
            qs_df.rename(columns={'equip_no': '机台', 'factory_date': '工厂日期', 'product_no': '配方编号',
                                  'classes': '班次', 'plan_classes_uid': '计划编号', 'begin_time': '开始时间',
                                  'end_time': '结束时间', 'plan_trains': '设定车次', 'actual_trains': '实际车次',
                                  'control_mode': '本远控', 'operating_type': '手自动', 'plan_weight': '计划重量(kg)',
                                  'actual_weight': '实际重量(kg)', 'evacuation_time': '排胶时间(s)',
                                  'evacuation_temperature': '排胶温度', 'evacuation_energy': '排胶能量(kW.h)',
                                  'operation_user': '操作人', 'mixer_time': '密炼时间(s)', 'interval_time': '间隔时间(s)',
                                  'ai_power': 'AI值'},
                         inplace=True)
            qs_df.to_excel(writer, sheet_name='Sheet1', index=False)
            writer.save()
            bio.seek(0)
            response = FileResponse(bio)
            response['Content-Type'] = 'application/octet-stream'
            response['Content-Disposition'] = 'attachment;filename="mm.xlsx"'
            return response

        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            data_list = serializer.data
            data_list.append({'version': 'v2'})
            return self.get_paginated_response(data_list)

        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)


@method_decorator([api_recorder], name="dispatch")
class CurveInformationList(mixins.ListModelMixin, mixins.RetrieveModelMixin,
                           GenericViewSet):
    """工艺曲线信息"""
    queryset = EquipStatus.objects.filter()
    permission_classes = (IsAuthenticatedOrReadOnly,)
    pagination_class = SinglePageNumberPagination
    serializer_class = CurveInformationSerializer
    filter_backends = [DjangoFilterBackend, OrderingFilter]

    def get_queryset(self):
        feed_back_id = self.request.query_params.get('feed_back_id')
        try:
            tfb_obk = TrainsFeedbacks.objects.get(id=feed_back_id)
            irc_queryset = EquipStatus.objects.filter(equip_no=tfb_obk.equip_no,
                                                      plan_classes_uid=tfb_obk.plan_classes_uid,
                                                      product_time__gte=tfb_obk.begin_time,
                                                      product_time__lte=tfb_obk.end_time).order_by('product_time')
        except:
            raise ValidationError('车次产出反馈或车次报表工艺曲线数据表没有数据')

        return irc_queryset


@method_decorator([api_recorder], name="dispatch")
class MixerInformationList(mixins.ListModelMixin, mixins.RetrieveModelMixin,
                           GenericViewSet):
    """密炼信息"""
    permission_classes = (IsAuthenticatedOrReadOnly,)
    filter_backends = [DjangoFilterBackend, OrderingFilter]
    serializer_class = MixerInformationSerializer2

    def get_queryset(self):
        feed_back_id = self.request.query_params.get('feed_back_id')
        params = self.request.query_params
        equip_no = params.get("equip_no", None)
        try:
            tfb_obk = TrainsFeedbacks.objects.get(id=feed_back_id)
            irm_queryset = ProcessFeedback.objects.filter(plan_classes_uid=tfb_obk.plan_classes_uid,
                                                          equip_no=tfb_obk.equip_no,
                                                          product_no=tfb_obk.product_no,
                                                          current_trains=tfb_obk.actual_trains
                                                          )
        except:
            raise ValidationError('车次产出反馈或胶料配料标准步序详情没有数据')
        return irm_queryset


@method_decorator([api_recorder], name="dispatch")
class WeighInformationList(mixins.ListModelMixin, mixins.RetrieveModelMixin,
                           GenericViewSet):
    """称量信息"""
    permission_classes = (IsAuthenticatedOrReadOnly,)
    filter_backends = [DjangoFilterBackend, OrderingFilter]
    serializer_class = WeighInformationSerializer2

    def get_queryset(self):
        params = self.request.query_params
        equip_no = params.get("equip_no", None)

        feed_back_id = self.request.query_params.get('feed_back_id')
        try:
            tfb_obk = TrainsFeedbacks.objects.get(id=feed_back_id)
            irw_queryset = ExpendMaterial.objects.filter(equip_no=tfb_obk.equip_no,
                                                         plan_classes_uid=tfb_obk.plan_classes_uid,
                                                         product_no=tfb_obk.product_no,
                                                         trains=tfb_obk.actual_trains, delete_flag=False)
        except:
            raise ValidationError('车次产出反馈或车次报表材料重量没有数据')
        return irw_queryset


@method_decorator([api_recorder], name="dispatch")
class AlarmLogList(mixins.ListModelMixin, mixins.RetrieveModelMixin,
                   GenericViewSet):
    """报警信息"""
    queryset = AlarmLog.objects.filter()
    permission_classes = (IsAuthenticatedOrReadOnly,)
    serializer_class = AlarmLogSerializer
    filter_backends = [DjangoFilterBackend, OrderingFilter]

    def get_queryset(self):
        feed_back_id = self.request.query_params.get('feed_back_id')
        try:
            tfb_obk = TrainsFeedbacks.objects.get(id=feed_back_id)
            al_queryset = AlarmLog.objects.filter(equip_no=tfb_obk.equip_no,
                                                  product_time__gte=tfb_obk.begin_time,
                                                  product_time__lte=tfb_obk.end_time).order_by('product_time')
        except:
            raise ValidationError('报警日志没有数据')

        return al_queryset


@method_decorator([api_recorder], name="dispatch")
class MaterialOutputView(APIView):
    permission_classes = (IsAuthenticated,)

    def get(self, request, *args, **kwargs):
        param = request.query_params
        material_no = param.get('material_no')
        query_unit = param.get('query_unit')
        groups = ['product_no']
        filters = {'product_time__year': datetime.datetime.now().strftime('%Y')}
        unit = '吨/天'
        num = 365
        if material_no:
            filters.update(product_no=material_no)
        if not query_unit:
            query_unit = 'day'
            num = 365
            unit = '吨/天'
        if query_unit == 'year':
            num = 1
            unit = '吨/年'
        if query_unit == 'month':
            num = 12
            unit = '吨/月'
        if query_unit == 'classes':
            temp_set = TrainsFeedbacks.objects.values('classes').annotate().values_list('classes', flat=True).distinct()
            num = len(set([x if x in ['早班', '中班', '夜班'] else '夜班' for x in temp_set]))
            unit = '吨/班'
        if query_unit == 'hour':
            num = 365 * 24
            unit = '吨/小时'
        product_set = TrainsFeedbacks.objects.values(*groups).filter(**filters).annotate(
            all_weight=Sum('actual_weight'))
        data = [{"material_no": x.get("product_no"), 'output': round(x.get("all_weight", 0) / num, 2)} for x in
                product_set]
        ret = {
            "time": datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            "unit": unit,
            "data": data
        }
        return Response(ret)


@method_decorator([api_recorder], name="dispatch")
class EquipProductRealView(APIView):
    permission_classes = (IsAuthenticated,)

    def _instance_prepare(self, ret, equip_no):
        _ = EquipStatus.objects.filter(equip_no=equip_no).order_by('product_time').last()
        if not _:
            return
        plan_uid = _.plan_classes_uid
        try:
            plan = ProductClassesPlan.objects.filter(plan_classes_uid=plan_uid).last()
            plan_trains = plan.plan_trains
            product_no = plan.product_batching.stage_product_batch_no
        except:
            raise ValidationError(f"计划:{plan_uid}缺失")
        temp = {
            "equip_no": _.equip_no,
            "temperature": _.temperature,
            "rpm": _.rpm,
            "energy": _.energy,
            "power": _.power,
            "pressure": _.pressure,
            "status": _.status,
            "material_no": product_no,
            "current_trains": _.current_trains,
            "plan_trains": plan_trains,
            "actual_trains": _.current_trains,
            "product_time": _.product_time.strftime('%Y-%m-%d %H:%M:%S'),
        }
        ret['data'].append(temp)

    def get(self, request, *args, **kwargs):
        equip_no = request.query_params.get("equip_no")
        ret = {
            "data": []
        }
        if equip_no:
            self._instance_prepare(ret, equip_no)
        else:
            self._another(ret)
        return Response(ret)

    def _another(self, ret):
        temp_set = EquipStatus.objects.values("equip_no").annotate(id=Max('id')).values('id', 'equip_no')
        for temp in temp_set:
            _ = EquipStatus.objects.get(id=temp.get('id'))
            try:
                plan = ProductClassesPlan.objects.filter(plan_classes_uid=_.plan_classes_uid).last()
                plan_trains = plan.plan_trains
                product_no = plan.product_batching.stage_product_batch_no
            except Exception as e:
                raise ValidationError(f"错误信息:{e}")
            new = {
                "equip_no": _.equip_no,
                "temperature": _.temperature,
                "rpm": _.rpm,
                "energy": _.energy,
                "power": _.power,
                "pressure": _.pressure,
                "status": _.status,
                "material_no": product_no,
                "current_trains": _.current_trains,
                "plan_trains": plan_trains,
                "actual_trains": _.current_trains,
                "product_time": _.product_time.strftime('%Y-%m-%d %H:%M:%S'),
            }
            ret["data"].append(new)


@method_decorator([api_recorder], name="dispatch")
class MaterialTankStatusList(APIView):
    def get(self, request):
        """机台编号和罐编号"""
        mts_set = MaterialTankStatus.objects.values('equip_no', 'tank_no').distinct()
        return Response({"results": mts_set})


@method_decorator([api_recorder], name="dispatch")
class WeekdayProductStatisticsView(APIView):

    def get(self, request):
        """"""
        # 计算上一周是几号到几号
        params = request.query_params

        unit = params.get("unit")
        value = params.get("value")
        query_type = params.get("type")
        if params:
            if unit != "day" or query_type != "week" or value != "lastweek":
                raise ValidationError("暂不支持该粒度查询，敬请期待")
        temp = datetime.date.today().isoweekday()
        monday = datetime.date.today() - datetime.timedelta(days=(6 + temp))
        sunday = datetime.date.today() - datetime.timedelta(temp)
        # 相对简单的查询 存在脏数据会导致结果误差 另一种方案是先根据uid分类取最终值的id，再统计相对麻烦  后续补充
        temp_list = list(TrainsFeedbacks.objects.filter(factory_date__gte=monday, factory_date__lte=sunday).
                         values("equip_no", "factory_date"). \
                         annotate(all_weight=Sum("actual_weight")).order_by("factory_date").values("equip_no",
                                                                                                   "factory_date",
                                                                                                   "all_weight"))
        # __week=4查当周  3，2，1 上周 上上周， 上上上周
        # temp_list = list(
        #     TrainsFeedbacks.objects.filter(factory_date__week=3).values("equip_no", "factory_date"). \
        #         annotate(all_weight=Sum("actual_weight")).order_by("factory_date").values("equip_no", "factory_date",
        #                                                                                   "all_weight"))
        ret = {"Z%02d" % x: [] for x in range(1, 16)}
        day_week_map = {1: "周一", 2: "周二", 3: "周三", 4: "周四", 5: "周五", 6: "周六", 7: "周日"}
        for data in temp_list:
            try:
                index = data["factory_date"].isoweekday()
            except:
                continue
            ret[data["equip_no"]].append({day_week_map[index]: str(data["all_weight"] / 1000)})
        return Response(ret)


@method_decorator([api_recorder], name="dispatch")
class ProductionStatisticsView(APIView):

    def get(self, request):
        params = request.query_params
        unit = params.get("unit")
        value = params.get("value")
        query_type = params.get("type")
        ret = None
        if not value:
            raise ValidationError("value参数必传")
        try:
            if query_type == "year":
                my_key = value[0:4]
                value = datetime.datetime.strptime(value, "%Y-%m").year
                temp_set = TrainsFeedbacks.objects.filter(factory_date__year=value)
                if unit == "day":
                    middle_list = list(temp_set.values("factory_date").annotate(all_weight=Sum("actual_weight"))
                                       .order_by("factory_date").values("factory_date", "all_weight"))
                    ret = {my_key: [{_["factory_date"].strftime('%Y-%m-%d'): str(_["all_weight"] / 1000)} for _ in
                                    middle_list]}
                elif unit == "month":
                    middle_list = list(temp_set.annotate(month=TruncMonth('factory_date')).values("month")
                                       .annotate(all_weight=Sum("actual_weight") / 1000).order_by("factory_date")
                                       .values("month", "all_weight"))
                    ret = {value: [{_["month"].strftime('%Y-%m'): str(_["all_weight"])} for _ in middle_list]}
            elif query_type == "month":
                my_key = value[0:7]
                month = datetime.datetime.strptime(value, "%Y-%m").month
                year = datetime.datetime.strptime(value, "%Y-%m").year
                temp_set = TrainsFeedbacks.objects.filter(factory_date__month=month, factory_date__year=year)
                if unit == "day":
                    middle_list = list(temp_set.values("factory_date").annotate(all_weight=Sum("actual_weight"))
                                       .order_by("factory_date").values("factory_date", "all_weight"))
                    ret = {my_key: [{_["factory_date"].strftime('%Y-%m-%d'): str(_["all_weight"] / 1000)} for _ in
                                    middle_list]}
        except Exception as e:
            raise ValidationError(f"参数错误，请检查是否符合接口标准: {e}")
        if not ret:
            raise ValidationError("参数错误，请检查是否符合接口标准")
        else:
            return Response(ret)


@method_decorator([api_recorder], name="dispatch")
class DayCapacityView(APIView):

    def get(self, request):
        params = request.query_params
        value = params.get("value")
        query_type = params.get("type")
        if not value:
            raise ValidationError("value值必传，")
        try:
            factory_date = datetime.datetime.strptime(value, "%Y-%m-%d")
        except:
            raise ValidationError("请输入%Y-%m-%d格式的日期")
        temp_set = None
        if query_type == "week":
            monday = factory_date - timedelta(days=factory_date.weekday())
            sunday = factory_date + timedelta(days=6 - factory_date.weekday())
            # __week=4查当周  3，2，1 上周 上上周， 上上上周
            temp_set = TrainsFeedbacks.objects.filter(factory_date__range=(monday, sunday))
        elif query_type == "month":
            month = factory_date.month
            temp_set = TrainsFeedbacks.objects.filter(factory_date__month=month, factory_date__year=factory_date.year)
        elif query_type == "day":
            temp_set = TrainsFeedbacks.objects.filter(factory_date=factory_date)
        elif query_type == "year":
            year = factory_date.year
            temp_set = TrainsFeedbacks.objects.filter(factory_date__year=year)
        if temp_set is None:
            raise ValidationError("参数错误")
        temp_list = list(temp_set.values("equip_no", "product_no").annotate(output=Sum("actual_weight")))
        temp_list.sort(key=lambda x: (x.get("equip_no", "product_no")))
        for x in temp_list:
            x["output"] = str(x["output"] / 1000)
        ret = {"results": temp_list}
        return Response(ret)


@method_decorator([api_recorder], name="dispatch")
class PlanInfoReal(APIView):
    """密炼状态信息"""

    def get(self, request):
        type = self.request.query_params.get('type', None)
        value = self.request.query_params.get('value', None)
        if not all([type, value]):
            raise ValidationError('参数不全，type和value都得传')
        try:
            production_factory_date = datetime.datetime.strptime(value, "%Y-%m-%d")
        except:
            raise ValidationError('时间格式不正确')
        if type == 'day':
            filter_dict = {'factory_date': value, 'delete_flag': False}
        elif type == 'week':
            this_week_start = production_factory_date - timedelta(days=production_factory_date.weekday())  # 当天坐在的周的周一
            this_week_end = production_factory_date + timedelta(days=6 - production_factory_date.weekday())  # 当天所在周的周日
            filter_dict = {'factory_date__lte': this_week_end.date(), 'factory_date__gte': this_week_start.date(),
                           'delete_flag': False}
        elif type == 'month':
            filter_dict = {'factory_date__month': production_factory_date.month,
                           'factory_date__year': production_factory_date.year,
                           'delete_flag': False}
        elif type == 'year':
            filter_dict = {'factory_date__year': production_factory_date.year,
                           'delete_flag': False}
        else:
            raise ValidationError('type只能传day，week，month，year')
        tfb_set = TrainsFeedbacks.objects.filter(**filter_dict).values(
            'equip_no', 'plan_classes_uid', 'product_no').annotate(
            actual_trains=Max('actual_trains'), plan_trains=Max('plan_trains'), start_time=Min('begin_time'),
            end_time=Max('end_time'))
        for tfb_obj in tfb_set:
            tfb_obj['actual_trains'] = str(tfb_obj['actual_trains'])
            tfb_obj['plan_trains'] = str(tfb_obj['plan_trains'])
            tfb_obj['start_time'] = tfb_obj['start_time'].strftime('%Y-%m-%d %H:%M:%S')
            tfb_obj['end_time'] = tfb_obj['end_time'].strftime('%Y-%m-%d %H:%M:%S')
        return Response({'resluts': tfb_set})


# /api/v1/production/equip-info-real/
@method_decorator([api_recorder], name="dispatch")
class EquipInfoReal(APIView):
    """设备状态信息"""

    def get(self, request):
        type = self.request.query_params.get('type', None)
        value = self.request.query_params.get('value', None)
        if not all([type, value]):
            raise ValidationError('参数不全，type和value都得传')
        try:
            production_factory_date = datetime.datetime.strptime(value, "%Y-%m-%d")
        except:
            raise ValidationError('时间格式不正确')
        now = datetime.datetime.now()
        if type == 'day':
            work_schedule_plan = WorkSchedulePlan.objects.filter(
                plan_schedule__day_time=production_factory_date,
                plan_schedule__work_schedule__work_procedure__global_name='密炼').first()
            if not work_schedule_plan:
                raise ValidationError('日期传参不正确')
            if now <= work_schedule_plan.end_time:
                end_time = now
            else:
                end_time = work_schedule_plan.end_time
            total_time = end_time - work_schedule_plan.start_time
            filter_dict = {"factory_date": value, 'delete_flag': False}
        elif type == 'week':
            this_week_start = production_factory_date - timedelta(days=production_factory_date.weekday())  # 当天坐在的周的周一
            this_week_end = production_factory_date + timedelta(days=6 - production_factory_date.weekday())  # 当天所在周的周日
            work_schedule_plan = WorkSchedulePlan.objects.filter(
                plan_schedule__day_time__gte=this_week_start.date(),
                plan_schedule__day_time__lte=this_week_end.date(),
                plan_schedule__work_schedule__work_procedure__global_name='密炼').values().aggregate(
                start_times=Min('start_time'),
                end_times=Max('end_time'))
            if not work_schedule_plan['end_times'] or not work_schedule_plan['start_times']:
                raise ValidationError('日期传参不正确')
            if now <= work_schedule_plan['end_times']:
                end_time = now
            else:
                end_time = work_schedule_plan['end_times']
            total_time = end_time - work_schedule_plan['start_times']
            filter_dict = {"factory_date__gte": this_week_start.date(), "factory_date__lte": this_week_end.date(),
                           'delete_flag': False}
        elif type == 'month':
            work_schedule_plan = WorkSchedulePlan.objects.filter(
                plan_schedule__day_time__month=production_factory_date.month,
                plan_schedule__day_time__year=production_factory_date.year,
                plan_schedule__work_schedule__work_procedure__global_name='密炼').values().aggregate(
                start_times=Min('start_time'),
                end_times=Max('end_time'))
            if not work_schedule_plan['end_times'] or not work_schedule_plan['start_times']:
                raise ValidationError('日期传参不正确')
            if now <= work_schedule_plan['end_times']:
                end_time = now
            else:
                end_time = work_schedule_plan['end_times']
            total_time = end_time - work_schedule_plan['start_times']
            filter_dict = {'factory_date__month': production_factory_date.month,
                           'factory_date__year': production_factory_date.year,
                           'delete_flag': False}
        elif type == 'year':
            work_schedule_plan = WorkSchedulePlan.objects.filter(
                plan_schedule__day_time__year=production_factory_date.year,
                plan_schedule__work_schedule__work_procedure__global_name='密炼').values().aggregate(
                start_times=Min('start_time'),
                end_times=Max('end_time'))
            if not work_schedule_plan['end_times'] or not work_schedule_plan['start_times']:
                raise ValidationError('日期传参不正确')
            if now <= work_schedule_plan['end_times']:
                end_time = now
            else:
                end_time = work_schedule_plan['end_times']
            total_time = end_time - work_schedule_plan['start_times']
            filter_dict = {'factory_date__year': production_factory_date.year,
                           'delete_flag': False}
        else:
            raise ValidationError('type只能传day，week，month，year')
        e_set = Equip.objects.filter(delete_flag=False).all()
        resluts = []
        for e_obj in e_set:
            tfb_set = TrainsFeedbacks.objects.filter(equip_no=e_obj.equip_no, **filter_dict
                                                     ).values(
                'equip_no', 'plan_classes_uid').annotate(actual_trains=Max('actual_trains'),
                                                         plan_trains=Max('plan_trains')).aggregate(
                plan_trains_sum=Sum('plan_trains'), actual_trains_sum=Sum('actual_trains'))
            plan_trains = tfb_set.pop("plan_trains_sum", None)
            actual_trains = tfb_set.pop("actual_trains_sum", None)

            tfb_set["plan_trains"] = str(plan_trains) if plan_trains else "0"
            tfb_set["actual_trains"] = str(actual_trains) if actual_trains else "0"
            tfb_set["equip_no"] = e_obj.equip_no
            try:
                if e_obj.equip_current_status_equip.status in ['故障', '维修开始', '维修结束', '空转']:
                    tfb_set['status'] = '故障'
                else:
                    tfb_set['status'] = e_obj.equip_current_status_equip.status
            except:
                tfb_set['status'] = '未知'
            epe_set = e_obj.equip_part_equip.filter(delete_flag=False).all()
            if not epe_set:
                # continue
                tfb_set["fault_time"] = total_time
            for epe_obj in epe_set:
                emo_set = epe_obj.equip_maintenance_order_part.filter(affirm_time__isnull=False, **filter_dict).all()
                if not emo_set:
                    continue
                for emo_obj in emo_set:
                    if not emo_obj.begin_time:
                        continue
                    wx_time = emo_obj.affirm_time - emo_obj.begin_time
                    total_time = total_time - wx_time
            tfb_set["fault_time"] = str(total_time)
            resluts.append(tfb_set)
        return Response({'resluts': resluts})


@method_decorator([api_recorder], name="dispatch")
class RuntimeRecordView(APIView):
    permission_classes = (IsAuthenticated, PermissionClass({'view': 'view_production_record'}))

    def get(self, request, *args, **kwargs):
        params = request.query_params
        classes = params.get("classes")
        factory_date = params.get("date", datetime.date.today())
        if classes in ["早班", "夜班"]:
            filters = {"classes": classes}
        else:
            filters = {}
        queryset = TrainsFeedbacks.objects.filter(factory_date=factory_date).filter(**filters)
        id_list = queryset.values('plan_classes_uid').annotate(max_id=Max('id')).values_list("max_id", flat=True)
        queryset = queryset.filter(id__in=id_list)
        data = queryset.values("equip_no", "product_no").annotate(plan_trains=Sum('plan_trains'),
                                                                  actual_trains=Sum('actual_trains'),
                                                                  begin_time=Min('begin_time'),
                                                                  end_time=Max('end_time')).values(
            "equip_no", "product_no", "plan_trains", "actual_trains", "begin_time", "end_time").order_by('equip_no')

        # 计算获取设备停机时间
        equip_set = EquipMaintenanceOrder.objects.filter(factory_date=factory_date, down_flag=True).values(
            "equip_part__equip__equip_no").annotate(stop_time=OSum((F('end_time') - F('begin_time')))).values(
            "equip_part__equip__equip_no", "stop_time")
        equip_data = {x.get("equip_part__equip__equip_no"): x.get("stop_time") for x in equip_set}
        equip_no_list = Equip.objects.filter(category__equip_type__global_name="密炼设备").values_list("equip_no",
                                                                                                   flat=True)
        equip_trains = {_: [0] for _ in equip_no_list}

        equip_list = ['Z01', 'Z02', 'Z03', 'Z04', 'Z05', 'Z06', 'Z07', 'Z08', 'Z09', 'Z10', 'Z11', 'Z12', 'Z13', 'Z14',
                      'Z15']
        if classes:
            if not isinstance(factory_date, datetime.date):
                factory_date = datetime.datetime.strptime(factory_date, '%Y-%m-%d').date()

            if factory_date == datetime.date.today():
                if not ProductionDailyRecords.objects.filter(factory_date=factory_date, classes=classes).first():

                    production_daily = ProductionDailyRecords.objects.create(factory_date=factory_date, classes=classes,
                                                                            equip_error_record = None,
                                                                            process_shutdown_record = None,
                                                                            production_shutdown_record = None,
                                                                            auxiliary_positions_record = None,
                                                                            shift_leader = None,
                    )

                    # 获取上一天的人员姓名
                    last_date = (factory_date + datetime.timedelta(days=-1)).strftime("%Y-%m-%d")
                    for equip in equip_list:
                        last = ProductionPersonnelRecords.objects.filter(equip_no=equip,
                                                                         production_daily__factory_date=last_date,
                                                                         production_daily__classes=classes).first()
                        if last:
                            ProductionPersonnelRecords.objects.create(equip_no=equip, production_daily=production_daily,
                                                                      feeding_post=last.feeding_post,
                                                                      extrusion_post=last.extrusion_post,
                                                                      collection_post=last.collection_post,
                                                                      )
                        else:
                            ProductionPersonnelRecords.objects.create(equip_no=equip, production_daily=production_daily)

        # 获取计划表里的计划车次
        if filters:
            plan_filter = {"work_schedule_plan__classes__global_name": classes}
        else:
            plan_filter = {}
        plan_set = ProductClassesPlan.objects.filter(work_schedule_plan__plan_schedule__day_time=factory_date,
                                                     work_schedule_plan__plan_schedule__work_schedule__work_procedure__global_name="密炼",
                                                     delete_flag=False, **plan_filter).values('equip__equip_no',
                                                                                              'product_batching__stage_product_batch_no').annotate(
            plan_trains=Sum('plan_trains')).values('equip__equip_no',
                                                   'product_batching__stage_product_batch_no', 'plan_trains')
        plan_data = {_.get('equip__equip_no') + _.get('product_batching__stage_product_batch_no'): _.get('plan_trains')
                     for _ in plan_set}
        results = []
        users = None
        for _ in data:
            equip_trains[_.get("equip_no")][0] += _.get("actual_trains")
            achieve_rate = round(_.get("actual_trains") / plan_data.get(_.get("equip_no") + _.get("product_no")), 4) \
                if plan_data.get(_.get("equip_no") + _.get("product_no")) else None
            equip_no = _.get("equip_no")
            if classes:
                users = ProductionPersonnelRecords.objects.filter(production_daily__factory_date=factory_date,
                                                           production_daily__classes=classes, equip_no=equip_no).first()
            else:
                users = None
            temp_dict = {"id": users.id if users else None,
                         "equip_no": equip_no,
                         "product_no": _.get("product_no"),
                         "plan_trains": plan_data.get(_.get("equip_no") + _.get("product_no")),
                         "actual_trains": _.get("actual_trains"),
                         "achieve_rate": achieve_rate,
                         "put_user": users.feeding_post if users else None,
                         "extrusion_user": users.extrusion_post if users else None,
                         "collection_user": users.collection_post if users else None,
                         "product_time": (_.get("end_time") - _.get("begin_time")).total_seconds(),
                         "trains_sum": equip_trains.get(_.get("equip_no")),
                         "start_rate": (24 * 60 * 60 - equip_data.get(_.get("equip_no"))) / 60 if equip_data.get(
                             _.get("equip_no")) else 1.0
                         }
            results.append(temp_dict)
        return Response({"results": results, "shift_leader": users.production_daily.shift_leader if users else None,
})


@method_decorator([api_recorder], name="dispatch")
class RuntimeRecordDetailView(APIView):
    permission_classes = (IsAuthenticated,)

    # 点击报表预览
    def get(self, request, *args, **kwargs):
        factory_date = self.request.query_params.get('date', datetime.date.today())
        classes = self.request.query_params.get('classes', None)
        if classes:
            obj = ProductionDailyRecords.objects.get(factory_date=factory_date, classes=classes)

            results = {
                'id': obj.id,
                'shift_leader': obj.shift_leader,
                'equip_error_record': obj.equip_error_record,
                'process_shutdown_record': obj.process_shutdown_record,
                'production_shutdown_record': obj.production_shutdown_record,
                'auxiliary_positions_record': obj.auxiliary_positions_record
            }

            try:
                obj = WorkSchedulePlan.objects.filter(start_time__startswith=factory_date,
                                                      classes__global_name=classes).first()
                group = obj.group.global_name
            except: group = None

            return Response({'group': group, 'results': results})
        return Response({})

    def post(self, request, *args, **kwargs):
        data = self.request.data
        main = data.get('main')
        detail = data.get('detail', None)
        shift_leader = data.get('shift_leader', None)
        feeding_post = data.get('feeding_post', None)
        extrusion_post = data.get('extrusion_post', None)
        collection_post = data.get('collection_post', None)
        equip_error_record = data.get('equip_error_record', None)
        process_shutdown_record = data.get('process_shutdown_record', None)
        production_shutdown_record = data.get('production_shutdown_record', None)
        auxiliary_positions_record = data.get('auxiliary_positions_record', None)
        if detail:
        # 修改人员姓名
            obj = ProductionPersonnelRecords.objects.filter(id=detail).first()
            if feeding_post:
                obj.feeding_post = feeding_post
            if extrusion_post:
                obj.extrusion_post = extrusion_post
            if collection_post:
                obj.collection_post = collection_post
            obj.save()
            results = {'feeding_post': obj.feeding_post,
                        'extrusion_post': obj.extrusion_post,
                        'collection_post': obj.collection_post}
            return Response(results)
        # 修改相关记录
        if main:
            obj1 = ProductionDailyRecords.objects.filter(id=main).first()
            if equip_error_record:
                obj1.equip_error_record = equip_error_record
            if process_shutdown_record:
                obj1.process_shutdown_record = process_shutdown_record
            if production_shutdown_record:
                obj1.production_shutdown_record = production_shutdown_record
            if auxiliary_positions_record:
                obj1.auxiliary_positions_record = auxiliary_positions_record
            obj1.save()
            results = {
                'shift_leader': obj1.shift_leader,
                'equip_error_record': obj1.equip_error_record,
                'process_shutdown_record': obj1.process_shutdown_record,
                'production_shutdown_record': obj1.production_shutdown_record,
                'auxiliary_positions_record': obj1.auxiliary_positions_record
                 }
            return Response(results)

        if shift_leader:
            factory_date = data.get("date")
            classes = data.get("classes")
            obj1 = ProductionDailyRecords.objects.filter(factory_date=factory_date, classes=classes).first()
            if factory_date and classes:
                obj1.shift_leader=shift_leader
                obj1.save()
            return Response({'shift_leader': obj1.shift_leader})


@method_decorator([api_recorder], name="dispatch")
class TrainsFixView(APIView):
    """修改收皮车次，支持指定托修改和批次修改"""
    # {"factory_date": "2021-07-14",
    #  "classes": "早班",
    #  "equip_no": "Z01",
    #  "product_no": "C-RE-J260-02",
    #  "begin_trains": 1,
    #  "end_trains": 6,
    #  "fix_num": 1,
    #  "lot_no": "AAJ1Z062021072313078"
    #  }

    def post(self, request):
        s = TrainsFixSerializer(data=self.request.data)
        s.is_valid(raise_exception=True)
        data = s.validated_data
        lot_no = data.get('lot_no')
        if lot_no:  # 修改特定托的车次
            pallet = PalletFeedbacks.objects.filter(lot_no=data['lot_no']).first()
            if not pallet:
                raise ValidationError('未找到收皮数据！')
            data['product_no'] = pallet.product_no
            data['classes'] = pallet.classes
            data['equip_no'] = pallet.equip_no
            data['factory_date'] = pallet.factory_date
            if PalletFeedbacks.objects.exclude(
                    lot_no=data['lot_no']).filter(equip_no=data['equip_no'],
                                                  product_no=data['product_no'],
                                                  classes=data['classes'],
                                                  factory_date=data['factory_date']
                                                  ).filter(Q(begin_trains__lte=data['begin_trains'],
                                                           end_trains__gte=data['begin_trains']) |
                                                           Q(begin_trains__lte=data['end_trains'],
                                                           end_trains__gte=data['end_trains']) |
                                                           Q(begin_trains__gte=data['begin_trains'],
                                                             end_trains__lte=data['end_trains'])):
                raise ValidationError('车次重复，请确认后重试！')
            pallet.begin_trains = data['begin_trains']
            pallet.end_trains = data['end_trains']
            pallet.save()
            MaterialTestOrder.objects.filter(lot_no=data['lot_no']).delete()
            lot_nos = [data['lot_no']]
        else:
            fix_num = data['fix_num']
            if data['begin_trains'] + fix_num <= 0:
                raise ValidationError('修改后车次不可为0！')
            if not PalletFeedbacks.objects.filter(equip_no=data['equip_no'],
                                                  product_no=data['product_no'],
                                                  classes=data['classes'],
                                                  factory_date=data['factory_date'],
                                                  begin_trains=data['begin_trains']):
                raise ValidationError('开始车次必须为一托开始车次！')
            if not PalletFeedbacks.objects.filter(equip_no=data['equip_no'],
                                                  product_no=data['product_no'],
                                                  classes=data['classes'],
                                                  factory_date=data['factory_date'],
                                                  end_trains=data['end_trains']):
                raise ValidationError('结束车次必须为一托结束车次！')
            pallet_data = PalletFeedbacks.objects.filter(equip_no=data['equip_no'],
                                                         product_no=data['product_no'],
                                                         classes=data['classes'],
                                                         factory_date=data['factory_date'],
                                                         ).filter(Q(begin_trains__lte=data['begin_trains'],
                                                                  end_trains__gte=data['begin_trains']) |
                                                                  Q(begin_trains__lte=data['end_trains'],
                                                                  end_trains__gte=data['end_trains']) |
                                                                  Q(begin_trains__gte=data['begin_trains'],
                                                                  end_trains__lte=data['end_trains'])
                                                                  ).order_by('begin_trains')
            if not pallet_data:
                raise ValidationError('未找到改批次收皮数据！')
            lot_nos = set(pallet_data.values_list('lot_no', flat=True))
            if PalletFeedbacks.objects.filter(equip_no=data['equip_no'],
                                              product_no=data['product_no'],
                                              classes=data['classes'],
                                              factory_date=data['factory_date']
                                              ).exclude(id__in=pallet_data.values_list('id', flat=True)).filter(
                    Q(begin_trains__lte=data['begin_trains']+fix_num, end_trains__gte=data['begin_trains']+fix_num) |
                    Q(begin_trains__lte=data['end_trains']+fix_num, end_trains__gte=data['end_trains']+fix_num) |
                    Q(begin_trains__gte=data['begin_trains']+fix_num, end_trains__lte=data['end_trains']+fix_num)
            ).exists():
                raise ValidationError('修改后车次信息重复！')
            for pallet in pallet_data:
                pallet.begin_trains += fix_num
                pallet.end_trains += fix_num
                pallet.save()

            MaterialTestOrder.objects.filter(lot_no__in=lot_nos).delete()
        MaterialDealResult.objects.filter(lot_no__in=lot_nos).delete()
        return Response('修改成功')


@method_decorator([api_recorder], name="dispatch")
class PalletTrainsBatchFixView(ListAPIView, UpdateAPIView):
    """
        收皮车次批量修改
    """
    queryset = PalletFeedbacks.objects.filter(delete_flag=False).order_by('begin_trains', 'begin_time')
    permission_classes = (IsAuthenticated,)
    serializer_class = PalletFeedbacksBatchModifySerializer
    filter_backends = [DjangoFilterBackend, ]
    filter_class = PalletFeedbacksFilter
    pagination_class = None

    @atomic()
    def put(self, request, *args, **kwargs):
        data = self.request.data
        if not isinstance(data, list):
            raise ValidationError('参数错误！')
        for item in data:
            instance = PalletFeedbacks.objects.get(id=item['id'])
            s = PalletFeedbacksBatchModifySerializer(instance=instance, data=item, context={'request': request})
            s.is_valid(raise_exception=True)
            s.save()
            MaterialTestOrder.objects.filter(lot_no=instance.lot_no).delete()
            MaterialDealResult.objects.filter(lot_no=instance.lot_no).delete()
        return Response('修改成功')


@method_decorator([api_recorder], name="dispatch")
class ProductPlanRealView(ListAPIView):
    queryset = ProductClassesPlan.objects.filter(delete_flag=False).order_by('id')
    serializer_class = ProductPlanRealViewSerializer
    permission_classes = (IsAuthenticated,)
    filter_backends = [DjangoFilterBackend, ]
    filter_class = ProductClassesPlanFilter
    pagination_class = None

    def list(self, request, *args, **kwargs):
        ret = {}
        auto = int(self.request.query_params.get('auto'))
        manual = int(self.request.query_params.get('manual'))
        equip_no = self.request.query_params.get('equip_no')
        day_time = self.request.query_params.get('day_time', datetime.datetime.now().date())
        queryset = self.filter_queryset(self.get_queryset())
        queryset = queryset.filter(work_schedule_plan__plan_schedule__day_time=day_time)
        if manual:
            add_data = ManualInputTrains.objects.filter(factory_date=day_time, equip_no=equip_no).values(
                'factory_date', 'equip_no', 'product_no', 'classes', 'actual_trains')
        else:
            add_data = []
        for classes in ['早班', '中班', '夜班']:
            page = queryset.filter(work_schedule_plan__classes__global_name=classes)
            data = self.get_serializer(page, many=True).data
            data = list(filter(lambda x: x['begin_time'] is not None, data))
            data.sort(key=lambda x: x['begin_time'])
            ret[classes] = data if auto else []

        for item in add_data:
            tfb_obj = TrainsFeedbacks.objects.filter(product_no=item['product_no'],
                                                     factory_date=item['factory_date']).order_by('id').first()
            if tfb_obj:
                plan_obj = self.queryset.filter(plan_classes_uid=tfb_obj.plan_classes_uid).first()
            else:
                plan_obj = None
            last_obj = TrainsFeedbacks.objects.filter(product_no=item['product_no'],
                                                     factory_date=item['factory_date']).order_by('id').last()
            begin_time = tfb_obj.begin_time.strftime("%Y-%m-%d %H:%M:%S") if tfb_obj else None
            kwargs = {
                'classes': item['classes'],
                'plan_trains': plan_obj.plan_trains if plan_obj else None,
                'actual_trains': last_obj.actual_trains if last_obj else None,
                'product_no': item['product_no'],
                'begin_time': begin_time
            }
            if kwargs in ret[item['classes']]:
                index = ret[item['classes']].index(kwargs)
                kwargs.update({'actual_trains': kwargs['actual_trains'] + item['actual_trains']})
                ret[item['classes']][index] = kwargs
            else:
                kwargs.update({'actual_trains': item['actual_trains']})
                ret[item['classes']].append(kwargs)

        return Response(ret)


@method_decorator([api_recorder], name="dispatch")
class ProductBatchInfo(APIView):

    def get(self, request):
        lot_no = self.request.query_params.get('lot_no')
        if not lot_no:
            return Response({'success': False, 'data': {}, 'message': '请携带正确的条码号查询！'})
        pallet_feedback = PalletFeedbacks.objects.filter(lot_no=lot_no).first()
        if not pallet_feedback:
            return Response({'success': False, 'data': {}, 'message': '未找到该条码信息！'})

        pallet_data = PalletFeedbacks.objects.filter(lot_no=lot_no).first()
        plan = ProductClassesPlan.objects.filter(plan_classes_uid=pallet_data.plan_classes_uid).first()
        if not plan:
            group = ''
        else:
            group = plan.work_schedule_plan.group.global_name

        ins = MaterialDealResult.objects.filter(lot_no=lot_no).first()
        if not ins:
            data = {
                'product_no': pallet_data.product_no,
                'equip_no': pallet_data.equip_no,
                'classes': pallet_data.classes,
                'group': group,
                'factory_date': pallet_data.factory_date,
                'weight': pallet_feedback.actual_weight,
                'trains': '/'.join([str(i) for i in range(pallet_data.begin_trains, pallet_data.end_trains + 1)]),
                'test_result': '未检测',
            }
            return Response({'success': True, 'data': data, 'message': '查询成功！'})

        data = {
            'product_no': pallet_data.product_no,
            'equip_no': pallet_data.equip_no,
            'classes': pallet_data.classes,
            'group': group,
            'factory_date': pallet_data.factory_date,
            'weight': pallet_feedback.actual_weight,
            'trains': '/'.join([str(i) for i in range(pallet_data.begin_trains, pallet_data.end_trains + 1)]),
            'test_result': ins.test_result,
        }

        ret = []
        test_orders = MaterialTestOrder.objects.filter(lot_no=lot_no,
                                                       product_no=ins.product_no
                                                       ).order_by('actual_trains')
        for test_order in test_orders:
            max_result_ids = list(test_order.order_results.values(
                'test_indicator_name', 'data_point_name'
            ).annotate(max_id=Max('id')).values_list('max_id', flat=True))
            test_results = MaterialTestResult.objects.filter(id__in=max_result_ids,
                                                             is_judged=True).order_by('test_indicator_name',
                                                                                      'data_point_name')
            for test_result in test_results:
                if test_result.level == 1:
                    result = '合格'
                elif test_result.is_passed:
                    result = 'pass'
                else:
                    result = '不合格'
                indicator = MaterialDataPointIndicator.objects.filter(
                    data_point__name=test_result.data_point_name,
                    material_test_method__material__material_name=ins.product_no,
                    level=1).first()
                if indicator:
                    upper_limit = indicator.upper_limit
                    lower_limit = indicator.lower_limit
                else:
                    upper_limit = lower_limit = None
                ret.append(
                    {
                        'data_point_name': test_result.data_point_name,
                        'result': result,
                        'value': test_result.value,
                        'test_indicator_name': test_result.test_indicator_name,
                        'train': test_order.actual_trains,
                        'upper_limit': upper_limit,
                        'lower_limit': lower_limit
                    }
                )
        data['test_info'] = ret
        return Response({'success': True, 'data': data, 'message': '查询成功！'})


@method_decorator([api_recorder], name="dispatch")
class RubberCannotPutinReasonView(APIView):
    permission_classes = (IsAuthenticated,)

    def export_xls(self, columns, result, details):
        response = HttpResponse(content_type='application/vnd.ms-excel')
        filename = '不入库原因统计'
        response['Content-Disposition'] = u'attachment;filename= ' + filename.encode('gbk').decode(
            'ISO-8859-1') + '.xls'
        # 创建一个文件对象
        wb = xlwt.Workbook(encoding='utf8')
        # 创建一个sheet对象
        sheet = wb.add_sheet('统计', cell_overwrite_ok=True)
        sheet2 = wb.add_sheet('明细', cell_overwrite_ok=True)
        style = xlwt.XFStyle()
        style.alignment.wrap = 1
        columns.insert(0, '不入库原因')
        # columns.append('合计')
        # 写入文件标题
        for idx, sheet_head in enumerate(columns):
            sheet.write(0, idx, sheet_head)
        for idx, sheet_head in enumerate(['不入库原因','工厂日期','机台','配方名','车次','收皮条码','托盘号', '登录时间']):
            sheet2.write(0, idx, sheet_head)
        # 写入数据
        data_row = 1
        for i in result:
            for col_num, data_key in enumerate(columns):
                sheet.write(data_row, col_num, i[data_key])
            data_row += 1

        data_row2 = 1
        for item in details:
            sheet2.write(data_row2, 0, item['reason_name'])
            sheet2.write(data_row2, 1, item['factory_date'])
            sheet2.write(data_row2, 2, item['machine_no'])
            sheet2.write(data_row2, 3, item['production_no'])
            sheet2.write(data_row2, 4, item['trains'])
            sheet2.write(data_row2, 5, item['lot_no'])
            sheet2.write(data_row2, 6, item['pallet_no'])
            sheet2.write(data_row2, 7, item['input_datetime'])
            data_row2 += 1

        # 写出到IO
        output = BytesIO()
        wb.save(output)
        # 重新定位到开始
        output.seek(0)
        response.write(output.getvalue())
        return response

    def get(self, request):
        s_time = self.request.query_params.get('s_time')
        e_time = self.request.query_params.get('e_time')
        factory_date = self.request.query_params.get('factory_date')
        equip = self.request.query_params.get('equip')
        export = self.request.query_params.get('export')
        equip_list = ['Z%.2d' % i for i in range(1, 16)]
        queryset = RubberCannotPutinReason.objects.all()
        if equip:
            result = RubberCannotPutinReasonSerializer(queryset.filter(factory_date__date=factory_date, machine_no=equip), many=True).data
        elif s_time:
            start = datetime.datetime.strptime(s_time, "%Y-%m-%d")
            end = datetime.datetime.strptime(e_time, "%Y-%m-%d")
            """获取两个日期之间的所有日期"""
            delta = end - start
            time_list = [(start + timedelta(days=i)).strftime("%Y年%m月%d日") for i in range(delta.days + 1)]
            dic = {}
            queryset = queryset.filter(factory_date__date__range=(s_time, e_time))
            temp = queryset.values('reason_name', 'factory_date__date').annotate(
                num=Count('id'))
            for item in temp:
                t = item['factory_date__date'].strftime("%Y年%m月%d日")
                if dic.get(item['reason_name']):
                    if dic[item['reason_name']][t]:
                        dic[item['reason_name']][t] += item['num']
                    else:
                        dic[item['reason_name']][t] = item['num']
                    dic[item['reason_name']]['count'] += item['num']
                else:
                    dic[item['reason_name']] = {'不入库原因': item['reason_name'], 'count': 0}
                    for time in time_list:
                        dic[item['reason_name']].update({time: None})
                    dic[item['reason_name']][t] = item['num']
                    dic[item['reason_name']]['count'] += item['num']
            result = dic.values()
            details = RubberCannotPutinReasonSerializer(queryset, many=True).data
            if export:
                return self.export_xls(time_list, result, details)
            # count = {}  # 底部数量统计
            # for i in result:
            #     for key, value in i.items():
            #         if key[1:2].isdecimal() and value:
            #             if count.get(key):
            #                 count[key] += value
            #             else:
            #                 count[key] = value
        elif factory_date:
            dic = {}
            temp = queryset.filter(factory_date__date=factory_date).values('reason_name', 'machine_no').annotate(
                num=Count('id'))
            for item in temp:
                if dic.get(item['reason_name']):
                    if dic[item['reason_name']][item['machine_no']]:
                        dic[item['reason_name']][item['machine_no']] += item['num']
                    else:
                        dic[item['reason_name']][item['machine_no']] = item['num']
                    dic[item['reason_name']]['count'] += item['num']
                else:
                    dic[item['reason_name']] = {'不入库原因': item['reason_name'], 'count': 0}
                    for equip in equip_list:
                        dic[item['reason_name']].update({equip: None})
                    dic[item['reason_name']][item['machine_no']] = item['num']
                    dic[item['reason_name']]['count'] += item['num']
            result = dic.values()
        return Response({'results': result})


@method_decorator([api_recorder], name="dispatch")
class MachineTargetValue(APIView):
    permission_classes = (IsAuthenticated,)

    def get(self, request):
        target_month = self.request.query_params.get('target_month')
        equip_no = self.request.query_params.get('equip_no')
        if not target_month:
            raise ValidationError('请选择月份')
        results = []
        queryset = list(MachineTargetYieldSettings.objects.filter(target_month=target_month).order_by('day', 'id').values())
        equip_list = ['Z01', 'Z02', 'Z03', 'Z04', 'Z05', 'Z06', 'Z07', 'Z08', 'Z09', 'Z10', 'Z11', 'Z12', 'Z13', 'Z14', 'Z15', '190E']
        if not queryset:
            if not equip_no:
                for equip_no in equip_list:
                    if equip_no == '190E':
                        k = 'E190'
                    else:
                        k = equip_no
                    results.append({'equip_no': equip_no,
                                    'target_weight': '',
                                    'max_weight': ''})
        else:
            if not equip_no:
                data = queryset[-1]
                for equip_no in equip_list:
                    if equip_no == '190E':
                        k = 'E190'
                    else:
                        k = equip_no
                    results.append({'equip_no': equip_no,
                                    'target_weight': data.get(k),
                                    'max_weight': data.get('{}_max'.format(k))})
            else:
                for i in queryset:
                    results.append({'id': i.get('id'),
                                    'equip_no': equip_no,
                                    'day': i.get('day'),
                                    'classes': i.get('classes'),
                                    'target_weight': i.get(equip_no),
                                    'max_weight': i.get(f'{equip_no}_max')})

        return Response({'results': results})

    def post(self, request):
        rid = self.request.data.get('id')
        target_month = self.request.data.get('target_month')
        target_data = self.request.data.get('data')
        data = {}
        if rid:
            equip_no = self.request.data.get('equip_no')
            target_weight = self.request.data.get('target_weight')
            max_weight = self.request.data.get('max_weight')
            MachineTargetYieldSettings.objects.filter(id=rid).update(**{equip_no: target_weight, f'{equip_no}_max': max_weight})
            return Response('ok')
        for item in target_data:
            equip_no = item['equip_no']
            if equip_no == '190E':
                k = 'E190'
            else:
                k = equip_no
            target_weight = item['target_weight']
            max_weight = item['max_weight']
            data[k] = target_weight
            data['{}_max'.format(k)] = max_weight
        # 获取设定值(存在则更新，否则创建)
        exist_month = MachineTargetYieldSettings.objects.filter(target_month=target_month).order_by('id')
        if not exist_month:
            data.update(day=1, classes='早班', target_month=target_month)
            MachineTargetYieldSettings.objects.create(**data)
        else:
            last_setting = exist_month.last()
            if last_setting.day == 1 and last_setting.classes == '早班':
                day, classes = 1, '早班'
            else:
                res = get_current_factory_date()
                if len(res) < 2:
                    raise ValidationError(f'{target_month}密炼排班数据录入不全')
                day, classes = res.get('factory_date').day, res.get('classes')
            MachineTargetYieldSettings.objects.update_or_create(defaults=data, **{'target_month': target_month, 'day': day, 'classes': classes})
        return Response('ok')


@method_decorator([api_recorder], name="dispatch")
class MonthlyOutputStatisticsReport(APIView):
    queryset = TrainsFeedbacks.objects.filter(Q(~Q(equip_no='Z04')) | Q(equip_no='Z04', operation_user='Mixer1'))
    permission_classes = (IsAuthenticated,)

    def get(self, request, *args, **kwargs):

        st = self.request.query_params.get('st')
        et = self.request.query_params.get('et')
        if not all([st, et]):
            raise ValidationError('请选择日期范围查询！')
        try:
            e_time = datetime.datetime.strptime(et, '%Y-%m-%d')
            s_time = datetime.datetime.strptime(st, '%Y-%m-%d')
        except Exception:
            raise ValidationError('日期错误！')
        diff = e_time - s_time
        diff_days = diff.days + 1
        if diff_days > 31:
            raise ValidationError('搜索日期跨度不得超过一个月！')

        st_month = st[:7]
        et_month = et[:7]
#         if st_month == et_month:
#             month_target = MachineTargetYieldSettings.objects.filter(target_month=st_month).order_by('-id')
#             if not month_target:
#                 equip_target = {}
#             else:
#                 equip_target = month_target.values('Z01', 'Z02', 'Z03', 'Z04', 'Z05',
#                                                    'Z06', 'Z07', 'Z08', 'Z09', 'Z10',
#                                                    'Z11', 'Z12', 'Z13', 'Z14', 'Z15')[0]
#                 equip_target = {k: v*diff_days*2 for k, v in equip_target.items()}
#         else:
#             equip_target = {}
#             st_month_target = MachineTargetYieldSettings.objects.filter(target_month=st_month).order_by('-id')
#             et_month_target = MachineTargetYieldSettings.objects.filter(target_month=et_month).order_by('-id')
#             if st_month_target and et_month_target:
#                 st_month_days = calendar.monthrange(int(st_month[:4]), int(st_month[5:7]))[1] - int(st[8:10]) + 1
#                 et_month_days = int(et[8:10])
#                 st_month_equip_target = st_month_target.values('Z01', 'Z02', 'Z03', 'Z04', 'Z05',
#                                                                'Z06', 'Z07', 'Z08', 'Z09', 'Z10',
#                                                                'Z11', 'Z12', 'Z13', 'Z14', 'Z15')[0]
#                 et_month_equip_target = et_month_target.values('Z01', 'Z02', 'Z03', 'Z04', 'Z05',
#                                                                'Z06', 'Z07', 'Z08', 'Z09', 'Z10',
#                                                                'Z11', 'Z12', 'Z13', 'Z14', 'Z15')[0]
#                 for equip_no, st_v in st_month_equip_target.items():
#                     et_v = et_month_equip_target.get(equip_no, 0)
#                     equip_target[equip_no] = st_v * st_month_days * 2 + et_v * et_month_days * 2
#
#         # 区间时间范围内机台每日、每班次最大生产数据
#         equip_max_output_data = TrainsFeedbacks.objects.exclude(
#             Q(product_no__icontains='XCJ') |
#             Q(product_no__icontains='洗车胶') |
#             Q(operation_user='Mixer2') |
#             Q(product_no__icontains='WUMING')).filter(
#             factory_date__gte=st,
#             factory_date__lte=et
#         ).values('equip_no', 'factory_date', 'classes'
#                  ).annotate(w=Sum('plan_weight')/1000, t=Count('id')).order_by('equip_no', 'w', 't')
#         equip_max_classes_dict, equip_max = {}, {}
#         for item in equip_max_output_data:
#             equip_no = item['equip_no']
#             factory_date = item['factory_date']
#             classes = item['classes']
#             w = item['w']
#             _max_num = equip_max.get(item['equip_no'], item['t'])
#             t = item['t'] if _max_num < item['t'] else _max_num
#             equip_max[item['equip_no']] = t
#             equip_max_classes_dict[equip_no] = {'factory_date': factory_date, 'weight': w, 'classes': classes, 'trains': t}
#
#         sql = """
#         select temp2.EQUIP_NO, w2/1000, FACTORY_DATE, CLASSES from(
# select
#        EQUIP_NO,
#        max(w) as w2
# from (
# select EQUIP_NO,
#        sum(PLAN_WEIGHT) as w
# from TRAINS_FEEDBACKS where FACTORY_DATE is not null group by FACTORY_DATE, EQUIP_NO, CLASSES) temp
# group by EQUIP_NO) temp2
# inner join (select FACTORY_DATE,
#        EQUIP_NO,
#        CLASSES,
#        sum(PLAN_WEIGHT) as w3
# from TRAINS_FEEDBACKS where FACTORY_DATE is not null group by FACTORY_DATE, EQUIP_NO, CLASSES) temp3
# on temp2.EQUIP_NO=temp3.EQUIP_NO and temp2.w2=temp3.w3;"""
#         cursor = connection.cursor()
#         cursor.execute(sql)
#         query_data = cursor.fetchall()
#         equip_history_max_classes_dict = {}
#         for item in query_data:
#             equip_history_max_classes_dict[item[0]] = {'factory_date': item[2],
#                                                        'weight': item[1]/2 if item[0]=='Z04' else item[1],
#                                                        'classes': item[3]}
#
#         # 机台产量
#         queryset = TrainsFeedbacks.objects.exclude(
#             Q(product_no__icontains='XCJ') |
#             Q(product_no__icontains='洗车胶') |
#             Q(operation_user='Mixer2') |
#             Q(product_no__icontains='WUMING')
#         ).filter(
#             factory_date__gte=st,
#             factory_date__lte=et
#         ).values('equip_no').annotate(total_weight=Sum('plan_weight')/1000,
#                                       total_trains=Count('id')).order_by('equip_no')
#         # 历史最高车次(班组)
#         equip_history_data = TrainsFeedbacks.objects.filter(factory_date__isnull=False).values('equip_no', 'factory_date', 'classes')\
#             .annotate(trains=Count('id')).values('equip_no', 'trains')
#         classes_equip_history_data = {}
#         for i in equip_history_data:
#             qty = classes_equip_history_data.get(i['equip_no'], 0)
#             if qty < i['trains']:
#                 classes_equip_history_data[i['equip_no']] = i['trains']
#
#         # 补充机台所选时间内的产量最大值
#         for item in queryset:
#             group = ""
#             max_weight = ""
#             max_classes_trains = ""
#             history_group = ""
#             history_max_weight = ""
#             history_max_trains = ""
#             equip_max_classes_data = equip_max_classes_dict.get(item['equip_no'])
#             if equip_max_classes_data:
#                 max_weight = round(equip_max_classes_data['weight'], 2)
#                 max_classes_trains = equip_max_classes_data.get('trains', '')
#                 work_schedule_plan = WorkSchedulePlan.objects.filter(
#                     plan_schedule__work_schedule__work_procedure__global_name='密炼',
#                     plan_schedule__day_time=equip_max_classes_data['factory_date'],
#                     classes__global_name=equip_max_classes_data['classes']
#                 ).first()
#                 if work_schedule_plan:
#                     group = work_schedule_plan.group.global_name
#             equip_history_max_classes_data = equip_history_max_classes_dict.get(item['equip_no'])
#             if equip_history_max_classes_data:
#                 history_max_weight = round(equip_history_max_classes_data['weight'], 2)
#                 history_max_trains = classes_equip_history_data.get(item['equip_no'])
#                 work_schedule_plan = WorkSchedulePlan.objects.filter(
#                     plan_schedule__work_schedule__work_procedure__global_name='密炼',
#                     plan_schedule__day_time=equip_history_max_classes_data['factory_date'],
#                     classes__global_name=equip_history_max_classes_data['classes']
#                 ).first()
#                 if work_schedule_plan:
#                     history_group = work_schedule_plan.group.global_name
#             item['target'] = equip_target.get(item['equip_no'], 0)
#             item['group'] = group
#             item['max_classes_trains'] = max_classes_trains
#             item['max_weight'] = max_weight
#             item['history_group'] = history_group
#             item['history_max_weight'] = history_max_weight
#             item['history_max_trains'] = history_max_trains
#             item['total_weight'] = round(item['total_weight'], 2)
        # 只保留段数信息
        queryset = []
        wl_stage_output = {'wl': {'name': "wl", 'value': 0}}
        jl_stage_output = {'jl': {'name': "jl", 'value': 0}}
        # 各段次产量
        qs = TrainsFeedbacks.objects.exclude(
            Q(product_no__icontains='XCJ') |
            Q(product_no__icontains='洗车胶') |
            Q(operation_user='Mixer2') |
            Q(product_no__icontains='WUMING')
        ).filter(
            factory_date__gte=st,
            factory_date__lte=et
        )
        stage_production_queryset = qs.values('product_no').annotate(total_weight=Sum('plan_weight')/1000).order_by('product_no')
        # 当月190E除去洗车胶外的所有产量
        queryset_190e = Equip190EWeight.objects.exclude(setup__specification__in=('洗车胶', 'XCJ', 'WUMING')).filter(factory_date__gte=st, factory_date__lte=et)
        # 按胶料编码分组，计算190E总生产重量
        total_queryset_190e_dict = dict(queryset_190e.values('setup').annotate(stage=F('setup__state'), sum_weight=Sum(F('setup__weight')*F('qty')/1000, output_field=DecimalField())).values_list('stage', 'sum_weight'))
        # 按胶料编码分组，计算手动总生产重量
        total_manual_input_trains = ManualInputTrains.objects.exclude(
            Q(product_no__icontains='XCJ') |
            Q(product_no__icontains='洗车胶') |
            Q(product_no__icontains='-WUMING-')
        ).filter(factory_date__gte=st, factory_date__lte=et)
        total_manual_input_trains_dict = dict(total_manual_input_trains.values('product_no').annotate(weight=Sum(F('weight') * F('actual_trains')/1000, output_field=DecimalField())).values_list('product_no', 'weight'))

        # 密炼产量
        for item in stage_production_queryset:
            try:
                stage = item['product_no'].split('-')[1]
            except Exception:
                continue
            weight = round(item['total_weight'], 2)
            if stage in ('RE', 'FM', 'RFM'):
                jl_stage_output['jl']['value'] += weight
                if stage not in jl_stage_output:
                    jl_stage_output[stage] = {'name': stage, 'value': weight}
                else:
                    jl_stage_output[stage]['value'] += weight
            else:
                wl_stage_output['wl']['value'] += weight
                if stage not in wl_stage_output:
                    wl_stage_output[stage] = {'name': stage, 'value': weight}
                else:
                    wl_stage_output[stage]['value'] += weight
        # 190e产量
        for k, v in total_queryset_190e_dict.items():
            stage, weight = k, round(v, 2)
            if stage in ('RE', 'FM', 'RFM'):
                jl_stage_output['jl']['value'] += weight
                if stage not in jl_stage_output:
                    jl_stage_output[stage] = {'name': stage, 'value': weight}
                else:
                    jl_stage_output[stage]['value'] += weight
            else:
                wl_stage_output['wl']['value'] += weight
                if stage not in wl_stage_output:
                    wl_stage_output[stage] = {'name': stage, 'value': weight}
                else:
                    wl_stage_output[stage]['value'] += weight
        # 手动生产产量
        for k, v in total_manual_input_trains_dict.items():
            try:
                stage, weight = k.split('-')[1], round(v, 2)
            except Exception:
                continue
            if stage in ('RE', 'FM', 'RFM'):
                jl_stage_output['jl']['value'] += weight
                if stage not in jl_stage_output:
                    jl_stage_output[stage] = {'name': stage, 'value': weight}
                else:
                    jl_stage_output[stage]['value'] += weight
            else:
                wl_stage_output['wl']['value'] += weight
                if stage not in wl_stage_output:
                    wl_stage_output[stage] = {'name': stage, 'value': weight}
                else:
                    wl_stage_output[stage]['value'] += weight
        wl_data = wl_stage_output.values()
        jl_data = jl_stage_output.values()
        wl_sort_rules = {'1MB': 1, '2MB': 2, '3MB': 3, '4MB': 4, 'HMB': 5, 'CMB': 6, 'RMB': 7, 'wl': 99999}
        jl_sort_rules = {'RE': 1, 'FM': 2, 'RFM': 3, 'jl': 99999}
        wl_data = sorted(wl_data, key=lambda d: wl_sort_rules.get(d['name'], 999))
        jl_data = sorted(jl_data, key=lambda d: jl_sort_rules.get(d['name'], 999))

        # 计算段数
        month_total_dict = dict(qs.values(
            'factory_date__day'
        ).annotate(weight=Sum(
            'plan_weight', output_field=DecimalField()) / 1000
                   ).values_list('factory_date__day', 'weight'))
        fm_total_dict = dict(qs.filter(product_no__icontains='-FM-').values('factory_date__day').annotate(
            weight=Sum('plan_weight', output_field=DecimalField()) / 1000).values_list('factory_date__day', 'weight'))

        queryset_190e = Equip190EWeight.objects.exclude(
            setup__specification__in=('洗车胶', 'XCJ', 'WUMING')).filter(
            factory_date__gte=st,
            factory_date__lte=et)
        total_queryset_190e_dict = dict(queryset_190e.values(
            'factory_date__day').annotate(
            sum_weight=Sum(F('setup__weight') * F('qty') / 1000, output_field=DecimalField())).values_list(
            'factory_date__day', 'sum_weight'))
        total_manual_input_trains = ManualInputTrains.objects.exclude(
            Q(product_no__icontains='XCJ') |
            Q(product_no__icontains='洗车胶') |
            Q(product_no__icontains='-WUMING-')
        ).filter(factory_date__gte=st, factory_date__lte=et)
        total_manual_input_trains_dict = dict(total_manual_input_trains.values(
            'factory_date__day'
        ).annotate(
            weight=Sum(F('weight') * F('actual_trains') / 1000, output_field=DecimalField())
        ).values_list('factory_date__day', 'weight'))

        queryset_190e = Equip190EWeight.objects.exclude(
            setup__specification__in=('洗车胶', 'XCJ', 'WUMING')).filter(
            factory_date__gte=st, factory_date__lte=et)
        equip_190e_weight = queryset_190e.filter(setup__state='FM').values('factory_date__day').annotate(
            sum_weight=Sum(F('setup__weight') * F('qty') / 1000, output_field=DecimalField()))
        fm_queryset_190e_dict = dict(equip_190e_weight.values_list('factory_date__day', 'sum_weight'))

        out_queryset = OuterMaterial.objects.filter(
            factory_date__gte=st,
            factory_date__lte=et).values('factory_date__day').annotate(
            weight=Sum('weight', output_field=DecimalField()))
        out_queryset_dict = dict(out_queryset.values_list('factory_date__day', 'weight'))

        total_manual_input_trains = ManualInputTrains.objects.exclude(
            Q(product_no__icontains='XCJ') |
            Q(product_no__icontains='洗车胶') |
            Q(product_no__icontains='-WUMING-')
        ).filter(factory_date__gte=st, factory_date__lte=et)

        total_manual_input_trains_final_dict = dict(
            total_manual_input_trains.filter(
                product_no__icontains='-FM-'
            ).values('factory_date__day').annotate(
                weight=Sum(F('weight') * F('actual_trains')/1000, output_field=DecimalField())
            ).values_list('factory_date__day', 'weight'))

        total_weight = sum(list(month_total_dict.values())) \
                       + sum(list(total_queryset_190e_dict.values())) \
                       + sum(list(total_manual_input_trains_dict.values()))
        total_fm_weight = sum(list(fm_total_dict.values())) \
                          + sum(list(fm_queryset_190e_dict.values())) \
                          + sum(list(out_queryset_dict.values())) \
                          + sum(list(total_manual_input_trains_final_dict.values()))

        ds = "" if not total_fm_weight or not total_weight else round(total_weight / total_fm_weight, 2)

        return Response({'result': queryset, 'wl': wl_data, 'jl': jl_data, 'ds': ds})


@method_decorator([api_recorder], name="dispatch")
class MonthlyOutputStatisticsReportDetail(APIView):
    permission_classes = (IsAuthenticated,)

    def get(self, request):
        st = self.request.query_params.get('st')
        et = self.request.query_params.get('et')
        stage = self.request.query_params.get('state')
        equip_no = self.request.query_params.get('equip')
        product_no = self.request.query_params.get('space')
        if not all([st, et]):
            raise ValidationError('请选择日期范围查询！')
        try:
            e_time = datetime.datetime.strptime(et, '%Y-%m-%d')
            s_time = datetime.datetime.strptime(st, '%Y-%m-%d')
        except Exception:
            raise ValidationError('日期错误！')
        diff = e_time - s_time
        if diff.days > 31:
            raise ValidationError('搜索日期跨度不得超过一个月！')
        if not stage:
            raise ValidationError('请选择段次！')
        filter_kwargs = {'factory_date__gte': st, 'factory_date__lte': et, 'product_no__icontains': '-{}-'.format(stage)}
        query_set = TrainsFeedbacks.objects.exclude(
            Q(product_no__icontains='XCJ') |
            Q(product_no__icontains='洗车胶') |
            Q(operation_user='Mixer2') |
            Q(product_no__icontains='WUMING')
        ).filter(**filter_kwargs)

        total_weight = query_set.aggregate(tw=Sum('plan_weight')/1000)['tw']

        if equip_no:
            query_set = query_set.filter(equip_no=equip_no)
        if product_no:
            query_set = query_set.filter(product_no__icontains='-{}-'.format(product_no))

        equip_group_data = query_set.values('product_no', 'equip_no', 'factory_date').annotate(
            total_weight=Sum('plan_weight')/1000, total_trains=Count('id')).order_by('factory_date')
        ret = {}
        for item in equip_group_data:
            try:
                product_no = item['product_no'].split('-')[2]
            except Exception:
                continue
            equip_no = item['equip_no']
            factory_date = item['factory_date'].strftime("%Y-%m-%d")
            weight = item['total_weight']
            trains = item['total_trains']

            key = '{}-{}-{}'.format(equip_no, product_no, factory_date)
            if key in ret:
                ret[key]['value'] += trains
                ret[key]['weight'] += weight
                ret[key]['ratio'] = round(ret[key]['weight'] / total_weight * 100, 2)
            else:
                ret[key] = {'equip_no': equip_no,
                            'ratio': round(weight/total_weight*100, 2),
                            'space': product_no,
                            'time': factory_date,
                            'value': trains,
                            'weight': weight}
        return Response({'result': ret.values()})


@method_decorator([api_recorder], name="dispatch")
class DailyProductionCompletionReport(APIView):
    permission_classes = (IsAuthenticated,)
    pagination_class = None

    def export(self, month, days, data1, data2):
        response = HttpResponse(content_type='application/vnd.ms-excel')
        filename = '月产量完成报表'
        response['Content-Disposition'] = u'attachment;filename= ' + filename.encode('gbk').decode(
            'ISO-8859-1') + '.xls'
        # 创建一个文件对象
        wb = xlwt.Workbook(encoding='utf8')
        # 创建一个sheet对象
        sheet1 = wb.add_sheet('月产量完成', cell_overwrite_ok=True)
        sheet2 = wb.add_sheet('班次车数', cell_overwrite_ok=True)
        sheet3 = wb.add_sheet('月产量吨位', cell_overwrite_ok=True)
        title1 = ['项目/日期', '汇总', '平均']
        title2 = ['序号', '规格', '段数', '机台', '型号', '班别']
        title3 = ['序号', '规格', '段数', '机台', '型号', '班别']
        for day in range(1, days + 1):
            title1.append(f'{day}日')
            title2.append(f'{month}/{day}')
            title3.append(f'{month}/{day}')
        # title1.append('汇总')
        title2.append('汇总')
        title3.append('汇总')

        style = xlwt.XFStyle()
        style.alignment.wrap = 1

        # 写入文件标题
        for col_num in range(len(title1)):
            sheet1.write(0, col_num, title1[col_num])
        for col_num in range(len(title2)):
            sheet2.write(0, col_num, title2[col_num])
        for col_num in range(len(title3)):
            sheet3.write(0, col_num, title3[col_num])
        # 写入数据
        for index, dic in enumerate(data1):
            for key, value in dic.items():
                if key == 'name':
                    sheet1.write(index + 1, 0, dic['name'])
                elif key == 'weight':
                    sheet1.write(index + 1, 1, dic['weight'])
                elif key == 'avg':
                    sheet1.write(index + 1, 2, dic['avg'])
                else:
                    sheet1.write(index + 1, int(key[:-1])+2, value)
        for index, dic in enumerate(data2):
            sheet2.write(index + 1, 0, index + 1)
            sheet3.write(index + 1, 0, index + 1)
            for key, value in dic.items():
                if key == '规格':
                    sheet2.write(index + 1, 1, value)
                    sheet3.write(index + 1, 1, value)
                elif key == '段数':
                    sheet2.write(index + 1, 2, value)
                    sheet3.write(index + 1, 2, value)
                elif key == '机台':
                    sheet2.write(index + 1, 3, value)
                    sheet3.write(index + 1, 3, value)
                elif key == '机型':
                    sheet2.write(index + 1, 4, value)
                    sheet3.write(index + 1, 4, value)
                elif key == '班别':
                    sheet2.write(index + 1, 5, value)
                    sheet3.write(index + 1, 5, value)
                elif key == '汇总_qty':
                    sheet2.write(index + 1, len(title2) - 1, value)
                elif key == '汇总_weight':
                    sheet3.write(index + 1, len(title3) - 1, value)
                else:
                    day, s = key.split('_')
                    if s == 'qty':
                        sheet2.write(index + 1, int(day) + 5, value)
                    else:
                        sheet3.write(index + 1, int(day) + 5, value)
        # 写出到IO
        output = BytesIO()
        wb.save(output)
        # 重新定位到开始
        output.seek(0)
        response.write(output.getvalue())
        return response

    def get(self, request):
        params = self.request.query_params
        date = params.get('date')
        td_flag = params.get('td_flag', 'N')  # 默认
        if not date:
            raise ValidationError('请选择月份！')
        try:
            year = int(date.split('-')[0])
            month = int(date.split('-')[1])
        except Exception:
            raise ValidationError('请输入正确的月份!')
        days = calendar.monthrange(year, month)[1]
        exclude_today_flat = False  # 总计是否去掉当日产量
        if td_flag != 'Y' and month == datetime.datetime.now().month and year == datetime.datetime.now().year:
            exclude_today_flat = True
        now_day = datetime.datetime.now().day
        results = {
            'name_1': {'name': '混炼胶实际完成(吨)', 'weight': 0},  # CMB HMB 1MB~4MB
            'name_2': {'name': '终炼胶实际完成(吨)', 'weight': 0},  # FM + 190E终炼产量
            'name_3': {'name': '外发无硫料(吨)', 'weight': 0},  # 人工输入
            'name_4': {'name': '实际完成数-1(吨)', 'weight': 0},  # 190E终炼产量 + FM + 外发无硫料*0.7
            'name_5': {'name': '实际完成数-2(吨)', 'weight': 0},  # 190E终炼产量 + FM + 外发无硫料
            'name_6': {'name': '实际生产工作日数', 'weight': 0},
            'name_10': {'name': '190E实际生产工作日数', 'weight': 0},
            'name_7': {'name': '日均完成量-1（吨）', 'weight': 0},
            'name_8': {'name': '日均完成量-2（吨）', 'weight': 0},
            'name_9': {'name': '实际生产机台数', 'weight': 0},
            'name_11': {'name': '单机台效率-1（吨/台）', 'weight': 0},
            'name_12': {'name': '单机台效率-2（吨/台）', 'weight': 0},
            'name_13': {'name': '每日段数', 'weight': 0},
            'name_14': {'name': '吨耗时（分钟/吨）', 'weight': 0},
            'name_15': {'name': '吨胶密炼耗能（KWH/吨）', 'weight': 0}
        }

        # 除去洗车胶的总车次报表数据
        total_queryset = TrainsFeedbacks.objects.exclude(Q(product_no__icontains='XCJ') |
                                                         Q(product_no__icontains='洗车胶') |
                                                         Q(operation_user='Mixer2') |
                                                         Q(product_no__icontains='WUMING')
                                                         ).filter(factory_date__year=year, factory_date__month=month)
        # 按日期分组，计算每日总生产重量
        month_total_dict = dict(total_queryset.values(
                'factory_date__day'
            ).annotate(weight=Sum(
                'plan_weight', output_field=DecimalField())/1000
            ).values_list('factory_date__day', 'weight'))

        # 当月混炼实际完成吨  CMB HMB 1MB~4MB
        queryset1 = total_queryset.filter(Q(product_no__icontains='-CMB-') |
                                          Q(product_no__icontains='-HMB-') |
                                          Q(product_no__icontains='-1MB-') |
                                          Q(product_no__icontains='-2MB-') |
                                          Q(product_no__icontains='-3MB-') |
                                          Q(product_no__icontains='-4MB-'))
        mix_queryset = queryset1.values('factory_date__day').annotate(weight=Sum('plan_weight', output_field=DecimalField())/1000)

        # 当月终炼实际完成（FM段次）  FM
        queryset2 = total_queryset.filter(product_no__icontains='-FM-')
        fin_queryset = queryset2.values('factory_date__day').annotate(weight=Sum('plan_weight', output_field=DecimalField())/1000)
        fm_total_dict = dict(fin_queryset.values_list('factory_date__day', 'weight'))

        # 当月190E除去洗车胶外的所有产量
        queryset_190e = Equip190EWeight.objects.exclude(
            setup__specification__in=('洗车胶', 'XCJ', 'WUMING')).filter(
            factory_date__year=year, factory_date__month=month)
        # 按日期分组，计算190E每日总生产重量
        total_queryset_190e_dict = dict(queryset_190e.values(
            'factory_date__day').annotate(
                sum_weight=Sum(F('setup__weight')*F('qty')/1000, output_field=DecimalField())).values_list('factory_date__day', 'sum_weight'))
        # 按日期分组，计算190E每日终炼产量  FM
        equip_190e_weight = queryset_190e.filter(setup__state='FM').values('factory_date__day').annotate(
                sum_weight=Sum(F('setup__weight')*F('qty')/1000, output_field=DecimalField()))
        # 按日期分组，计算190E每日混炼产量  CMB HMB 1MB~4MB
        equip_190e_mixin_weight = queryset_190e.filter(
            setup__state__in=('1MB', '2MB', '3MB', '4MB', 'CMB', 'HMB')
        ).values('factory_date__day').annotate(sum_weight=Sum(F('setup__weight') * F('qty') / 1000, output_field=DecimalField()))
        fm_queryset_190e_dict = dict(equip_190e_weight.values_list('factory_date__day', 'sum_weight'))

        # 按日期分组，计算每日手动总生产重量
        total_manual_input_trains = ManualInputTrains.objects.exclude(
            Q(product_no__icontains='XCJ') |
            Q(product_no__icontains='洗车胶') |
            Q(product_no__icontains='-WUMING-')
        ).filter(factory_date__year=year, factory_date__month=month)
        total_manual_input_trains_dict = dict(total_manual_input_trains.values(
                'factory_date__day'
            ).annotate(
            weight=Sum(F('weight') * F('actual_trains')/1000, output_field=DecimalField())
        ).values_list('factory_date__day', 'weight'))

        # 当月每日手动总生产混炼实际完成吨  CMB HMB 1MB~4MB
        total_manual_input_trains_mixin_dict = dict(
            total_manual_input_trains.filter(
                Q(product_no__icontains='-CMB-') |
                Q(product_no__icontains='-HMB-') |
                Q(product_no__icontains='-1MB-') |
                Q(product_no__icontains='-2MB-') |
                Q(product_no__icontains='-3MB-') |
                Q(product_no__icontains='-4MB-')
            ).values('factory_date__day').annotate(
                weight=Sum(F('weight') * F('actual_trains')/1000, output_field=DecimalField())
            ).values_list('factory_date__day', 'weight'))

        # 当月每日手动总生产终炼实际完成吨  FM
        total_manual_input_trains_final_dict = dict(
            total_manual_input_trains.filter(
                product_no__icontains='-FM-'
            ).values('factory_date__day').annotate(
                weight=Sum(F('weight') * F('actual_trains')/1000, output_field=DecimalField())
            ).values_list('factory_date__day', 'weight'))

        # 190E产量曲线 加硫、无硫、所有、终炼
        data_190e = {
            'dates': [i for i in range(1, days+1)],
            'wl': [0] * days,
            'jl': [0] * days,
            'total': [total_queryset_190e_dict.get(i, 0) for i in range(1, days+1)],
            'fm': [fm_queryset_190e_dict.get(i, 0) for i in range(1, days+1)]
        }

        # 每日实际工作天数
        actual_working_day_dict = dict(
            ActualWorkingDay.objects.filter(
                factory_date__year=year,
                factory_date__month=month,
                num__gt=0).values_list('factory_date__day', 'num'))

        # 190E每日实际工作天数
        actual_working_day_190e_dict = dict(
            ActualWorkingDay190E.objects.filter(
                factory_date__year=year,
                factory_date__month=month,
                num__gt=0).values_list('factory_date__day', 'num'))

        # 每日实际工作机台数
        actual_working_equip_dict = dict(
            ActualWorkingEquip.objects.filter(
                factory_date__year=year,
                factory_date__month=month,
                num__gt=0).values_list('factory_date__day', 'num'))

        # 外发无硫料（吨）
        out_queryset = OuterMaterial.objects.filter(
            factory_date__year=year,
            factory_date__month=month).values('factory_date__day').annotate(weight=Sum('weight', output_field=DecimalField()))
        out_queryset_dict = dict(out_queryset.values_list('factory_date__day', 'weight'))

        time_consume_dict = dict(TrainsFeedbacks.objects.filter(
            factory_date__year=year,
            factory_date__month=month).values('factory_date').annotate(
            time_consume=OSum((F('end_time') - F('begin_time')))).values_list('factory_date__day', 'time_consume'))
        energy_consume_data = TrainsFeedbacks.objects.filter(
            factory_date__year=year,
            factory_date__month=month).extra().values('factory_date__day', 'equip_no').annotate(
            energy_consume=Sum('evacuation_energy'),
            sum_weight=Avg('plan_weight')
        )
        energy_consume_dict = {}
        for e in energy_consume_data:
            energy_consume = 0 if not e['energy_consume'] else e['energy_consume']
            equip_no = e['equip_no']
            if equip_no == 'Z01':
                evacuation_energy = energy_consume / 10
            elif equip_no == 'Z02':
                evacuation_energy = energy_consume / 0.6
            elif equip_no == 'Z04':
                evacuation_energy = energy_consume * 0.28 * float(e['sum_weight']) / 1000
            elif equip_no == 'Z12':
                evacuation_energy = energy_consume / 5.3
            elif equip_no == 'Z13':
                evacuation_energy = energy_consume / 31.7
            else:
                evacuation_energy = energy_consume
            energy_consume_dict[e['factory_date__day']] = energy_consume_dict.get(e['factory_date__day'], 0) + evacuation_energy
        for item in mix_queryset:
            mixin_weight = round(item['weight'], 2)
            if not(item['factory_date__day'] == now_day and exclude_today_flat):
                results['name_1']['weight'] += mixin_weight
            results['name_1'][f"{item['factory_date__day']}日"] = mixin_weight
        for item in equip_190e_mixin_weight:
            weight = round(item['sum_weight'], 2)
            if not(item['factory_date__day'] == now_day and exclude_today_flat):
                results['name_1']['weight'] += weight
            results['name_1'][f"{item['factory_date__day']}日"] = results['name_1'].get(f"{item['factory_date__day']}日", 0) + weight
            data_190e['wl'][item['factory_date__day']-1] = weight
        for d, v in total_manual_input_trains_mixin_dict.items():
            v = round(v, 2)
            if not(d == now_day and exclude_today_flat):
                results['name_1']['weight'] += v
            results['name_1'][f"{d}日"] = results['name_1'].get(f"{d}日", 0) + v
        for item in equip_190e_weight:
            weight = round(item['sum_weight'], 2)
            results['name_2'][f"{item['factory_date__day']}日"] = results['name_2'].get(f"{item['factory_date__day']}日", 0) + weight
            results['name_4'][f"{item['factory_date__day']}日"] = results['name_4'].get(f"{item['factory_date__day']}日", 0) + weight
            results['name_5'][f"{item['factory_date__day']}日"] = results['name_5'].get(f"{item['factory_date__day']}日", 0) + weight
            data_190e['jl'][item['factory_date__day']-1] = weight
            if not(item['factory_date__day'] == now_day and exclude_today_flat):
                results['name_2']['weight'] += weight
                results['name_4']['weight'] += weight
                results['name_5']['weight'] += weight
        for item in fin_queryset:
            final_weight = round(item['weight'], 2)
            if not(item['factory_date__day'] == now_day and exclude_today_flat):
                results['name_2']['weight'] += final_weight
                results['name_4']['weight'] += final_weight
                results['name_5']['weight'] += final_weight
            results['name_2'][f"{item['factory_date__day']}日"] = results['name_2'].get(f"{item['factory_date__day']}日", 0) + final_weight
            results['name_4'][f"{item['factory_date__day']}日"] = results['name_4'].get(f"{item['factory_date__day']}日", 0) + final_weight
            results['name_5'][f"{item['factory_date__day']}日"] = results['name_5'].get(f"{item['factory_date__day']}日", 0) + final_weight
        for d, v in total_manual_input_trains_final_dict.items():
            v = round(v, 2)
            if not(d == now_day and exclude_today_flat):
                results['name_2']['weight'] += v
                results['name_4']['weight'] += v
                results['name_5']['weight'] += v
            results['name_2'][f"{d}日"] = results['name_2'].get(f"{d}日", 0) + v
            results['name_4'][f"{d}日"] = results['name_4'].get(f"{d}日", 0) + v
            results['name_5'][f"{d}日"] = results['name_5'].get(f"{d}日", 0) + v
        for item in out_queryset:
            results['name_3'][f"{item['factory_date__day']}日"] = round(item['weight'], 2)
            results['name_4'][f"{item['factory_date__day']}日"] = results['name_4'].get(f"{item['factory_date__day']}日", 0) + round((item['weight']) * decimal.Decimal(0.7), 2)
            results['name_5'][f"{item['factory_date__day']}日"] = results['name_5'].get(f"{item['factory_date__day']}日", 0) + round(item['weight'], 2)
            if not(item['factory_date__day'] == now_day and exclude_today_flat):
                results['name_3']['weight'] += round(item['weight'], 2)
                results['name_4']['weight'] += round((item['weight']) * decimal.Decimal(0.7), 2)
                results['name_5']['weight'] += round(item['weight'], 2)
        for k, v in actual_working_day_dict.items():
            v = round(v, 2)
            results['name_6'][f"{k}日"] = v
            if not(k == now_day and exclude_today_flat):
                results['name_6']['weight'] = round(results['name_6']['weight'] + v, 2)
        for k, v in actual_working_equip_dict.items():
            v = round(v, 2)
            results['name_9'][f"{k}日"] = v
            if not(k == now_day and exclude_today_flat):
                results['name_9']['weight'] = round(results['name_9']['weight'] + v, 2)
        for k, v in actual_working_day_190e_dict.items():
            v = round(v, 2)
            results['name_10'][f"{k}日"] = v
            if not(k == now_day and exclude_today_flat):
                results['name_10']['weight'] = round(results['name_10']['weight'] + v, 2)
        if len(results['name_6']) - 2 != 0:
            for key, value in results['name_4'].items():
                if key[0].isdigit():
                    if results['name_6'].get(key):
                        results['name_7'][key] = round(results['name_4'][key] / decimal.Decimal(results['name_6'][key]), 2)
                        results['name_8'][key] = round(results['name_5'][key] / decimal.Decimal(results['name_6'][key]), 2)
            results['name_7']['weight'] = 0 if results['name_6']['weight'] == 0 else round(results['name_4']['weight'] / decimal.Decimal(results['name_6']['weight']), 2)
            results['name_8']['weight'] = 0 if results['name_6']['weight'] == 0 else round(results['name_5']['weight'] / decimal.Decimal(results['name_6']['weight']), 2)
        if exclude_today_flat:
            # 190E当月总天数
            month_working_days_190e = float(sum(actual_working_day_190e_dict.values()) -
                                            actual_working_day_190e_dict.get(now_day, 0))
            # 当月总天数
            month_working_days = float(sum(actual_working_day_dict.values()) -
                                            actual_working_day_dict.get(now_day, 0))
            # 当月总生产机台数
            actual_working_equips = float(sum(actual_working_equip_dict.values()) -
                                            actual_working_equip_dict.get(now_day, 0))
        else:
            # 190E当月总天数
            month_working_days_190e = float(sum(actual_working_day_190e_dict.values()))
            # 当月总天数
            month_working_days = float(sum(actual_working_day_dict.values()))
            # 当月总生产机台数
            actual_working_equips = float(sum(actual_working_equip_dict.values()))

        # 计算每日单机台效率-1
        for k, v in results['name_4'].items():
            if k[0].isdigit():
                ds = actual_working_equip_dict.get(int(k[:-1]))
                if not ds:
                    continue
                value = round(float(v) / ds, 2)
                results['name_11'][k] = value
        results['name_11']['weight'] = 0 if not actual_working_equips else round(float(results['name_4']['weight']) / actual_working_equips, 2)

        # 计算每日单机台效率-2
        for k, v in results['name_5'].items():
            if k[0].isdigit():
                ds = actual_working_equip_dict.get(int(k[:-1]))
                if ds:
                    value = round(float(v) / ds, 2)
                    results['name_12'][k] = value
                time_consume = time_consume_dict.get(int(k[:-1]))
                if time_consume:
                    time_value = round(float(time_consume.total_seconds()) / 60 / float(v), 2)
                    results['name_14'][k] = time_value
                    if not (k == now_day and exclude_today_flat):
                        results['name_14']['weight'] = round(results['name_14']['weight'] + time_value, 2)
                energy_consume = energy_consume_dict.get(int(k[:-1]))
                if energy_consume:
                    energy_value = round(float(energy_consume) / float(v), 2)
                    results['name_15'][k] = energy_value
                    if not (k == now_day and exclude_today_flat):
                        results['name_15']['weight'] = round(results['name_15']['weight'] + energy_value, 2)
        results['name_12']['weight'] = 0 if not actual_working_equips else round(float(results['name_5']['weight']) / actual_working_equips, 2)

        # 计算平均值
        for item in results.values():
            if item['name'] == '190E实际生产工作日数':
                l1 = len(actual_working_day_190e_dict)
                if exclude_today_flat:
                    if actual_working_day_190e_dict.get(now_day):
                        l1 -= 1
                item['avg'] = "" if not month_working_days_190e else round(
                    float(item['weight']) / l1, 2)
            elif item['name'] == '实际生产机台数':
                l2 = len(actual_working_equip_dict)
                if exclude_today_flat:
                    if actual_working_equip_dict.get(now_day):
                        l2 -= 1
                item['avg'] = "" if not actual_working_equips else round(
                    float(item['weight']) / l2, 2)
            elif item['name'] == '实际生产工作日数':
                l3 = len(actual_working_day_dict)
                if exclude_today_flat:
                    if actual_working_day_dict.get(now_day):
                        l3 -= 1
                item['avg'] = "" if not month_working_days else round(
                    float(item['weight']) / l3, 2)
            elif item['name'] in ('单机台效率-1（吨/台）', '单机台效率-2（吨/台）', '每日段数'):
                item['avg'] = item['weight']
            elif item['name'] == '吨耗时（分钟/吨）':
                length = len(results['name_14']) - 2
                item['avg'] = '' if not length else round(item['weight'] / length, 2)
            elif item['name'] == '吨胶密炼耗能（KWH/吨）':
                length = len(results['name_15']) - 2
                item['avg'] = '' if not length else round(item['weight'] / length, 2)
            else:
                item['avg'] = "" if month_working_days <= 0 else round(float(item['weight']) / month_working_days, 2)
        try:
            results['name_11']['avg'] = round(results['name_4']['avg'] / results['name_9']['avg'], 2)
        except Exception:
            results['name_11']['avg'] = ''
        try:
            results['name_12']['avg'] = round(results['name_5']['avg'] / results['name_9']['avg'], 2)
        except Exception:
            results['name_12']['avg'] = ''

        # 计算190E每日段数
        cnt = 0
        sum_ds = 0
        for idx in range(1, days):
            fm = data_190e['fm'][idx]
            total = data_190e['total'][idx]
            if not total:
                continue
            try:
                ds = total/fm
            except Exception:
                ds = 0
            sum_ds += ds
            cnt += 1

        if exclude_today_flat:
            total_190e_jl = sum(data_190e['jl']) - data_190e['jl'][now_day-1]
            # length1 = len([i for i in data_190e['jl'] if i > 0])
            # if data_190e['jl'][now_day-1]:
            #     length1 -= 1
            total_190e_wl = sum(data_190e['wl']) - data_190e['wl'][now_day-1]
            # length2 = len([i for i in data_190e['wl'] if i > 0])
            # if data_190e['wl'][now_day-1]:
            #     length2 -= 1
            total_190e_fm = sum(data_190e['fm']) - data_190e['fm'][now_day-1]
            total_190e_total = sum(data_190e['total']) - data_190e['total'][now_day-1]
        else:
            total_190e_jl = sum(data_190e['jl'])
            # length1 = len([i for i in data_190e['jl'] if i > 0])
            total_190e_wl = sum(data_190e['wl'])
            # length2 = len([i for i in data_190e['wl'] if i > 0])
            total_190e_fm = sum(data_190e['fm'])
            total_190e_total = sum(data_190e['total'])

        avg_190e = {'jl': 0 if not month_working_days_190e else round(float(total_190e_jl) / float(month_working_days_190e), 2),
                    'wl': 0 if not month_working_days_190e else round(float(total_190e_wl) / float(month_working_days_190e), 2),
                    'ds': 0 if not total_190e_fm else round(total_190e_total / total_190e_fm, 2)}

        # 计算每日总产量段数
        cnt2 = 0
        sum_ds2 = 0
        for t_day in range(1, days+1):
            t_weight = float(month_total_dict.get(t_day, 0)) \
                       + float(total_queryset_190e_dict.get(t_day, 0)) \
                       + float(total_manual_input_trains_dict.get(t_day, 0))
            if not t_weight:
                continue
            fm_weight = float(fm_total_dict.get(t_day, 0)) \
                        + float(fm_queryset_190e_dict.get(t_day, 0)) \
                        + float(out_queryset_dict.get(t_day, 0)) \
                        + float(total_manual_input_trains_final_dict.get(t_day, 0))
            try:
                ds2 = t_weight / fm_weight
            except Exception:
                ds2 = 0
            if not (exclude_today_flat and t_day == now_day):
                sum_ds2 += ds2
                cnt2 += 1
            results['name_13']['{}日'.format(str(t_day))] = round(ds2, 2)

        # 计算段数总计和平均值
        if exclude_today_flat:
            total_weight = sum(list(month_total_dict.values())) \
                           + sum(list(total_queryset_190e_dict.values())) \
                           + sum(list(total_manual_input_trains_dict.values())) \
                           - month_total_dict.get(now_day, 0)\
                           - total_queryset_190e_dict.get(now_day, 0)\
                           - total_manual_input_trains_dict.get(now_day, 0)
            total_fm_weight = sum(list(fm_total_dict.values())) \
                              + sum(list(fm_queryset_190e_dict.values())) \
                              + sum(list(out_queryset_dict.values())) \
                              + sum(list(total_manual_input_trains_final_dict.values())) \
                              - fm_total_dict.get(now_day, 0)\
                              - fm_queryset_190e_dict.get(now_day, 0)\
                              - out_queryset_dict.get(now_day, 0)\
                              - total_manual_input_trains_final_dict.get(now_day, 0)
        else:
            total_weight = sum(list(month_total_dict.values())) \
                           + sum(list(total_queryset_190e_dict.values())) \
                           + sum(list(total_manual_input_trains_dict.values()))
            total_fm_weight = sum(list(fm_total_dict.values())) \
                              + sum(list(fm_queryset_190e_dict.values())) \
                              + sum(list(out_queryset_dict.values())) \
                              + sum(list(total_manual_input_trains_final_dict.values()))
        results['name_13']['weight'] = "" if not total_fm_weight or not total_weight else round(total_weight/total_fm_weight, 2)
        results['name_13']['avg'] = 0 if not cnt2 else round(sum_ds2/cnt2, 2)
        results['name_14']['weight'] = '' if not results['name_5']['weight'] else round(sum([i.total_seconds() / 60 for i in time_consume_dict.values()]) / float(results['name_5']['weight']), 2)
        results['name_15']['weight'] = '' if not results['name_5']['weight'] else round(sum(energy_consume_dict.values()) / float(results['name_5']['weight']), 2)
        # 总产量加硫、无硫、段数平均值
        avg_results = {'jl': results['name_2']['avg'],
                       'wl': results['name_1']['avg'],
                       'ds': results['name_13']['weight']}
        if self.request.query_params.get('export', None):
            results2 = {}
            equip_query = Equip.objects.filter(
                category__equip_type__global_name='密炼设备').values('equip_no', 'category__category_name')
            equip_dic = {item['equip_no']: item['category__category_name'] for item in equip_query}
            data2 = TrainsFeedbacks.objects.exclude(operation_user='Mixer2').filter(
                factory_date__year=year,
                factory_date__month=month).values(
                'factory_date__day', 'product_no', 'equip_no', 'classes'
            ).annotate(actual_trains=Count('id'), weight=Sum('plan_weight')/1000).order_by('-classes')
            for item in data2:
                try:
                    state = item['product_no'].split("-")[1]
                    space = item['product_no'].split("-")[2]
                    key = f'{space}_{state}_{item["equip_no"]}_{item["classes"]}'
                except:
                    continue
                weight = round(float(item['weight']), 3)
                if results2.get(key):
                    results2[key][f'{item["factory_date__day"]}_qty'] = results2[key].get(
                        f'{item["factory_date__day"]}_qty', 0) + item['actual_trains']
                    results2[key][f'{item["factory_date__day"]}_weight'] = round(
                        results2[key].get(f'{item["factory_date__day"]}_weight', 0) + weight, 3)
                    results2[key]['汇总_weight'] += weight
                    results2[key]['汇总_qty'] += item['actual_trains']
                else:
                    results2[key] = {'规格': space, '段数': state, '机台': item["equip_no"],
                                     '机型': equip_dic[item["equip_no"]], '班别': item['classes'],
                                     '汇总_qty': item['actual_trains'], '汇总_weight': weight,
                                     f'{item["factory_date__day"]}_qty': item['actual_trains'],
                                     f'{item["factory_date__day"]}_weight': weight}
            equip_190e_result = {}
            equip_190e_queryset = Equip190EWeight.objects.filter(
                factory_date__year=year,
                factory_date__month=month
            ).values('factory_date__day', 'setup__specification', 'setup__state', 'classes').annotate(
                sum_weight=Sum(F('setup__weight')*F('qty')/1000, output_field=DecimalField()),
                total_trains=Sum('qty')).order_by('-classes')
            for item in equip_190e_queryset:
                weight = item['sum_weight']
                product_no = item['setup__specification']
                stage = item['setup__state']
                classes = item['classes']
                key = '{}_{}_{}'.format(
                    product_no, stage, classes
                )
                if equip_190e_result.get(key):
                    equip_190e_result[key][f'{item["factory_date__day"]}_qty'] = equip_190e_result[key].get(
                        f'{item["factory_date__day"]}_qty', 0) + item['total_trains']
                    equip_190e_result[key][f'{item["factory_date__day"]}_weight'] = round(
                        equip_190e_result[key].get(f'{item["factory_date__day"]}_weight', 0) + weight, 3)
                    equip_190e_result[key]['汇总_weight'] += weight
                    equip_190e_result[key]['汇总_qty'] += item['total_trains']
                else:
                    equip_190e_result[key] = {'规格': product_no, '段数': stage, '机台': '190E',
                                              '机型': '190E', '班别': classes,
                                              '汇总_qty': item['total_trains'], '汇总_weight': weight,
                                               f'{item["factory_date__day"]}_qty': item['total_trains'],
                                               f'{item["factory_date__day"]}_weight': weight}

            manual_input_queryset = ManualInputTrains.objects.filter(
                factory_date__year=year,
                factory_date__month=month
            ).values('factory_date__day',
                     'equip_no',
                     'product_no',
                     'classes'
                     ).annotate(
                weight=Sum(F('weight') * F('actual_trains') / 1000, output_field=DecimalField()),
                total_trains=Sum('actual_trains')
            ).order_by('-classes')
            for item in manual_input_queryset:
                try:
                    state = item['product_no'].split("-")[1]
                    space = item['product_no'].split("-")[2]
                    key = f'{space}_{state}_{item["equip_no"]}_{item["classes"]}'
                except:
                    continue
                weight = round(float(item['weight']), 3)
                if results2.get(key):
                    results2[key][f'{item["factory_date__day"]}_qty'] = results2[key].get(
                        f'{item["factory_date__day"]}_qty', 0) + item['total_trains']
                    results2[key][f'{item["factory_date__day"]}_weight'] = round(
                        results2[key].get(f'{item["factory_date__day"]}_weight', 0) + weight, 3)
                    results2[key]['汇总_weight'] += weight
                    results2[key]['汇总_qty'] += item['total_trains']
                else:
                    results2[key] = {'规格': space, '段数': state, '机台': item["equip_no"],
                                     '机型': equip_dic[item["equip_no"]], '班别': item['classes'],
                                     '汇总_qty': item['total_trains'], '汇总_weight': weight,
                                     f'{item["factory_date__day"]}_qty': item['total_trains'],
                                     f'{item["factory_date__day"]}_weight': weight}
            excel_sheet_1_data = list(results.values())
            excel_sheet_23_data = sorted(
                list(results2.values()), key=itemgetter('机台', '规格', '段数')
            ) + sorted(
                list(equip_190e_result.values()), key=itemgetter('机台', '规格', '段数')
            )
            return self.export(month, days, excel_sheet_1_data, excel_sheet_23_data)

        return Response({'results': results.values(),
                         'data_190e': data_190e,
                         'avg_190e': avg_190e,
                         'avg_results': avg_results
                         })

    def post(self, request):
        # 190E机台产量录入
        factory_date = self.request.data.get('factory_date', None)
        classes = self.request.data.get('classes', None)
        data = self.request.data.get('data', [])
        date = self.request.data.get('date')
        outer_data = self.request.data.get('outer_data', [])  # 外发无硫料
        working_data = self.request.data.get('working_data', [])  # 实际工作日期
        working_190e_data = self.request.data.get('working_190e_data', [])  # 实际工作日期
        equip_data = self.request.data.get('equip_data', [])  # 实际工作日期
        if data:
            serializer = Equip190EWeightSerializer(data=data, many=True)
            serializer.is_valid(raise_exception=True)
            Equip190EWeight.objects.filter(factory_date=factory_date, classes=classes).delete()
            for item in serializer.validated_data:
                Equip190EWeight.objects.create(
                    **{'setup': item['setup'],
                       'factory_date': factory_date,
                        'classes': classes,
                        'qty': item['qty']})
        if not data:
            Equip190EWeight.objects.filter(factory_date=factory_date, classes=classes).delete()

        if date:
            date_split = date.split('-')
            OuterMaterial.objects.filter(factory_date__year=date_split[0],
                                         factory_date__month=date_split[1]).delete()
            ActualWorkingDay.objects.filter(factory_date__year=date_split[0],
                                            factory_date__month=date_split[1]).delete()
            ActualWorkingEquip.objects.filter(factory_date__year=date_split[0],
                                              factory_date__month=date_split[1]).delete()
            ActualWorkingDay190E.objects.filter(factory_date__year=date_split[0],
                                                factory_date__month=date_split[1]).delete()
            for item in outer_data:
                try:
                    OuterMaterial.objects.update_or_create(
                        defaults={'weight': item['weight']},
                        factory_date=item['factory_date'])
                except Exception:
                    pass
            for item in working_data:
                try:
                    ActualWorkingDay.objects.update_or_create(
                        defaults={'num': item['num']},
                        factory_date=item['factory_date'])
                except Exception:
                    pass
            for item in equip_data:
                try:
                    ActualWorkingEquip.objects.update_or_create(
                        defaults={'num': item['num']},
                        factory_date=item['factory_date'])
                except Exception:
                    pass
            for item in working_190e_data:
                try:
                    ActualWorkingDay190E.objects.update_or_create(
                        defaults={'num': item['num']},
                        factory_date=item['factory_date'])
                except Exception:
                    pass
        return Response('ok')


@method_decorator([api_recorder], name="dispatch")
class MonthlyProductionCompletionReport(APIView):
    permission_classes = (IsAuthenticated,)
    pagination_class = None

    def export(self, month, days, data1, data2):
        response = HttpResponse(content_type='application/vnd.ms-excel')
        filename = '年产量完成报表'
        response['Content-Disposition'] = u'attachment;filename= ' + filename.encode('gbk').decode(
            'ISO-8859-1') + '.xls'
        # 创建一个文件对象
        wb = xlwt.Workbook(encoding='utf8')
        # 创建一个sheet对象
        sheet1 = wb.add_sheet('年产量完成', cell_overwrite_ok=True)
        sheet2 = wb.add_sheet('班次车数', cell_overwrite_ok=True)
        sheet3 = wb.add_sheet('年产量吨位', cell_overwrite_ok=True)
        title1 = ['项目/日期', '汇总', '平均']
        title2 = ['序号', '规格', '段数', '机台', '型号', '班别']
        title3 = ['序号', '规格', '段数', '机台', '型号', '班别']
        for day in range(1, days + 1):
            title1.append(f'{day}月')
            title2.append(f'{month}/{day}')
            title3.append(f'{month}/{day}')
        # title1.append('汇总')
        title2.append('汇总')
        title3.append('汇总')

        style = xlwt.XFStyle()
        style.alignment.wrap = 1

        # 写入文件标题
        for col_num in range(len(title1)):
            sheet1.write(0, col_num, title1[col_num])
        for col_num in range(len(title2)):
            sheet2.write(0, col_num, title2[col_num])
        for col_num in range(len(title3)):
            sheet3.write(0, col_num, title3[col_num])
        # 写入数据
        for index, dic in enumerate(data1):
            for key, value in dic.items():
                if key == 'name':
                    sheet1.write(index + 1, 0, dic['name'])
                elif key == 'weight':
                    sheet1.write(index + 1, 1, dic['weight'])
                elif key == 'avg':
                    sheet1.write(index + 1, 2, dic['avg'])
                else:
                    sheet1.write(index + 1, int(key[:-1])+2, value)
        for index, dic in enumerate(data2):
            sheet2.write(index + 1, 0, index + 1)
            sheet3.write(index + 1, 0, index + 1)
            for key, value in dic.items():
                if key == '规格':
                    sheet2.write(index + 1, 1, value)
                    sheet3.write(index + 1, 1, value)
                elif key == '段数':
                    sheet2.write(index + 1, 2, value)
                    sheet3.write(index + 1, 2, value)
                elif key == '机台':
                    sheet2.write(index + 1, 3, value)
                    sheet3.write(index + 1, 3, value)
                elif key == '机型':
                    sheet2.write(index + 1, 4, value)
                    sheet3.write(index + 1, 4, value)
                elif key == '班别':
                    sheet2.write(index + 1, 5, value)
                    sheet3.write(index + 1, 5, value)
                elif key == '汇总_qty':
                    sheet2.write(index + 1, len(title2) - 1, value)
                elif key == '汇总_weight':
                    sheet3.write(index + 1, len(title3) - 1, value)
                else:
                    day, s = key.split('_')
                    if s == 'qty':
                        sheet2.write(index + 1, int(day) + 5, value)
                    else:
                        sheet3.write(index + 1, int(day) + 5, value)
        # 写出到IO
        output = BytesIO()
        wb.save(output)
        # 重新定位到开始
        output.seek(0)
        response.write(output.getvalue())
        return response

    def get(self, request):
        params = self.request.query_params
        date = params.get('date')
        td_flag = params.get('td_flag', 'N')  # 默认
        if not date:
            raise ValidationError('请选择年份！')
        try:
            year = int(date)
        except Exception:
            raise ValidationError('请输入正确的年份!')
        now_time, days = datetime.datetime.now(), 12
        exclude_today_flat = False  # 总计是否去掉当月产量
        if td_flag != 'Y' and now_time.year == year:
            exclude_today_flat = True
        now_day = now_time.month
        results = {
            'name_1': {'name': '混炼胶实际完成(吨)', 'weight': 0},  # CMB HMB 1MB~4MB
            'name_2': {'name': '终炼胶实际完成(吨)', 'weight': 0},  # FM + 190E终炼产量
            'name_3': {'name': '外发无硫料(吨)', 'weight': 0},  # 人工输入
            'name_4': {'name': '实际完成数-1(吨)', 'weight': 0},  # 190E终炼产量 + FM + 外发无硫料*0.7
            'name_5': {'name': '实际完成数-2(吨)', 'weight': 0},  # 190E终炼产量 + FM + 外发无硫料
            'name_6': {'name': '实际生产工作日数', 'weight': 0},
            'name_10': {'name': '190E实际生产工作日数', 'weight': 0},
            'name_7': {'name': '月均完成量-1（吨）', 'weight': 0},
            'name_8': {'name': '月均完成量-2（吨）', 'weight': 0},
            'name_9': {'name': '实际生产机台数', 'weight': 0},
            'name_11': {'name': '单机台效率-1（吨/台）', 'weight': 0},
            'name_12': {'name': '单机台效率-2（吨/台）', 'weight': 0},
            'name_13': {'name': '每月段数', 'weight': 0},
            'name_14': {'name': '吨耗时（分钟/吨）', 'weight': 0},
            'name_15': {'name': '吨胶密炼耗能（KWH/吨）', 'weight': 0}
        }

        # 除去洗车胶的总车次报表数据
        total_queryset = TrainsFeedbacks.objects.exclude(Q(product_no__icontains='XCJ') |
                                                         Q(product_no__icontains='洗车胶') |
                                                         Q(operation_user='Mixer2') |
                                                         Q(product_no__icontains='WUMING')
                                                         ).filter(factory_date__year=year)
        # 按月分组，计算每月总生产重量
        month_total_dict = dict(total_queryset.values(
                'factory_date__month'
            ).annotate(weight=Sum(
                'plan_weight', output_field=DecimalField())/1000
            ).values_list('factory_date__month', 'weight'))

        # 当月混炼实际完成吨  CMB HMB 1MB~4MB
        queryset1 = total_queryset.filter(Q(product_no__icontains='-CMB-') |
                                          Q(product_no__icontains='-HMB-') |
                                          Q(product_no__icontains='-1MB-') |
                                          Q(product_no__icontains='-2MB-') |
                                          Q(product_no__icontains='-3MB-') |
                                          Q(product_no__icontains='-4MB-'))
        mix_queryset = queryset1.values('factory_date__month').annotate(weight=Sum('plan_weight', output_field=DecimalField())/1000)

        # 当月终炼实际完成（FM段次）  FM
        queryset2 = total_queryset.filter(product_no__icontains='-FM-')
        fin_queryset = queryset2.values('factory_date__month').annotate(weight=Sum('plan_weight', output_field=DecimalField())/1000)
        fm_total_dict = dict(fin_queryset.values_list('factory_date__month', 'weight'))

        # 当月190E除去洗车胶外的所有产量
        queryset_190e = Equip190EWeight.objects.exclude(
            setup__specification__in=('洗车胶', 'XCJ', 'WUMING')).filter(factory_date__year=year)
        # 按月分组，计算190E每月总生产重量
        total_queryset_190e_dict = dict(queryset_190e.values(
            'factory_date__month').annotate(
                sum_weight=Sum(F('setup__weight')*F('qty')/1000, output_field=DecimalField())).values_list('factory_date__month', 'sum_weight'))
        # 按月分组，计算190E每月终炼产量  FM
        equip_190e_weight = queryset_190e.filter(setup__state='FM').values('factory_date__month').annotate(
                sum_weight=Sum(F('setup__weight')*F('qty')/1000, output_field=DecimalField()))
        # 按月分组，计算190E每月混炼产量  CMB HMB 1MB~4MB
        equip_190e_mixin_weight = queryset_190e.filter(
            setup__state__in=('1MB', '2MB', '3MB', '4MB', 'CMB', 'HMB')
        ).values('factory_date__month').annotate(sum_weight=Sum(F('setup__weight') * F('qty') / 1000, output_field=DecimalField()))
        fm_queryset_190e_dict = dict(equip_190e_weight.values_list('factory_date__month', 'sum_weight'))

        # 按月分组，计算每月手动总生产重量
        total_manual_input_trains = ManualInputTrains.objects.exclude(
            Q(product_no__icontains='XCJ') |
            Q(product_no__icontains='洗车胶') |
            Q(product_no__icontains='-WUMING-')
        ).filter(factory_date__year=year)
        total_manual_input_trains_dict = dict(total_manual_input_trains.values(
                'factory_date__month'
            ).annotate(
            weight=Sum(F('weight') * F('actual_trains')/1000, output_field=DecimalField())
        ).values_list('factory_date__month', 'weight'))

        # 当月每日手动总生产混炼实际完成吨  CMB HMB 1MB~4MB
        total_manual_input_trains_mixin_dict = dict(
            total_manual_input_trains.filter(
                Q(product_no__icontains='-CMB-') |
                Q(product_no__icontains='-HMB-') |
                Q(product_no__icontains='-1MB-') |
                Q(product_no__icontains='-2MB-') |
                Q(product_no__icontains='-3MB-') |
                Q(product_no__icontains='-4MB-')
            ).values('factory_date__month').annotate(
                weight=Sum(F('weight') * F('actual_trains')/1000, output_field=DecimalField())
            ).values_list('factory_date__month', 'weight'))

        # 当月每日手动总生产终炼实际完成吨  FM
        total_manual_input_trains_final_dict = dict(
            total_manual_input_trains.filter(
                product_no__icontains='-FM-'
            ).values('factory_date__month').annotate(
                weight=Sum(F('weight') * F('actual_trains')/1000, output_field=DecimalField())
            ).values_list('factory_date__month', 'weight'))

        # 190E产量曲线 加硫、无硫、所有、终炼
        data_190e = {
            'dates': [i for i in range(1, days+1)],
            'wl': [0] * days,
            'jl': [0] * days,
            'total': [total_queryset_190e_dict.get(i, 0) for i in range(1, days+1)],
            'fm': [fm_queryset_190e_dict.get(i, 0) for i in range(1, days+1)]
        }

        # 每日实际工作天数
        actual_working_day_dict = dict(
            ActualWorkingDay.objects.filter(
                factory_date__year=year,
                num__gt=0).values_list('factory_date__month', 'num'))

        # 190E每日实际工作天数
        actual_working_day_190e_dict = dict(
            ActualWorkingDay190E.objects.filter(
                factory_date__year=year,
                num__gt=0).values_list('factory_date__month', 'num'))

        # 每日实际工作机台数
        actual_working_equip_dict = dict(
            ActualWorkingEquip.objects.filter(
                factory_date__year=year,
                num__gt=0).values_list('factory_date__month', 'num'))

        # 外发无硫料（吨）
        out_queryset = OuterMaterial.objects.filter(
            factory_date__year=year).values('factory_date__month').annotate(weight=Sum('weight', output_field=DecimalField()))
        out_queryset_dict = dict(out_queryset.values_list('factory_date__month', 'weight'))

        time_consume_dict = dict(TrainsFeedbacks.objects.filter(
            factory_date__year=year).values('factory_date').annotate(
            time_consume=OSum((F('end_time') - F('begin_time')))).values_list('factory_date__month', 'time_consume'))
        energy_consume_data = TrainsFeedbacks.objects.filter(
            factory_date__year=year).extra().values('factory_date__month', 'equip_no').annotate(
            energy_consume=Sum('evacuation_energy'),
            sum_weight=Avg('plan_weight')
        )
        energy_consume_dict = {}
        for e in energy_consume_data:
            energy_consume = 0 if not e['energy_consume'] else e['energy_consume']
            equip_no = e['equip_no']
            if equip_no == 'Z01':
                evacuation_energy = energy_consume / 10
            elif equip_no == 'Z02':
                evacuation_energy = energy_consume / 0.6
            elif equip_no == 'Z04':
                evacuation_energy = energy_consume * 0.28 * float(e['sum_weight']) / 1000
            elif equip_no == 'Z12':
                evacuation_energy = energy_consume / 5.3
            elif equip_no == 'Z13':
                evacuation_energy = energy_consume / 31.7
            else:
                evacuation_energy = energy_consume
            energy_consume_dict[e['factory_date__month']] = energy_consume_dict.get(e['factory_date__month'], 0) + evacuation_energy
        for item in mix_queryset:
            mixin_weight = round(item['weight'], 2)
            if not(item['factory_date__month'] == now_day and exclude_today_flat):
                results['name_1']['weight'] += mixin_weight
            results['name_1'][f"{item['factory_date__month']}月"] = mixin_weight
        for item in equip_190e_mixin_weight:
            weight = round(item['sum_weight'], 2)
            if not(item['factory_date__month'] == now_day and exclude_today_flat):
                results['name_1']['weight'] += weight
            results['name_1'][f"{item['factory_date__month']}月"] = results['name_1'].get(f"{item['factory_date__month']}月", 0) + weight
            data_190e['wl'][item['factory_date__month']-1] = weight
        for d, v in total_manual_input_trains_mixin_dict.items():
            v = round(v, 2)
            if not(d == now_day and exclude_today_flat):
                results['name_1']['weight'] += v
            results['name_1'][f"{d}月"] = results['name_1'].get(f"{d}月", 0) + v
        for item in equip_190e_weight:
            weight = round(item['sum_weight'], 2)
            results['name_2'][f"{item['factory_date__month']}月"] = results['name_2'].get(f"{item['factory_date__month']}月", 0) + weight
            results['name_4'][f"{item['factory_date__month']}月"] = results['name_4'].get(f"{item['factory_date__month']}月", 0) + weight
            results['name_5'][f"{item['factory_date__month']}月"] = results['name_5'].get(f"{item['factory_date__month']}月", 0) + weight
            data_190e['jl'][item['factory_date__month']-1] = weight
            if not(item['factory_date__month'] == now_day and exclude_today_flat):
                results['name_2']['weight'] += weight
                results['name_4']['weight'] += weight
                results['name_5']['weight'] += weight
        for item in fin_queryset:
            final_weight = round(item['weight'], 2)
            if not(item['factory_date__month'] == now_day and exclude_today_flat):
                results['name_2']['weight'] += final_weight
                results['name_4']['weight'] += final_weight
                results['name_5']['weight'] += final_weight
            results['name_2'][f"{item['factory_date__month']}月"] = results['name_2'].get(f"{item['factory_date__month']}月", 0) + final_weight
            results['name_4'][f"{item['factory_date__month']}月"] = results['name_4'].get(f"{item['factory_date__month']}月", 0) + final_weight
            results['name_5'][f"{item['factory_date__month']}月"] = results['name_5'].get(f"{item['factory_date__month']}月", 0) + final_weight
        for d, v in total_manual_input_trains_final_dict.items():
            v = round(v, 2)
            if not(d == now_day and exclude_today_flat):
                results['name_2']['weight'] += v
                results['name_4']['weight'] += v
                results['name_5']['weight'] += v
            results['name_2'][f"{d}月"] = results['name_2'].get(f"{d}月", 0) + v
            results['name_4'][f"{d}月"] = results['name_4'].get(f"{d}月", 0) + v
            results['name_5'][f"{d}月"] = results['name_5'].get(f"{d}月", 0) + v
        for item in out_queryset:
            results['name_3'][f"{item['factory_date__month']}月"] = round(item['weight'], 2)
            results['name_4'][f"{item['factory_date__month']}月"] = results['name_4'].get(f"{item['factory_date__month']}月", 0) + round((item['weight']) * decimal.Decimal(0.7), 2)
            results['name_5'][f"{item['factory_date__month']}月"] = results['name_5'].get(f"{item['factory_date__month']}月", 0) + round(item['weight'], 2)
            if not(item['factory_date__month'] == now_day and exclude_today_flat):
                results['name_3']['weight'] += round(item['weight'], 2)
                results['name_4']['weight'] += round((item['weight']) * decimal.Decimal(0.7), 2)
                results['name_5']['weight'] += round(item['weight'], 2)
        for k, v in actual_working_day_dict.items():
            v = round(v, 2)
            results['name_6'][f"{k}月"] = v
            if not(k == now_day and exclude_today_flat):
                results['name_6']['weight'] = round(results['name_6']['weight'] + v, 2)
        for k, v in actual_working_equip_dict.items():
            v = round(v, 2)
            results['name_9'][f"{k}月"] = v
            if not(k == now_day and exclude_today_flat):
                results['name_9']['weight'] = round(results['name_9']['weight'] + v, 2)
        for k, v in actual_working_day_190e_dict.items():
            v = round(v, 2)
            results['name_10'][f"{k}月"] = v
            if not(k == now_day and exclude_today_flat):
                results['name_10']['weight'] = round(results['name_10']['weight'] + v, 2)
        if len(results['name_6']) - 2 != 0:
            for key, value in results['name_4'].items():
                if key[0].isdigit():
                    if results['name_6'].get(key):
                        results['name_7'][key] = round(results['name_4'][key] / decimal.Decimal(results['name_6'][key]), 2)
                        results['name_8'][key] = round(results['name_5'][key] / decimal.Decimal(results['name_6'][key]), 2)
            results['name_7']['weight'] = 0 if results['name_6']['weight'] == 0 else round(results['name_4']['weight'] / decimal.Decimal(results['name_6']['weight']), 2)
            results['name_8']['weight'] = 0 if results['name_6']['weight'] == 0 else round(results['name_5']['weight'] / decimal.Decimal(results['name_6']['weight']), 2)
        if exclude_today_flat:
            # 190E当月总天数
            month_working_days_190e = float(sum(actual_working_day_190e_dict.values()) -
                                            actual_working_day_190e_dict.get(now_day, 0))
            # 当月总天数
            month_working_days = float(sum(actual_working_day_dict.values()) -
                                            actual_working_day_dict.get(now_day, 0))
            # 当月总生产机台数
            actual_working_equips = float(sum(actual_working_equip_dict.values()) -
                                            actual_working_equip_dict.get(now_day, 0))
        else:
            # 190E当月总天数
            month_working_days_190e = float(sum(actual_working_day_190e_dict.values()))
            # 当月总天数
            month_working_days = float(sum(actual_working_day_dict.values()))
            # 当月总生产机台数
            actual_working_equips = float(sum(actual_working_equip_dict.values()))

        # 计算每日单机台效率-1
        for k, v in results['name_4'].items():
            if k[0].isdigit():
                ds = actual_working_equip_dict.get(int(k[:-1]))
                if not ds:
                    continue
                value = round(float(v) / ds, 2)
                results['name_11'][k] = value
        results['name_11']['weight'] = 0 if not actual_working_equips else round(float(results['name_4']['weight']) / actual_working_equips, 2)

        # 计算每日单机台效率-2
        for k, v in results['name_5'].items():
            if k[0].isdigit():
                ds = actual_working_equip_dict.get(int(k[:-1]))
                if ds:
                    value = round(float(v) / ds, 2)
                    results['name_12'][k] = value
                time_consume = time_consume_dict.get(int(k[:-1]))
                if time_consume:
                    time_value = round(float(time_consume.total_seconds()) / 60 / float(v), 2)
                    results['name_14'][k] = time_value
                    if not (k == now_day and exclude_today_flat):
                        results['name_14']['weight'] = round(results['name_14']['weight'] + time_value, 2)
                energy_consume = energy_consume_dict.get(int(k[:-1]))
                if energy_consume:
                    energy_value = round(float(energy_consume) / float(v), 2)
                    results['name_15'][k] = energy_value
                    if not (k == now_day and exclude_today_flat):
                        results['name_15']['weight'] = round(results['name_15']['weight'] + energy_value, 2)
        results['name_12']['weight'] = 0 if not actual_working_equips else round(float(results['name_5']['weight']) / actual_working_equips, 2)

        # 计算平均值
        for item in results.values():
            if item['name'] == '190E实际生产工作日数':
                l1 = len(actual_working_day_190e_dict)
                if exclude_today_flat:
                    if actual_working_day_190e_dict.get(now_day):
                        l1 -= 1
                item['avg'] = "" if not month_working_days_190e else round(
                    float(item['weight']) / l1, 2)
            elif item['name'] == '实际生产机台数':
                l2 = len(actual_working_equip_dict)
                if exclude_today_flat:
                    if actual_working_equip_dict.get(now_day):
                        l2 -= 1
                item['avg'] = "" if not actual_working_equips else round(
                    float(item['weight']) / l2, 2)
            elif item['name'] == '实际生产工作日数':
                l3 = len(actual_working_day_dict)
                if exclude_today_flat:
                    if actual_working_day_dict.get(now_day):
                        l3 -= 1
                item['avg'] = "" if not month_working_days else round(
                    float(item['weight']) / l3, 2)
            elif item['name'] in ('单机台效率-1（吨/台）', '单机台效率-2（吨/台）', '每日段数'):
                item['avg'] = item['weight']
            elif item['name'] == '吨耗时（分钟/吨）':
                length = len(results['name_14']) - 2
                item['avg'] = '' if not length else round(item['weight'] / length, 2)
            elif item['name'] == '吨胶密炼耗能（KWH/吨）':
                length = len(results['name_15']) - 2
                item['avg'] = '' if not length else round(item['weight'] / length, 2)
            else:
                item['avg'] = "" if month_working_days <= 0 else round(float(item['weight']) / month_working_days, 2)
        try:
            results['name_11']['avg'] = round(results['name_4']['avg'] / results['name_9']['avg'], 2)
        except Exception:
            results['name_11']['avg'] = ''
        try:
            results['name_12']['avg'] = round(results['name_5']['avg'] / results['name_9']['avg'], 2)
        except Exception:
            results['name_12']['avg'] = ''

        # 计算190E每日段数
        cnt = 0
        sum_ds = 0
        for idx in range(1, days):
            fm = data_190e['fm'][idx]
            total = data_190e['total'][idx]
            if not total:
                continue
            try:
                ds = total/fm
            except Exception:
                ds = 0
            sum_ds += ds
            cnt += 1

        if exclude_today_flat:
            total_190e_jl = sum(data_190e['jl']) - data_190e['jl'][now_day-1]
            # length1 = len([i for i in data_190e['jl'] if i > 0])
            # if data_190e['jl'][now_day-1]:
            #     length1 -= 1
            total_190e_wl = sum(data_190e['wl']) - data_190e['wl'][now_day-1]
            # length2 = len([i for i in data_190e['wl'] if i > 0])
            # if data_190e['wl'][now_day-1]:
            #     length2 -= 1
            total_190e_fm = sum(data_190e['fm']) - data_190e['fm'][now_day-1]
            total_190e_total = sum(data_190e['total']) - data_190e['total'][now_day-1]
        else:
            total_190e_jl = sum(data_190e['jl'])
            # length1 = len([i for i in data_190e['jl'] if i > 0])
            total_190e_wl = sum(data_190e['wl'])
            # length2 = len([i for i in data_190e['wl'] if i > 0])
            total_190e_fm = sum(data_190e['fm'])
            total_190e_total = sum(data_190e['total'])

        avg_190e = {'jl': 0 if not month_working_days_190e else round(float(total_190e_jl) / float(month_working_days_190e), 2),
                    'wl': 0 if not month_working_days_190e else round(float(total_190e_wl) / float(month_working_days_190e), 2),
                    'ds': 0 if not total_190e_fm else round(total_190e_total / total_190e_fm, 2)}

        # 计算每日总产量段数
        cnt2 = 0
        sum_ds2 = 0
        for t_day in range(1, days+1):
            t_weight = float(month_total_dict.get(t_day, 0)) \
                       + float(total_queryset_190e_dict.get(t_day, 0)) \
                       + float(total_manual_input_trains_dict.get(t_day, 0))
            if not t_weight:
                continue
            fm_weight = float(fm_total_dict.get(t_day, 0)) \
                        + float(fm_queryset_190e_dict.get(t_day, 0)) \
                        + float(out_queryset_dict.get(t_day, 0)) \
                        + float(total_manual_input_trains_final_dict.get(t_day, 0))
            try:
                ds2 = t_weight / fm_weight
            except Exception:
                ds2 = 0
            if not (exclude_today_flat and t_day == now_day):
                sum_ds2 += ds2
                cnt2 += 1
            results['name_13']['{}月'.format(str(t_day))] = round(ds2, 2)

        # 计算段数总计和平均值
        if exclude_today_flat:
            total_weight = sum(list(month_total_dict.values())) \
                           + sum(list(total_queryset_190e_dict.values())) \
                           + sum(list(total_manual_input_trains_dict.values())) \
                           - month_total_dict.get(now_day, 0)\
                           - total_queryset_190e_dict.get(now_day, 0)\
                           - total_manual_input_trains_dict.get(now_day, 0)
            total_fm_weight = sum(list(fm_total_dict.values())) \
                              + sum(list(fm_queryset_190e_dict.values())) \
                              + sum(list(out_queryset_dict.values())) \
                              + sum(list(total_manual_input_trains_final_dict.values())) \
                              - fm_total_dict.get(now_day, 0)\
                              - fm_queryset_190e_dict.get(now_day, 0)\
                              - out_queryset_dict.get(now_day, 0)\
                              - total_manual_input_trains_final_dict.get(now_day, 0)
        else:
            total_weight = sum(list(month_total_dict.values())) \
                           + sum(list(total_queryset_190e_dict.values())) \
                           + sum(list(total_manual_input_trains_dict.values()))
            total_fm_weight = sum(list(fm_total_dict.values())) \
                              + sum(list(fm_queryset_190e_dict.values())) \
                              + sum(list(out_queryset_dict.values())) \
                              + sum(list(total_manual_input_trains_final_dict.values()))
        results['name_13']['weight'] = "" if not total_fm_weight or not total_weight else round(total_weight/total_fm_weight, 2)
        results['name_13']['avg'] = 0 if not cnt2 else round(sum_ds2/cnt2, 2)
        results['name_14']['weight'] = '' if not results['name_5']['weight'] else round(sum([i.total_seconds() / 60 for i in time_consume_dict.values()]) / float(results['name_5']['weight']), 2)
        results['name_15']['weight'] = '' if not results['name_5']['weight'] else round(sum(energy_consume_dict.values()) / float(results['name_5']['weight']), 2)
        # 总产量加硫、无硫、段数平均值
        avg_results = {'jl': results['name_2']['avg'],
                       'wl': results['name_1']['avg'],
                       'ds': results['name_13']['weight']}
        if self.request.query_params.get('export', None):
            results2 = {}
            equip_query = Equip.objects.filter(
                category__equip_type__global_name='密炼设备').values('equip_no', 'category__category_name')
            equip_dic = {item['equip_no']: item['category__category_name'] for item in equip_query}
            data2 = TrainsFeedbacks.objects.exclude(operation_user='Mixer2').filter(
                factory_date__year=year).values(
                'factory_date__month', 'product_no', 'equip_no', 'classes'
            ).annotate(actual_trains=Count('id'), weight=Sum('plan_weight')/1000).order_by('-classes')
            for item in data2:
                try:
                    state = item['product_no'].split("-")[1]
                    space = item['product_no'].split("-")[2]
                    key = f'{space}_{state}_{item["equip_no"]}_{item["classes"]}'
                except:
                    continue
                weight = round(float(item['weight']), 3)
                if results2.get(key):
                    results2[key][f'{item["factory_date__month"]}_qty'] = results2[key].get(
                        f'{item["factory_date__month"]}_qty', 0) + item['actual_trains']
                    results2[key][f'{item["factory_date__month"]}_weight'] = round(
                        results2[key].get(f'{item["factory_date__month"]}_weight', 0) + weight, 3)
                    results2[key]['汇总_weight'] += weight
                    results2[key]['汇总_qty'] += item['actual_trains']
                else:
                    results2[key] = {'规格': space, '段数': state, '机台': item["equip_no"],
                                     '机型': equip_dic[item["equip_no"]], '班别': item['classes'],
                                     '汇总_qty': item['actual_trains'], '汇总_weight': weight,
                                     f'{item["factory_date__month"]}_qty': item['actual_trains'],
                                     f'{item["factory_date__month"]}_weight': weight}
            equip_190e_result = {}
            equip_190e_queryset = Equip190EWeight.objects.filter(
                factory_date__year=year
            ).values('factory_date__month', 'setup__specification', 'setup__state', 'classes').annotate(
                sum_weight=Sum(F('setup__weight')*F('qty')/1000, output_field=DecimalField()),
                total_trains=Sum('qty')).order_by('-classes')
            for item in equip_190e_queryset:
                weight = item['sum_weight']
                product_no = item['setup__specification']
                stage = item['setup__state']
                classes = item['classes']
                key = '{}_{}_{}'.format(
                    product_no, stage, classes
                )
                if equip_190e_result.get(key):
                    equip_190e_result[key][f'{item["factory_date__month"]}_qty'] = equip_190e_result[key].get(
                        f'{item["factory_date__month"]}_qty', 0) + item['total_trains']
                    equip_190e_result[key][f'{item["factory_date__month"]}_weight'] = round(
                        equip_190e_result[key].get(f'{item["factory_date__month"]}_weight', 0) + weight, 3)
                    equip_190e_result[key]['汇总_weight'] += weight
                    equip_190e_result[key]['汇总_qty'] += item['total_trains']
                else:
                    equip_190e_result[key] = {'规格': product_no, '段数': stage, '机台': '190E',
                                              '机型': '190E', '班别': classes,
                                              '汇总_qty': item['total_trains'], '汇总_weight': weight,
                                               f'{item["factory_date__month"]}_qty': item['total_trains'],
                                               f'{item["factory_date__month"]}_weight': weight}

            manual_input_queryset = ManualInputTrains.objects.filter(
                factory_date__year=year
            ).values('factory_date__month',
                     'equip_no',
                     'product_no',
                     'classes'
                     ).annotate(
                weight=Sum(F('weight') * F('actual_trains') / 1000, output_field=DecimalField()),
                total_trains=Sum('actual_trains')
            ).order_by('-classes')
            for item in manual_input_queryset:
                try:
                    state = item['product_no'].split("-")[1]
                    space = item['product_no'].split("-")[2]
                    key = f'{space}_{state}_{item["equip_no"]}_{item["classes"]}'
                except:
                    continue
                weight = round(float(item['weight']), 3)
                if results2.get(key):
                    results2[key][f'{item["factory_date__month"]}_qty'] = results2[key].get(
                        f'{item["factory_date__month"]}_qty', 0) + item['total_trains']
                    results2[key][f'{item["factory_date__month"]}_weight'] = round(
                        results2[key].get(f'{item["factory_date__month"]}_weight', 0) + weight, 3)
                    results2[key]['汇总_weight'] += weight
                    results2[key]['汇总_qty'] += item['total_trains']
                else:
                    results2[key] = {'规格': space, '段数': state, '机台': item["equip_no"],
                                     '机型': equip_dic[item["equip_no"]], '班别': item['classes'],
                                     '汇总_qty': item['total_trains'], '汇总_weight': weight,
                                     f'{item["factory_date__month"]}_qty': item['total_trains'],
                                     f'{item["factory_date__month"]}_weight': weight}
            excel_sheet_1_data = list(results.values())
            excel_sheet_23_data = sorted(
                list(results2.values()), key=itemgetter('机台', '规格', '段数')
            ) + sorted(
                list(equip_190e_result.values()), key=itemgetter('机台', '规格', '段数')
            )
            return self.export(year, days, excel_sheet_1_data, excel_sheet_23_data)

        return Response({'results': results.values(),
                         'data_190e': data_190e,
                         'avg_190e': avg_190e,
                         'avg_results': avg_results
                         })


@method_decorator([api_recorder], name="dispatch")
class Equip190EViewSet(ModelViewSet):
    queryset = Equip190E.objects.filter(delete_flag=False).order_by('id')
    serializer_class = Equip190ESerializer
    filter_backends = (DjangoFilterBackend,)
    filter_class = Equip190EFilter
    pagination_class = None
    FILE_NAME = '190E机台规格信息'
    EXPORT_FIELDS_DICT = {
        "规格": "specification",
        "段次": "state",
        "重量(kg)": "weight",
    }

    def get_queryset(self):
        """获取specification, state字段相同时，id最大的那一条"""
        ids = self.queryset.values('specification', 'state').annotate(id=Max('id')).values_list('id', flat=True)
        return self.queryset.filter(id__in=ids).order_by('id')

    def list(self, request, *args, **kwargs):
        if self.request.query_params.get('export'):
            return gen_template_response(self.EXPORT_FIELDS_DICT, [], self.FILE_NAME)
        if self.request.query_params.get('search'):
            return Response({'results': list(set(self.queryset.values_list('specification', flat=True)))})
        if self.request.query_params.get('detail'):
            factory_date = self.request.query_params.get('factory_date')
            classes = self.request.query_params.get('classes')
            instance = Equip190EWeight.objects.filter(factory_date=factory_date, classes=classes)
            serializer = Equip190EWeightSerializer(instance=instance, many=True)
            return Response({'results': serializer.data})
        return super().list(request, *args, **kwargs)

    @action(methods=['post'], detail=False, permission_classes=[], url_path='import_xlsx',
            url_name='import_xlsx')
    def import_xlsx(self, request):
        excel_file = request.FILES.get('file', None)
        if not excel_file:
            raise ValidationError('文件不可为空！')
        cur_sheet = get_cur_sheet(excel_file)
        if cur_sheet.ncols != 3:
            raise ValidationError('导入文件数据错误！')
        data = get_sheet_data(cur_sheet)
        parts_list = []
        for item in data:
            if not item[0]:
                raise ValidationError('规格不可为空')
            if not item[1]:
                raise ValidationError('段数不可为空')
            if not item[2]:
                raise ValidationError('重量不可为空')
            parts_list.append({
                'specification': item[0],
                'state': item[1],
                'weight': item[2]
            })

        s = Equip190ESerializer(data=parts_list, many=True, context={'request': request})
        s.is_valid(raise_exception=True)
        if len(s.validated_data) < 1:
            raise ValidationError('没有可导入的数据')
        data = s.validated_data
        for item in data:
            Equip190E.objects.filter(delete_flag=False).update_or_create(defaults=item,
                                               specification=item['specification'],
                                               state=item['state'],
                                               weight=item['weight']
                                               )
        return Response(f'成功导入{len(s.validated_data)}条数据')

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        instance.delete_flag = True
        instance.save()
        return Response(status=status.HTTP_204_NO_CONTENT)


@method_decorator([api_recorder], name="dispatch")
class SummaryOfMillOutput(APIView):
    permission_classes = (IsAuthenticated,)

    def get(self, request):
        factory_date = self.request.query_params.get('factory_date')
        auto = int(self.request.query_params.get('auto'))
        manual = int(self.request.query_params.get('manual'))
        # 统计机台的机型

        queryset = Equip.objects.filter(category__equip_type__global_name='密炼设备').values('equip_no', 'category__category_name')
        dic = {item['equip_no']: item['category__category_name'] for item in queryset}
        state_list = GlobalCode.objects.filter(global_type__type_name='胶料段次').values('global_name')
        state_list = [item['global_name'] for item in state_list]
        equip_list = Equip.objects.filter(category__equip_type__global_name='密炼设备').order_by('equip_no').values('equip_no')

        results = {}
        count = {'equip_no': None, 'type': '合计', 'count': 0}
        for equip in equip_list:
            equip = equip['equip_no']
            results[f"{equip}_pt"] = {'equip_no': f"{equip}({dic.get(equip)})", 'type': '普通胶车数', 'count': 0}
            results[f"{equip}_dj"] = {'equip_no': f"{equip}({dic.get(equip)})", 'type': '丁基胶车数', 'count': 0}
            results[f"{equip}_xj"] = {'equip_no': f"{equip}({dic.get(equip)})", 'type': '小计', 'count': 0}

            for state in state_list:
                results[f"{equip}_pt"][f"{state}-早"] = 0
                results[f"{equip}_pt"][f"{state}-晚"] = 0
                results[f"{equip}_dj"][f"{state}-早"] = 0
                results[f"{equip}_dj"][f"{state}-晚"] = 0
                results[f"{equip}_xj"][f"{state}-早"] = 0
                results[f"{equip}_xj"][f"{state}-晚"] = 0
                count[f"{state}-早"] = 0
                count[f"{state}-晚"] = 0
        if auto:
            data1 = TrainsFeedbacks.objects.filter(factory_date=factory_date).values('equip_no',
                                                                                    'product_no', 'classes').annotate(
                actual_trains=Count('actual_trains')).values('equip_no', 'product_no', 'classes', 'actual_trains')
        else:
            data1 = []
        if manual:
            data2 = ManualInputTrains.objects.filter(factory_date=factory_date).values(
                'equip_no', 'product_no', 'classes', 'actual_trains')
        else:
            data2 = []
        data = list(data1) + list(data2)
        dj = ProductInfoDingJi.objects.filter(delete_flag=False, is_use=True).values('product_no')
        for item in data:
            equip = item['equip_no']
            state = item['product_no'].split('-')[1]
            classes = '早' if item['classes'] == '早班' else '晚'
            actual_trains = item['actual_trains']
            if state in state_list:
                # 判断是否是丁基胶
                if item['product_no'] in [item['product_no'] for item in dj]:
                    results[f"{equip}_dj"][f"{state}-{classes}"] += actual_trains
                    results[f"{equip}_dj"]['count'] += actual_trains
                else:
                    results[f"{equip}_pt"][f"{state}-{classes}"] += actual_trains
                    results[f"{equip}_pt"]['count'] += actual_trains
                results[f"{equip}_xj"][f"{state}-{classes}"] += actual_trains
                results[f"{equip}_xj"]['count'] += actual_trains
                count[f"{state}-{classes}"] += actual_trains
                count['count'] += actual_trains
        for dic in results.values():
            for key, value in dic.items():
                dic[key] = None if not dic[key] else dic[key]
        return Response({'results': results.values(), 'count': count, 'state_list': state_list})


@method_decorator([api_recorder], name="dispatch")
class SummaryOfWeighingOutput(APIView):
    permission_classes = (IsAuthenticated,)

    def concat_user_package(self, equip_no, result, factory_date, users, work_times, user_result, qty_data):
        dic = {'equip_no': equip_no, 'hj': 0}
        plan_model, report_basic = [JZPlan, JZReportBasic] if equip_no in JZ_EQUIP_NO else [Plan, ReportBasic]
        data = plan_model.objects.using(equip_no).filter(actno__gt=1, date_time__istartswith=factory_date).values('date_time', 'grouptime').annotate(count=Sum('actno'))
        for item in data:
            date = item['date_time']
            day = int(date.split('-')[2])    # 2  早班
            classes = item['grouptime']  # 早班/ 中班 / 夜班
            filter_classes = classes if equip_no not in JZ_EQUIP_NO else ('早' if classes == '早班' else ('晚' if classes == '夜班' else '中'))
            dic[f'{day}{classes}'] = item['count']
            dic['hj'] = dic.get('hj', 0) + item['count']
            names = users.get(f'{day}-{classes}-{equip_no}')
            if names:
                status = names.pop('status', None)
                key_dic = {}
                for name, section_list in names.items():
                    for section in section_list:
                        key = f"{name}_{day}_{classes}_{section}"
                        work_time = work_times.get(f'{day}-{classes}-{equip_no}').get(name + '_' + section, [])
                        if len(work_time) < 2:
                            continue
                        st, et = work_time[:2]
                        if f'{day}-{st}-{et}' in qty_data:
                            num = qty_data[f'{day}-{st}-{et}']
                        else:
                            c_num = report_basic.objects.using(equip_no).filter(starttime__gte=work_time[0], savetime__lte=work_time[1], grouptime=filter_classes).aggregate(num=Count('id'))['num']
                            num = c_num if c_num else 0  # 是否需要去除为0的机台再取平均
                            qty_data[f'{day}-{st}-{et}'] = num
                        # 车数计算：当天产量 / 12小时 * 实际工作时间 -> 修改为根据考勤时间计算
                        if f"{name}_{day}_{classes}" not in key_dic:
                            key_dic[f"{name}_{day}_{classes}"] = key
                        else:
                            key = key_dic[f"{name}_{day}_{classes}"]
                        if user_result.get(key):
                            user_result[key][equip_no] = user_result[key].get(equip_no, 0) + num
                        else:
                            user_result[key] = {equip_no: num}
                        if status == '调岗':
                            user_result[key]['status'] = status
        result.append(dic)

    def get(self, request):
        factory_date = self.request.query_params.get('factory_date')
        export = self.request.query_params.get('export')
        year = int(factory_date.split('-')[0])
        month = int(factory_date.split('-')[1])
        if export:  # 导出员工实际考勤
            try:
                days, export_data = actual_clock_data(factory_date, '生产配料')
            except Exception as e:
                logger.error(e.args[0])
                raise ValidationError('导出实际考勤数据异常')
            if not export_data:
                raise ValidationError('无考勤数据可以导出')
            EXPORT_FIELDS_DICT = {'姓名': 'username', '岗位': 'section', '班组': 'group', '机台': 'equip'}
            add_key = {'/'.join(i.split('-')[1:]): i for i in days}
            EXPORT_FIELDS_DICT.update(add_key)
            return gen_template_response(EXPORT_FIELDS_DICT, export_data.values(), f'{factory_date}配料间实际考勤数据', sheet_name=factory_date, handle_str=True)
        equip_list = Equip.objects.filter(category__equip_type__global_name='称量设备').values_list('equip_no', flat=True)
        # 获取人员包数
        u_name = self.request.query_params.get('name')
        day = int(self.request.query_params.get('day', 0))
        classes = self.request.query_params.get('classes')
        filter_kwargs = {}
        if all([u_name, day, classes]):
            filter_kwargs = {'user__username': u_name, 'factory_date__day': day, 'classes': classes}
        result = []
        result1 = {}
        users = {}
        work_times = {}
        user_result = {}
        user_package = {}
        price_obj = SetThePrice.objects.first()
        if not price_obj:
            raise ValidationError('请先去添加细料/硫磺单价')
        # 查询称量分类下当前月上班的所有员工
        user_list = EmployeeAttendanceRecords.objects.filter(
            Q(factory_date__year=year, factory_date__month=month, equip__in=equip_list) &
            Q(end_date__isnull=False, begin_date__isnull=False) & ~Q(is_use='废弃'), ~Q(clock_type='密炼'), **filter_kwargs)\
            .values('user__username', 'factory_date__day', 'group', 'classes', 'section', 'equip', 'calculate_begin_date', 'calculate_end_date', 'status')
        if filter_kwargs:  # 获取包数
            data = user_list.order_by('equip')
            user_total = {}
            for i in data:
                section, equip_no, st, et, classes = i.get('section'), i.get('equip'), i.get('calculate_begin_date'), i.get('calculate_end_date'), i.get('classes')
                plan_model, report_basic = [JZPlan, JZReportBasic] if equip_no in JZ_EQUIP_NO else [Plan, ReportBasic]
                if equip_no in JZ_EQUIP_NO:
                    classes = '早' if classes == '早班' else ('晚' if classes == '夜班' else '中')
                num = report_basic.objects.using(equip_no).filter(starttime__gte=st, savetime__lte=et, grouptime=classes).aggregate(num=Count('id'))['num']
                if not num:
                    continue
                key = f"{equip_no}-{section}"
                unit = price_obj.xl if equip_no.startswith('F') else price_obj.lh
                equip_data = user_package.get(key)
                if equip_data:
                    equip_data['num'] += num
                else:
                    user_package[key] = {'section': section, 'num': num, 'equip_no': equip_no, 'unit': unit}
                user_total[equip_no] = user_total.get(equip_no, 0) + num
            return Response({'detail': user_package.values(), 'user_total': user_total})
        # 岗位系数
        section_dic = {}
        section_info = PerformanceJobLadder.objects.filter(delete_flag=False, type='生产配料').values('type', 'name', 'coefficient', 'post_standard', 'post_coefficient')
        for item in section_info:
            section_dic[f"{item['name']}_{item['type']}"] = [item['coefficient'], item['post_standard'], item['post_coefficient']]
        # 员工类别
        independent = {}
        independent_lst = IndependentPostTemplate.objects.filter(date_time=factory_date).values('name', 'status', 'work_type')
        for item in independent_lst:
            independent[item['name']] = {'status': item['status'], 'work_type': item['work_type']}
        if not independent:
            raise ValidationError(f'请添加{factory_date}员工类别')
        # 员工类别系数
        employee_type = GlobalCode.objects.filter(global_type__type_name='员工类别', global_type__use_flag=True, use_flag=True).values('global_no', 'global_name')
        employee_type_dic = {dic['global_no']: dic['global_name'] for dic in employee_type}
        # 员工独立上岗系数
        coefficient = GlobalCode.objects.filter(global_type__type_name='是否独立上岗系数', global_type__use_flag=True, use_flag=True).values('global_no', 'global_name')
        coefficient_dic = {dic['global_no']: dic['global_name'] for dic in coefficient}
        if not coefficient or not employee_type_dic:
            raise ValidationError(f'请在公共变量中添加员工独立上岗系数或员工类别系数')

        for item in user_list:
            key = f"{item['factory_date__day']}-{item['classes']}-{item['equip']}"
            if users.get(key):
                work_times[key][item['user__username'] + '_' + item['section']] = [item['calculate_begin_date'], item['calculate_end_date']]
                users[key][item['user__username']] = [item['section']] + ([] if not users[key].get(item['user__username']) else users[key][item['user__username']])
            else:
                work_times[key] = {item['user__username'] + '_' + item['section']: [item['calculate_begin_date'], item['calculate_end_date']]}
                users[key] = {item['user__username']: [item['section']]}
                if item['status'] == '调岗':
                    users[key]['status'] = '调岗'

        # 机台产量统计
        qty_data, t_num = {}, 32
        pool = ThreadPool(t_num)
        for equip_no in equip_list:
            pool.apply_async(self.concat_user_package, args=(equip_no, result, factory_date, users, work_times, user_result, qty_data))
        pool.close()
        pool.join()
        sort_res = sorted(result, key=lambda x: x['equip_no'])
        # 普通员工
        permissions_list = self.request.user.permissions_list
        xl_permission = permissions_list.get('summary_of_weighing_output', [])
        if 'save' not in xl_permission:
            return Response({'results': sort_res})
        for key, value in user_result.items():
            """
            key: test_1_早班_主控
            value: {'F03': 109, 'F02': 100,}
            """
            name, day, classes, section = key.split('_')
            # trans_flag = user_list.filter(status='调岗', user__username=name, classes=classes, factory_date=(factory_date + '-' + '%02d' % int(day)))
            trans_flag = value.pop('status', None)
            equip = list(value.keys())[0]
            type = '生产配料'
            if trans_flag:
                equip, count_ = equip, sum(value.values())
            else:
                if section_dic[f"{section}_{type}"][1] == 1:  # 最大值
                    equip, count_ = sorted(value.items(), key=lambda kv: (kv[1], kv[0]))[-1]
                else:  # 平均值  是否需要去除为0的机台再取平均
                    equip, count_ = equip, sum(value.values()) / len(value)
            # 细料/硫磺单价'
            unit_price = price_obj.xl if equip in ['F01', 'F02', 'F03'] else price_obj.lh
            # 员工类别系数
            a, w_coefficient = float(coefficient_dic.get('是', 1)), 1
            if independent.get(name):
                if independent.get(name).get('status') != 1:
                    a = float(coefficient_dic.get('否'))
                work_type = independent.get(name).get('work_type')
                w_coefficient = float(employee_type_dic.get(work_type)) if work_type else w_coefficient
            coefficient = section_dic[f"{section}_{type}"][0] / 100
            post_coefficient = section_dic[f"{section}_{type}"][2] / 100
            price = round(count_ * coefficient * post_coefficient * unit_price * a * w_coefficient, 2)
            xl = price if equip in ['F01', 'F02', 'F03'] else 0
            lh = price if equip in ['S01', 'S02'] else 0

            if result1.get(name):
                result1[name][f"{day}{classes}"] = price
                result1[name][f"{day}{classes}_count"] = count_
                result1[name]['xl'] = round(result1[name].get('xl', 0) + xl, 2)
                result1[name]['lh'] = round(result1[name].get('lh', 0) + lh, 2)
            else:
                result1[name] = {'name': name, f"{day}{classes}": price, f"{day}{classes}_count": count_, 'xl': round(xl, 2), 'lh': round(lh, 2)}
        return Response({'results': sort_res, 'users': result1.values(), 'user_result': user_result})


@method_decorator([api_recorder], name="dispatch")
class EmployeeAttendanceRecordsView(APIView):
    permission_classes = (IsAuthenticated,)

    def get(self, request):
        date = self.request.query_params.get('date')
        clock_type = self.request.query_params.get('clock_type', '密炼')
        name = self.request.query_params.get('name', '')
        year = int(date.split('-')[0])
        month = int(date.split('-')[1])
        # 获取班组
        this_month_start = datetime.datetime(year, month, 1)
        if month == 12:
            this_month_end = datetime.datetime(year + 1, 1, 1) - timedelta(days=1)
        else:
            this_month_end = datetime.datetime(year, month + 1, 1) - timedelta(days=1)
        if clock_type == "密炼":
            group = WorkSchedulePlan.objects.filter(start_time__date__gte=this_month_start,
                                                    start_time__date__lte=this_month_end,
                                                    plan_schedule__work_schedule__work_procedure__global_name='密炼')\
                .values('group__global_name', 'start_time__date').order_by('start_time')
        else:
            basic_info = WeightClassPlan.objects.filter(target_month=date, delete_flag=False)
            if not basic_info:
                return Response({'results': [], 'group_list': [], 'export_flag': False, 'state': 0,
                                 'audit_user': None, 'user_groups': [], 'approve_user': None,
                                 's_choice': [], 'm_choice': []})
                # raise ValidationError(f'未找到当月{clock_type}排班信息')
            groups = basic_info.values_list('classes', flat=True).distinct()
            days = basic_info.last().weight_class_details.all().values_list('factory_date', flat=True)
            dis_group = set([i.split('/')[0] for i in groups])
            group = [{'start_time__date': s_day, 'group__global_name': s_group.split('/')[0]} for s_day in days for s_group in dis_group]
        group_list = []
        for key, group in groupby(list(group), key=lambda x: x['start_time__date']):
            group_list.append([item['group__global_name'] for item in group])

        # 获取当前登录人员的班组
        user_groups = get_user_group(self.request.user.username, clock_type)
        results = {}
        data = EmployeeAttendanceRecords.objects.filter(
            Q(Q(end_date__isnull=True, begin_date__isnull=True) | Q(begin_date__isnull=False, end_date__isnull=False)) &
            Q(factory_date__year=year, factory_date__month=month, user__username__icontains=name) &
            ~Q(is_use='废弃'), group__in=user_groups, clock_type=clock_type
            ).values(
            'equip', 'section', 'group', 'factory_date__day', 'user__username', 'actual_time', 'record_status')
        for item in data:
            equip = item['equip']
            section = item['section']
            if not results.get(f'{equip}_{section}'):
                results[f'{equip}_{section}'] = {'equip': equip, 'section': section}
            value = item['user__username'] if item['actual_time'] == 12 else '%s(%.2f)' % (item['user__username'], item['actual_time'])
            detail = {'name': value, 'color': item['record_status']}
            data = results[f'{equip}_{section}'].get(f"{item['factory_date__day']}{item['group']}")
            if data:
                # 同一天同班组同一人同岗位同机台时间累加
                for i in data:
                    split_name = re.split('[(|)]', i['name'])
                    if split_name[0] == item['user__username']:
                        total_time = item['actual_time'] + float(split_name[1] if len(split_name) != 1 else 12)
                        name = item['user__username'] if int(total_time) == 12 else '%s(%.2f)' % (item['user__username'], total_time)
                        i.update(name=name)
                        break
                else:
                    data.append(detail)
            else:
                results[f'{equip}_{section}'][f"{item['factory_date__day']}{item['group']}"] = [detail]
        res = list(results.values())
        for item in res:
            item['equip'] = '' if not item.get('equip') else item['equip']
            item['sort'] = 2 if not item.get('equip') else 1
        results_sort = sorted(list(results.values()), key=lambda x: (x['sort'], x['equip']))
        # 当月考勤是否全部审核审批(已审核审批不能添加考勤数据)
        audit_user, approve_user, state = None, None, 0  # state 0 空白 1 确认中 2 已审批 3已审核
        attendance_data = EmployeeAttendanceRecords.objects.filter(~Q(is_use='废弃'),
                                                                   Q(begin_date__isnull=False, end_date__isnull=False)
                                                                   | Q(begin_date__isnull=True, end_date__isnull=True),
                                                                   factory_date__in=days_cur_month_dates(),
                                                                   clock_type=clock_type)
        if attendance_data:
            a_records = attendance_data.exclude(opera_flag__in=[2, 3])
            if a_records:
                state = 1
            else:
                basic_info = AttendanceResultAudit.objects.filter(date=date, clock_type=clock_type).order_by('id')
                audit_approve = basic_info.last()
                if audit_approve and audit_approve.result:
                    if audit_approve.audit_user:  # 审核完成
                        state, audit_user = 3, audit_approve.audit_user
                        approve_info = basic_info.filter(id=audit_approve.id - 1).last()
                        if approve_info:
                            approve_user = approve_info.approve_user
                    else:  # 审批完成
                        state, approve_user = 2, audit_approve.approve_user
        # 增加能否导出的标记  08-19:默认可以导出[去除审核以后才能导出的限制]
        export_flag = True
        # 返回岗位与机台关联
        s_choice = list(PerformanceJobLadder.objects.filter(type=clock_type, relation=1, delete_flag=False).values_list('name', flat=True).distinct())
        m_choice = list(PerformanceJobLadder.objects.filter(type=clock_type, relation=2, delete_flag=False).values_list('name', flat=True).distinct())
        return Response({'results': results_sort, 'group_list': group_list, 'export_flag': export_flag, 'state': state,
                         'audit_user':  audit_user, 'user_groups': user_groups, 'approve_user': approve_user,
                         's_choice': s_choice, 'm_choice': m_choice})

    # 导入出勤记录
    @atomic
    def post(self, request):
        date = self.request.data.get('date')  # 2022-02
        excel_file = request.FILES.get('file', None)
        if not excel_file:
            raise ValidationError('文件不可为空！')
        cur_sheet = get_cur_sheet(excel_file)
        rows = cur_sheet.nrows
        cols = cur_sheet.ncols
        if cur_sheet.row_values(0)[1] != '日期' or cur_sheet.row_values(1)[1] != '班次' or cur_sheet.row_values(2)[1] != '班别':
            raise ValidationError("导入的格式有误")
        # 获取班组
        year = int(date.split('-')[0])
        month = int(date.split('-')[1])
        if (datetime.date(year, month + 1, 1) - datetime.timedelta(days=1)).day != (cols - 2) / 2:
            raise ValidationError("导入的格式有误")
        this_month_start = datetime.datetime(year, month, 1)
        if month == 12:
            this_month_end = datetime.datetime(year + 1, 1, 1) - timedelta(days=1)
        else:
            this_month_end = datetime.datetime(year, month + 1, 1) - timedelta(days=1)

        group = WorkSchedulePlan.objects.filter(start_time__date__gte=this_month_start,
                                                start_time__date__lte=this_month_end).values('group__global_name', 'classes__global_name', 'start_time__day').order_by('start_time')
        group_dic = {}
        classes_dic = {}
        for item in group:
            day = item['start_time__day']
            if group_dic.get(day):
                group_dic[day].append(item['group__global_name'])
            else:
                group_dic[day] = [item['group__global_name']]
            if classes_dic.get(day):
                classes_dic[day].append(item['classes__global_name'])
            else:
                classes_dic[day] = [item['classes__global_name']]
        group_list, classes_list = list(group_dic.values()), list(classes_dic.values())
        start_row = 3
        equip_list = []
        for i in range(start_row, rows):
            equip_list.append(cur_sheet.cell(i, 0).value)
        for i in equip_list:
            index = equip_list.index(i)
            if not equip_list[index]:
                equip_list[index] = equip_list[index - 1]
        section_list = []
        for i in range(start_row, rows):
            section_list.append(cur_sheet.cell(i, 1).value)
        rows_num = cur_sheet.nrows  # sheet行数
        if rows_num <= start_row:
            raise ValidationError('没有可导入的数据')
        ret = [None] * (rows_num - start_row)
        for i in range(start_row, rows_num):
            ret[i - start_row] = cur_sheet.row_values(i)[2:]
        data = ret
        records_lst = []
        # 判断出勤记录是否存在，存在就更新
        records = list(EmployeeAttendanceRecords.objects.filter(factory_date__year=year, factory_date__month=month).values(
            'user', 'section', 'factory_date', 'group', 'equip'))
        for names in data:
            index = data.index(names)
            for i, name in enumerate(names):
                if name:
                    day = (i // 2) + 1
                    date_ = f'{date}-{day}'
                    user = User.objects.filter(username=name).first()
                    if not user:
                        raise ValidationError(f'系统中不存在{name}用户')
                    section = section_list[index]
                    group = group_list[day - 1][i % 2]
                    classes = classes_list[day - 1][i % 2]
                    equip = equip_list[index]
                    if equip in ['辅助', '班长']:
                        equips = [None]
                    elif ',' in equip:
                        equips = equip.split(',')
                    else:
                        equips = [equip]
                    for equip in equips:
                        dic = {'user': user, 'section': section, 'factory_date': date_, 'group': group, 'classes': classes, 'equip': equip, 'work_time': 12, 'actual_time': 12}
                        if dic in records:
                            continue
                        records_obj = EmployeeAttendanceRecords(**dic)
                        records_lst.append(records_obj)
        EmployeeAttendanceRecords.objects.bulk_create(records_lst)
        return Response(f'导入成功')


@method_decorator([api_recorder], name="dispatch")
class EmployeeAttendanceRecordsExport(ViewSet):
    permission_classes = (IsAuthenticated,)

    # 导出模板
    def export(self, request):
        response = HttpResponse(content_type='application/vnd.ms-excel')
        filename = '员工出勤记录表'
        response['Content-Disposition'] = u'attachment;filename= ' + filename.encode('gbk').decode(
            'ISO-8859-1') + '.xls'
        # 创建一个文件对象
        wb = xlwt.Workbook(encoding='utf8')
        # 创建一个sheet对象
        sheet = wb.add_sheet('sheet1', cell_overwrite_ok=True)
        alignment = xlwt.Alignment()
        alignment.horz = xlwt.Alignment.HORZ_CENTER
        alignment.vert = xlwt.Alignment.VERT_CENTER
        style = xlwt.XFStyle()
        style.alignment = alignment

        # 工厂排班计划
        date = self.request.query_params.get('date')
        year = int(date.split('-')[0])
        month = int(date.split('-')[1])
        this_month_start = datetime.datetime(year, month, 1)
        if month == 12:
            this_month_end = datetime.datetime(year + 1, 1, 1) - timedelta(days=1)
        else:
            this_month_end = datetime.datetime(year, month + 1, 1) - timedelta(days=1)
        # 获取班组
        group = WorkSchedulePlan.objects.filter(start_time__date__gte=this_month_start,
                                                start_time__date__lte=this_month_end).values('group__global_name', 'classes__global_name', 'start_time__day').order_by('start_time')
        group_dic = {}
        classes_dic = {}
        for item in group:
            day = item['start_time__day']
            if group_dic.get(day):
                group_dic[day].append(item['group__global_name'])
            else:
                group_dic[day] = [item['group__global_name']]
            if classes_dic.get(day):
                classes_dic[day].append(item['classes__global_name'])
            else:
                classes_dic[day] = [item['classes__global_name']]
        group_list, classes_list = list(group_dic.values()), list(classes_dic.values())


        # 添加标题
        # sheet.write_merge(开始行, 结束行, 开始列, 结束列, 'My merge', style)
        sheet.write_merge(0, 2, 0, 0, '机台', style)
        sheet.write_merge(0, 0, 1, 1, '日期', style)
        sheet.write_merge(1, 1, 1, 1, '班次', style)
        sheet.write_merge(2, 2, 1, 1, '班别', style)
        for i in range(len(group_list)):
            sheet.write_merge(0, 0, 2 * (i + 1), 2 * (i + 1) + 1, f'{i+1}日', style)
            sheet.write_merge(1, 1, 2 * (i + 1), 2 * (i + 1), group_list[i][0], style)
            sheet.write_merge(1, 1, 2 * (i + 1) + 1, 2 * (i + 1) + 1, group_list[i][1], style)

            sheet.write_merge(2, 2, 2 * (i + 1), 2 * (i + 1), classes_list[i][0], style)
            sheet.write_merge(2, 2, 2 * (i + 1) + 1, 2 * (i + 1) + 1, classes_list[i][1], style)

        # index = 2
        # for i in equip_list:
        #     sheet.write_merge(index, index, 0, 0, i['equip_no'], style)
        #     for section in section_list:
        #         if i['equip_no'].startswith('Z') and section['type'] == '密炼':
        #             sheet.write_merge(index, index, 1, 1, section['name'], style)
        #             index += 1
        #         if i['equip_no'].startswith('S') and section['type'] == '硫磺称量':
        #             sheet.write_merge(index, index, 1, 1, section['name'], style)
        #             index += 1
        #         if i['equip_no'].startswith('F') and section['type'] == '细料称量':
        #             sheet.write_merge(index, index, 1, 1, section['name'], style)
        #             index += 1

                # 从几个开始 （）
            # index = (list(equip_list).index(i) + 1) * 4      :  4,  8, 12
            # sheet.write_merge(index - 2, index - 2, 0, 0, i['equip_no'] + f"({dic.get(i['equip_no'], '')})", style)
            # sheet.write_merge(index - 2, index - 2, 0, 0, i['equip_no'], style)
            # sheet.write_merge(index - 2, index - 2, 1, 1, '主投', style)
            # sheet.write_merge(index - 1, index - 1, 1, 1, '辅投', style)
            # sheet.write_merge(index, index, 1, 1, '挤出', style)
            # sheet.write_merge(index + 1, index + 1, 1, 1, '收皮', style)

        # 写出到IO
        output = BytesIO()
        wb.save(output)
        # 重新定位到开始
        output.seek(0)
        response.write(output.getvalue())
        return response


@method_decorator([api_recorder], name="dispatch")
class PerformanceJobLadderViewSet(ModelViewSet):
    queryset = PerformanceJobLadder.objects.filter(delete_flag=False).order_by('code')
    serializer_class = PerformanceJobLadderSerializer
    permission_classes = (IsAuthenticated,)
    filter_backends = (DjangoFilterBackend,)
    filter_class = PerformanceJobLadderFilter

    def list(self, request, *args, **kwargs):
        if self.request.query_params.get('all'):
            if self.request.query_params.get('weight'):
                check_type = self.request.query_params.get('type', '')
                res = PerformanceJobLadder.objects.filter(delete_flag=False, type=check_type).values('name').annotate(c=Count('id')).values('name')
            else:
                res = self.filter_queryset(self.get_queryset()).values('id', 'type', 'name')
            return Response({'results': res})
        return super().list(request, *args, **kwargs)


@method_decorator([api_recorder], name="dispatch")
class PerformanceUnitPriceView(APIView):
    permission_classes = (IsAuthenticated,)

    def get(self, request):
        results = {}
        state = self.request.query_params.get('state')
        category_lst = ['E580', 'F370', 'GK320', 'GK255', 'GK400', 'fz']
        if state:  # 历史价格变化
            state_data = PerformanceUnitPrice.objects.filter(state=state, equip_type__in=category_lst).order_by('target_month', 'equip_type').values('equip_type', 'pt', 'dj', 'target_month')
            for i in state_data:
                target_month, equip_type, pt, dj = i['target_month'], i['equip_type'], i['pt'], i['dj']
                if target_month in results:
                    results[target_month].update({f"{equip_type}_pt": pt, f"{equip_type}_dj": dj})
                else:
                    results[target_month] = {'target_month': target_month, f"{equip_type}_pt": pt, f"{equip_type}_dj": dj}
        else:
            state_lst = GlobalCode.objects.filter(global_type__type_name='胶料段次').values('global_name')
            # category_lst = EquipCategoryAttribute.objects.filter(delete_flag=False).values('category_no')
            for state in state_lst:
                state = state['global_name']
                results[state] = {'state': state}
                for category in category_lst:
                    results[state][f"{category}_pt"] = None
                    results[state][f"{category}_dj"] = None
            p = PerformanceUnitPrice.objects.filter(target_month__isnull=False).order_by('target_month').last()
            if p:
                queryset = PerformanceUnitPrice.objects.filter(target_month=p.target_month).values('state', 'equip_type', 'pt', 'dj')
                for item in queryset:
                    results[f"{item['state']}"][f"{item['equip_type']}_pt"] = item['pt']
                    results[f"{item['state']}"][f"{item['equip_type']}_dj"] = item['dj']
        return Response({'result': results.values()})

    @atomic
    def post(self, request):
        data = self.request.data  # list
        username = self.request.user.username
        category_lst = ['E580', 'F370', 'GK320', 'GK255', 'GK400', 'fz']
        # 当前日期
        factory_date = get_current_factory_date()['factory_date']
        for item in data:
            for category in category_lst:
                pt = item[f"{category}_pt"]
                dj = item[f"{category}_dj"]
                PerformanceUnitPrice.objects.update_or_create(defaults={'state': item['state'], 'equip_type': category,
                                                                        'pt': pt, 'dj': dj, 'opera_user': username,
                                                                        'target_month': factory_date},
                                                              state=item['state'], equip_type=category, target_month=factory_date)

        return Response('添加成功')


@method_decorator([api_recorder], name="dispatch")
class ProductInfoDingJiViewSet(ModelViewSet):
    queryset = ProductInfoDingJi.objects.filter(delete_flag=False)
    serializer_class = ProductInfoDingJiSerializer
    permission_classes = (IsAuthenticated,)
    filter_backends = (DjangoFilterBackend,)
    filter_class = ProductInfoDingJiFilter

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        instance.delete_flag = True
        instance.delete_date = datetime.datetime.now()
        instance.delete_user = self.request.user
        instance.save()
        return Response(status=status.HTTP_204_NO_CONTENT)


@method_decorator([api_recorder], name="dispatch")
class SetThePriceViewSet(ModelViewSet):
    queryset = SetThePrice.objects.all()
    serializer_class = SetThePriceSerializer
    permission_classes = (IsAuthenticated,)


@method_decorator([api_recorder], name="dispatch")
class PerformanceSummaryView(APIView):
    permission_classes = (IsAuthenticated,)

    def get_unit(self, date):
        price_dic, index_list = {}, []
        # 当月1号是否存在单价设置
        day_one = PerformanceUnitPrice.objects.filter(target_month=f"{date}-01").last()
        if not day_one:
            p = PerformanceUnitPrice.objects.filter(target_month__lt=f"{date}-01").order_by('target_month').last()
            if p:
                now_units = PerformanceUnitPrice.objects.filter(target_month=p.target_month).values('state', 'equip_type', 'pt', 'dj')
                s_data = {}
                for item in now_units:
                    s_data[f"{item['equip_type']}_{item['state']}"] = {'pt': item['pt'], 'dj': item['dj']}
                price_dic = {1: s_data}
                index_list.append(1)
        month_data = PerformanceUnitPrice.objects.filter(target_month__startswith=date).order_by('target_month').values('state', 'equip_type', 'pt', 'dj', 'target_month')
        for i in month_data:
            day = i['target_month'].day
            if day in price_dic:
                price_dic[day][f"{i['equip_type']}_{i['state']}"] = {'pt': i['pt'], 'dj': i['dj']}
            else:
                price_dic[day] = {f"{i['equip_type']}_{i['state']}": {'pt': i['pt'], 'dj': i['dj']}}
            if day not in index_list:
                index_list.append(day)
        return price_dic, index_list

    def concat_user_train(self, key, detail, equip_dic, price_info, index_list, dj_list, date, qty_data, r_time):
        day, group, section, equip, classes, user_name = key.split('_')
        _day = bisect(index_list, int(day)) - 1
        price_dic = price_info.get(index_list[_day], {})
        calculate_begin, calculate_end, standard_begin, standard_end = detail.pop('calculate_begin_date'), detail.pop('calculate_end_date'), detail.pop('standard_begin_date'), detail.pop('standard_end_date')
        begin_date = (standard_begin - timedelta(minutes=r_time)) if calculate_begin <= standard_begin else calculate_begin
        end_date = (standard_end - timedelta(minutes=r_time)) if calculate_end >= standard_end else calculate_end
        k = f"{equip}-{classes}-{begin_date.strftime('%Y-%m-%d-%H-%M-%S')}-{end_date.strftime('%Y-%m-%d-%H-%M-%S')}"
        _data = qty_data.get(k)
        if _data:
            # 更新单价
            for i in _data:
                _i = i.split('_')
                if _i[-1] == 'unit':
                    m, n = '_'.join(_i[:2]), i[2]
                    if price_dic.get(m) and price_dic.get(m).get(n):
                        _data[i] = price_dic.get(m).get(n)
            detail.update(_data)
        else:
            equip_kwargs = {'equip_no': equip, 'classes': classes}
            add_qty = ManualInputTrains.objects.filter(**equip_kwargs, factory_date=f'{date}-' + '%02d' % int(day)).values('id', 'actual_trains', 'product_no')
            if equip == 'Z04':
                equip_kwargs['operation_user'] = 'Mixer1'
            equip_kwargs.update({'begin_time__gte': begin_date, 'end_time__lte': end_date, 'factory_date__day': day})
            query_set = TrainsFeedbacks.objects.filter(**equip_kwargs).values('product_no').annotate(actual_trains=Count('id')).values('actual_trains', 'product_no')
            ready_data = list(query_set) + list(add_qty)
            for item in ready_data:
                m_id, equip_type = item.get('id'), equip_dic.get(equip)
                try:
                    state = item['product_no'].split('-')[1]
                    if state in ['XCJ', 'DJXCJ']:
                        if item['product_no'].split('-')[0] == 'FM':
                            state = 'RFM'
                        elif item['product_no'].split('-')[0] == 'WL':
                            state = 'RMB'
                        else:
                            continue
                except:
                    continue
                if not price_dic.get(f"{equip_type}_{state}"):
                    # PerformanceUnitPrice.objects.create(state=state, equip_type=equip_type, dj=1.2, pt=1.1)
                    price_dic.update({f"{equip_type}_{state}": {'pt': 1.2, 'dj': 1.1}, f"fz_{state}": {'pt': 1.2, 'dj': 1.1}})
                # 判断是否是丁基胶
                frame_type = 'dj' if item['product_no'] in dj_list else 'pt'
                # 根据工作时长求机台的产量
                work_time = detail['actual_time']
                equip_type = 'fz' if section in ['三楼粉料', '吊料', '出库叉车', '叉车', '一楼叉车', '密炼叉车', '二楼出库'] else equip_type
                unit = price_dic.get(f"{equip_type}_{state}").get(frame_type)
                now_qty = (item['actual_trains'] / 12 * work_time) if m_id else item['actual_trains']
                if qty_data.get(k):
                    if qty_data[k].get(f"{equip_type}_{state}_{frame_type}_qty"):
                        qty_data[k][f"{equip_type}_{state}_{frame_type}_qty"] = round(qty_data[k][f"{equip_type}_{state}_{frame_type}_qty"] + now_qty, 2)
                    else:
                        qty_data[k].update({f"{equip_type}_{state}_{frame_type}_qty": round(now_qty, 2), f"{equip_type}_{state}_{frame_type}_unit": unit})
                else:
                    qty_data[k] = {f"{equip_type}_{state}_{frame_type}_qty": round(now_qty, 2), f"{equip_type}_{state}_{frame_type}_unit": unit}
            if qty_data.get(k):
                detail.update(qty_data[k])

    def get(self, request):
        date = self.request.query_params.get('date')
        export = self.request.query_params.get('export')
        name_d = self.request.query_params.get('name_d')
        day_d = self.request.query_params.get('day_d')
        group_d = self.request.query_params.get('group_d')
        ccjl = self.request.query_params.get('ccjl')
        year = int(date.split('-')[0])
        month = int(date.split('-')[1])
        if export:  # 导出员工实际考勤
            try:
                days, export_data = actual_clock_data(date, '密炼')
            except Exception as e:
                logger.error(e.args[0])
                raise ValidationError('导出实际考勤数据异常')
            if not export_data:
                raise ValidationError('无考勤数据可以导出')
            EXPORT_FIELDS_DICT = {'姓名': 'username', '岗位': 'section', '班组': 'group', '机台': 'equip'}
            add_key = {'/'.join(i.split('-')[1:]): i for i in days}
            EXPORT_FIELDS_DICT.update(add_key)
            return gen_template_response(EXPORT_FIELDS_DICT, export_data.values(), f'{date}密炼实际考勤数据', sheet_name=date, handle_str=True)
        # 员工独立上岗系数
        coefficient = GlobalCode.objects.filter(global_type__type_name='是否独立上岗系数', global_type__use_flag=True, use_flag=True).values('global_no', 'global_name')
        coefficient_dic = {dic['global_no']: dic['global_name'] for dic in coefficient}
        # 员工类别
        employee_type = GlobalCode.objects.filter(global_type__type_name='员工类别', global_type__use_flag=True, use_flag=True).values('global_no', 'global_name')
        employee_type_dic = {dic['global_no']: dic['global_name'] for dic in employee_type}
        # 超产奖励系数
        coefficient1 = GlobalCode.objects.filter(global_type__type_name='超产单价', global_type__use_flag=True, use_flag=True).values('global_no', 'global_name')
        coefficient1_dic = {dic['global_no']: dic['global_name'] for dic in coefficient1}
        if not coefficient or not coefficient1:
            raise ValidationError('请先去添加独立上岗或超产奖励系数')
        # 员工考勤记录
        section_info = {}
        for item in PerformanceJobLadder.objects.filter(type='密炼').values('name', 'coefficient', 'post_standard', 'post_coefficient', 'type'):
            section_info[item['name']] = {'coefficient': item['coefficient'], 'post_standard': item['post_standard'],
                                          'post_coefficient': item['post_coefficient'], 'type': item['type']}
        section_list = set(section_info.keys())
        kwargs = {
            'factory_date__year': year,
            'factory_date__month': month,
            'section__in': section_list
        }
        kwargs2 = {
            'factory_date__year': year,
            'factory_date__month': month,
        }
        if name_d:
            kwargs['user__username__icontains'] = name_d
        if day_d:
            kwargs['factory_date__day'] = day_d
            kwargs2['factory_date__day'] = day_d
        user_query = EmployeeAttendanceRecords.objects.filter(Q(**kwargs) & ~Q(is_use__in=['废弃', '驳回']),
                                                              end_date__isnull=False, begin_date__isnull=False,
                                                              clock_type='密炼').order_by('user', 'factory_date', 'equip')\
            .values_list('user__username', 'section', 'factory_date__day', 'group', 'equip', 'actual_time', 'classes',
                         'calculate_begin_date', 'calculate_end_date', 'standard_begin_date', 'standard_end_date')
        user_dic = {}
        equip_dic = {}
        equip_list = Equip.objects.filter(category__equip_type__global_name='密炼设备', equip_no__startswith='Z').values('category__category_no', 'equip_no')
        for item in equip_list:
            equip_dic[item['equip_no']] = 'GK400' if item['category__category_no'].startswith('GK400') else item['category__category_no']
        for item in user_query:
            key = f"{item[2]}_{item[3]}_{item[1]}_{item[4]}_{item[6]}_{item[0]}"  # 1_A班_挤出_Z01_早班
            if user_dic.get(key):  # 可能出现调岗后又换回来的情况，两次时间累加
                user_dic[key]['actual_time'] = user_dic[key].get('actual_time', 0) + item[5]
            else:
                user_dic[key] = {'name': item[0], 'section': item[1], 'day': item[2], 'group': item[3],
                                 'equip': item[4], 'actual_time': item[5], 'classes': item[6],
                                 'calculate_begin_date': item[7], 'calculate_end_date': item[8],
                                 'standard_begin_date': item[9], 'standard_end_date': item[10]}
        group = WorkSchedulePlan.objects.filter(start_time__year=year, start_time__month=month,
                                                plan_schedule__work_schedule__work_procedure__global_name='密炼')\
            .values_list('group__global_name', 'classes__global_name', 'start_time__day').order_by('start_time')
        group_list = []
        for key, g in groupby(list(group), key=lambda x: x[2]):
            group_list.append([item[0] for item in g])
        # 密炼交接班产量弹性时间
        r_instance = GlobalCode.objects.filter(global_type__type_name='密炼交接班产量弹性时间', global_type__use_flag=True, use_flag=True).last()
        try:
            r_time = int(r_instance.global_name)
        except:
            r_time = 0
        # 单价
        price_info, index_list = self.get_unit(date)
        if not price_info:
            raise ValidationError(f'{date}单价设置异常')
        dj_list = ProductInfoDingJi.objects.filter(is_use=True).values_list('product_name', flat=True)
        qty_data, t_num = {}, 4
        pool = ThreadPool(t_num)
        for key, detail in user_dic.items():
            pool.apply_async(self.concat_user_train, args=(key, detail, equip_dic, price_info, index_list, dj_list, date, qty_data, r_time))
        pool.close()
        pool.join()
        results1 = {}
        # 是否独立上岗
        independent = {}
        independent_lst = IndependentPostTemplate.objects.all().values('name', 'status', 'work_type')
        for item in independent_lst:
            independent[item['name']] = {'status': item['status'], 'work_type': item['work_type']}
        if not independent:
            raise ValidationError(f'请添加员工类别')
        # 取每个机台设定的目标值
        max_setting = MachineTargetYieldSettings.objects.filter(target_month=date).order_by('day', 'classes').values()
        if not max_setting:
            raise ValidationError('请先完成当月的机台目标值设定')
        # 处理机台目标值
        target_info = {f"{i['day']}-{i['classes']}": i for i in max_setting}
        # 获取该月人员密炼完成率
        ratio_info = FinishRatio.objects.filter(target_month=date).values('username', 'ratio')
        ratio = {i['username']: i['ratio'] for i in ratio_info}
        # 计算薪资
        for k, v in user_dic.items():
            name, day, group, section = v['name'], v['day'], v['group'], v['section']
            key = f"{name}_{day}_{group}_{section}"  # 合并岗位
            if results1.get(key):
                results1[key].append(v)
            else:
                results1[key] = [v]
        ccjl_detail = {}  # 按月统计超产奖励
        ccjl_flag = {}  # 超产次数
        # 绩效详情
        if day_d and group_d:
            start_with = f"{name_d}_{day_d}_{group_d}"
            results1 = {k: v for k, v in results1.items() if k.startswith(start_with)}
            ccjl_dic = {}
            results = {}
            results_sort = {}
            equip_qty = {}
            equip_price = {}
            # 完成率
            s_ratio = ratio.get(name_d, 0)
            hj = {'name': '产量工资合计', 'price': 0, 'ccjl': 0}
            for item in results1.values():
                for dic in item:
                    section, equip = dic['section'], dic['equip']
                    if len(dic) == 7:
                        continue
                    if not equip_qty.get(section):
                        equip_qty[section] = {}
                        equip_price[section] = {}
                    _dic = {}
                    for k in dic.keys():
                        if k.split('_')[-1] == 'qty':
                            _dic = {'_'.join(i.split('_')[1:]): dic[i] for i in dic if i.split('_')[-1] in ['qty', 'unit']} if not _dic else _dic
                            state = k.split('_')[1]
                            type1 = k.split('_')[2]
                            qty = _dic.get(f"{state}_{type1}_qty")  # 数量
                            unit = _dic.get(f"{state}_{type1}_unit")  # 单价
                            equip_qty[section][equip] = equip_qty[section].get(equip, 0) + qty
                            equip_price[section][equip] = equip_price[section].get(equip, 0) + round(qty * unit, 2)

                            hj[state] = round(hj.get(state, 0) + qty * unit, 2)
                            type2 = '普通' if type1 == 'pt' else '丁基'
                            if results.get(f"{equip}_{type2}_{section}_1"):
                                results[f"{equip}_{type2}_{section}_1"].update({state: dic[k]})
                                results[f"{equip}_{type2}_{section}_2"].update({state: _dic[f"{state}_{type1}_unit"]})
                                results[f"{equip}_{type2}_{section}_3"].update({state: section})
                            else:
                                results[f"{equip}_{type2}_{section}_1"] = {'name': f"{equip}{type2}-车数", state: dic[k]}
                                results[f"{equip}_{type2}_{section}_2"] = {'name': f"{equip}{type2}-单价", state: _dic[f"{state}_{type1}_unit"]}
                                results[f"{equip}_{type2}_{section}_3"] = {'name': f"{equip}{type2}-岗位", state: section}

            for i in sorted(results):
                results_sort[i] = results.pop(i)
            a = float(coefficient_dic.get('是'))  # 是否独立上岗
            w_coefficient = 1  # 员工类别系数
            if independent.get(name_d):
                if independent.get(name_d).get('status') != 1:
                    a = float(coefficient_dic.get('否'))
                work_type = independent.get(name_d).get('work_type')
                w_coefficient = float(employee_type_dic.get(work_type))
            # 计算超产奖励
            res = get_current_factory_date(select_date=f"{date}-{'%02d' % int(day_d)}", group=group_d)
            s_target_setting = target_info.get(f"{day_d}-{res.get('classes')}")
            target_setting = s_target_setting if s_target_setting else list(target_info.values())[0]
            for section, equip_dic in equip_qty.items():
                coefficient = section_info[section]['coefficient'] / 100
                post_coefficient = section_info[section]['post_coefficient'] / 100
                post_standard = section_info[section]['post_standard']  # 1最大值 2 平均值
                post_coefficient = 1 if len(equip_qty) > 1 else post_coefficient

                price = 0
                for equip, qty in equip_dic.items():
                    m, s = target_setting.get(equip + '_max'), target_setting.get(equip)
                    if m and s:
                        if qty <= s:
                            price = 0
                        elif qty > m:
                            price = (m - s) * float(coefficient1_dic.get('超过目标产量部分')) + (qty - m) * float(
                                coefficient1_dic.get('超过最高值部分'))
                        elif s < qty <= m:
                            price = (qty - s) * float(coefficient1_dic.get('超过目标产量部分'))
                        if price == 0:
                            continue
                        if section in ['班长', '机动', '三楼粉料', '吊料', '出库叉车', '叉车', '一楼叉车', '密炼叉车', '二楼出库']:
                            _s = 0.2 if section == '班长' else 0.15
                            ccjl_dic[equip] = round(ccjl_dic.get(equip, 0) + price * a * w_coefficient * _s, 2)
                        else:
                            ccjl_dic[equip] = round(ccjl_dic.get(equip, 0) + price * a * w_coefficient * coefficient, 2)
                # 岗位不同，计算超产奖励的方式不同
                # if section in ['班长', '机动']:
                #     s_ratio = 1
                #     hj['ccjl'] += round(sum(ccjl_dic.values()) / len(ccjl_dic), 2) if ccjl_dic.values() else 0
                # elif section in ['三楼粉料', '吊料', '出库叉车', '叉车', '一楼叉车', '密炼叉车', '二楼出库']:
                #     hj['ccjl'] += round(sum(ccjl_dic.values()) * 0.2 * coefficient, 2) if ccjl_dic.values() else 0
                # else:
                #     if len(equip_dic.values()) > 1:
                #         if post_standard == 1:  # 最大值
                #             hj['ccjl'] += round(max(ccjl_dic.values()), 2) if ccjl_dic.values() else 0
                #         else:
                #             hj['ccjl'] += round(sum(ccjl_dic.values()) / (len(results_sort) // 3),
                #                                2) if ccjl_dic.values() else 0
                #     else:
                #         hj['ccjl'] = round(max(ccjl_dic.values()), 2) if ccjl_dic.values() else 0
                if section in ['班长', '机动', '三楼粉料', '吊料', '出库叉车', '叉车', '一楼叉车', '密炼叉车', '二楼出库']:
                    s_ratio = 1
                    hj['ccjl'] = str(round(sum(ccjl_dic.values()) / len(ccjl_dic), 2) if ccjl_dic.values() else 0)
                # elif section in ['三楼粉料', '吊料', '出库叉车', '叉车', '一楼叉车', '密炼叉车', '二楼出库']:
                #     hj['ccjl'] = str(round(sum(ccjl_dic.values()) * 0.2 * coefficient, 2) if ccjl_dic.values() else 0)
                else:
                    hj['ccjl'] = '、'.join([str(i) for i in ccjl_dic.values()])
                k = equip_price[section].values()
                if post_standard == 1:
                    hj['price'] = hj.get('price', 0) + (round(max(k) * post_coefficient * coefficient * a * w_coefficient, 2) if k else 0)
                else:
                    hj['price'] = hj.get('price', 0) + (round(sum(k) / len(k) * post_coefficient * coefficient * a * w_coefficient, 2) if k else 0)
                hj['price'] = round(hj['price'] * s_ratio, 2)
            return Response({'results': results_sort.values(), 'hj': hj, 'all_price': hj['price'], '超产奖励': hj['ccjl'], 'group_list': group_list})

        results = {}
        ccjl_date = {}
        equip_qty = {}
        equip_price = {}

        for item in list(results1.values()):
            section, name, day, group, classes = item[0].get('section'), item[0].get('name'), item[0].get('day'), item[0].get('group'), item[0].get('classes')
            s_ratio = ratio.get(name, 0) if section not in ['班长', '机动', '三楼粉料', '吊料', '出库叉车', '叉车', '一楼叉车', '密炼叉车', '二楼出库'] else 1
            key = f"{name}_{day}_{section}"
            _date = f"{date}-{day}"
            a = float(coefficient_dic.get('是'))  # 是否独立上岗系数
            aa = '是'  # 是否独立上岗
            w_coefficient = 1  # 员工类别系数
            if independent.get(name):
                if independent.get(name).get('status') != 1:
                    a = float(coefficient_dic.get('否'))
                    aa = '否'
                work_type = independent.get(name).get('work_type')
                w_coefficient = float(employee_type_dic.get(work_type))
            coefficient = section_info[section]['coefficient'] / 100
            post_coefficient = section_info[section]['post_coefficient'] / 100
            post_standard = section_info[section]['post_standard']  # 1最大值 2 平均值
            for dic in item:
                equip = dic.get('equip')
                if len(dic) == 7:
                    continue
                _dic = {}
                for k in dic.keys():
                    if k.split('_')[-1] == 'qty':
                        _dic = {'_'.join(i.split('_')[1:]): dic[i] for i in dic if i.split('_')[-1] in ['qty', 'unit']} if not _dic else _dic
                        state = k.split('_')[1]
                        type1 = k.split('_')[2]
                        qty = _dic.get(f"{state}_{type1}_qty")  # 数量
                        unit = _dic.get(f"{state}_{type1}_unit")  # 单价
                        # 统计车数
                        if equip_qty.get(key):
                            equip_qty[key][equip] = equip_qty[key].get(equip, 0) + qty
                        else:
                            equip_qty[key] = {equip: qty}
                        if equip_price.get(key):
                            equip_price[key][equip] = equip_price[key].get(equip, 0) + qty * unit
                        else:
                            equip_price[key] = {equip: qty * unit}
            # 计算薪资
            if not equip_price.get(key):
                price = 0
            else:
                if post_standard == 1:  # 最大值
                    price = round(max(equip_price.get(key).values()), 2)
                else:  # 平均值
                    price = round(
                        sum(equip_price.get(key).values()) / len(equip_price.get(key)), 2)
                price = round(price * coefficient * a * post_coefficient * w_coefficient * s_ratio, 2)
            if results.get(name):
                results[name][f"{day}_{group}"] = round(results[name].get(f"{day}_{group}", 0) + price, 2)
                results[name]['hj'] = results[name].get('hj', 0) + price
                results[name]['all'] = results[name].get('all', 0) + price
            else:
                results[name] = {'name': name, '超产奖励': 0, '是否定岗': aa, 'hj': price, 'all': price,
                                 f"{day}_{group}": price}
            # 计算超产奖励
            p_dic = {}
            if not equip_qty.get(key):
                continue
            s_target_setting = target_info.get(f"{day}-{classes}")
            target_setting = s_target_setting if s_target_setting else list(target_info.values())[0]
            for equip, qty in equip_qty[key].items():
                m, s = target_setting.get(equip + '_max'), target_setting.get(equip)
                if m and s:
                    if qty <= s:
                        price = 0
                    elif qty > m:
                        price = (m - s) * float(coefficient1_dic.get('超过目标产量部分')) + (qty - m) * float(
                            coefficient1_dic.get('超过最高值部分'))
                    elif s < qty <= m:
                        price = (qty - s) * float(coefficient1_dic.get('超过目标产量部分'))
                    if price == 0:
                        continue
                    # 记录超产奖励明细[班长和机动一类计算方式、其他人一类]
                    s_ccjl = ccjl_detail.get(name)
                    if section in ['班长', '机动', '三楼粉料', '吊料', '出库叉车', '叉车', '一楼叉车', '密炼叉车', '二楼出库']:
                        p_dic[equip] = round(price * a * w_coefficient, 2)  # 员工类别系数、是否独立上岗系数
                        if s_ccjl:
                            if equip in s_ccjl:
                                s_ccjl[equip] = s_ccjl[equip] + p_dic[equip]
                            else:
                                s_ccjl[equip] = p_dic[equip]
                            c_time = ccjl_flag.get(f"{name}-{_date}")  # 记录当天第几次超产
                            if c_time:
                                ccjl_date[name][_date]['price'].append(p_dic[equip])
                                ccjl_flag[f"{name}-{_date}"] = ccjl_flag.get(f"{name}-{_date}", 0) + 1
                            else:
                                ccjl_flag[f"{name}-{_date}"] = 1
                                ccjl_date[name][_date] = {'date': _date, 'price': [p_dic[equip]]}
                        else:
                            ccjl_detail[name] = {equip: p_dic[equip]}
                            ccjl_flag[f"{name}-{_date}"] = 1
                            ccjl_date[name] = {_date: {'date': _date, 'price': [p_dic[equip]]}}
                    else:
                        p_dic[equip] = round(price * a * w_coefficient * coefficient, 2)  # 绩效需要乘绩效系数、员工类别系数、是否独立上岗系数
                        if s_ccjl:
                            c_time = ccjl_flag.get(f"{name}-{_date}")
                            if c_time:
                                ccjl_detail[name][c_time + 1] = ccjl_detail[name].get(c_time + 1, 0) + p_dic[equip]
                                ccjl_date[name][_date]['price'].append(p_dic[equip])
                                ccjl_flag[f"{name}-{_date}"] = ccjl_flag.get(f"{name}-{_date}", 0) + 1
                            else:
                                ccjl_detail[name][1] = ccjl_detail[name].get(1, 0) + p_dic[equip]
                                ccjl_flag[f"{name}-{_date}"] = 1
                                ccjl_date[name][_date] = {'date': _date, 'price': [p_dic[equip]]}
                        else:
                            ccjl_detail[name] = {1: p_dic[equip]}
                            ccjl_flag[f"{name}-{_date}"] = 1
                            ccjl_date[name] = {_date: {'date': _date, 'price': [p_dic[equip]]}}
            if section in ['班长', '机动', '三楼粉料', '吊料', '出库叉车', '叉车', '一楼叉车', '密炼叉车', '二楼出库']:
                _s = 0.2 if section == '班长' else 0.15
                p = round(sum(p_dic.values()) / len(p_dic) * _s, 2) if p_dic.values() else 0
                s_ccjl = ccjl_detail.get(name)
                if s_ccjl:
                    if ccjl_detail[name].get('all'):
                        ccjl_detail[name]['all'] = round(ccjl_detail[name]['all'] + p, 2)
                    else:
                        ccjl_detail[name]['all'] = p

            # elif section in ['三楼粉料', '吊料', '出库叉车', '叉车', '一楼叉车', '密炼叉车', '二楼出库']:
            #     p = round(sum(p_dic.values()) * 0.2 * coefficient, 2) if p_dic.values() else 0
            # else:
            #     if len(equip_qty[key].values()) > 1:
            #         if post_standard == 1:
            #             p = round(max(p_dic.values()), 2) if p_dic.values() else 0
            #         else:
            #             p = round(sum(p_dic.values()) / len(equip_qty[key].values()), 2) if p_dic.values() else 0
            #     else:
            #         p = round(max(p_dic.values()), 2) if p_dic.values() else 0
            # if p > 0:
            #     _d = f"{year}-{month}-{day}"
            #     if ccjl_dic.get(name):
            #         if ccjl_dic[name].get(_d):
            #             ccjl_dic[name][_d]['price'].append(p)
            #         else:
            #             ccjl_dic[name][_d] = {"date": _d, 'price': [p]}
            #     else:
            #         ccjl_dic[name] = {_d: {"date": _d, 'price': [p]}}
        if ccjl:  # 超产奖励详情
            res = ccjl_date.get(f"{name_d}", {}).values()
            len_data = [len(i['price']) for i in res]
            return Response({'results': res, 'max_id': max(len_data) if len_data else 0})
        for item in list(results.values()):
            # 其他奖惩+生产补贴
            additional = SubsidyInfo.objects.filter(name=item['name'], type__in=[1, 2], date__year=year,
                                                    date__month=month).aggregate(others=Sum('price', filter=Q(type=1)),
                                                                                 prods=Sum('price', filter=Q(type=2)))
            item['其他奖惩'] = additional['others'] if additional['others'] else 0
            item['生产补贴'] = additional['prods'] if additional['prods'] else 0
            item['all'] = item.get('all', 0) + (item['其他奖惩'] + item['生产补贴'])
            # 乘员工类别系数
            if independent.get(item['name']):
                work_type = independent[item['name']].get('work_type')
            else:
                work_type = '正常'
            # v = float(employee_type_dic.get(work_type, 1)) 其他奖惩和生产补贴直接累加
            v = 1
            item['work_type'] = work_type
            item['all'] = round(item['all'] * v, 2)
            item['hj'] = round(item['hj'], 2)
            # 并入月超产奖励
            s_ccjl = ccjl_detail.get(item['name'])
            # 存在员工升班长或者班长降员工场景(分别计算)
            if not s_ccjl:
                p = 0
            else:
                common_p = dict(filter(lambda x: isinstance(x[0], int), s_ccjl.items()))
                leader_p = s_ccjl.get('all', 0)
                p = leader_p + round(max(common_p.values(), default=0), 2)
            # p = 0 if not s_ccjl else (round(max(s_ccjl.values()), 2) if 'all' not in s_ccjl else s_ccjl.get('all'))
            item['超产奖励'] = round(p, 2)
            item['all'] = round(item['all'] + p, 2)
        return Response({'results': results.values(), 'group_list': group_list})


@method_decorator([api_recorder], name="dispatch")
class PerformanceSubsidyViewSet(ModelViewSet):
    queryset = SubsidyInfo.objects.order_by('id')
    serializer_class = SubsidyInfoSerializer
    permission_classes = (IsAuthenticated,)
    filter_backends = (DjangoFilterBackend,)
    filter_class = SubsidyInfoFilter
    pagination_class = None

    @atomic
    def create(self, request, *args, **kwargs):
        for data in request.data:
            price = data.get('price') if data.get('price') else 0
            desc = data.get('desc') if data.get('desc') else None
            if data.get('id'):
                SubsidyInfo.objects.filter(id=data.get('id')).update(price=price, desc=desc)
            else:
                serializer = SubsidyInfoSerializer(data=data)
                serializer.is_valid(raise_exception=True)
                serializer.save()
        return Response(status=status.HTTP_200_OK)


@method_decorator([api_recorder], name="dispatch")
class IndependentPostTemplateView(APIView):
    permission_classes = (IsAuthenticated,)

    def get(self, request):
        date = self.request.query_params.get('date')
        export = self.request.query_params.get('export')
        year = int(date.split('-')[0])
        month = int(date.split('-')[1])
        # 导出是否独立上岗模版

        if export:
            names = EmployeeAttendanceRecords.objects.filter(factory_date__year=year, factory_date__month=month, clock_type='密炼').values_list('user__username', flat=True).distinct()
            if names:
                data = [{'name': name, 'work_type': '正常', 'state': '是'} for name in list(names)]
                return gen_template_response({'姓名': 'name', '员工类别': 'work_type', '是否独立上岗': 'state'}, data, '是否独立上岗模版')
            raise ValidationError('当月没有考勤记录')
        return Response()

    @atomic
    def post(self, request):
        date = self.request.data.get('date')
        excel_file = request.FILES.get('file', None)
        if not excel_file:
            raise ValidationError('文件不可为空！')
        cur_sheet = get_cur_sheet(excel_file)
        rows = cur_sheet.nrows
        lst = []
        for i in range(1, rows):
            name = cur_sheet.cell(i, 0).value
            work_type = cur_sheet.cell(i, 1).value
            value = cur_sheet.cell(i, 2).value
            if value == '是':
                status = 1
            elif value == '否':
                status = 0
            else:
                raise ValidationError('导入的格式有误')

            if IndependentPostTemplate.objects.filter(date_time=date, name=name).exists():
                IndependentPostTemplate.objects.filter(date_time=date, name=name).update(status=status, work_type=work_type)
            else:
                lst.append(IndependentPostTemplate(name=name, status=status, date_time=date, work_type=work_type))
        IndependentPostTemplate.objects.bulk_create(lst)
        return Response(f'导入成功')


@method_decorator([api_recorder], name="dispatch")
class AttendanceGroupSetupViewSet(ModelViewSet):
    queryset = AttendanceGroupSetup.objects.all()
    serializer_class = AttendanceGroupSetupSerializer
    permission_classes = (IsAuthenticated,)
    filter_backends = (DjangoFilterBackend,)
    filter_class = AttendanceGroupSetupFilter
    option = True

    def get_queryset(self):
        attendance_users = self.request.query_params.get('attendance_users', None)
        queryset = self.filter_queryset(self.queryset)
        if attendance_users:
            queryset_ids = [obj.id for obj in queryset if attendance_users not in ','.join(list(obj.users.all().values_list('username', flat=True)))]
            queryset = queryset.exclude(id__in=queryset_ids)
        return queryset

    def get_section(self, section_name, section_id=None):  # 判断用户是不是设备科或生产科的成员
        if not self.option:
            return True
        if section_id:
            if Section.objects.filter(pk=section_id, name=section_name).exists():
                self.option = False
                return True
            a = Section.objects.filter(pk=section_id).first()
            if a:
                parent_id = a.parent_section_id
                self.get_section(section_name, parent_id)
        else:
            return False

    @action(methods=['get'], detail=False, permission_classes=[IsAuthenticated, ], url_path='in_group', url_name='in_group')
    def in_group(self, request):  # 判断用户是否在考勤组
        name = self.request.user.username
        state = False
        for obj in AttendanceGroupSetup.objects.all():
            if name in obj.principal:
                state = True
                break
            if name in obj.users.all().values_list('username', flat=True):
                state = True
                break

        return Response({'status': state})

    @action(methods=['get'], detail=False, permission_classes=[IsAuthenticated, ], url_path='is_section', url_name='is_section')
    def is_section(self, request):  # 判断用户是否是设备科门或生产科
        section_id = self.request.user.section_id
        for section in ['设备科', '生产科']:
            self.option = True
            self.get_section(section, section_id)
            if not self.option:
                return Response({'section': section, "permissions": self.request.user.permissions_list})
        return Response({'section': None, "permissions": self.request.user.permissions_list})


@method_decorator([api_recorder], name="dispatch")
class AttendanceClockViewSet(ModelViewSet):
    queryset = EmployeeAttendanceRecords.objects.all()
    serializer_class = EmployeeAttendanceRecordsSerializer
    permission_classes = (IsAuthenticated,)

    def save_attendance_clock_detail(self, name, equip_list, data, clock_type):
        AttendanceClockDetail.objects.create(
            name=name,
            equip=','.join(equip_list),
            group=data.get('group'),
            classes=data.get('classes'),
            section=data.get('section'),
            work_type=data.get('status'),
            clock_type=clock_type
        )

    def send_message(self, user, content):
        phone = user.phone_number
        ding_api = DinDinAPI()
        ding_uid = ding_api.get_user_id(phone)
        ding_api.send_message([ding_uid], content, attendance=True)

    def get_user_group(self, user_obj, now=None, apply=None):
        username =user_obj.username
        # 获取登陆用户所在考勤组
        flag, clock_type = get_user_weight_flag(user_obj)
        attendance_group_obj = None
        for obj in AttendanceGroupSetup.objects.filter(type=clock_type):
            if username in obj.principal:
                attendance_group_obj = obj
                break
            if username in obj.users.all().values_list('username', flat=True):
                attendance_group_obj = obj
                break

        if not attendance_group_obj:
            raise ValidationError(f'当前用户未添加至{clock_type}考勤组')
        group_type = attendance_group_obj.type   # 密炼/细料称量/硫磺称量
        if group_type == '密炼':
            equip_type = '密炼设备'
            equip_list = Equip.objects.filter(category__equip_type__global_name=equip_type, equip_no__startswith='Z').values_list('equip_no', flat=True)
        else:
            equip_type = '称量设备'
            equip_list = Equip.objects.filter(category__equip_type__global_name=equip_type).values_list('equip_no',  flat=True)
        section_list = PerformanceJobLadder.objects.filter(delete_flag=False, type=group_type).values_list('name', flat=True)

        # 获取当前时间的工厂日期
        now = now if now else (get_virtual_time() + timedelta(minutes=attendance_group_obj.lead_time))
        current_work_schedule_plan = WorkSchedulePlan.objects.filter(
            start_time__lte=now,
            end_time__gte=now,
            plan_schedule__work_schedule__work_procedure__global_name='密炼'
        ).first()
        if current_work_schedule_plan:
            date_now = str(current_work_schedule_plan.plan_schedule.day_time)
        else:
            date_now = str(now.date())
        if group_type == '密炼':
            # 获取班次班组
            queryset = WorkSchedulePlan.objects.filter(plan_schedule__day_time=date_now,
                                                       plan_schedule__work_schedule__work_procedure__global_name='密炼')\
                .values('group__global_name', 'classes__global_name')
            group_list = [{'group': item['group__global_name'], 'classes': item['classes__global_name']} for item in queryset]
        else:
            r_queryset = WeightClassPlanDetail.objects.filter(weight_class_plan__user=user_obj,
                                                              weight_class_plan__delete_flag=False)
            queryset = r_queryset.filter(factory_date=date_now).last()
            if queryset.class_code == '休' and not apply:  # 进入考勤页面时异常记录日志,补卡时忽略
                b_date = (datetime.datetime.strptime(date_now, '%Y-%m-%d') - timedelta(days=1)).strftime('%Y-%m-%d')
                last_obj = EmployeeAttendanceRecords.objects.filter(user=self.request.user, factory_date=b_date, end_date__isnull=True,
                                                                    clock_type=clock_type).order_by('factory_date', 'id').last()
                if last_obj:
                    queryset = r_queryset.filter(factory_date=b_date).last()
                    date_now = b_date
                else:
                    raise ValidationError('该称量员工今天休息')
            u_group = queryset.weight_class_plan.classes.split('/')[0] if queryset.weight_class_plan.classes.split('/') else None
            u_class = get_work_time(queryset.class_code, date_now).keys()
            group_list = [] if not all([u_group, u_class]) else [{'group': u_group, 'classes': i} for i in u_class]
        return attendance_group_obj, list(section_list), list(equip_list), date_now, group_list

    def list(self, request, *args, **kwargs):
        username = self.request.user.username
        id_card_num = self.request.user.id_card_num
        apply = self.request.query_params.get('apply', None)
        select_time = self.request.query_params.get('select_time')
        time_now = get_virtual_time()
        date_now = None
        try:
            attendance_group_obj, section_list, equip_list, date_now, group_list = self.get_user_group(self.request.user, select_time, apply)
        except Exception as e:
            # 查询审批不返回异常
            if apply:
                group_list, equip_list, section_list, principal = [], [], [], ''
            else:
                raise ValidationError(e.args[0])
        else:
            group_list, equip_list, section_list, principal = group_list, equip_list, section_list, attendance_group_obj.principal
        equip_list.sort()
        # 获取单选和多选机台的岗位
        s_choice, m_choice = [], []
        if equip_list:
            keyword = set([i[0] for i in equip_list if i.startswith('Z')])
            equip_type = '密炼' if 'Z' in keyword else '生产配料'
            s_choice = list(PerformanceJobLadder.objects.filter(type=equip_type, relation=1).values_list('name', flat=True).distinct())
            m_choice = list(PerformanceJobLadder.objects.filter(type=equip_type, relation=2).values_list('name', flat=True).distinct())
        results = {
            # 'ids': ids,  # 进行中的id，前端打卡传这个过来
            'username': username,
            'id_card_num': id_card_num,
            'group_list': group_list,
            'equip_list': equip_list,
            'section_list': section_list,
            's_choice': s_choice,
            'm_choice': m_choice,
            'principal': principal,  # 前端根据这个判断是否显示审批
        }
        if apply:  # 补卡/加班
            return Response({'results': results})

        # 判断最后一条的工厂时间是不是当天，是的话说明是正在进行中的
        last_obj = EmployeeAttendanceRecords.objects.filter(user=self.request.user, clock_type=attendance_group_obj.type).order_by('factory_date', 'id').last()
        if attendance_group_obj.type == '密炼':
            if last_obj:
                key_second = (last_obj.standard_end_date - last_obj.standard_begin_date).total_seconds()
                flat = False
                report = EmployeeAttendanceRecords.objects.filter(begin_date=last_obj.begin_date, clock_type='密炼',
                                                                  user_id=last_obj.user_id).values_list('equip', 'id')
                ids, equips = [item[1] for item in report], [item[0] for item in report]
                results['equips'] = sorted(list(set(equips))) if last_obj.section in (s_choice + m_choice) else []

                if str(last_obj.factory_date) == date_now:
                    # begin_time, end_time = get_standard_time(username, date_now)
                    begin_time, end_time = last_obj.standard_begin_date, last_obj.standard_end_date
                    if not begin_time:
                        raise ValidationError('未找到排班信息')
                    p_date_now = datetime.datetime.strptime(date_now, '%Y-%m-%d')
                    if end_time.hour > 12:  # 白班
                        # if time_now < p_date_now + datetime.timedelta(days=1, hours=2):  # 直到明天凌晨两点，显示当前的打卡记录
                        if time_now < end_time + datetime.timedelta(minutes=attendance_group_obj.leave_time):
                            results['state'] = 3  # 进行中的
                            results['ids'] = ids
                            results['begin_date'] = datetime.datetime.strftime(last_obj.begin_date, '%H:%M:%S') if last_obj.begin_date else None
                            results['end_date'] = datetime.datetime.strftime(last_obj.end_date, '%H:%M:%S') if last_obj.end_date else None
                            if last_obj.section in results['section_list']:
                                results['section_list'].remove(last_obj.section)
                            results['section_list'].insert(0, last_obj.section)  # 放到第一位显示
                            group, classes = last_obj.group, last_obj.classes
                            if {'group': group, 'classes': classes} in results['group_list']:
                                results['group_list'].remove({'group': group, 'classes': classes})
                            results['group_list'].insert(0, {'group': group, 'classes': classes})
                        else:
                            flat = True
                    else:  # 夜班
                        # if time_now < p_date_now + datetime.timedelta(days=1, hours=14):
                        if time_now < end_time + datetime.timedelta(minutes=attendance_group_obj.leave_time):
                            results['state'] = 3
                            results['ids'] = ids
                            results['begin_date'] = datetime.datetime.strftime(last_obj.begin_date, '%H:%M:%S') if last_obj.begin_date else None
                            results['end_date'] = datetime.datetime.strftime(last_obj.end_date, '%H:%M:%S') if last_obj.end_date else None
                            if last_obj.section in results['section_list']:
                                results['section_list'].remove(last_obj.section)
                            results['section_list'].insert(0, last_obj.section)  # 放到第一位显示
                            group, classes = last_obj.group, last_obj.classes
                            if {'group': group, 'classes': classes} in results['group_list']:
                                results['group_list'].remove({'group': group, 'classes': classes})
                            results['group_list'].insert(0, {'group': group, 'classes': classes})
                        else:
                            flat = True
                else:
                    if datetime.datetime.strptime(date_now, '%Y-%m-%d').date() > last_obj.factory_date + timedelta(days=1) \
                            or (not last_obj.end_date and (time_now - last_obj.begin_date).total_seconds() >
                                key_second + (attendance_group_obj.lead_time + attendance_group_obj.leave_time) * 60) \
                            or (last_obj.end_date and (time_now - last_obj.end_date).total_seconds() > attendance_group_obj.leave_time * 60):
                        flat = True
                    else:
                        if datetime.datetime.strptime(date_now, '%Y-%m-%d').date() == last_obj.factory_date + timedelta(days=1) and not last_obj.end_date:
                            begin_time, end_time, date_now = last_obj.standard_begin_date, last_obj.standard_end_date, str(last_obj.factory_date)
                        else:
                            begin_time, end_time = get_standard_time(username, date_now)
                        if not begin_time:
                            raise ValidationError('未找到排班信息')
                        p_date_now = datetime.datetime.strptime(date_now, '%Y-%m-%d')
                        if end_time.hour > 12:  # 白班
                            # if time_now < p_date_now + datetime.timedelta(days=1, hours=2):  # 直到明天凌晨两点，显示当前的打卡记录
                            if time_now < end_time + datetime.timedelta(minutes=attendance_group_obj.leave_time):
                                results['state'] = 3  # 进行中的
                                results['ids'] = ids
                                results['begin_date'] = datetime.datetime.strftime(last_obj.begin_date,
                                                                                   '%H:%M:%S') if last_obj.begin_date else None
                                results['end_date'] = datetime.datetime.strftime(last_obj.end_date,
                                                                                 '%H:%M:%S') if last_obj.end_date else None
                                if last_obj.section in results['section_list']:
                                    results['section_list'].remove(last_obj.section)
                                results['section_list'].insert(0, last_obj.section)  # 放到第一位显示
                                group, classes = last_obj.group, last_obj.classes
                                if {'group': group, 'classes': classes} in results['group_list']:
                                    results['group_list'].remove({'group': group, 'classes': classes})
                                results['group_list'].insert(0, {'group': group, 'classes': classes})
                            else:
                                flat = True
                        else:  # 夜班
                            # if time_now < p_date_now + datetime.timedelta(days=1, hours=14):
                            if time_now < end_time + datetime.timedelta(minutes=attendance_group_obj.leave_time):
                                results['state'] = 3
                                results['ids'] = ids
                                results['begin_date'] = datetime.datetime.strftime(last_obj.begin_date,
                                                                                   '%H:%M:%S') if last_obj.begin_date else None
                                results['end_date'] = datetime.datetime.strftime(last_obj.end_date,
                                                                                 '%H:%M:%S') if last_obj.end_date else None
                                if last_obj.section in results['section_list']:
                                    results['section_list'].remove(last_obj.section)
                                results['section_list'].insert(0, last_obj.section)  # 放到第一位显示
                                group, classes = last_obj.group, last_obj.classes
                                if {'group': group, 'classes': classes} in results['group_list']:
                                    results['group_list'].remove({'group': group, 'classes': classes})
                                results['group_list'].insert(0, {'group': group, 'classes': classes})
                            else:
                                flat = True
                if flat:
                    # 默认返回前一条的打卡记录，默认显示
                    results['state'] = 2  # 默认显示
                    if last_obj.section in results['section_list']:
                        results['section_list'].remove(last_obj.section)
                    results['section_list'].insert(0, last_obj.section)  # 放到第一位显示
                    now = get_virtual_time() + timedelta(minutes=attendance_group_obj.lead_time)
                    current_work_schedule_plan = WorkSchedulePlan.objects.filter(
                        start_time__lte=now,
                        end_time__gte=now,
                        plan_schedule__work_schedule__work_procedure__global_name='密炼'
                    ).first()
                    if current_work_schedule_plan:
                        date_now = str(current_work_schedule_plan.plan_schedule.day_time)
                    else:
                        date_now = str(now.date())
                    # 获取班次班组
                    queryset = WorkSchedulePlan.objects.filter(plan_schedule__day_time=date_now, start_time__lte=now,
                                                               end_time__gte=now,
                                                               plan_schedule__work_schedule__work_procedure__global_name='密炼') \
                        .values('group__global_name', 'classes__global_name')
                    group_list = [{'group': item['group__global_name'], 'classes': item['classes__global_name']} for item in
                                  queryset]
                    results['group_list'] = group_list
            else:
                results['state'] = 1  # 没有打卡记录 显示当前时间的班次班组
                now = get_virtual_time() + timedelta(minutes=attendance_group_obj.lead_time)
                current_work_schedule_plan = WorkSchedulePlan.objects.filter(
                    start_time__lte=now,
                    end_time__gte=now,
                    plan_schedule__work_schedule__work_procedure__global_name='密炼'
                ).first()
                if current_work_schedule_plan:
                    date_now = str(current_work_schedule_plan.plan_schedule.day_time)
                else:
                    date_now = str(now.date())
                # 获取班次班组
                queryset = WorkSchedulePlan.objects.filter(plan_schedule__day_time=date_now, start_time__lte=now,
                                                           end_time__gte=now,
                                                           plan_schedule__work_schedule__work_procedure__global_name='密炼') \
                    .values('group__global_name', 'classes__global_name')
                group_list = [{'group': item['group__global_name'], 'classes': item['classes__global_name']} for item in queryset]
                results['group_list'] = group_list
        else:
            if last_obj:
                flat = False
                key_second = (last_obj.standard_end_date - last_obj.standard_begin_date).total_seconds()
                report = EmployeeAttendanceRecords.objects.filter(begin_date=last_obj.begin_date, clock_type=attendance_group_obj.type,
                                                                  user_id=last_obj.user_id).values_list('equip', 'id')
                ids, equips = [item[1] for item in report], [item[0] for item in report]
                results['equips'] = sorted(list(set(equips))) if last_obj.section in (s_choice + m_choice) else []

                if str(last_obj.factory_date) == date_now:
                    begin_time, end_time = last_obj.standard_begin_date, last_obj.standard_end_date
                    p_date_now = datetime.datetime.strptime(date_now, '%Y-%m-%d')
                    if end_time.hour > 12:  # 白班
                        diff_hour = 18 if end_time.hour == 16 or begin_time.hour == 0 else 26  # 称量早8夜8显示问题
                        # if time_now < p_date_now + datetime.timedelta(hours=diff_hour):  # 直到明天凌晨两点，显示当前的打卡记录
                        if time_now < end_time + datetime.timedelta(minutes=attendance_group_obj.leave_time):
                            results['state'] = 3  # 进行中的
                            results['ids'] = ids
                            results['begin_date'] = datetime.datetime.strftime(last_obj.begin_date,
                                                                               '%H:%M:%S') if last_obj.begin_date else None
                            results['end_date'] = datetime.datetime.strftime(last_obj.end_date,
                                                                             '%H:%M:%S') if last_obj.end_date else None
                            if last_obj.section in results['section_list']:
                                results['section_list'].remove(last_obj.section)
                            results['section_list'].insert(0, last_obj.section)  # 放到第一位显示
                            group, classes = last_obj.group, last_obj.classes
                            if {'group': group, 'classes': classes} in results['group_list']:
                                results['group_list'].remove({'group': group, 'classes': classes})
                            results['group_list'].insert(0, {'group': group, 'classes': classes})
                        else:
                            flat = True
                    else:  # 夜班
                        # if time_now < p_date_now + datetime.timedelta(days=1, hours=14):
                        if time_now < end_time + datetime.timedelta(minutes=attendance_group_obj.leave_time):
                            results['state'] = 3
                            results['ids'] = ids
                            results['begin_date'] = datetime.datetime.strftime(last_obj.begin_date,
                                                                               '%H:%M:%S') if last_obj.begin_date else None
                            results['end_date'] = datetime.datetime.strftime(last_obj.end_date,
                                                                             '%H:%M:%S') if last_obj.end_date else None
                            if last_obj.section in results['section_list']:
                                results['section_list'].remove(last_obj.section)
                            results['section_list'].insert(0, last_obj.section)  # 放到第一位显示
                            group, classes = last_obj.group, last_obj.classes
                            if {'group': group, 'classes': classes} in results['group_list']:
                                results['group_list'].remove({'group': group, 'classes': classes})
                            results['group_list'].insert(0, {'group': group, 'classes': classes})
                        else:
                            flat = True
                else:
                    if datetime.datetime.strptime(date_now, '%Y-%m-%d').date() > last_obj.factory_date + timedelta(days=1) \
                            or (not last_obj.end_date and (time_now - last_obj.begin_date).total_seconds() >
                                key_second + (attendance_group_obj.lead_time + attendance_group_obj.leave_time) * 60) \
                            or (last_obj.end_date and (time_now - last_obj.end_date).total_seconds() > attendance_group_obj.leave_time * 60):
                        flat = True
                    else:
                        if datetime.datetime.strptime(date_now, '%Y-%m-%d').date() == last_obj.factory_date + timedelta(days=1) and not last_obj.end_date:
                            begin_time, end_time, date_now = last_obj.standard_begin_date, last_obj.standard_end_date, str(last_obj.factory_date)
                        else:
                            u_info = WeightClassPlanDetail.objects.filter(weight_class_plan__user__username=username,
                                                                          factory_date=date_now,
                                                                          weight_class_plan__delete_flag=False).last()
                            if not u_info or (u_info and u_info.class_code == '休'):
                                raise ValidationError('未找到排班信息')
                            res = get_work_time(u_info.class_code, date_now)
                            if not res or (res and not res.get(last_obj.classes)):
                                raise ValidationError('未找到排班信息')
                            begin_time, end_time = res[last_obj.classes]
                            begin_time, end_time = datetime.datetime.strptime(begin_time,
                                                                              '%Y-%m-%d %H:%M:%S'), datetime.datetime.strptime(
                                end_time, '%Y-%m-%d %H:%M:%S')
                        p_date_now = datetime.datetime.strptime(date_now, '%Y-%m-%d')
                        if end_time.hour > 12:  # 白班
                            # if time_now < p_date_now + datetime.timedelta(days=1, hours=2):  # 直到明天凌晨两点，显示当前的打卡记录
                            if time_now < end_time + datetime.timedelta(minutes=attendance_group_obj.leave_time):
                                results['state'] = 3  # 进行中的
                                results['ids'] = ids
                                results['begin_date'] = datetime.datetime.strftime(last_obj.begin_date,
                                                                                   '%H:%M:%S') if last_obj.begin_date else None
                                results['end_date'] = datetime.datetime.strftime(last_obj.end_date,
                                                                                 '%H:%M:%S') if last_obj.end_date else None
                                if last_obj.section in results['section_list']:
                                    results['section_list'].remove(last_obj.section)
                                results['section_list'].insert(0, last_obj.section)  # 放到第一位显示
                                group, classes = last_obj.group, last_obj.classes
                                if {'group': group, 'classes': classes} in results['group_list']:
                                    results['group_list'].remove({'group': group, 'classes': classes})
                                results['group_list'].insert(0, {'group': group, 'classes': classes})
                            else:
                                flat = True
                        else:  # 夜班
                            # if time_now < p_date_now + datetime.timedelta(days=1, hours=14):
                            if time_now < end_time + datetime.timedelta(minutes=attendance_group_obj.leave_time):
                                results['state'] = 3
                                results['ids'] = ids
                                results['begin_date'] = datetime.datetime.strftime(last_obj.begin_date,
                                                                                   '%H:%M:%S') if last_obj.begin_date else None
                                results['end_date'] = datetime.datetime.strftime(last_obj.end_date,
                                                                                 '%H:%M:%S') if last_obj.end_date else None
                                if last_obj.section in results['section_list']:
                                    results['section_list'].remove(last_obj.section)
                                results['section_list'].insert(0, last_obj.section)  # 放到第一位显示
                                group, classes = last_obj.group, last_obj.classes
                                if {'group': group, 'classes': classes} in results['group_list']:
                                    results['group_list'].remove({'group': group, 'classes': classes})
                                results['group_list'].insert(0, {'group': group, 'classes': classes})
                            else:
                                flat = True
                if flat:
                    # 默认返回前一条的打卡记录，默认显示
                    results['state'] = 2  # 默认显示
                    if last_obj.section in results['section_list']:
                        results['section_list'].remove(last_obj.section)
                    results['section_list'].insert(0, last_obj.section)  # 放到第一位显示
                    queryset = WeightClassPlanDetail.objects.filter(factory_date=date_now,
                                                                    weight_class_plan__delete_flag=False,
                                                                    weight_class_plan__user=self.request.user).last()
                    u_group = queryset.weight_class_plan.classes.split('/')[
                        0] if queryset.weight_class_plan.classes.split('/') else None
                    u_class = get_work_time(queryset.class_code, date_now).keys()
                    group_list = [] if not all([u_group, u_class]) else [{'group': u_group, 'classes': i} for i in u_class]
                    results['group_list'] = group_list
            else:
                results['state'] = 1  # 没有打卡记录 显示当前时间的班次班组
                queryset = WeightClassPlanDetail.objects.filter(factory_date=date_now,
                                                                weight_class_plan__delete_flag=False,
                                                                weight_class_plan__user=self.request.user).last()
                u_group = queryset.weight_class_plan.classes.split('/')[0] if queryset.weight_class_plan.classes.split('/') else None
                u_class = get_work_time(queryset.class_code, date_now).keys()
                group_list = [] if not all([u_group, u_class]) else [{'group': u_group, 'classes': i} for i in u_class]
                results['group_list'] = group_list
        return Response(results)

    @atomic
    def create(self, request, *args, **kwargs):
        user = self.request.user
        data = self.request.data  # {classes group equip_list ids section status}
        attendance_group_obj, section_list, equip_list, date_now, group_list = self.get_user_group(user)
        time_now = get_virtual_time()
        ids = data.pop('ids', None)
        equip_list = data.pop('equip_list', None)
        status = data.get('status')
        results = {
            'ids': [],
        }
        flag, clock_type = get_user_weight_flag(user)
        # 标准上下班时间
        date_now, group, classes = date_now, data['group'], data['classes']
        a_f = ApplyForExtraWork.objects.filter(Q(~Q(handling_result=1) | ~Q(f_handling_result=1) | ~Q(s_handling_result=1)), clock_type=clock_type,
                                               factory_date=date_now, user=user, group=group, classes=classes, section=data.get('section'))
        if a_f:
            raise ValidationError('加班申请尚未审批通过')
        last_record = EmployeeAttendanceRecords.objects.filter(user=user, end_date__isnull=True, clock_type=clock_type).order_by('factory_date').last()
        if last_record:
            key_second = (last_record.standard_end_date - last_record.standard_begin_date).total_seconds()
            if status == '离岗' and last_record.factory_date != datetime.datetime.strptime(date_now, '%Y-%m-%d').date() and not (last_record.begin_date and (time_now - last_record.begin_date).total_seconds() > key_second + (attendance_group_obj.lead_time + attendance_group_obj.leave_time) * 60):
                date_now, group, classes = [str(last_record.factory_date), last_record.group, last_record.classes]
        if clock_type == '密炼':
            if status in ['上岗', '调岗']:
                standard_begin_time, standard_end_time = get_standard_time(user.username, date_now, group=group, classes=classes)
                if not standard_begin_time:
                    raise ValidationError('未找到排班信息')
            # # 需要和加班申请做交集
            # extra_work = ApplyForExtraWork.objects.filter(factory_date=date_now, user=user, group=data['group'],
            #                                               classes=data['classes'], section=data['section'],
            #                                               handling_result=True, clock_type=clock_type).last()
            # if extra_work:
            #     standard_begin_time, standard_end_time = extra_work.begin_date, extra_work.end_date
            # 加班被拒绝的不可以打卡
            if status == '上岗':
                begin_date = time_now
                lead_time = datetime.timedelta(minutes=attendance_group_obj.lead_time)
                if time_now < standard_begin_time - lead_time:
                    raise ValidationError('未到可打卡时间')
                # 本次上岗打卡操作自动补上上次离岗打卡
                last_date = datetime.datetime.strptime(date_now, '%Y-%m-%d') - datetime.timedelta(days=1)
                last_date = datetime.datetime.strftime(last_date, '%Y-%m-%d')
                if standard_end_time.hour > 12:  # 白班
                    end_date = f"{str(last_date)} {str(standard_end_time)[11:]}"
                else:
                    end_date = f"{str(date_now)} {str(standard_end_time)[11:]}"
                # 自动补全之前的离岗时间
                factory_records = EmployeeAttendanceRecords.objects.filter(user=user, end_date__isnull=True, factory_date=last_date, clock_type=clock_type)
                for i in factory_records:
                    calculate_end_date = time_now if i.standard_begin_date < time_now < i.standard_end_date else i.standard_end_date
                    i.end_date = time_now
                    i.actual_end_date = calculate_end_date
                    i.calculate_end_date = calculate_end_date
                    i.save()
                for equip in equip_list:
                    calculate_begin_date = begin_date if standard_begin_time < begin_date < standard_end_time else standard_begin_time
                    obj = EmployeeAttendanceRecords.objects.create(
                        user=user,
                        equip=equip,
                        begin_date=begin_date,
                        actual_begin_date=calculate_begin_date,
                        factory_date=date_now,
                        standard_begin_date=standard_begin_time,
                        standard_end_date=standard_end_time,
                        calculate_begin_date=calculate_begin_date,
                        **data
                    )
                    results['ids'].append(obj.id)
            elif status == '离岗':  # 可重复打卡
                range_time = datetime.timedelta(minutes=attendance_group_obj.range_time)
                s_record = EmployeeAttendanceRecords.objects.filter(id__in=ids).first()
                begin_date, standard_begin_time, standard_end_time = s_record.begin_date, s_record.standard_begin_date, s_record.standard_end_date
                if time_now < begin_date + range_time:
                    raise ValidationError(f'上班{attendance_group_obj.range_time}分钟内不能打下班卡')
                # 下班半小时后不能再打卡
                if time_now > standard_end_time + datetime.timedelta(minutes=attendance_group_obj.leave_time):
                    raise ValidationError(f'下班超过{attendance_group_obj.leave_time}不可再打卡')
                end_date = time_now
                if end_date <= standard_begin_time:
                    raise ValidationError('未到班次标准上班时间不可打下班卡')
                # 计算实际工时(打卡时间)
                actual_begin_time = begin_date if standard_end_time > begin_date > standard_begin_time else standard_begin_time
                actual_end_time = end_date if end_date < standard_end_time else standard_end_time
                # if extra_work:
                #     actual_begin_time = actual_begin_time if actual_begin_time < extra_work.begin_date else extra_work.begin_date
                #     actual_end_time = actual_end_time if actual_end_time < extra_work.end_date else extra_work.end_date
                work_time = round((actual_end_time - actual_begin_time).seconds / 3600, 2)
                EmployeeAttendanceRecords.objects.filter(id__in=ids).update(
                    end_date=end_date, actual_end_date=actual_end_time, work_time=work_time, actual_time=work_time,
                    calculate_end_date=actual_end_time
                )
                results['ids'] = ids
            elif status == '调岗':
                if EmployeeAttendanceRecords.objects.filter(id__in=ids, end_date__isnull=True).exists():
                    obj_r = EmployeeAttendanceRecords.objects.filter(id__in=ids).first()
                    begin_date, before_begin_time, before_end_time = obj_r.begin_date, obj_r.standard_begin_date, obj_r.standard_end_date
                    end_date = time_now
                    if time_now < before_begin_time:
                        raise ValidationError('未到标准上班时间不可调岗')
                    if time_now >= obj_r.standard_end_date - timedelta(minutes=attendance_group_obj.lead_time):
                        raise ValidationError('超出可调岗时间范围')
                    work_time = round((end_date - begin_date).seconds / 3600, 2)
                    calculate_end_date = end_date if before_begin_time < end_date < before_end_time else before_end_time
                    EmployeeAttendanceRecords.objects.filter(id__in=ids).update(
                        end_date=end_date, actual_end_date=calculate_end_date, work_time=work_time, actual_time=work_time,
                        calculate_end_date=calculate_end_date
                    )
                else:
                    raise ValidationError('已打下班卡不可调岗')
                calculate_begin_date = time_now if standard_begin_time < time_now < standard_end_time else standard_begin_time
                for equip in equip_list:
                    obj = EmployeeAttendanceRecords.objects.create(
                        user=user,
                        equip=equip,
                        begin_date=time_now,
                        actual_begin_date=calculate_begin_date,
                        factory_date=date_now,
                        standard_begin_date=standard_begin_time,
                        standard_end_date=standard_end_time,
                        calculate_begin_date=calculate_begin_date,
                        **data
                    )
                    results['ids'].append(obj.id)
        else:
            if status in ['上岗', '调岗']:
                u_info = WeightClassPlanDetail.objects.filter(weight_class_plan__user=user, factory_date=date_now,
                                                              weight_class_plan__delete_flag=False).last()
                if not u_info or (u_info and u_info.class_code == '休'):
                    raise ValidationError('未找到排班信息')
                res = get_work_time(u_info.class_code, date_now)
                if not res or (res and not res.get(classes)):
                    raise ValidationError('未找到排班信息')
                begin_time, end_time = res[classes]
                standard_begin_time, standard_end_time = datetime.datetime.strptime(begin_time, '%Y-%m-%d %H:%M:%S'), datetime.datetime.strptime(end_time, '%Y-%m-%d %H:%M:%S')
                if not standard_begin_time:
                    raise ValidationError('未找到排班信息')
            # # 需要和加班申请做交集
            # extra_work = ApplyForExtraWork.objects.filter(factory_date=date_now, user=user, group=data['group'],
            #                                               classes=data['classes'], section=data['section'],
            #                                               handling_result=True, clock_type=clock_type).last()
            # if extra_work:
            #     standard_begin_time, standard_end_time = extra_work.begin_date, extra_work.end_date
            if status == '上岗':
                begin_date = time_now
                lead_time = datetime.timedelta(minutes=attendance_group_obj.lead_time)
                if time_now < standard_begin_time - lead_time:
                    raise ValidationError('未到可打卡时间')
                # 本次上岗打卡操作自动补上上次离岗打卡
                last_date = datetime.datetime.strptime(date_now, '%Y-%m-%d') - datetime.timedelta(days=1)
                last_date = datetime.datetime.strftime(last_date, '%Y-%m-%d')
                if standard_end_time.hour > 12:  # 白班
                    end_date = f"{str(last_date)} {str(standard_end_time)[11:]}"
                else:
                    end_date = f"{str(date_now)} {str(standard_end_time)[11:]}"
                # 自动补全之前的离岗时间
                factory_records = EmployeeAttendanceRecords.objects.filter(user=user, end_date__isnull=True,
                                                                           factory_date=last_date,
                                                                           clock_type=clock_type)
                for i in factory_records:
                    calculate_end_date = time_now if i.standard_begin_date < time_now < i.standard_end_date else i.standard_end_date
                    i.end_date = end_date
                    i.actual_end_date = calculate_end_date
                    i.calculate_end_date = calculate_end_date
                    i.save()
                for equip in equip_list:
                    calculate_begin_date = begin_date if standard_begin_time < begin_date < standard_end_time else standard_begin_time
                    obj = EmployeeAttendanceRecords.objects.create(
                        user=user,
                        equip=equip,
                        begin_date=begin_date,
                        actual_begin_date=calculate_begin_date,
                        factory_date=date_now,
                        clock_type=clock_type,
                        standard_begin_date=standard_begin_time,
                        standard_end_date=standard_end_time,
                        calculate_begin_date=calculate_begin_date,
                        **data
                    )
                    results['ids'].append(obj.id)
            elif status == '离岗':  # 可重复打卡
                range_time = datetime.timedelta(minutes=attendance_group_obj.range_time)
                s_record = EmployeeAttendanceRecords.objects.filter(id__in=ids).first()
                begin_date, standard_begin_time, standard_end_time = s_record.begin_date, s_record.standard_begin_date, s_record.standard_end_date
                if time_now < begin_date + range_time:
                    raise ValidationError(f'上班{attendance_group_obj.range_time}分钟内不能打下班卡')
                # 下班半小时后不能再打卡
                if time_now > standard_end_time + datetime.timedelta(minutes=attendance_group_obj.leave_time):
                    raise ValidationError(f'下班超过{attendance_group_obj.leave_time}不可再打卡')
                end_date = time_now
                if end_date <= standard_begin_time:
                    raise ValidationError('未到班次标准上班时间不可打下班卡')
                # 计算实际工时(打卡时间)
                actual_begin_time = begin_date if standard_end_time > begin_date > standard_begin_time else standard_begin_time
                actual_end_time = end_date if end_date < standard_end_time else standard_end_time
                # if extra_work:
                #     actual_begin_time = actual_begin_time if actual_begin_time < extra_work.begin_date else extra_work.begin_date
                #     actual_end_time = actual_end_time if actual_end_time < extra_work.end_date else extra_work.end_date
                work_time = round((actual_end_time - actual_begin_time).seconds / 3600, 2)
                EmployeeAttendanceRecords.objects.filter(id__in=ids).update(
                    end_date=end_date, actual_end_date=actual_end_time, work_time=work_time, actual_time=work_time,
                    calculate_end_date=actual_end_time
                )
                results['ids'] = ids
            elif status == '调岗':
                if EmployeeAttendanceRecords.objects.filter(id__in=ids, end_date__isnull=True).exists():
                    obj_r = EmployeeAttendanceRecords.objects.filter(id__in=ids).first()
                    begin_date, before_begin_time, before_end_time = obj_r.begin_date, obj_r.standard_begin_date, obj_r.standard_end_date
                    end_date = time_now
                    if time_now < before_begin_time:
                        raise ValidationError('未到标准上班时间不可调岗')
                    if time_now >= obj_r.standard_end_date - timedelta(minutes=attendance_group_obj.lead_time):
                        raise ValidationError('超出可调岗时间范围')
                    calculate_end_date = end_date if before_begin_time < end_date < before_end_time else before_end_time
                    work_time = round((calculate_end_date - obj_r.calculate_begin_date).seconds / 3600, 2)
                    EmployeeAttendanceRecords.objects.filter(id__in=ids).update(
                        end_date=end_date, actual_end_date=calculate_end_date, work_time=work_time, actual_time=work_time,
                        calculate_end_date=calculate_end_date
                    )
                else:
                    raise ValidationError('已打下班卡不可调岗')
                calculate_begin_date = time_now if standard_begin_time < time_now < standard_end_time else standard_begin_time
                for equip in equip_list:
                    obj = EmployeeAttendanceRecords.objects.create(
                        user=user,
                        equip=equip,
                        begin_date=time_now,
                        actual_begin_date=calculate_begin_date,
                        factory_date=date_now,
                        clock_type=clock_type,
                        standard_begin_date=standard_begin_time,
                        standard_end_date=standard_end_time,
                        calculate_begin_date=calculate_begin_date,
                        **data
                    )
                    results['ids'].append(obj.id)
        # 记录考勤打卡详情
        self.save_attendance_clock_detail(user.username, equip_list, data, clock_type)
        return Response({'results': results})

    @action(methods=['post'], detail=False, permission_classes=[], url_path='reissue_card', url_name='reissue_card')
    def reissue_card(self, request):
        """提交补卡申请"""
        data = self.request.data
        user = self.request.user
        username = user.username
        status = data.get('status')
        now_time = get_virtual_time()
        bk_date = data.get('bk_date', None)
        now = datetime.datetime.strptime(f'{bk_date}:00', '%Y-%m-%d %H:%M:%S')
        if status == '离岗':
            select_begin_date = data.pop('select_begin_date')
            p_select_begin_date = datetime.datetime.strptime(select_begin_date, '%Y-%m-%d %H:%M:%S')
            if bk_date <= select_begin_date or (now - p_select_begin_date).total_seconds() > 14 * 3600:
                raise ValidationError('补卡时间选择有误')
            check_date = datetime.datetime.strptime(select_begin_date, '%Y-%m-%d %H:%M:%S').date()
        else:
            check_date = now.date()
        # 补卡规则
        card_rules = {item['global_no']: item['global_name'] for item in GlobalCode.objects.filter(global_type__type_name='补卡规则',
                                                                                                   use_flag=True).values('global_no', 'global_name')}
        # 一个月内可补卡次数
        num = card_rules.get('一个月内补卡次数', None)
        # 可补卡时间
        card_time = card_rules.get('可补卡时间', None)
        if not all([num, card_time]):
            raise ValidationError('请先添加补卡规则')
        if FillCardApply.objects.filter(user=user, bk_date__year=now.year, bk_date__month=now.month,
                                        handling_result=True).count() == int(num):
            raise ValidationError(f'当月已提交了{num}次补卡申请')
        attendance_group_obj, section_list, equip_list, date_now, group_list = self.get_user_group(user, now)
        principal = attendance_group_obj.principal  # 考勤负责人
        clock_type = attendance_group_obj.type
        # 离岗时间
        equip_list = data.pop('equip_list')
        # 标准上下班时间
        if status == '离岗':
            r = EmployeeAttendanceRecords.objects.filter(begin_date__startswith=select_begin_date, classes=data.get('classes'), group=data.get('group'),
                                                        section=data.get('section'), clock_type=clock_type).last()
            if r:
                standard_begin_time, standard_end_time = r.standard_begin_date, r.standard_end_date
            else:
                raise ValidationError('根据所选数据未找到对应考勤记录')
        else:
            if clock_type == '密炼':
                standard_begin_time, standard_end_time = get_standard_time(user.username, date_now, group=data['group'], classes=data['classes'])
                if not standard_begin_time:
                    raise ValidationError('未找到排班信息')
            else:
                u_info = WeightClassPlanDetail.objects.filter(weight_class_plan__user=user, factory_date=date_now, weight_class_plan__delete_flag=False).last()
                if not u_info or (u_info and u_info.class_code == '休'):
                    raise ValidationError('未找到排班信息')
                res = get_work_time(u_info.class_code, date_now)
                if not res or (res and not res.get(data['classes'])):
                    raise ValidationError('未找到排班信息')
                begin_time, end_time = res[data['classes']]
                standard_begin_time, standard_end_time = datetime.datetime.strptime(begin_time, '%Y-%m-%d %H:%M:%S'), datetime.datetime.strptime(end_time, '%Y-%m-%d %H:%M:%S')
        if standard_end_time.hour > 12:  # 白班
            attendance_et = datetime.datetime.strptime(f"{date_now} {str(standard_end_time)[11:]}", '%Y-%m-%d %H:%M:%S')
            factory_date = datetime.datetime.strptime(date_now, '%Y-%m-%d').date()
        else:  # 夜班
            hours = now_time.time()
            if clock_type != '密炼' and status == '离岗':
                factory_date = r.factory_date
            else:
                if standard_end_time.time() < hours < standard_begin_time.time():
                    factory_date = (now_time - datetime.timedelta(days=1)).date()
                else:
                    factory_date = datetime.datetime.strptime(date_now, '%Y-%m-%d').date()
            factory_date1 = str(factory_date + datetime.timedelta(days=1))
            attendance_et = datetime.datetime.strptime(f"{factory_date1} {str(standard_end_time)[11:]}", '%Y-%m-%d %H:%M:%S')
        if now_time > attendance_et + datetime.timedelta(hours=int(card_time)):
            raise ValidationError(f'不可提交超过{card_time}小时的申请')
        flag, clock_type = get_user_weight_flag(user)
        if status == '上岗':
            # 判断是否有打卡记录
            if FillCardApply.objects.filter(Q(handling_result=1) | Q(Q(handling_result__isnull=True), ~Q(f_handling_result=0)),
                                            user=user, factory_date=factory_date, status=status, clock_type=clock_type).exists():
                raise ValidationError('当前已提交过上岗补卡申请')
            if EmployeeAttendanceRecords.objects.filter(user=user, factory_date=factory_date, status=status, clock_type=clock_type).exists():
                raise ValidationError('当天存在上岗打卡记录')
        elif status == '调岗':
            if not EmployeeAttendanceRecords.objects.filter(user=user, factory_date=factory_date, clock_type=clock_type,
                                                            status__in=['上岗', '调岗'], end_date__isnull=True).exists():
                raise ValidationError('请先提交当天的上岗申请')
            last = EmployeeAttendanceRecords.objects.filter(user=user, factory_date=factory_date, clock_type=clock_type,
                                                            status__in=['上岗', '调岗'], end_date__isnull=False)\
                .order_by('end_date').last()
            if last:  # 存在上岗或换岗的打卡，
                if now < last.end_date:
                    raise ValidationError('提交的补卡申请有误')
        elif status == '离岗':
            obj = EmployeeAttendanceRecords.objects.filter(
                user=user,
                factory_date=factory_date,
                end_date__isnull=True,
                section=data.get('section'),
                classes=data.get('classes'),
                group=data.get('group'),
                status__in=['上岗', '调岗'],
                clock_type=clock_type
            )
            if not obj:
                raise ValidationError('提交的补卡申请有误')

        # 记录补卡申请记录
        apply = FillCardApply.objects.create(
            user=self.request.user,
            equip=','.join(equip_list),
            apply_date=now_time,
            factory_date=factory_date,
            clock_type=clock_type,
            **data
        )
        # 钉钉提醒
        u = User.objects.filter(username=self.request.user.username, section__name__startswith='生产', section__isnull=False).last()
        if u:
            in_charge_user = u.section.in_charge_user
            if in_charge_user == u:
                principal_obj = [u.section.parent_section.in_charge_user] if u.section.parent_section and u.section.parent_section.in_charge_user else []
            else:
                principal_obj = [in_charge_user]
        else:
            principal_obj = User.objects.filter(username__in=principal.split(','))
        serializer_data = FillCardApplySerializer(apply).data
        content = {
            "title": f"{username}的补卡申请",
            "form": [
                {"key": "姓名:", "value": username},
                {"key": "班组:", "value": apply.group},
                {"key": "班次:", "value": apply.classes},
                {"key": "岗位:", "value": apply.section},
                {"key": "机台:", "value": apply.equip},
                {"key": "补卡时间:", "value": serializer_data.get('bk_date')},
                {"key": "补卡理由:", "value": apply.desc},
            ],
        }
        for principal in principal_obj:
            self.send_message(principal, content)
        return Response('消息发送给审批人')

    @action(methods=['post'], detail=False, permission_classes=[], url_path='overtime', url_name='overtime')
    def overtime(self, request):
        """提交加班申请"""
        # 加班也存在调岗的情况
        user = self.request.user
        data = self.request.data
        equip_list = data.pop('equip_list', None)
        equip = ','.join(equip_list)
        attendance_group_obj, section_list, equip_list, date_now, group_list = self.get_user_group(user, now=data.get('begin_date'), apply=1)
        principal = attendance_group_obj.principal  # 考勤负责人
        flag, clock_type = get_user_weight_flag(user)
        apply = ApplyForExtraWork.objects.create(
            user=self.request.user,
            equip=equip,
            factory_date=date_now,
            clock_type=clock_type,
            **data
        )
        # 钉钉提醒
        u = User.objects.filter(username=self.request.user.username, section__name__startswith='生产', section__isnull=False).last()
        if u:
            in_charge_user = u.section.in_charge_user
            if in_charge_user == u:
                principal_obj = [u.section.parent_section.in_charge_user] if u.section.parent_section and u.section.parent_section.in_charge_user else []
            else:
                principal_obj = [in_charge_user]
        else:
            principal_obj = User.objects.filter(username__in=principal.split(','))
        serializer_data = ApplyForExtraWorkSerializer(apply).data
        content = {
            "title": f"{user.username}的补卡申请",
            "form": [
                {"key": "姓名:", "value": user.username},
                {"key": "班组:", "value": apply.group},
                {"key": "班次:", "value": apply.classes},
                {"key": "岗位:", "value": apply.section},
                {"key": "机台:", "value": apply.equip},
                {"key": "加班开始时间:", "value": serializer_data.get('begin_date')},
                {"key": "加班结束时间:", "value": serializer_data.get('end_date')},
                {"key": "加班理由:", "value": apply.desc},
            ],
        }
        for principal in principal_obj:
            self.send_message(principal, content)
        return Response('消息发送给审批人')

    @action(methods=['get'], detail=False, permission_classes=[], url_path='default_value', url_name='default_value')
    def default_value(self, request):
        """获取补卡时默认数据显示"""
        state = self.request.query_params.get('state')
        user = self.request.user
        select_time = self.request.query_params.get('select_time')
        begin_flag = self.request.query_params.get('begin_flag')
        select_begin_date = self.request.query_params.get('select_begin_date')
        limit_time = get_virtual_time() - datetime.timedelta(hours=24)
        flag, clock_type = get_user_weight_flag(user)
        query_set = EmployeeAttendanceRecords.objects.filter(Q(user=user) & ~Q(is_use='废弃'), clock_type=clock_type)
        if begin_flag and state == '离岗':
            begin_times = query_set.filter(begin_date__gte=limit_time, end_date__isnull=True).values_list('begin_date', flat=True).distinct()
            new_begin_times = [i.strftime('%Y-%m-%d %H:%M:%S') for i in begin_times]
            return Response(new_begin_times)
        if select_begin_date:
            s_info = query_set.filter(begin_date__startswith=select_begin_date, end_date__isnull=True).last()
            if not s_info:
                raise ValidationError('所选上岗时间未匹配到考勤记录')
            group_list = [{'group': s_info.group, 'classes': s_info.classes}]
            equip_list = []
            section_list = [s_info.section]
        else:
            try:
                attendance_group_obj, section_list, equip_list, date_now, group_list = self.get_user_group(self.request.user, select_time)
            except Exception as e:
                group_list, equip_list, section_list, principal = [], [], [], ''
            else:
                group_list, equip_list, section_list, principal = group_list, equip_list, section_list, attendance_group_obj.principal
        if clock_type != '密炼' and state != '离岗':
            equip_list = Equip.objects.filter(category__equip_type__global_name='称量设备').values_list('equip_no', flat=True)
            section_list = PerformanceJobLadder.objects.filter(delete_flag=False, type=clock_type).values_list('name', flat=True)
            # 获取班次班组
            if not state:  # 加班取密炼班组
                queryset = WorkSchedulePlan.objects.filter(plan_schedule__day_time=select_time[:10] if select_time else None,
                                                           plan_schedule__work_schedule__work_procedure__global_name='密炼')\
                    .values('group__global_name', 'classes__global_name')
                group_list = [{'group': item['group__global_name'], 'classes': item['classes__global_name']} for item in queryset]
        res = {'group': group_list,
               'section': section_list,
               'equips': equip_list}
        equips = []
        if state == '上岗':
            obj = query_set.filter(status__in=['上岗', '调岗']).last()
            if obj:
                equips = query_set.filter(begin_date=obj.begin_date).values_list('equip', flat=True)
        elif state == '调岗':
            obj = query_set.filter(status__in=['上岗', '调岗']).last()
            if obj:
                equips = query_set.filter(begin_date=obj.begin_date).values_list('equip', flat=True)
        elif state == '离岗':
            obj = query_set.filter(status__in=['上岗', '调岗'], end_date__isnull=True).last()
            if obj:
                equips = query_set.filter(begin_date=obj.begin_date).values_list('equip', flat=True)
        if equips:
            res['equips'] = sorted(list(set(equips)))
        return Response({'results': res})


@method_decorator([api_recorder], name="dispatch")
class ReissueCardView(APIView):
    permission_classes = (IsAuthenticated,)
    queryset = FillCardApply.objects.order_by('-id')
    queryset2 = ApplyForExtraWork.objects.order_by('-id')

    def send_message(self, user, content):
        phone = user.phone_number
        ding_api = DinDinAPI()
        ding_uid = ding_api.get_user_id(phone)
        ding_api.send_message([ding_uid], content, attendance=True)

    def get(self, request):
        # 分页
        user = self.request.user
        page = self.request.query_params.get("page", 1)
        page_size = self.request.query_params.get("page_size", 10)
        state = self.request.query_params.get('state', 0)
        try:
            st = (int(page) - 1) * int(page_size)
            et = int(page) * int(page_size)
        except:
            raise ValidationError("page/page_size异常，请修正后重试")
        level = None
        if self.request.query_params.get('apply'):  # 查看自己的补卡申请
            section = Section.objects.filter(in_charge_user=user, name__startswith='生产').last()
            all_user = [user.username] + (list(section.section_users.filter(is_active=True).values_list('username', flat=True)) if section else [])
            data = self.queryset.filter(user__username__in=set(all_user)).order_by('-id')
            data2 = None
        else:  # 审批补卡申请
            res = get_user_level()
            detail = res.get(user.username)
            if not detail:
                raise ValidationError('该用户不是相关生产部门负责人')
            users, level = set(detail['users']), detail['level']
            # 补卡申请(未审批 + 已审批)
            if level == 1:
                data = self.queryset.filter(user__username__in=users)
                data2 = self.queryset2.filter(user__username__in=users)
            elif level == 2:
                data = self.queryset.filter(Q(f_approver__in=users, f_handling_result=1) | Q(user__username__in=users))
                data2 = self.queryset2.filter(Q(f_approver__in=users, f_handling_result=1) | Q(user__username__in=users))
            else:
                data = self.queryset.filter(user__in=[])  # 补卡无第三级审批
                data2 = self.queryset2.filter(Q(s_approver__in=users, s_handling_result=1) | Q(user__username__in=users))
        serializer = FillCardApplySerializer(data, many=True)
        serializer2 = ApplyForExtraWorkSerializer(data2, many=True)
        res = serializer.data + serializer2.data
        if int(state) == 1:  # 未审批
            if level == 1:
                res = [item for item in res if item.get('f_handling_result') == None]
            elif level == 2:
                res = [item for item in res if item.get('s_handling_result') == None]
            else:
                res = [item for item in res if item.get('handling_result') == None]
        elif int(state) == 2:
            if level == 1:
                res = [item for item in res if item.get('handling_result') != None or item.get('f_handling_result') != None]
            elif level == 2:
                res = [item for item in res if item.get('handling_result') != None or item.get('s_handling_result') != None]
            else:
                res = [item for item in res if item.get('handling_result') != None]
        count = len(res)
        results = sorted(list(res), key=lambda x: x['apply_date'], reverse=True)
        return Response({'results': results[st:et], 'count': count, 'level': level})

    @atomic
    def post(self, request):  # 处理补卡申请
        data = self.request.data
        user_name = self.request.user.username
        obj = FillCardApply.objects.filter(id=data.get('id')).first()
        flag, now_time = False, datetime.datetime.now()
        f_handling_suggestion, f_handling_result, s_handling_suggestion, s_handling_result = data.get('f_handling_suggestion'), data.get('f_handling_result'), data.get('s_handling_suggestion'), data.get('s_handling_result')
        if isinstance(f_handling_result, int):
            obj.f_handling_suggestion = f_handling_suggestion
            obj.f_handling_result = f_handling_result
            obj.f_approver = user_name
            obj.f_approver_time = now_time
        if isinstance(s_handling_result, int):
            obj.s_handling_suggestion = s_handling_suggestion
            obj.s_handling_result = s_handling_result
            obj.s_approver = user_name
            obj.s_approver_time = now_time
            obj.handling_suggestion = s_handling_suggestion
            obj.handling_result = s_handling_result
            obj.approver_time = now_time
            obj.approver = user_name
            flag = True
        obj.save()
        serializer_data = FillCardApplySerializer(obj).data
        user = obj.user
        date_time = obj.bk_date
        clock_type = serializer_data.get('clock_type')
        content = {
            "title": "补卡申请",
            "form": [
                {"key": "姓名:", "value": user.username},
                {"key": "班组:", "value": serializer_data.get('group')},
                {"key": "班次:", "value": serializer_data.get('classes')},
                {"key": "岗位:", "value": serializer_data.get('section')},
                {"key": "机台:", "value": serializer_data.get('equip')},
                {"key": "补卡时间:", "value": serializer_data.get('bk_date')},
                {"key": "补卡理由:", "value": serializer_data.get('desc')},
                {"key": "处理意见:", "value": serializer_data.get('handling_suggestion')},
                {"key": "处理结果:", "value": serializer_data.get('handling_result')}
            ],
        }
        self.send_message(user, content)
        if flag and s_handling_result == 1:  # 审批通过
            status = obj.status
            equips = serializer_data.get('equip')
            equip_list = equips.split(',')
            if status == '上岗':
                lst = []
                if clock_type == '密炼':
                    standard_begin_time, standard_end_time = get_standard_time(user.username, obj.factory_date, group=obj.group,
                                                                               classes=obj.classes)
                else:
                    u_info = WeightClassPlanDetail.objects.filter(weight_class_plan__user=user, factory_date=obj.factory_date,
                                                                  weight_class_plan__delete_flag=False).last()
                    if not u_info or (u_info and u_info.class_code == '休'):
                        raise ValidationError('未找到排班信息')
                    res = get_work_time(u_info.class_code, str(obj.factory_date))
                    if not res or (res and not res.get(obj.classes)):
                        raise ValidationError('未找到排班信息')
                    begin_time, end_time = res[obj.classes]
                    standard_begin_time, standard_end_time = datetime.datetime.strptime(begin_time, '%Y-%m-%d %H:%M:%S'), datetime.datetime.strptime(end_time, '%Y-%m-%d %H:%M:%S')
                calculate_begin_date = date_time if standard_begin_time < date_time < standard_end_time else standard_begin_time
                for equip in equip_list:
                    lst.append(
                        EmployeeAttendanceRecords(
                            user=user,
                            section=serializer_data.get('section'),
                            factory_date=serializer_data.get('factory_date'),
                            begin_date=serializer_data.get('bk_date'),
                            actual_begin_date=calculate_begin_date,
                            classes=serializer_data.get('classes'),
                            group=serializer_data.get('group'),
                            equip=equip,
                            status=status,
                            standard_begin_date=standard_begin_time,
                            standard_end_date=standard_end_time,
                            calculate_begin_date=calculate_begin_date,
                            clock_type=clock_type
                        ))
                EmployeeAttendanceRecords.objects.bulk_create(lst)
            elif status == '调岗':
                factory_r = EmployeeAttendanceRecords.objects.filter(user=user, factory_date=obj.factory_date, clock_type=clock_type,
                                                                     end_date__isnull=True, status__in=['上岗', '调岗'])
                if factory_r:
                    s_record = factory_r.first()
                    calculate_end_date = date_time if s_record.standard_begin_date < date_time < s_record.standard_end_date else s_record.standard_end_date
                    factory_r.update(end_date=date_time, actual_end_date=date_time, calculate_end_date=calculate_end_date)
                lst = []
                if clock_type == '密炼':
                    standard_begin_time, standard_end_time = get_standard_time(user.username, obj.factory_date,
                                                                               group=obj.group, classes=obj.classes)
                else:
                    u_info = WeightClassPlanDetail.objects.filter(weight_class_plan__user=user, factory_date=obj.factory_date,
                                                                  weight_class_plan__delete_flag=False).last()
                    if not u_info or (u_info and u_info.class_code == '休'):
                        raise ValidationError('未找到排班信息')
                    res = get_work_time(u_info.class_code, str(obj.factory_date))
                    if not res or (res and not res.get(obj.classes)):
                        raise ValidationError('未找到排班信息')
                    begin_time, end_time = res[obj.classes]
                    standard_begin_time, standard_end_time = datetime.datetime.strptime(begin_time, '%Y-%m-%d %H:%M:%S'), datetime.datetime.strptime(end_time, '%Y-%m-%d %H:%M:%S')
                calculate_begin_date = date_time if standard_begin_time < date_time < standard_end_time else standard_begin_time
                for equip in equip_list:
                    lst.append(
                        EmployeeAttendanceRecords(
                            user=user,
                            section=serializer_data.get('section'),
                            factory_date=serializer_data.get('factory_date'),
                            begin_date=serializer_data.get('bk_date'),
                            actual_begin_date=calculate_begin_date,
                            classes=serializer_data.get('classes'),
                            group=serializer_data.get('group'),
                            equip=equip,
                            status=status,
                            calculate_begin_date=calculate_begin_date,
                            standard_begin_date=standard_begin_time,
                            standard_end_date=standard_end_time,
                            clock_type=clock_type
                        ))
                EmployeeAttendanceRecords.objects.bulk_create(lst)
            elif status == '离岗':
                end_date = obj.bk_date
                dic = {
                    'user': obj.user,
                    'factory_date': serializer_data.get('factory_date'),
                    'end_date__isnull': True,
                    'section': serializer_data.get('section'),
                    'classes': serializer_data.get('classes'),
                    'group': serializer_data.get('group'),
                    'status__in': ['上岗', '调岗'],
                    'clock_type': clock_type
                }
                obj = EmployeeAttendanceRecords.objects.filter(**dic).first()
                if not obj:
                    raise ValidationError('提交的补卡参数有误')
                calculate_end_date = end_date if obj.standard_begin_date < end_date < obj.standard_end_date else obj.standard_end_date
                work_time = round((calculate_end_date - obj.calculate_begin_date).seconds / 3600, 2)
                EmployeeAttendanceRecords.objects.filter(**dic).update(end_date=end_date,
                                                                       actual_end_date=calculate_end_date,
                                                                       work_time=work_time,
                                                                       actual_time=work_time,
                                                                       calculate_end_date=calculate_end_date)
        return Response({'results': serializer_data})


@method_decorator([api_recorder], name="dispatch")
class OverTimeView(APIView):
    permission_classes = (IsAuthenticated,)
    queryset = ApplyForExtraWork.objects.order_by('-id')

    def send_message(self, user, content):
        phone = user.phone_number
        ding_api = DinDinAPI()
        ding_uid = ding_api.get_user_id(phone)
        ding_api.send_message([ding_uid], content, attendance=True)

    def get(self, request):
        # 分页
        user = self.request.user
        page = self.request.query_params.get("page", 1)
        page_size = self.request.query_params.get("page_size", 10)
        try:
            st = (int(page) - 1) * int(page_size)
            et = int(page) * int(page_size)
        except:
            raise ValidationError("page/page_size异常，请修正后重试")
        level = None
        if self.request.query_params.get('apply'):  # 查看自己的加班申请
            section = Section.objects.filter(in_charge_user=user, name__startswith='生产').last()
            all_user = [user.username] + (list(section.section_users.filter(is_active=True).values_list('username', flat=True)) if section else [])
            data = self.queryset.filter(user__username__in=set(all_user)).order_by('-id')
        else:  # 审批加班申请
            res = get_user_level()
            detail = res.get(user.username)
            if not detail:
                raise ValidationError('该用户不是相关生产部门负责人')
            users, level = set(detail['users']), detail['level']
            # 补卡申请(未审批 + 已审批)
            if level == 1:
                data = self.queryset.filter(user__username__in=users)
            elif level == 2:
                data = self.queryset.filter(Q(f_approver__in=users, f_handling_result=1) | Q(user__username__in=users))
            else:
                data = self.queryset.filter(Q(s_approver__in=users, s_handling_result=1) | Q(user__username__in=users))
        serializer = ApplyForExtraWorkSerializer(data, many=True)
        count = len(serializer.data)
        result = serializer.data[st:et]
        return Response({'results': result, 'count': count, 'level': level})

    @atomic
    def post(self, request):  # 处理加班申请
        data = self.request.data
        user_name = self.request.user.username
        obj = ApplyForExtraWork.objects.filter(id=data.get('id')).first()
        flag, now_time = False, datetime.datetime.now()
        f_handling_suggestion, f_handling_result, s_handling_suggestion, s_handling_result, handling_suggestion, handling_result = \
            data.get('f_handling_suggestion'), data.get('f_handling_result'), data.get('s_handling_suggestion'), data.get('s_handling_result'), \
            data.get('handling_suggestion'), data.get('handling_result')
        if isinstance(f_handling_result, int):
            obj.f_handling_suggestion = f_handling_suggestion
            obj.f_handling_result = f_handling_result
            obj.f_approver = user_name
            obj.f_approver_time = now_time
        if isinstance(s_handling_result, int):
            obj.s_handling_suggestion = s_handling_suggestion
            obj.s_handling_result = s_handling_result
            obj.s_approver = user_name
            obj.s_approver_time = now_time
        if isinstance(handling_result, int):
            obj.handling_suggestion = handling_suggestion
            obj.handling_result = handling_result
            obj.approver_time = now_time
            obj.approver = user_name
            flag = True
        obj.save()
        serializer_data = ApplyForExtraWorkSerializer(obj).data
        if flag:
            user = obj.user
            content = {
                "title": "加班申请",
                "form": [
                    {"key": "姓名:", "value": user.username},
                    {"key": "班组:", "value": serializer_data.get('group')},
                    {"key": "班次:", "value": serializer_data.get('classes')},
                    {"key": "岗位:", "value": serializer_data.get('section')},
                    {"key": "机台:", "value": serializer_data.get('equip')},
                    {"key": "加班开始时间:", "value": serializer_data.get('begin_date')},
                    {"key": "加班结束时间:", "value": serializer_data.get('end_date')},
                    {"key": "加班理由:", "value": serializer_data.get('desc')},
                    {"key": "处理意见:", "value": serializer_data.get('handling_suggestion')},
                    {"key": "处理结果:", "value": serializer_data.get('handling_result')}
                ],
            }
            self.send_message(user, content)
        # if data.get('handling_result'):  # 申请通过
        #     begin_date = obj.begin_date
        #     end_date = obj.end_date
        #     work_time = round((end_date - begin_date).seconds / 3600, 2)
        #     equips = serializer_data.get('equip')
        #     for equip in equips.split(','):
        #         EmployeeAttendanceRecords.objects.create(
        #             user=obj.user,
        #             section=obj.section,
        #             factory_date=serializer_data.get('factory_date'),
        #             begin_date=serializer_data.get('begin_date'),
        #             end_date=serializer_data.get('end_date'),
        #             work_time=work_time,
        #             actual_time=work_time,
        #             classes=serializer_data.get('classes'),
        #             group=serializer_data.get('group'),
        #             equip=equip,
        #             status='加班'
        #         )
        #     # 记录考勤打卡详情
        #     AttendanceClockDetail.objects.create(
        #         name=user.username,
        #         equip=obj.equip,
        #         group=obj.group,
        #         classes=obj.classes,
        #         section=obj.section,
        #         work_type='加班',
        #         date=datetime.date(date_time.year, date_time.month, date_time.day),
        #         date_time=date_time
        #     )
        return Response({'results': serializer_data})


@method_decorator([api_recorder], name="dispatch")
class AttendanceRecordSearch(APIView):
    permission_classes = (IsAuthenticated,)

    def get(self, request):
        username = self.request.user.username
        detail = self.request.query_params.get("detail")
        date = self.request.query_params.get('date')
        day = self.request.query_params.get('day')
        year = int(date.split('-')[0])
        month = int(date.split('-')[1])
        flag, clock_type = get_user_weight_flag(self.request.user)
        queryset = EmployeeAttendanceRecords.objects.filter(
            Q(Q(begin_date__isnull=False, end_date__isnull=False) | Q(begin_date__isnull=True, end_date__isnull=True))
            & ~Q(is_use='废弃') & Q(user=self.request.user), clock_type=clock_type).order_by('id')
        if day:  # 当前的上下班时间
            record = EmployeeAttendanceRecords.objects.filter(
                ~Q(is_use='废弃') & Q(user=self.request.user, factory_date=f"{date}-{day}"), clock_type=clock_type).order_by('begin_date')
            group_setup = None
            for obj in AttendanceGroupSetup.objects.filter(type=clock_type):
                if username in obj.principal:
                    group_setup = obj
                    break
                if username in obj.users.all().values_list('username', flat=True):
                    group_setup = obj
                    break
            if not group_setup:
                raise ValidationError(f'{username}不在考勤组')
            # 标准上下班时间
            s_fac_tory_date = f"{date}-{day if len(day) > 1 else ('0' + day)}"
            if clock_type == '密炼':
                standard_begin_time, standard_end_time = get_standard_time(username, s_fac_tory_date)
                if not standard_begin_time:
                    raise ValidationError('未找到排班信息')
            else:
                if record:
                    standard_begin_time, standard_end_time = record.last().standard_begin_date, record.last().standard_end_date
                else:
                    u_info = WeightClassPlanDetail.objects.filter(weight_class_plan__user__username=username,
                                                                  factory_date=s_fac_tory_date,
                                                                  weight_class_plan__delete_flag=False).last()
                    if not u_info or (u_info and u_info.class_code == '休'):
                        raise ValidationError('未找到排班信息')
                    res = get_work_time(u_info.class_code, s_fac_tory_date)
                    if not res:
                        raise ValidationError('未找到排班信息')
                    begin_time, end_time = list(res.values())[0]
                    standard_begin_time, standard_end_time = datetime.datetime.strptime(begin_time,'%Y-%m-%d %H:%M:%S'), datetime.datetime.strptime(end_time, '%Y-%m-%d %H:%M:%S')
            results = {
                    'attendance_st': str(standard_begin_time),
                    'attendance_et': str(standard_end_time)
                }
            if record:
                begin_date = record.first().begin_date
                end_date = record.last().end_date
                all_time = 0
                lst = []
                all_time_query = record.filter(Q(Q(begin_date__isnull=False, end_date__isnull=False) |
                                            Q(begin_date__isnull=True, end_date__isnull=True))).values()
                for query in all_time_query:
                    query.pop('equip')
                    query.pop('id')
                    if query not in lst:
                        lst.append(query)
                        all_time += query.get('actual_time')

                if begin_date:
                    work_time = [{'title': f"上班: {datetime.datetime.strftime(begin_date, '%Y-%m-%d %H:%M:%S')}"}]
                else:
                    work_time = []
                if record.filter(status='调岗'):
                    times = record.filter(status='调岗').values_list('begin_date', flat=True).order_by('begin_date')
                    if times:
                        lst = list(set(times))
                        lst.sort()
                        for t in lst:
                            work_time.append({'title': f"调岗: {datetime.datetime.strftime(t, '%Y-%m-%d %H:%M:%S')}"})
                if end_date:
                    work_time.append({'title': f"下班: {datetime.datetime.strftime(end_date, '%Y-%m-%d %H:%M:%S')}"})
                results['work_time'] = round(all_time, 2) if all_time else 0
                results['time'] = work_time
                return Response(results)
            return Response(results)
        # 汇总数据
        results = {
            'days': [],
            'work_times': 0,  # 总工时
            'cd': 0,  # 迟到
            'zt': 0,  # 早退
        }
        lst = []
        for item in queryset.filter(factory_date__year=year, factory_date__month=month).values():
            item.pop('equip', None)
            item.pop('id', None)
            day = int(str(item['factory_date']).split('-')[-1])
            record = queryset.filter(factory_date=item['factory_date'])
            if item not in lst:
                lst.append(item)
                results['work_times'] += item['actual_time']
                attendance_group_obj = None
                for obj in AttendanceGroupSetup.objects.filter(type=clock_type):
                    if username in obj.principal:
                        attendance_group_obj = obj
                        break
                    if username in obj.users.all().values_list('username', flat=True):
                        attendance_group_obj = obj
                        break
                if not attendance_group_obj:
                    raise ValidationError(f'{username}不在考勤组')
            if day not in results.get('days'):
                results['days'].append(day)
                begin_date = record.first().begin_date
                end_date = record.last().end_date
                if begin_date and end_date:  # 导入的没有上岗和离岗时间
                    # 标准上下班时间
                    standard_begin_time, standard_end_time = record.first().standard_begin_date, record.first().standard_end_date
                    if not standard_begin_time:
                        raise ValidationError('未找到排班信息')
                    if begin_date > datetime.datetime.strptime(
                            f"{str(item['factory_date'])} {str(standard_end_time)[11:]}",
                            '%Y-%m-%d %H:%M:%S'):
                        results['cd'] += 1
                    if end_date and end_date < datetime.datetime.strptime(
                            f"{str(item['factory_date'])} {str(standard_end_time)[11:]}",
                            '%Y-%m-%d %H:%M:%S'):
                        results['zt'] += 1
        if detail:
            equip_list = queryset.filter(factory_date__year=year, factory_date__month=month
                                         ).values('equip', 'section').annotate(work_times=Sum('actual_time')).values(
                'equip', 'section', 'actual_time')
            results['equip_list'] = equip_list
            work_detail = queryset.filter(factory_date__year=year, factory_date__month=month
                                          ).values('factory_date', 'classes', 'group', 'equip', 'section', 'actual_time')
            results['work_detail'] = work_detail
        else:
            days = queryset.values('factory_date').annotate(Count('id')).count()
            if days != 0:
                results['avg_times'] = round(results['work_times'] / days, 2)
            else:
                results['avg_times'] =0
        results['work_times'] = round(results['work_times'], 2)
        return Response({'results': results})


@method_decorator([api_recorder], name="dispatch")
class GroupClockDetailView(APIView):
    permission_classes = (IsAuthenticated,)

    def get(self, request):
        query_name = self.request.query_params.get('query_name')
        group = self.request.query_params.get('group', 'A班')
        select_date = self.request.query_params.get('select_date', datetime.datetime.now().strftime('%Y-%m-%d'))
        results = {}
        if not query_name:
            user_name = self.request.user.username
            # 查看人员是否为考勤负责人
            u_flag = User.objects.filter(is_superuser=True, username=user_name)
            if not u_flag:  # 是否为管理员
                group_sets = AttendanceGroupSetup.objects.filter(principal__icontains=user_name)
                if not group_sets:
                    raise ValidationError('当前用户不是考勤组负责人, 请联系管理员')
                clock_types = group_sets.values_list('type', flat=True).distinct()
            else:
                clock_types = GlobalCode.objects.filter(use_flag=True, global_type__use_flag=True, global_type__type_name='绩效计算岗位类别').values_list('global_name', flat=True).distinct()
            records = EmployeeAttendanceRecords.objects.filter(~Q(is_use='废弃'), clock_type__in=clock_types, factory_date=select_date, group=group).order_by('clock_type', 'section', 'equip', 'user__username', 'status')
            exist_r = []
            for s in records:
                name, section, status, clock_type, end_date, equip = s.user.username, s.section, s.status, s.clock_type, s.end_date, s.equip
                _key = f"{select_date}-{name}-{section}"
                real_name = name if status != '调岗' else f"{name}[{status}]"
                color = '' if not end_date else 'orange'
                s_type = results.get(clock_type)
                _s_info = {'name': real_name, 'color': color, 'equip': equip}
                if s_type:
                    s_section = s_type.get(section)
                    if s_section:
                        names = [i for i in s_section if i['name'] == real_name]
                        if _key in exist_r:
                            if names:
                                if equip in names[0]['equip']:
                                    continue
                                _equip = names[0]['equip'] + f'/{equip}'
                                if len(_equip.split('/')) >= 15 and clock_type == '密炼':
                                    _equip = 'Z01~Z15'
                                names[0]['equip'] = _equip
                            else:
                                s_section += [_s_info]
                        else:
                            s_section += [_s_info]
                    else:
                        s_type[section] = [_s_info]
                else:
                    results[clock_type] = {section: [_s_info]}
                exist_r.append(_key)
        else:  # 某位员工当班组打卡明细
            clock_type = self.request.query_params.get('clock_type')
            section = self.request.query_params.get('section')
            user_name = re.split('\[|-', query_name)[0]
            records = EmployeeAttendanceRecords.objects.filter(~Q(is_use='废弃'), user__username=user_name, factory_date=select_date,
                                                               group=group, clock_type=clock_type, section=section).order_by('begin_date', 'equip')
            for s in records:
                section, status, equip = s.section, s.status, s.equip
                begin_date = s.begin_date.strftime('%Y-%m-%d %H:%M:%S') if s.begin_date else None
                end_date = s.end_date.strftime('%Y-%m-%d %H:%M:%S') if s.end_date else None
                s_status = results.get(status)
                if s_status:
                    if equip in s_status['equip']:
                        continue
                    s_status['equip'] += f",{equip}"
                else:
                    results[status] = {'id': s.id, 'user_name': user_name, 'section': section, 'equip': equip, 'status': status,
                                       'begin_date': begin_date, "end_date": end_date}
            results = results.values()
        return Response(results)


@method_decorator([api_recorder], name="dispatch")
class AttendanceTimeStatisticsViewSet(ModelViewSet):
    queryset = EmployeeAttendanceRecords.objects.filter(
        Q(end_date__isnull=False, begin_date__isnull=False) |
        Q(end_date__isnull=True, begin_date__isnull=True)).order_by('begin_date')
    serializer_class = EmployeeAttendanceRecordsSerializer
    permission_classes = (IsAuthenticated,)

    def list(self, request, *args, **kwargs):
        clock_type = self.request.query_params.get('clock_type')
        work_time = self.request.query_params.get('work_time')  # 获取当天的班次和时间信息
        if work_time:
            username = self.request.query_params.get('name')  # 获取上下班时间
            res = get_classes_plan(work_time, clock_type, username)
            return Response(res)
        classes_handle = self.request.query_params.get('classes_handle')  # 班次提交过滤数据
        if classes_handle:
            filter_kwargs = {'clock_type': clock_type}
            date = self.request.query_params.get('date')
            classes = self.request.query_params.get('classes')
            equip = self.request.query_params.get('equip')
            section = self.request.query_params.get('section')
            group = self.request.query_params.get('group')
            if date:
                filter_kwargs['factory_date'] = date
            if classes:
                filter_kwargs['classes'] = classes
            if group:
                filter_kwargs['group'] = group
            else:
                filter_kwargs['group__in'] = get_user_group(self.request.user.username, clock_type)
            if equip:
                filter_kwargs['equip__in'] = equip.split(',')
            else:  # 所有机台
                if clock_type == '密炼':
                    equip_type = '密炼设备'
                    equip_info = list(Equip.objects.filter(category__equip_type__global_name=equip_type, use_flag=True, equip_no__startswith='Z').values_list('equip_no', flat=True))
                else:
                    equip_type = '称量设备'
                    equip_info = list(Equip.objects.filter(category__equip_type__global_name=equip_type, use_flag=True).values_list('equip_no', flat=True))
                filter_kwargs['equip__in'] = equip_info
            if section:
                filter_kwargs['section__in'] = section.split(',')
            else:
                section_list = list(PerformanceJobLadder.objects.filter(delete_flag=False, type=clock_type).values_list('name', flat=True))
                filter_kwargs['section__in'] = section_list
            queryset = self.get_queryset().filter(~Q(is_use='废弃'), **filter_kwargs)
            data = self.get_serializer(queryset, many=True).data
            records = {}
            # 处理多机台
            for i in data:
                prefix = f"{i['user']}-{i['section']}-{i['group']}-{i['actual_time']}"
                if prefix not in records:
                    i['id'] = [i['id']]
                    records[prefix] = i
                else:
                    records[prefix]['id'] = records[prefix]['id'] + [i['id']]
                    records[prefix]['equip'] = records[prefix]['equip'] + ',' + i['equip']
            return Response({'results': records.values()})
        else:
            name = self.request.query_params.get('name')
            date = self.request.query_params.get('date')
            year, month = int(date.split('-')[0]), int(date.split('-')[-1])
            queryset = self.get_queryset().filter(factory_date__year=year, factory_date__month=month, user__username=name, clock_type=clock_type)
            data = self.get_serializer(queryset, many=True).data
            principal, id_card_num = None, None
            if data:
                user = User.objects.filter(username=name).first()
                id_card_num = user.id_card_num
                principal_obj = None
                for obj in AttendanceGroupSetup.objects.filter(type=clock_type):
                    if user.username in obj.principal:
                        principal_obj = obj
                        break
                    if user.username in obj.users.all().values_list('username', flat=True):
                        principal_obj = obj
                        break
                principal = principal_obj.principal if principal_obj else ''
            return Response({'results': data, 'principal': principal, 'id_card_num': id_card_num})

    @atomic
    def create(self, request, *args, **kwargs):
        report_list = self.request.data.get('report_list', [])
        confirm_list = self.request.data.get('confirm_list', [])
        reject_list = self.request.data.get('reject_list', [])
        abandon_list = self.request.data.get('abandon_list', [])
        opera_user = self.request.user.username
        clock_type = self.request.data.get('clock_type')
        factory_date, opera_type, delete_ids = None, None, []
        if report_list:  # 添加考勤数据
            # 处理机台数据
            s_data, create_data = report_list[0], []
            equip = s_data.get('equip', [])
            for s_equip in equip:
                i = deepcopy(s_data)
                i['equip'] = s_equip
                create_data.append(i)
            serializer = self.get_serializer(data=create_data, many=True)
            serializer.is_valid(raise_exception=True)
            serializer.save()
        elif confirm_list:  # 确认某一天的考勤数据
            opera_type = '确认'
            for item in confirm_list:  # #141414黑色
                ids, is_use, factory_date, is_check = item['id'], item.get('is_use'), item.get('factory_date'), item.get('is_check', False)
                id_param = [ids] if isinstance(ids, int) else ids
                EmployeeAttendanceRecords.objects.filter(pk__in=id_param)\
                    .update(actual_time=item.get('actual_time', 0), is_use=is_use, record_status='#141414', opera_flag=1,
                            actual_begin_date=item.get('actual_begin_date'), actual_end_date=item.get('actual_end_date'),
                            calculate_begin_date=item.get('actual_begin_date'), calculate_end_date=item.get('actual_end_date'),
                            is_check=is_check)
        elif reject_list:  # 审批驳回某一天的数据 #DA1F27 红色
            opera_type = '单天驳回'
            id_list = []
            for item in reject_list:
                factory_date = item.get('factory_date')
                s_record = EmployeeAttendanceRecords.objects.filter(id=item['id']).last()
                if s_record.opera_flag in [2, 3]:
                    raise ValidationError('已经审核/审批的数据不可驳回')
                id_list.append(item['id'])
            EmployeeAttendanceRecords.objects.filter(id__in=id_list).update(is_use='驳回', record_status='#DA1F27',
                                                                            opera_flag=0)
        elif abandon_list:  # 废弃
            opera_type = '废弃'
            for item in abandon_list:
                ids, is_use, factory_date = item['id'], item.get('is_use'), item.get('factory_date')
                id_param = [ids] if isinstance(ids, int) else ids
                EmployeeAttendanceRecords.objects.filter(pk__in=id_param).update(is_use=is_use, opera_flag=0)
                delete_ids += id_param
        else:
            raise ValidationError('不支持的操作')
        # 记录操作履历
        if factory_date and opera_type:
            if delete_ids:
                c_data = [EmployeeAttendanceRecordsLog(**{'opera_user': opera_user, 'opera_type': opera_type,
                                                          'record_date': factory_date, 'delete_id': d_id,
                                                          'clock_type': clock_type})
                          for d_id in delete_ids]
            else:
                c_data = [EmployeeAttendanceRecordsLog(**{'opera_user': opera_user, 'opera_type': opera_type,
                                                          'record_date': factory_date, 'clock_type': clock_type})]
            EmployeeAttendanceRecordsLog.objects.bulk_create(c_data)
        return Response('ok')

    @atomic
    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)

        if getattr(instance, '_prefetched_objects_cache', None):
            # If 'prefetch_related' has been applied to a queryset, we need to
            # forcibly invalidate the prefetch cache on the instance.
            instance._prefetched_objects_cache = {}
        # 废弃操作履历
        EmployeeAttendanceRecordsLog.objects.create(**{'opera_user': self.request.user.username, 'delete_id': instance.id,
                                                       'opera_type': '废弃', 'record_date': instance.factory_date,
                                                       'clock_type': instance.clock_type})

        return Response(serializer.data)


@method_decorator([api_recorder], name="dispatch")
class EmployeeAttendanceRecordsLogViewSet(mixins.ListModelMixin, GenericViewSet):
    queryset = EmployeeAttendanceRecordsLog.objects.order_by('-opera_time')
    serializer_class = EmployeeAttendanceRecordsLogSerializer
    filter_backends = (DjangoFilterBackend,)
    filter_class = EmployeeAttendanceRecordsLogFilter
    permission_classes = (IsAuthenticated,)


@method_decorator([api_recorder], name="dispatch")
class AttendanceClockDetailViewSet(ModelViewSet):
    queryset = AttendanceClockDetail.objects.order_by('id')
    serializer_class = AttendanceClockDetailSerializer
    filter_backends = (DjangoFilterBackend,)
    filter_class = AttendanceClockDetailFilter
    pagination_class = None
    permission_classes = (IsAuthenticated,)


@method_decorator([api_recorder], name="dispatch")
class AttendanceResultAuditView(APIView):
    permission_classes = (IsAuthenticated, )

    def get(self, request):
        date = self.request.query_params.get('date')
        audit = self.request.query_params.get('audit', None)
        approve = self.request.query_params.get('approve', None)
        clock_type = self.request.query_params.get('clock_type')
        if not all([date, clock_type]):
            raise ValidationError('缺少参数[日期与类别必选]')
        kwargs = {'clock_type': clock_type, 'date': date}
        if audit:
            kwargs['audit_user__isnull'] = False
        if approve:
            kwargs['approve_user__isnull'] = False
        records = AttendanceResultAudit.objects.filter(**kwargs).order_by('-created_date').values('date', 'approve_user', 'audit_user', 'result', 'result_desc')
        data = []
        for s_record in records:
            approve_user, audit_user, s_result = s_record.get('approve_user'), s_record.get('audit_user'), s_record.get('result')
            opera_type = '审批' if approve_user else '审核'
            opera_user = approve_user if approve_user else audit_user
            opera_result = '同意' if s_result else '驳回'
            s_record.update({'opera_user': opera_user, 'opera_type': opera_type, 'opera_result': opera_result})
            data.append(s_record)
        return Response({"results": data})

    @atomic
    def post(self, request):
        data = self.request.data
        overall = data.pop('overall', None)  # 整体提交
        audit = data.pop('audit', None)  # 审核人
        approve = data.pop('approve', None)  # 审批人
        clock_type = data.get('clock_type', '密炼')
        is_user = self.request.user.username
        opera_type = None
        attendance_data = EmployeeAttendanceRecords.objects.filter(~Q(is_use='废弃'),
                                                                   Q(begin_date__isnull=False, end_date__isnull=False)
                                                                   | Q(begin_date__isnull=True, end_date__isnull=True),
                                                                   factory_date__in=days_cur_month_dates(),
                                                                   clock_type=clock_type)
        if not attendance_data:
            raise ValidationError('暂无本月考勤数据需要处理')
        s_record = attendance_data.last()  # 获取已确认的某条数据
        not_overall = attendance_data.exclude(record_status='#141414')  # 非整体提交数据
        if overall:  # #141414 黑色
            opera_type = '整体提交'
            if not_overall:
                if s_record.opera_type != 0:
                    raise ValidationError('未操作数据才可以确认')
                attendance_data.update(record_status='#141414', opera_flag=1)
            else:
                raise ValidationError('请勿重复提交')
        else:
            # 未整体提交的考勤数据不能审核、审批
            if not_overall:
                raise ValidationError('存在未确认的考勤数据, 请处理后再进行审批/审核操作')
            opera_flag = 0
            if approve:
                s_record = attendance_data.exclude(opera_flag__in=[2, 3]).last()
                if not s_record:
                    raise ValidationError('考勤数据已经审批完成')
                if s_record.opera_flag != 1:
                    raise ValidationError('确认的考勤数据才可审批')
                opera_type, opera_flag, data['approve_user'] = '审批', 2, is_user
            if audit:
                s_record = attendance_data.exclude(opera_flag__in=[3]).last()
                if not s_record:
                    raise ValidationError('考勤数据已经审核完成')
                if s_record.opera_flag != 2:
                    raise ValidationError('审批通过的考勤数据才可审核')
                opera_type, opera_flag, data['audit_user'] = '审核', 3, is_user

            AttendanceResultAudit.objects.create(**data)
            # 审核或审批不通过,当月考勤数据全为红色 #DA1F27 红色
            if not data.get('result'):
                attendance_data.update(record_status='#DA1F27', opera_flag=0)
            else:
                attendance_data.update(opera_flag=opera_flag)
        # 记录履历
        EmployeeAttendanceRecordsLog.objects.create(**{'opera_user': is_user, 'opera_type': opera_type, 'clock_type': clock_type})
        return Response('ok')


class MaterialExpendSummaryView(APIView):
    permission_classes = (IsAuthenticated, PermissionClass({'view': 'view_material_expend_summary'}))

    def get(self, request):
        start_time = self.request.query_params.get('s_time')
        end_time = self.request.query_params.get('e_time')
        filter_material_type = self.request.query_params.get('material_type')
        material_name = self.request.query_params.get('material_name')
        equip_no = self.request.query_params.get('equip_no')
        product_no = self.request.query_params.get('product_no')
        if not all([start_time, end_time]):
            raise ValidationError('请选择日期范围进行查询')
        try:
            e_time = datetime.datetime.strptime(end_time, '%Y-%m-%d')
            s_time = datetime.datetime.strptime(start_time, '%Y-%m-%d')
        except Exception:
            raise ValidationError('日期错误！')
        diff = e_time - s_time
        if diff.days > 30:
            raise ValidationError('搜索日期跨度不得超过一个月！')
        qs = ProductClassesPlan.objects.filter(
            work_schedule_plan__plan_schedule__day_time__gte=s_time,
            work_schedule_plan__plan_schedule__day_time__lte=e_time)
        class_uid_date_dict = dict(qs.values_list('plan_classes_uid', 'work_schedule_plan__plan_schedule__day_time'))
        plan_uids = list(qs.values_list(
            'plan_classes_uid', flat=True))
        queryset = ExpendMaterial.objects.filter(plan_classes_uid__in=plan_uids)
        if equip_no:
            queryset = queryset.filter(equip_no=equip_no.strip())
        if product_no:
            queryset = queryset.filter(product_no__icontains=product_no.strip())
        if material_name:
            queryset = queryset.filter(material_name__icontains=material_name.strip())
        data = queryset.values('equip_no', 'product_no', 'material_name', 'plan_classes_uid').annotate(actual_weight=Sum('actual_weight')/100).order_by('product_no', 'equip_no', 'material_name')
        material_type_dict = dict(Material.objects.values_list('material_name', 'material_type__global_name'))
        days = date_range(s_time, e_time)
        ret = {}
        material_weight_dict = {}
        for item in data:
            material_name = item['material_name'].rstrip('-C').rstrip('-X')
            equip_no = item['equip_no']
            product_no = item['product_no']
            key = '{}-{}-{}'.format(equip_no, material_name, product_no)
            weight = item['actual_weight']
            material_type = material_type_dict.get(item['material_name'], 'UN_KNOW')
            factory_date = class_uid_date_dict[item['plan_classes_uid']].strftime("%Y-%m-%d")
            if key not in ret:
                ret[key] = {
                    'material_type': material_type,
                    'material_name': material_name,
                    'equip_no': equip_no,
                    'product_no': product_no,
                    factory_date: weight,
                    'total_weight': weight
                }
            else:
                if factory_date in ret[key]:
                    ret[key][factory_date] += weight
                else:
                    ret[key][factory_date] = weight
                ret[key]['total_weight'] += weight
            material_weight_dict[material_name] = material_weight_dict.get(material_name, 0) + weight
        result = ret.values()
        if filter_material_type:
            result = list(filter(lambda x: x['material_type'] == filter_material_type.strip(), result))
        result = sorted(result, key=itemgetter('material_type', 'material_name', 'equip_no'))  #  按多个字段排序
        return Response({'days': days, 'data': result, 'material_weight_dict': material_weight_dict})


@method_decorator([api_recorder], name="dispatch")
class ShiftTimeSummaryView(APIView):
    permission_classes = (IsAuthenticated, PermissionClass({'view': 'view_shift_time_summary'}))

    def get(self, request):
        st = self.request.query_params.get('st', None)
        et = self.request.query_params.get('et', None)
        classes = self.request.query_params.get('classes', None)
        if not st or not et:
            raise ValidationError('缺少查询起止时间参数')
        filter_kwargs = {
            'factory_date__gte': st,
            'factory_date__lte': et
        }
        if classes:
            filter_kwargs['classes'] = classes
        results = {}
        group = WorkSchedulePlan.objects.filter(start_time__date__gte=st, start_time__date__lte=et). \
            values_list('group__global_name','classes__global_name', 'start_time__date').order_by('start_time')
        group_dic = {f'{item[2]}_{item[1]}': item[0] for item in group}
        queryset = TrainsFeedbacks.objects.filter(**filter_kwargs).values('classes', 'factory_date', 'equip_no').\
            annotate(begin=Min('begin_time'), end=Max('end_time')).order_by('begin')
        gcs = dict(GlobalCode.objects.filter(
            global_type__type_name='交接班时间标准').values_list('global_name', 'description'))

        for item in queryset:
            factory_date = item['factory_date']
            equip_no = item['equip_no']
            standard_time = int(gcs.get(equip_no, 0))
            if item['classes'] == '早班':
                s = datetime.datetime.strptime(f'{str(factory_date)} 08:00:00', '%Y-%m-%d %H:%M:%S')
                e = datetime.datetime.strptime(f'{str(factory_date)} 20:00:00', '%Y-%m-%d %H:%M:%S')
                # 计算耗时
                time_consuming = round((item['begin'] - s + e - item['end']).total_seconds() / 60, 2)
                key = f'{factory_date}_早班'
            elif item['classes'] == '夜班':
                s = datetime.datetime.strptime(f'{str(factory_date)} 20:00:00', '%Y-%m-%d %H:%M:%S')
                e = datetime.datetime.strptime(f'{str(factory_date)} 08:00:00', '%Y-%m-%d %H:%M:%S') + datetime.timedelta(days=1)
                time_consuming = round((item['begin'] - s + e - item['end']).total_seconds() / 60, 2)
                key = f'{factory_date}_夜班'
            else:
                continue
            if not results.get(key, None):
                results[key] = {'consuming': 0, 'abnormal': 0, 'factory_date': factory_date,
                                'classes': key.split('_')[-1],
                                'group': group_dic.get(f'{str(factory_date)}_{key.split("_")[-1]}', None)}
            results[key][f'{equip_no}_time_consuming'] = time_consuming if abs(time_consuming) <= 20 else None
            results[key][f'{equip_no}_time_abnormal'] = time_consuming if abs(time_consuming) > 20 else None
            results[key]['consuming'] += time_consuming if abs(time_consuming) <= 20 else 0
            results[key]['abnormal'] += time_consuming if abs(time_consuming) > 20 else 0
            results[key][f'{equip_no}_rate'] = None if not time_consuming else round((standard_time / 60 / time_consuming) * 100, 2)
        res = list(results.values())
        for item in res:
            equip_count = (len(item) - 5) // 2
            item['consuming'] = round(item['consuming'] / equip_count, 2)
            item['abnormal'] = round(item['abnormal'] / equip_count, 2)
        # 增加分页
        page = self.request.query_params.get('page', 1)
        page_size = self.request.query_params.get('page_size', 10)
        try:
            begin = (int(page) - 1) * int(page_size)
            end = int(page) * int(page_size)
        except:
            raise ValidationError("page/page_size异常，请修正后重试")
        else:
            if end >= 10000:
                page_result, total_page = res[begin:], 1
            else:
                if begin not in range(0, 99999):
                    raise ValidationError("page/page_size值异常")
                if end not in range(0, 99999):
                    raise ValidationError("page/page_size值异常")
                page_result, total_page = res[begin: end], math.ceil(len(res) / int(page_size))
        return Response({'results': page_result, 'equip_sts_time': gcs, 'total_data': len(res), 'total_page': total_page})


@method_decorator([api_recorder], name="dispatch")
class ShiftTimeSummaryDetailView(APIView):

    def get(self, request):
        params = self.request.query_params
        factory_date = params.get('factory_date', None)
        classes = params.get('classes', None)
        equip_no = params.get('equip_no', None)
        if not factory_date or not classes or not equip_no:
            raise ValidationError('查询参数缺失')
        query = TrainsFeedbacks.objects.filter(factory_date=factory_date,
                                               classes=classes, equip_no=equip_no)
        if not query:
            return Response({'results': None})
        begin = query.order_by('begin_time').first()
        end = query.order_by('end_time').last()
        return Response({'results': {'begin': datetime.datetime.strftime(begin.begin_time, '%Y-%m-%d %H:%M:%S'),
                                     'end': datetime.datetime.strftime(end.end_time, '%Y-%m-%d %H:%M:%S')}})


@method_decorator([api_recorder], name="dispatch")
class RubberFrameRepairView(APIView):
    """胶架维修记录"""
    permission_classes = (IsAuthenticated, PermissionClass({'view': 'view_rubber_frame_repair',
                                                            'add': 'add_rubber_frame_repair'}))

    def get(self, request):
        date_time = self.request.query_params.get('date_time')
        results = {}
        query_set = RubberFrameRepair.objects.filter(date_time=date_time)
        if query_set:
            max_times = query_set.aggregate(max_times=Max('times'))['max_times']
            instance = query_set.filter(times=max_times).last()
            results['details'] = json.loads(instance.content)
        else:  # 返回空格式
            details = [{'name': '待维修胶架发出量', '总计': None}, {'name': '已维修胶架数量', '总计': None},
                       {'name': '待维修胶架数量', '总计': None}, {'name': '输送人员', '总计': None},
                       {'name': '确认人员', '总计': None}]
            data = days_cur_month_dates(date_time)
            update_data = {i: None for i in range(1, len(data) + 1)}
            for detail in details:
                detail.update(update_data)
            results['details'] = details
        results['date_time'] = date_time
        return Response({'results': results})

    @atomic
    def post(self, request):
        date_time = self.request.data.get('date_time')
        details = self.request.data.get('details')
        save_user = self.request.user.username
        if not all([date_time, details]):
            raise ValidationError('参数异常')
        # 获取最新保存次数
        max_times = RubberFrameRepair.objects.filter(date_time=date_time).aggregate(max_times=Max('times'))['max_times']
        times = 1 if not max_times else max_times + 1
        RubberFrameRepair.objects.create(date_time=date_time, content=json.dumps(details), times=times, save_user=save_user)
        return Response('保存成功')


@method_decorator([api_recorder], name="dispatch")
class RubberFrameRepairSummaryView(APIView):
    """胶架维修记录汇总"""
    permission_classes = (IsAuthenticated, PermissionClass({'view': 'view_rubber_frame_repair_summary'}))

    def get(self, request):
        select_date = self.request.query_params.get('select_date')
        if not select_date:
            raise ValidationError('未选择日期')
        send_num, repaired, waited = {"name": "待维修胶架发出量"}, {"name": "已维修胶架数量"}, {"name": "待维修胶架数量"}
        for i in range(1, 13):
            s_column = f'{i}月'
            date_time = select_date + '-' + '%02d' % i
            month_data = RubberFrameRepair.objects.filter(date_time=date_time).order_by('id').last()
            if month_data:
                content = json.loads(month_data.content)[:3]
                self.handle_data(content, s_column, send_num, repaired, waited)
            else:
                send_num[s_column], repaired[s_column], waited[s_column] = 0, 0, 0
        results = [send_num, repaired, waited]
        # results = {'待维修胶架发出量': send_num, '已维修胶架数量': repaired, '待维修胶架数量': waited}
        return Response(results)

    def handle_data(self, content, s_column, send_num, repaired, waited):
        for index, s_content in enumerate(content):
            s_total = s_content['总计'] if s_content['总计'] else 0
            if index == 0:
                send_num[s_column] = s_total
                send_num['总计'] = s_total + (send_num['总计'] if '总计' in send_num else 0)
            elif index == 1:
                repaired[s_column] = s_total
                repaired['总计'] = s_total + (repaired['总计'] if '总计' in repaired else 0)
            else:
                waited[s_column] = s_total
                waited['总计'] = s_total + (waited['总计'] if '总计' in waited else 0)


@method_decorator([api_recorder], name="dispatch")
class ToolManageAccountView(APIView):
    """工装管理台帐"""
    permission_classes = (IsAuthenticated, PermissionClass({'view': 'view_tool_manage_account',
                                                            'add': 'add_tool_manage_account'}))

    def get(self, request):
        year = self.request.query_params.get('year', '')[:4]
        results = {}
        query_set = ToolManageAccount.objects.filter(year__startswith=year)
        if query_set:
            max_record = query_set.aggregate(max_record=Max('save_date'))['max_record']
            instance = query_set.filter(save_date=max_record).last()
            results.update({'day': json.loads(instance.day), 'content': json.loads(instance.content)})
        else:  # 返回空格式
            results.update({'day': [], 'content': []})
        results['year'] = year
        return Response({'results': results})

    @atomic
    def post(self, request):
        year = self.request.data.get('year')
        day = self.request.data.get('day')
        content = self.request.data.get('content')
        save_user = self.request.user.username
        if not all([year, day, content]):
            raise ValidationError('参数异常')
        ToolManageAccount.objects.create(year=year, day=json.dumps(day), content=json.dumps(content), save_user=save_user)
        return Response('保存成功')


@method_decorator([api_recorder], name='dispatch')
class WeightClassPlanViewSet(ModelViewSet):

    queryset = WeightClassPlan.objects.filter(delete_flag=False)
    serializer_class = WeightClassPlanSerializer
    permission_classes = (IsAuthenticated,)
    pagination_class = None
    FILE_NAME = '称量排班'
    EXPORT_FIELDS_DICT = {
        "班别": "classes",
        "姓名": "username",
        "岗位": "station",
    }

    def get_queryset(self):
        target_month = self.request.query_params.get('target_month')
        classes = self.request.query_params.get('classes')
        username = self.request.query_params.get('username')
        filter_kwargs = {}
        if target_month:
            filter_kwargs['target_month'] = target_month
        if classes:
            filter_kwargs['classes'] = classes
        if username:
            filter_kwargs['user__username__icontains'] = username
        return self.queryset.filter(**filter_kwargs)

    def get_serializer_class(self):
        if self.action == 'update':
            return WeightClassPlanUpdateSerializer
        else:
            return WeightClassPlanSerializer

    def list(self, request, *args, **kwargs):
        target_month = self.request.query_params.get('target_month')
        queryset = self.filter_queryset(self.get_queryset())
        if self.request.query_params.get('export'):
            date_list = days_cur_month_dates(target_month)
            add_key = {}
            for i in date_list:
                m, d = i.split('-')[-2:]
                m = m[-1] if m.startswith('0') else m
                d = d[-1] if d.startswith('0') else d
                add_key[f'{m}/{d}'] = f'{m}/{d}'
            self.EXPORT_FIELDS_DICT.update(add_key)
            data = self.get_serializer(queryset, many=True).data
            for j in data:
                weight_class_details = j.pop('weight_class_details')
                j.update(weight_class_details)
            return gen_template_response(self.EXPORT_FIELDS_DICT, data, self.FILE_NAME, sheet_name=target_month, handle_str=True)

        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)

        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)

    def destroy(self, request, *args, **kwargs):
        try:
            instance = self.get_object()
        except Exception as e:
            raise ValidationError('未找到对应数据,请刷新页面后重试')
        instance.delete_flag = 1
        instance.save()
        return Response('操作成功')

    @atomic
    @action(methods=['post'], detail=False, permission_classes=[IsAuthenticated], url_path='import-xlsx', url_name='import_xlsx')
    def import_xlsx(self, request):
        excel_file = request.FILES.get('file', None)
        if not excel_file:
            raise ValidationError('文件不可为空！')
        cur_sheet = get_cur_sheet(excel_file)
        data = get_sheet_data(cur_sheet)
        if not data:
            raise ValidationError('导入文件无内容')
        target_month = cur_sheet.name
        try:
            date_list = days_cur_month_dates(target_month)
            now_date = get_current_factory_date()['factory_date'].strftime('%Y-%m-%d')
            check = datetime.datetime.strptime(target_month, '%Y-%m')
            if target_month < now_date[:7]:
                raise ValidationError(f'导入异常: 只能导入当月及以后的数据')
            date_index = date_list.index(now_date) if now_date in date_list else 0
            for index, item in enumerate(data):
                if len(item[3:]) != len(date_list):
                    raise ValidationError(f'导入异常: 导入天数与月份不符')
                # 检查班别、姓名、岗位、排班
                if not GlobalCode.objects.filter(use_flag=True, global_type__use_flag=True, global_type__type_name='配料间排班组别', global_name=item[0]):
                    raise ValidationError(f'导入异常: 班别{item[0]}不存在')
                user = User.objects.filter(is_active=True, username=item[1]).last()
                if not user:
                    raise ValidationError(f'导入异常: 姓名{item[1]}不存在')
                if not PerformanceJobLadder.objects.filter(name=item[2]):
                    raise ValidationError(f'导入异常: 岗位{item[2]}不存在')
                class_codes = set([i for i in item[3:] if i and i != '休'])
                class_codes_set = set(GlobalCode.objects.filter(use_flag=True, global_type__use_flag=True, global_type__type_name='配料间排班详细分类').values_list('global_name', flat=True).distinct())
                if class_codes - class_codes_set:
                    raise ValidationError(f"导入异常: 排班代码异常{','.join(list(class_codes - class_codes_set))}")
                self.execute_import(date_index, target_month, item, user, date_list)
        except Exception as e:
            logger.error(e.args[0])
            error_msg = e.args[0] if '导入异常' in e.args[0] else '导入数据异常'
            raise ValidationError(error_msg)
        return Response('导入成功')

    def execute_import(self, index, target_month, item, user, date_list):
        classes, username, station = item[0], item[1], item[2]
        record = WeightClassPlan.objects.filter(delete_flag=False, target_month=target_month, user__username=username, classes=classes).last()
        if record:
            update_item = item[index + 3:]
            details = record.weight_class_details.all().order_by('factory_date')[index:]
            for j, content in enumerate(details):
                class_code = None if not update_item[j] else update_item[j]
                content.class_code = class_code
                content.save()
            if record.station != station:
                record.station = station
                record.save()
        else:
            instance = WeightClassPlan.objects.create(target_month=target_month, classes=classes, station=station, user=user,
                                                      delete_flag=False)
            create_item = item[3:]
            for k, content in enumerate(date_list):
                if k < index:
                    create_data = {'weight_class_plan': instance, 'factory_date': content, 'class_code': None}
                else:
                    create_data = {'weight_class_plan': instance, 'factory_date': content, 'class_code': create_item[k] if create_item[k] else None}
                WeightClassPlanDetail.objects.create(**create_data)


@method_decorator([api_recorder], name='dispatch')
class ShiftProductionSummaryView(APIView):
    """班次产量统计"""

    def get(self, request):
        target_month = self.request.query_params.get('target_month')
        group_name = self.request.query_params.get('group_name')
        if not target_month:
            raise ValidationError('请选择月份！')
        month_split = target_month.split('-')
        year = int(month_split[0])
        month = int(month_split[1])
        schedule_data = WorkSchedulePlan.objects.filter(
            plan_schedule__work_schedule__work_procedure__global_name='密炼',
            plan_schedule__day_time__year=year,
            plan_schedule__day_time__month=month,
        ).order_by('start_time').values('plan_schedule__day_time',
                                        'classes__global_name',
                                        'group__global_name')
        production_data = TrainsFeedbacks.objects.filter(
            factory_date__year=year,
            factory_date__month=month
        ).values('equip_no', 'factory_date', 'classes').annotate(total_trains=Count('id')).order_by('equip_no', 'factory_date')
        equip_target_data = MachineTargetYieldSettings.objects.filter(target_month=target_month).order_by('-id').values()
        target_data = {}
        if equip_target_data:
            target_data = equip_target_data[0]
        schedule_dict = {}
        for i in schedule_data:
            k = '{}-{}'.format(i['plan_schedule__day_time'].strftime("%m/%d"), i['classes__global_name'][0])
            schedule_dict[k] = i['group__global_name'][0]

        working_days = ActualWorkingDay.objects.filter(
            factory_date__year=year, factory_date__month=month).aggregate(days=Sum('num'))['days']
        working_days = 0 if not working_days else working_days
        down_days_dict = dict(EquipDownDetails.objects.filter(
            delete_flag=False,
            factory_date__year=year,
            factory_date__month=month
        ).values('equip_no').annotate(days=Sum('times')/60/24).values_list('equip_no', 'days'))
        if month == datetime.datetime.now().month and year == datetime.datetime.now().year:
            now_date = get_current_factory_date()['factory_date']
            group_schedule_days = WorkSchedulePlan.objects.filter(
                plan_schedule__work_schedule__work_procedure__global_name='密炼',
                plan_schedule__day_time__year=year,
                plan_schedule__day_time__month=month,
                plan_schedule__day_time__lte=now_date,
                group__global_name=group_name
            ).count()
        else:
            group_schedule_days = WorkSchedulePlan.objects.filter(
                plan_schedule__work_schedule__work_procedure__global_name='密炼',
                plan_schedule__day_time__year=year,
                plan_schedule__day_time__month=month,
                group__global_name=group_name
            ).count()
        group_down_days_dict = dict(EquipDownDetails.objects.filter(
            delete_flag=False,
            factory_date__year=year,
            factory_date__month=month,
            group=group_name
        ).values('equip_no').annotate(days=Sum('times') / 60 / 24).values_list('equip_no', 'days'))
        # 获取机台历史最高产量和班组
        history_info = self.get_max_values(target_month)
        equip_production_data_dict = {i: {'equip_no': i,
                                          'total_trains': 0,
                                          'target_trains': target_data.get(i, 0),
                                          'days': working_days - down_days_dict.get(i, 0),
                                          'group_days': group_schedule_days - group_down_days_dict.get(i, 0),
                                          'max_trains': history_info.get(i, {}).get('max_trains', 0),
                                          'max_group': history_info.get(i, {}).get('max_group', ''),
                                          } for i in
                                      list(Equip.objects.filter(
                                          category__equip_type__global_name="密炼设备"
                                      ).order_by('equip_no').values_list("equip_no", flat=True))}
        for d in production_data:
            equip_no = d['equip_no']
            k = '{}-{}'.format(d['factory_date'].strftime("%m/%d"), d['classes'][0])
            key = '{}-{}'.format(k, schedule_dict[k])
            trains = d['total_trains'] // 2 if equip_no == 'Z04' else d['total_trains']
            equip_production_data_dict[equip_no][key] = trains
            equip_production_data_dict[equip_no]['total_trains'] += trains
            max_trains = equip_production_data_dict[equip_no]['max_trains']
            if not max_trains or max_trains < trains:
                equip_production_data_dict[equip_no]['max_trains'] = trains
                equip_production_data_dict[equip_no]['max_group'] = schedule_dict[k]
                HistoryProductionGroup.objects.filter(target_month=target_month, equip_no=equip_no).update(max_trains=trains, max_group=schedule_dict[k])
        # 折线图
        equip_list, day_trains, day_ratio = [], [], []
        for j in equip_production_data_dict.values():
            equip_list.append(j['equip_no'])
            day_trains.append(j['total_trains'])
            ratio = 0 if not j['days'] or j['target_trains'] else round(j['total_trains'] / j['days'] / j['target_trains'] / 2 * 100, 2)
            day_ratio.append(ratio)
        return Response({'table_head': schedule_dict,
                         'data': equip_production_data_dict.values(),
                         'equip_list': equip_list,
                         'day_trains': day_trains,
                         'day_ratio': day_ratio})

    @atomic
    def get_max_values(self, target_month):
        history_info = HistoryProductionGroup.objects.filter(target_month=target_month).values('equip_no', 'max_trains', 'max_group')
        if history_info:
            return {i['equip_no']: {'max_trains': i['max_trains'], 'max_group': i['max_group']} for i in history_info}
        else:
            last_info = HistoryProductionGroup.objects.filter(target_month__lte=target_month).order_by('id').last()
            if last_info:
                create_data = HistoryProductionGroup.objects.filter(target_month=last_info.target_month).values('equip_no', 'max_trains', 'max_group')
                for i in create_data:
                    i['target_month'] = target_month
            else:
                create_data = [{'equip_no': i, 'max_trains': 0, 'max_group': '', 'target_month': target_month}
                               for i in list(Equip.objects.filter(category__equip_type__global_name="密炼设备")
                                             .order_by('equip_no').values_list("equip_no", flat=True))]
            HistoryProductionGroup.objects.bulk_create([HistoryProductionGroup(**i) for i in create_data])
            return {i['equip_no']: {'max_trains': i['max_trains'], 'max_group': i['max_group']} for i in create_data}


@method_decorator([api_recorder], name='dispatch')
class EquipDownDetailView(APIView):
    """机台停机明细"""

    def post(self, request):
        excel_file = request.FILES.get('file', None)
        if not excel_file:
            raise ValidationError('文件不可为空！')
        cur_sheet = get_cur_sheet(excel_file)
        if cur_sheet.ncols != 8:
            raise ValidationError('导入文件数据错误！')
        data = get_sheet_data(cur_sheet, start_row=1)
        detail_list = []
        for item in data:
            try:
                delta = datetime.timedelta(days=item[0])
                date_1 = datetime.datetime.strptime('1899-12-30', '%Y-%m-%d') + delta
                factory_date = datetime.datetime.strftime(date_1, '%Y-%m-%d')
            except Exception:
                raise ValidationError('密炼日期格式错误！')
            user_data = {
                "factory_date": factory_date,
                "classes": item[1],
                "group": item[2],
                "equip_no": item[3],
                "down_reason": item[4],
                "down_type": item[5],
                "details": item[6],
                "times": item[7],
            }
            detail_list.append(user_data)
        s = EquipDownDetailsImportSerializer(data=detail_list, many=True)
        if not s.is_valid():
            for i in s.errors:
                if i:
                    raise ValidationError(list(i.values())[0])
        s.save()
        return Response('ok')

    def get(self, request):
        target_month = self.request.query_params.get('target_month')
        if not target_month:
            raise ValidationError('请选择月份！')
        month_split = target_month.split('-')
        year = month_split[0]
        month = month_split[1]
        schedule_data = WorkSchedulePlan.objects.filter(
            plan_schedule__work_schedule__work_procedure__global_name='密炼',
            plan_schedule__day_time__year=year,
            plan_schedule__day_time__month=month,
        ).order_by('start_time').values('plan_schedule__day_time',
                                        'classes__global_name',
                                        'group__global_name')
        schedule_dict = {}
        for i in schedule_data:
            k = '{}-{}'.format(i['plan_schedule__day_time'].strftime("%m/%d"), i['classes__global_name'][0])
            schedule_dict[k] = i['group__global_name'][0]
        down_data = EquipDownDetails.objects.filter(
            delete_flag=False,
            factory_date__year=year,
            factory_date__month=month
        ).values('factory_date', 'classes', 'equip_no').annotate(t=Sum('times')).order_by('equip_no', 'factory_date')
        production_data = TrainsFeedbacks.objects.filter(
            factory_date__year=year,
            factory_date__month=month
        ).values('equip_no', 'factory_date', 'classes').annotate(total_trains=Count('id'))

        max_pd_data_dict = {}
        for i in production_data:
            eqn = i['equip_no']
            fd = i['factory_date']
            cs = i['classes']
            trs = i['total_trains']//2 if eqn == 'Z04' else i['total_trains']
            key = '{}-{}'.format(fd.strftime("%m/%d"), cs[0])
            if key not in max_pd_data_dict:
                max_pd_data_dict[key] = {'equip_no': eqn, 'max_trains': trs}
            else:
                if trs > max_pd_data_dict[key]['max_trains']:
                    max_pd_data_dict[key] = {'equip_no': eqn, 'max_trains': trs}

        down_data_dict = {i: {'equip_no': i} for i in
                          list(Equip.objects.filter(
                              category__equip_type__global_name="密炼设备"
                          ).order_by('equip_no').values_list("equip_no", flat=True))}
        for d in down_data:
            equip_no = d['equip_no']
            factory_date = d['factory_date']
            classes = d['classes']
            times = d['t']
            k = '{}-{}'.format(factory_date.strftime("%m/%d"), classes[0])
            key = '{}-{}'.format(k, schedule_dict[k])
            down_data_dict[equip_no][key] = times
        data = list(down_data_dict.values())
        max_data = {'equip_no': 'MAX产量机台'}
        for k, v in max_pd_data_dict.items():
            max_data['{}-{}'.format(k, schedule_dict[k])] = v['equip_no']
        data.append(max_data)
        return Response({'table_head': schedule_dict,
                         'data': data})


@method_decorator([api_recorder], name="dispatch")
class EquipDownAnalysisView(APIView):
    """机台故障停机时间分析与汇总"""

    permission_classes = (IsAuthenticated,)

    def get(self, request):
        all_equip = self.request.query_params.get('all_equip')
        group_flag = self.request.query_params.get('group_flag')
        select_date = self.request.query_params.get('select_date')  # 2022-12-06
        if not all_equip:
            equip_no = self.request.query_params.get('equip_no')  # Z01
            group = self.request.query_params.get('group')
            filter_kwargs = {}
            if group:
                filter_kwargs['group'] = group
            queryset = EquipDownDetails.objects.filter(delete_flag=False, equip_no=equip_no, factory_date=select_date, **filter_kwargs).order_by('id')
            data = {'results': queryset.values()}
            if group_flag:
                group_classes = WorkSchedulePlan.objects.filter(plan_schedule__day_time=select_date,
                                                                plan_schedule__work_schedule__work_procedure__global_name='密炼') \
                    .values('classes__global_name', 'group__global_name')
                data.update({'group_classes': group_classes})
        else:  # 查看所有机台设定值
            equip_list, _column, data = all_equip.split(','), 5, {'results': []}
            tables = math.ceil(len(equip_list) / _column)
            if tables:
                for _i in range(tables):
                    res, max_info, equip_info = {}, {}, {}
                    _equip_title = equip_list[_i * 5:(_i + 1) * 5]
                    queryset = EquipDownDetails.objects.filter(delete_flag=False, equip_no__in=_equip_title, factory_date=select_date).order_by('equip_no', 'id').values()
                    for i in queryset:
                        equip_info[i['equip_no']] = equip_info.get(i['equip_no'], []) + [i]
                        max_info[i['equip_no']] = max_info.get(i['equip_no'], 0) + 1
                    if max_info:
                        max_num = max(max_info.values())
                        for equip_no in _equip_title:
                            _s_info = equip_info.get(equip_no, [])
                            for i in range(max_num):
                                _s_data = {f"{equip_no}-begin_time": _s_info[i]['begin_time'] if _s_info[i: i+1] else '',
                                           f"{equip_no}-end_time": _s_info[i]['end_time'] if _s_info[i: i+1] else '',
                                           f"{equip_no}-times": _s_info[i]['times'] if _s_info[i: i+1] else '',
                                           f"{equip_no}-down_reason": _s_info[i]['down_reason'] if _s_info[i: i+1] else '',
                                           f"{equip_no}-down_type": _s_info[i]['down_type'] if _s_info[i: i+1] else ''}
                                index_data = res.get(i)
                                if index_data:
                                    res[i].update(_s_data)
                                else:
                                    res[i] = _s_data
                        data['results'].append({'headers': _equip_title, 'values': res.values()})
        return Response(data)

    @atomic
    def post(self, request):
        user_name = self.request.user.username
        excel_file = self.request.FILES.get('file', None)
        if excel_file:
            try:
                df = pd.read_excel(excel_file, index_col=None)
                for i in df.values:
                    if not all([i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7]]):
                        continue
                    s_date = i[0].strftime('%Y-%m-%d')
                    classes = i[1] if '班' in i[1] else f'{i[1]}班'
                    group = i[2] if '班' in i[2] else f'{i[2]}班'
                    begin_time, end_time = i[4].split('-')
                    _s_data = {'factory_date': s_date, 'classes': classes, 'group': group, 'equip_no': i[5], 'down_reason': i[7],
                               'down_type': i[6], 'begin_time': begin_time, 'end_time': end_time, 'times': i[3], 'save_user': user_name}
                    instance = EquipDownDetails.objects.filter(factory_date=s_date, classes=classes, group=group, equip_no=i[5], down_type=i[6], down_reason=i[7]).last()
                    if instance:
                        instance.times = i[3]
                        instance.begin_time = begin_time
                        instance.end_time = end_time
                        instance.update_user = user_name
                        instance.update_time = datetime.datetime.now()
                        instance.save()
                    else:
                        EquipDownDetails.objects.create(**_s_data)
            except Exception as e:
                logger.error(f'导入异常: {e.args[0]}')
                raise ValidationError('导入异常')
            return Response('导入数据成功')
        data = self.request.data.get('set_data')
        delete_id = self.request.data.get('delete_ids')
        factory_date = self.request.data.get('factory_date')
        equip_no = self.request.data.get('equip_no')
        group = self.request.data.get('group')
        classes = self.request.data.get('classes')
        # 删除物料
        if delete_id:
            EquipDownDetails.objects.filter(id=delete_id).update(delete_flag=True, update_user=user_name)
            return Response('删除成功')
        if not data:
            raise ValidationError('未录入数据')
        for s_data in data:
            rid, begin_time, end_time = s_data.get('id'), s_data.get('begin_time'), s_data.get('end_time')
            # 班次班组
            _s_data = {'begin_time': begin_time, 'end_time': end_time, 'factory_date': factory_date, 'equip_no': equip_no,
                       'down_reason': s_data.get('down_reason'), 'down_type': s_data.get('down_type'), 'times': round(s_data.get('times', 0), 2),
                       'group': group, 'classes': classes}
            if rid:
                _s_data['update_user'] = user_name
            else:
                _s_data['save_user'] = user_name
            EquipDownDetails.objects.update_or_create(defaults=_s_data, **{'id': rid})
        return Response('录入成功')


@method_decorator([api_recorder], name="dispatch")
class EquipDownSummaryView(APIView):
    """机台故障停机时间分析与汇总"""

    # permission_classes = (IsAuthenticated,)

    def get(self, request):
        select_type = self.request.query_params.get('select_type', '1')
        st = self.request.query_params.get('st')  # 2022-12-06
        et = self.request.query_params.get('et')  # 2022-12-06
        if not all([st, et]):
            now_date = datetime.datetime.now().strftime('%Y-%m-%d')
            st = et = now_date
        equip_no = self.request.query_params.get('equip_no')  # Z01
        group = self.request.query_params.get('group')
        classes = self.request.query_params.get('classes')
        filter_kwargs = {'factory_date__gte': st, 'factory_date__lte': et}
        if equip_no:
            filter_kwargs['equip_no'] = equip_no
        if group:
            filter_kwargs['group'] = group
        if classes:
            filter_kwargs['classes'] = classes
        results = []
        if select_type == '1':  # 各班机台停机时间汇总
            queryset = EquipDownDetails.objects.filter(delete_flag=False, **filter_kwargs).order_by('factory_date', 'equip_no', 'begin_time', 'end_time').values()
            total_times = list(set([i['times'] for i in queryset]))
            sorted_times = sorted(total_times, key=lambda x: x, reverse=True)
            for i in queryset:  # 更新排名
                _index = sorted_times.index(i['times']) + 1
                i.update(**{'index': _index})
                results.append(i)
        else:  # 生产分析汇总
            detail = self.request.query_params.get('detail')
            if detail:  # 停机时间汇总(详情)
                down_type = self.request.query_params.get('down_type')
                filter_kwargs['down_type'] = down_type
                down_detail = EquipDownDetails.objects.filter(delete_flag=False, **filter_kwargs).values('down_reason').annotate(total_times=Sum('times')).values('down_reason', 'total_times').order_by('-total_times')
                totals = sum([i['total_times'] for i in down_detail])
                for j in down_detail:
                    _s_data = {'down_reason': j.get('down_reason'), 'down_time': round(j.get('total_times'), 2), 'ratio': round((j.get('total_times', 0) / totals) * 100, 2)}
                    results.append(_s_data)
                if down_detail:
                    results.append({'down_reason': '总计', 'down_time': totals, 'ratio': 100})
            else:
                # 统计停机类别时间
                results = {}
                all_info = list(EquipDownDetails.objects.filter(delete_flag=False, **filter_kwargs).values('down_type').annotate(all_times=Sum('times'), index=F('down_type')).values('index', 'all_times'))
                stop_times = round(sum([i['all_times'] for i in all_info]), 2)
                for k in all_info:
                    k['index'] = k['index'] + '-' + '总时间/min'
                    k['ratio'] = round(k['all_times'] / stop_times * 100, 2)
                    k['super'] = True
                # 产量统计
                product_dict, begin_time, end_time, total_spends, record_equip = {}, None, None, 0, []
                product_info = TrainsFeedbacks.objects.filter(~Q(operation_user='Mixer2'), **filter_kwargs).order_by('equip_no', 'factory_date', 'created_date', 'actual_trains')
                equips = len(set(product_info.values_list('equip_no', flat=True)))
                days = (datetime.datetime.strptime(et, '%Y-%m-%d') - datetime.datetime.strptime(st, '%Y-%m-%d')).days + 1
                # days = len(set(product_info.values_list('factory_date', flat=True)))
                for j in product_info:
                    product_no, equip_no = j.product_no, j.equip_no
                    _p = product_dict.get(product_no)
                    if equip_no not in record_equip:
                        begin_time, end_time = j.begin_time, j.end_time
                        record_equip.append(equip_no)
                        continue
                    _interval_time = (j.begin_time - end_time).total_seconds()
                    interval_time = _interval_time if _interval_time > 0 else 20
                    mixer_time = (j.end_time - j.begin_time).total_seconds()
                    begin_time, end_time = j.begin_time, j.end_time
                    if interval_time >= 20:
                        continue
                    if not _p:
                        actual_trains = 1
                        origin_mixer = mixer_time
                        origin_interval = interval_time
                    else:
                        actual_trains = _p.get('actual_trains', 0) + 1
                        origin_mixer = _p.get('origin_mixer', 0) + mixer_time
                        origin_interval = _p.get('origin_interval', 0) + interval_time
                    s_mixer_time = round(origin_mixer / actual_trains, 2)
                    s_interval_time = round(origin_interval / actual_trains, 2)
                    all_time = round(origin_mixer + origin_interval, 2)
                    total_spends = round(total_spends + mixer_time + interval_time, 2)
                    product_dict[product_no] = {'product_no': product_no, 'actual_trains': actual_trains, 's_mixer_time': s_mixer_time,
                                                's_interval_time': s_interval_time, 'origin_mixer': origin_mixer, 'origin_interval': origin_interval,
                                                'all_times': all_time}
                # 胶料时间占比
                for k, v in product_dict.items():
                    v.update(**{'ratio': round(v.get('all_times', 0) / total_spends * 100, 2)})
                day_times = days * 2 * 12 * 60 * equips
                m_total_spends = round(total_spends / 60, 2)
                total_times = round(m_total_spends + stop_times, 2)
                missing_time = round(day_times - total_times, 2)
                summary_time = [{'index': '运行-总时间/min', 'all_times': m_total_spends, 'ratio': round(m_total_spends / total_times * 100, 2) if total_times else 0}]
                stop_time = [{'index': '停机-总时间/min', 'all_times': stop_times, 'ratio': round(stop_times / total_times * 100, 2) if total_times else 0}]
                all_times = [{'index': '总时间/min', 'all_times': total_times, 'ratio': round(total_times / day_times * 100, 2) if day_times else 0}]
                miss_times = [{'index': '缺失时间/min', 'all_times': missing_time, 'ratio': round(missing_time / day_times * 100, 2) if day_times else 0}]
                results = {'details': product_dict.values(), 'summary': summary_time + all_info + stop_time + all_times + miss_times}
        return Response({'results': results})


@method_decorator([api_recorder], name="dispatch")
class EquipDownSummaryTableView(APIView):
    """机台故障停机时间分析与汇总"""

    # permission_classes = (IsAuthenticated,)

    def get(self, request):
        equip_no = self.request.query_params.get('equip_no')
        table_flag = self.request.query_params.get('table_flag')  # 导出时展示的各个机台的数据
        group_flag = self.request.query_params.get('group_flag')  # 传入表示班组停机类型汇总
        st = self.request.query_params.get('st')  # 2022-12-06
        et = self.request.query_params.get('et')  # 2022-12-06
        export = self.request.query_params.get('export')
        if not all([st, et]):
            now_date = datetime.datetime.now().strftime('%Y-%m-%d')
            st = et = now_date
        results = {}
        if group_flag:
            equip_list = equip_no.split(',')
            # 03-31 停机原因->停机类型
            queryset = EquipDownDetails.objects.filter(delete_flag=False, factory_date__lte=et, factory_date__gte=st, equip_no__in=equip_list) \
                .values('group', 'down_type').annotate(total_times=Sum('times')).order_by('group').values('group', 'down_type', 'total_times')
            if queryset:
                titles = list(set(queryset.values_list('down_type', flat=True)))
                groups = list(set(queryset.values_list('group', flat=True)))
                groups.sort()
                data = {}
                for i in queryset:
                    group, down_type, total_times = i['group'], i['down_type'], i['total_times']
                    if f'{group}-data' not in data:
                        data[f'{group}-data'] = {title: 0 for title in titles}
                        data[f'{group}-data'].update({'group': group, down_type: total_times})
                    else:
                        data[f'{group}-data'].update({down_type: total_times})
                # 图表数据
                for k in data:
                    s_data = deepcopy(data[k])
                    group = s_data.pop('group')
                    results.update({group: s_data.values()})
                # 表格数据
                execl_data, reason, ratio, total_times, temp = deepcopy(list(data.values())), {}, {}, 0, {}
                for j in execl_data:
                    group, all_times = j.pop('group'), sum(j.values())
                    for down_type, times in j.items():
                        total_times += times
                        if down_type not in reason:
                            reason[down_type] = {'down_reason': down_type, group: times, '总计': times}
                            ratio[down_type] = {'down_reason': down_type, group: round(times / all_times * 100, 2) if all_times else 0}
                        else:
                            reason[down_type].update({group: times, '总计': reason[down_type]['总计'] + times})
                            ratio[down_type].update({group: round(times / all_times * 100, 2) if all_times else 0})
                        temp[down_type] = temp.get(down_type, 0) + times
                        # # 数据  12-15前端计算总计
                        # if '总计' not in reason:
                        #     reason['总计'] = {'down_reason': '总计', group: all_times, '总计': times}
                        # else:
                        #     if group not in reason['总计']:
                        #         reason['总计'].update({group: all_times})
                        #     reason['总计']['总计'] = reason['总计']['总计'] + times
                for m in ratio:
                    ratio[m]['总计'] = round(temp[m] / total_times * 100, 2) if total_times else 0
                results.update({'reason': reason.values(), 'ratio': ratio.values(), 'titles': titles, 'groups': groups, 'details': data.values()})
            return Response({'results': results})
        if export:
            equips_data = EquipDownDetails.objects.filter(delete_flag=False, factory_date__gte=st, factory_date__lte=et) \
                .values('equip_no', 'down_reason').annotate(total_times=Sum('times')).values('equip_no', 'down_reason', 'total_times') \
                .order_by('equip_no', '-total_times')
            if not equips_data:
                raise ValidationError('无数据可以导出')
            res, temp_data = {}, {}
            for i in equips_data:
                equip_no, down_reason, total_times = i['equip_no'], i['down_reason'], i['total_times']
                equip_info = res.get(equip_no)
                if not equip_info:
                    res[equip_no] = [{'down_reason': down_reason, 'times': total_times, 'equip_no': equip_no}]
                    temp_data[f"{equip_no}_times"] = total_times
                    temp_data[f"{equip_no}_cnt"] = 1
                else:
                    if temp_data[f"{equip_no}_cnt"] >= 10:
                        continue
                    res[equip_no] = equip_info + [{'down_reason': down_reason, 'times': total_times, 'equip_no': equip_no}]
                    temp_data[f"{equip_no}_times"] += total_times
                    temp_data[f"{equip_no}_cnt"] += 1
            data = []
            for k, v in res.items():  # 比例
                s_times = temp_data.get(f"{equip_no}_times")
                for j in v:
                    j['ratio'] = round(j['times'] / s_times * 100, 2)
                    data.append(j)
            file_name = '各机台图表(TOP10)'
            export_fields_dict = {"机台": "equip_no", "异常原因": "down_reason", "分钟数": "times", "百分比": "ratio"}
            return gen_template_response(export_fields_dict, data, file_name, handle_str=True)
        if not equip_no:  # 所有密炼机TOP10停机原因汇总(总min)
            titles, details, ratios = [], [], []
            equips_data = EquipDownDetails.objects.filter(delete_flag=False, factory_date__gte=st, factory_date__lte=et).values('down_reason').annotate(total_times=Sum('times')).values('down_reason', 'total_times').order_by('-total_times')
            all_times = sum([i['total_times'] for i in equips_data][:10])
            for i in equips_data[:10]:
                titles.append(i['down_reason'])
                details.append(i['total_times'])
                ratios.append(round(i['total_times'] / all_times * 100, 2))
            results.update(**{'titles': titles, 'details': details, 'ratios': ratios, 'times': all_times})
        else:  # 单机台密炼机TOP10停机原因汇总(总min)
            filter_kwargs = {}
            if not table_flag:
                filter_kwargs['equip_no'] = equip_no
            equips_data = EquipDownDetails.objects.filter(delete_flag=False, factory_date__gte=st, factory_date__lte=et, **filter_kwargs)\
                .values('equip_no', 'down_reason').annotate(total_times=Sum('times')).values('equip_no', 'down_reason', 'total_times')\
                .order_by('equip_no', '-total_times')
            for i in equips_data:
                equip_no, down_reason, total_times = i['equip_no'], i['down_reason'], i['total_times']
                equip_info = results.get(equip_no)
                if not equip_info:
                    results[equip_no] = {'times': total_times, 'cnt': 1, 'titles': [down_reason], 'details': [total_times], 'equip_no': equip_no}
                else:
                    if equip_info['cnt'] >= 10:
                        continue
                    equip_info['times'] = equip_info['times'] + total_times
                    equip_info['cnt'] = equip_info['cnt'] + 1
                    equip_info['titles'].append(down_reason)
                    equip_info['details'].append(total_times)
            for k, v in results.items():  # 比例
                v['ratios'] = list(map(lambda x: round(x / v['times'] * 100, 2), v['details']))
            if table_flag:
                results = results.values()
            else:
                if results:
                    results = results.get(equip_no, {})
        return Response({'results': results})


@method_decorator([api_recorder], name='dispatch')
class GroupProductionSummary(APIView):

    def get(self, request):
        target_month = self.request.query_params.get('target_month')
        if not target_month:
            raise ValidationError('请选择月份！')
        td_flag = self.request.query_params.get('td_flag')  # 是否包含当天
        month_split = target_month.split('-')
        year = int(month_split[0])
        month = int(month_split[1])
        production_data = TrainsFeedbacks.objects.filter(
            factory_date__year=year,
            factory_date__month=month
        ).values('equip_no', 'factory_date', 'classes').annotate(total_trains=Count('id'))
        now_date = get_current_factory_date()['factory_date']
        if month == datetime.datetime.now().month and year == datetime.datetime.now().year:
            filter_kwargs = {'plan_schedule__day_time__lte': now_date} if td_flag else {'plan_schedule__day_time__lt': now_date}
            schedule_queryset = WorkSchedulePlan.objects.filter(
                plan_schedule__work_schedule__work_procedure__global_name='密炼',
                plan_schedule__day_time__year=year,
                plan_schedule__day_time__month=month,
                **filter_kwargs
            )
        else:
            schedule_queryset = WorkSchedulePlan.objects.filter(
                plan_schedule__work_schedule__work_procedure__global_name='密炼',
                plan_schedule__day_time__year=year,
                plan_schedule__day_time__month=month,
            )
        group_schedule_data = schedule_queryset.values('group__global_name').annotate(cnt=Count('id'))
        date_classes_dict = {'{}-{}'.format(i.plan_schedule.day_time.strftime("%m-%d"), i.classes.global_name): i.group.global_name for i in schedule_queryset}
        if td_flag:
            down_data = EquipDownDetails.objects.filter(
                factory_date__year=year,
                factory_date__month=month,
                down_type__in=['计划停机', '计划检修']
            ).values('group', 'equip_no').annotate(s=Sum('times'))
        else:
            down_data = EquipDownDetails.objects.filter(
                ~Q(factory_date=now_date),
                factory_date__year=year,
                factory_date__month=month,
                down_type__in=['计划停机', '计划检修']
            ).values('group', 'equip_no').annotate(s=Sum('times'))
        equip_target_data = MachineTargetYieldSettings.objects.filter(target_month=target_month).order_by('-id').values()
        target_data = {}
        if equip_target_data:
            target_data = equip_target_data[0]
        group_data_dict = {i: {'equip_no': i, 'target_trains': target_data.get(i, 0)} for i in
                           list(Equip.objects.filter(
                               category__equip_type__global_name="密炼设备"
                           ).order_by('equip_no').values_list("equip_no", flat=True))}

        for p in production_data:
            equip_no = p['equip_no']
            trains = p['total_trains'] if equip_no != 'Z04' else p['total_trains']//2
            gp = date_classes_dict.get('{}-{}'.format(p['factory_date'].strftime("%m-%d"), p['classes']))
            if not gp:
                continue
            gp_key = 'trains_{}'.format(gp)
            if gp_key not in group_data_dict[equip_no]:
                group_data_dict[equip_no][gp_key] = trains
            else:
                group_data_dict[equip_no][gp_key] += trains

        for s in group_schedule_data:
            s_key = 'days_{}'.format(s['group__global_name'])
            for equip_no in group_data_dict.keys():
                group_data_dict[equip_no][s_key] = s['cnt']

        for d in down_data:
            equip_no = d['equip_no']
            gp = d['group']
            times = d['s']
            d_key = 'down_{}'.format(gp)
            if d_key not in group_data_dict[equip_no]:
                group_data_dict[equip_no][d_key] = times
            else:
                group_data_dict[equip_no][d_key] += times
        return Response(group_data_dict.values())


@method_decorator([api_recorder], name='dispatch')
class TimeEnergyConsuming(APIView):

    def get(self, request):
        st = self.request.query_params.get('st')
        et = self.request.query_params.get('et')
        if not all([st, et]):
            raise ValidationError('请选择开始结束时间！')
        # 产量数据，按每个规格和车次数量排序
        production_data = TrainsFeedbacks.objects.filter(
            factory_date__gte=st,
            factory_date__lte=et
        ).values('product_no', 'equip_no').annotate(cnt=Count('id'),
                                                    actual_weight=Max('actual_weight')/100,
                                                    evacuation_energy=Avg('evacuation_energy'),
                                                    consum_time=OSum((F('end_time') - F('begin_time'))),
                                                    avg_interval_time=Avg('interval_time'),
                                                    ).order_by('product_no', 'cnt')
        item_dict = {}
        # 设备机台对应机型字典数据
        equip_dev_type_dict = dict(Equip.objects.filter(
            category__equip_type__global_name="密炼设备"
        ).values_list("equip_no", 'category_id'))
        pt_dict = {}
        pt = GlobalCode.objects.filter(global_type__type_name='配方类别').order_by('id').values('global_no', 'global_name')
        for item in pt:
            type_name = item['global_no']
            for i in item['global_name'].split(','):
                pt_dict[i] = type_name
        # 每个规格，找到生产车次最多的机台数据
        for item in production_data:
            try:
                pb_split = item['product_no'].split('-')
                stage, product_no, version = pb_split[1], pb_split[2], pb_split[3]
                recipe_no = '{}-{}'.format(product_no, version)
            except Exception:
                continue
            if stage not in ['CMB', 'HMB', '1MB', '2MB', '3MB', 'RMB', 'FM']:
                continue
            if item['equip_no'] == 'Z04':
                evacuation_energy = None if not item['evacuation_energy'] else item['evacuation_energy'] * 2
            else:
                evacuation_energy = item['evacuation_energy']
            actual_weight = item['actual_weight']
            try:
                consum_time = item['consum_time'].total_seconds()//item['cnt']
                if not consum_time or consum_time <= 50 or consum_time >= 500:
                    consum_time = 150
            except Exception:
                consum_time = 150
            avg_interval_time = item['avg_interval_time']
            if not avg_interval_time or avg_interval_time <= 5 or avg_interval_time >= 30:
                avg_interval_time = 15
            if recipe_no not in item_dict:
                item_dict[recipe_no] = {stage: {'devoted_weight': actual_weight,
                                                'actual_weight': actual_weight,
                                                'evacuation_energy': evacuation_energy,
                                                'consum_time': consum_time+avg_interval_time,
                                                'equip_no': item['equip_no'],
                                                }}
            else:
                item_dict[recipe_no][stage] = {'devoted_weight': actual_weight,
                                                'actual_weight': actual_weight,
                                                'evacuation_energy': evacuation_energy,
                                                'consum_time': consum_time+avg_interval_time,
                                                'equip_no': item['equip_no']}
        # 写入excel表格
        wb = load_workbook('xlsx_template/energy_consume.xlsx')
        sheet = wb.worksheets[0]
        sheet.title = '吨耗时(吨耗能){}-{}'.format('.'.join(st.split('-')[1:]), '.'.join(et.split('-')[1:]))
        data_row = 4
        stage_idx = {'CMB': {7: 'equip_no', 8: 'devoted_weight', 9: 'actual_weight', 10: 'evacuation_energy', 11: 'consum_time'},
                     'HMB': {14: 'equip_no', 15: 'devoted_weight', 16: 'actual_weight', 17: 'evacuation_energy', 18: 'consum_time'},
                     '1MB': {21: 'equip_no', 22: 'devoted_weight', 23: 'actual_weight', 24: 'evacuation_energy', 25: 'consum_time'},
                     '2MB': {28: 'equip_no', 29: 'devoted_weight', 30: 'actual_weight', 31: 'evacuation_energy', 32: 'consum_time'},
                     '3MB': {35: 'equip_no', 36: 'devoted_weight', 37: 'actual_weight', 38: 'evacuation_energy', 39: 'consum_time'},
                     'RMB': {42: 'equip_no', 43: 'devoted_weight', 44: 'actual_weight', 45: 'evacuation_energy', 46: 'consum_time'},
                     'FM': {49: 'equip_no', 50: 'devoted_weight', 51: 'actual_weight', 52: 'evacuation_energy', 53: 'consum_time'},
                     }
        for recipe_no, stage_data in item_dict.items():
            pn_split = recipe_no.split('-')
            try:
                re_result = re.match(r'[A-Z]+', pn_split[0])
                if not re_result:
                    recipe_type = '未知'
                else:
                    recipe_type = pt_dict.get(re_result.group(), '未知')
            except Exception:
                recipe_type = '未知'
            if recipe_type == '未知':
                continue
            sheet.cell(data_row, 1).value = recipe_type
            sheet.cell(data_row, 2).value = data_row - 3
            sheet.cell(data_row, 4).value = pn_split[0]
            sheet.cell(data_row, 5).value = pn_split[1]
            stage_length = 0
            for k, v in stage_data.items():
                equip_no = v['equip_no']
                if equip_no == 'Z04':
                    if k in ('CMB', 'HMB'):
                        stage_length += 0.5
                    else:
                        stage_length += 2
                else:
                    stage_length += 1
                #  查询配方数据，如果查到配方则补充收皮重量为配方重量，以及补充每个段次所投入上段次的重量
                product_batching = ProductBatching.objects.filter(
                    batching_type=2,
                    dev_type_id=equip_dev_type_dict.get(equip_no),
                    stage_product_batch_no__endswith='{}-{}'.format(k, recipe_no)
                ).order_by('id').last()
                if product_batching:
                    v['actual_weight'] = product_batching.batching_weight
                    v['devoted_weight'] = product_batching.batching_weight
                    devoted_material = product_batching.batching_details.filter(
                        type=1,
                        delete_flag=False,
                        material__material_no__endswith=recipe_no).first()
                    if devoted_material:
                        v['devoted_weight'] = devoted_material.actual_weight

                for idx, field_name in stage_idx.get(k, {}).items():
                    if field_name == 'evacuation_energy':  # 计算耗电量，每个机台都不一样
                        if not v['evacuation_energy']:
                            evacuation_energy = None
                        elif equip_no == 'Z01':
                            evacuation_energy = int(v['evacuation_energy'] / 10)
                        elif equip_no == 'Z02':
                            evacuation_energy = int(v['evacuation_energy'] / 0.6)
                        elif equip_no == 'Z04':
                            evacuation_energy = int(v['evacuation_energy'] * 0.28 * float(v['actual_weight']) / 1000)
                        elif equip_no == 'Z12':
                            evacuation_energy = int(v['evacuation_energy'] / 5.3)
                        elif equip_no == 'Z13':
                            evacuation_energy = int(v['evacuation_energy'] / 31.7)
                        else:
                            evacuation_energy = int(v['evacuation_energy'])
                        sheet.cell(data_row, idx).value = evacuation_energy
                    else:
                        sheet.cell(data_row, idx).value = v[field_name]
            sheet.cell(data_row, 6).value = stage_length
            data_row += 1
        output = BytesIO()
        wb.save(output)
        # 重新定位到开始
        output.seek(0)
        response = HttpResponse(content_type='application/vnd.ms-excel')
        filename = '吨耗时吨耗能'
        response['Content-Disposition'] = u'attachment;filename= ' + filename.encode('gbk').decode(
            'ISO-8859-1') + '.xls'
        response.write(output.getvalue())
        return response


@method_decorator([api_recorder], name='dispatch')
class RubberLogView(APIView):

    permission_classes = (IsAuthenticated,)

    @atomic
    def get(self, request):
        target_month = self.request.query_params.get('target_month')
        display_type = self.request.query_params.get('display_type')  # 存在表示汇总表
        title, _temp, max_times = self.display_content(target_month, display_type)
        flag = _temp.pop('add', None)
        if flag:
            results = _temp.values()
        else:
            if not display_type and max_times:  # 保存合计到汇总表
                # 记录当月汇总
                year, month = target_month.split('-')
                exist_record = RubberLogSummary.objects.filter(target_month=year, day=month, times=max_times)
                if not exist_record:
                    _add_info = []
                    for k, v in _temp.get('合计', {}).items():
                        if k == 'day':
                            continue
                        supplier_name, r_type, r_size = k.split('-')
                        if supplier_name in ['合计', '总计']:
                            continue
                        _s_data = {'target_month': year, 'day': month, 'times': max_times, 'supplier_name': supplier_name, 'r_type': r_type,
                                   'r_size': r_size, 'nums': v}
                        _add_info.append(RubberLogSummary(**_s_data))
                    RubberLogSummary.objects.bulk_create(_add_info)
            else:  # 补全数据(每月缺失的品牌)
                for i in _temp.values():
                    # 获取所有品牌
                    _u_data, day = {}, i.get('day')
                    names = set([k.split('-')[0] for k in i.keys() if k.split('-')[0] not in ['day', '合计', '总计']])
                    diff = set(title) - names
                    if not diff:
                        continue
                    for name in diff:
                        if isinstance(day, int) or day == '合计':
                            _u_data.update({f"{name}-进-大": 0, f"{name}-进-小": 0, f"{name}-出-大": 0, f"{name}-出-小": 0})
                        else:
                            _u_data.update({f"{name}": 0})
                    i.update(_u_data)
            results = sorted(_temp.values(), key=lambda x: x['day'])
        return Response({'results': results, 'title': title})

    @atomic
    def post(self, request):
        target_month = self.request.data.get('target_month')
        details = self.request.data.get('details')
        if not all([target_month, details]):
            raise ValidationError('参数异常')
        username = self.request.user.username
        now_time = datetime.datetime.now()
        max_times = RubberLog.objects.filter(target_month=target_month).aggregate(max_times=Max('times'))['max_times']
        times = 1 + (0 if not max_times else max_times)
        add_info = []
        for i in details:
            day = i.pop('day')
            if day in ['合计', '结余(大)', '结余(小)']:
                continue
            for k, v in i.items():
                supplier_name, r_type, r_size = k.split('-')
                if not supplier_name or supplier_name in ['合计', '总计']:
                    continue
                _s_data = {'target_month': target_month, 'day': day, 'created_datetime': now_time, 'created_username': username, 'times': times,
                           'supplier_name': supplier_name, 'r_type': r_type, 'r_size': r_size, 'nums': v if v else 0}
                add_info.append(RubberLog(**_s_data))
        RubberLog.objects.bulk_create(add_info)
        return Response('保存成功')

    def display_content(self, target_month, display_type):
        title, _temp, exist_day = [], {}, []
        days = [int(i.split('-')[-1]) for i in days_cur_month_dates(target_month)] + [32, 33, 34] if not display_type else []
        if not display_type:
            max_times = RubberLog.objects.filter(target_month=target_month).aggregate(max_times=Max('times'))['max_times']
            if not max_times:
                max_times = 0
            query_set = RubberLog.objects.filter(target_month=target_month, times=max_times).values('day', 'supplier_name', 'r_type', 'r_size', 'nums')
        else:
            max_times = RubberLogSummary.objects.filter(target_month=target_month).values('day').annotate(s_times=Max('times')).values_list('day', 's_times')
            if not max_times:
                max_times = []
            query_set = []
            for i in max_times:
                s_data = RubberLogSummary.objects.filter(target_month=target_month, day=i[0], times=i[1]).values('day', 'supplier_name', 'r_type', 'r_size', 'nums')
                query_set.extend(s_data)
        if query_set:
            # 32:合计 33:结余(大) 34:结余(小)
            for i in query_set:
                if i['supplier_name'] not in title:
                    title.append(i['supplier_name'])
                key = f"{i['supplier_name']}-{i['r_type']}-{i['r_size']}"
                total_key = f"总计-{i['r_type']}-{i['r_size']}"
                if _temp.get(i['day']):
                    _temp[i['day']][key] = i['nums']
                else:
                    _temp[i['day']] = {'day': i['day'], key: i['nums']}
                # 横向合计
                if total_key in _temp[i['day']]:
                    _temp[i['day']][total_key] = _temp[i['day']][total_key] + i['nums']
                else:
                    _temp[i['day']][total_key] = i['nums']
                # 纵向合计
                z_key1 = f"{i['supplier_name']}-{i['r_type']}-{i['r_size']}"
                if '合计' in _temp:
                    if z_key1 not in _temp['合计']:
                        _temp['合计'][z_key1] = i['nums']
                    else:
                        _temp['合计'][z_key1] += i['nums']
                    if f"合计-{i['r_type']}-{i['r_size']}" in _temp['合计']:
                        _temp['合计'][f"合计-{i['r_type']}-{i['r_size']}"] += i['nums']
                        _temp['合计'][f"总计-{i['r_type']}-{i['r_size']}"] += i['nums']
                    else:
                        _temp['合计'][f"合计-{i['r_type']}-{i['r_size']}"] = i['nums']
                        _temp['合计'][f"总计-{i['r_type']}-{i['r_size']}"] = i['nums']
                else:
                    _temp['合计'] = {'day': 32, z_key1: i['nums'], f"合计-{i['r_type']}-{i['r_size']}": i['nums'], f"总计-{i['r_type']}-{i['r_size']}": i['nums']}
                # 结余
                _num = -i['nums'] if i['r_type'] == '出' else i['nums']
                if f"结余({i['r_size']})" in _temp:
                    if f"{i['supplier_name']}" not in _temp[f"结余({i['r_size']})"]:
                        _temp[f"结余({i['r_size']})"][f"{i['supplier_name']}"] = _num
                    else:
                        _temp[f"结余({i['r_size']})"][f"{i['supplier_name']}"] += _num
                    _temp[f"结余({i['r_size']})"][f"总计"] += _num
                else:
                    _temp[f"结余({i['r_size']})"] = {'day': 33 if i['r_size'] == '大' else 34, f"总计": _num, f"{i['supplier_name']}": _num}
                # 补全月份
                if not display_type:
                    exist_day.append(i['day'])
            if exist_day:
                other_day = set(days[:-3]) - set(exist_day)
                for s_day in other_day:
                    _temp[s_day] = {'day': s_day}
        else:
            if not display_type:  # 月份没有数据新增行头(非汇总表)
                _temp = {day: {'day': day} for day in days}
                _temp['add'] = True
        return title, _temp, max_times


@method_decorator([api_recorder], name='dispatch')
class RecentRecipeName(APIView):

    def get(self, request):
        st = self.request.query_params.get('st')
        et = self.request.query_params.get('et')
        product_no = self.request.query_params.get('product_no')
        stage_product_batch_nos = ProductBatching.objects.filter(
            used_type=4,
            stage_product_batch_no__icontains='-FM-{}-'.format(product_no)
        ).order_by('used_time').values_list('stage_product_batch_no', flat=True)
        for stage_product_batch_no in stage_product_batch_nos:
            if PalletFeedbacks.objects.filter(
                    factory_date__lte=et,
                    factory_date__gte=st,
                    product_no=stage_product_batch_no).exists():
                return Response(stage_product_batch_no)
        raise ValidationError('该规格启用FM段次配方未找到!')